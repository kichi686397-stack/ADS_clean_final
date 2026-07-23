# -*- coding: utf-8 -*-
"""
同步 Meta Lead Ads 表单客户留言数据（增量版 / 每次新建 Excel 版）。

放置位置：ADS/lead_sync/sync_leads.py
输出文件：ADS/lead_sync/leads_YYYYMMDD_HHMMSS.xlsx
去重文件：ADS/lead_sync/leads_seen.json
状态文件：ADS/lead_sync/sync_state.json
日志文件：ADS/日志/sync_leads_YYYYMMDD_HHMMSS.log

默认逻辑：
- 第一次运行：只拉最近 30 天
- 后续运行：按每个 form 的上次同步时间增量拉取
- 使用 leads_seen.json 按 lead_id 去重，避免重复写入
- 每次运行都会新建一个 Excel，不再追加到旧 leads.xlsx
- Excel 中只保留一个“客户留言”数据表和一个“同步摘要”表，不再按产品分 sheet
- 表单自定义问题会展开为“字段_xxx”列，不再只塞进 raw_fields
- 支持按时间段导出：指定开始/结束日期时，不受增量状态和 leads_seen 去重影响
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import requests
from openpyxl import Workbook, load_workbook

API_VERSION = "v23.0"
GRAPH_BASE = f"https://graph.facebook.com/{API_VERSION}"

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
CACHE_DIR = BASE_DIR / "_cache"
LOG_DIR = BASE_DIR / "日志"
FORMS_LIBRARY = CACHE_DIR / "forms_library.json"
TOKEN_FILE = BASE_DIR / "token.txt"
OUTPUT_XLSX_PREFIX = "leads"
SEEN_FILE = SCRIPT_DIR / "leads_seen.json"
STATE_FILE = SCRIPT_DIR / "sync_state.json"

DEFAULT_INITIAL_DAYS = 30
SYNC_BUFFER_SECONDS = 300

ALL_SHEET_NAME = "客户留言"
SUMMARY_SHEET_NAME = "同步摘要"

DEFAULT_HEADERS = [
    "id",
    "created_time",
    "ad_id",
    "ad_name",
    "adset_id",
    "adset_name",
    "campaign_id",
    "campaign_name",
    "form_id",
    "form_name",
    "is_organic",
    "platform",
    "type_of_cooperation",
    "industry_/_application",
    "full_name",
    "country",
    "phone_number",
    "email",
    "lead_status",
]

# 输出字段固定成 Meta / Odoo 导入更容易识别的格式。
# 多语言表单里的自定义问题会归一到 type_of_cooperation 和 industry_/_application，
# 避免因为 EN/ES/PT/FR/KR/JP 等不同问题名被拆成很多列。
FIELD_ALIASES = {
    "full_name": [
        "full_name", "name", "your_name", "first_name", "last_name",
        "姓名", "姓名/氏名", "お名前", "氏名", "名前", "성함", "이름",
        "nombre", "nombre_completo", "nom", "nome", "nome_completo", "nombre_y_apellido",
    ],
    "phone_number": [
        "phone_number", "phone", "mobile_phone_number", "telephone", "tel",
        "电话号码", "电话", "手机号", "電話番号", "電話", "연락처", "전화번호",
        "teléfono", "telefono", "número_de_teléfono", "numéro_de_téléphone", "telefone",
    ],
    "email": ["email", "邮箱", "email_address", "work_email", "e-mail", "メール", "이메일", "correo", "correo_electrónico", "courriel"],
    "country": ["country", "country_id", "region", "国家", "国", "國家", "국가", "país", "pais", "pays", "paese"],
    "company": ["company", "company_name", "公司", "公司名称", "会社", "会社名", "회사", "회사명", "empresa", "société", "societe", "azienda"],
    "job_title": ["job_title", "title", "position", "职位", "職位", "役職", "직책", "cargo", "poste", "funzione"],
    "type_of_cooperation": [
        "type_of_cooperation", "cooperation_type", "cooperation", "collaboration_type",
        "合作类型", "合作方式", "合作意向", "協力タイプ", "협력 유형",
        "tipo_de_cooperación", "tipo_de_cooperacion", "tipo_de_colaboración", "tipo_de_colaboracion",
        "tipo_de_cooperação", "tipo_de_cooperacao", "type_de_coopération", "type_de_cooperation",
        "tipo_di_collaborazione", "art_der_zusammenarbeit", "тип_сотрудничества",
    ],
    "industry_/_application": [
        "industry_/_application", "industry/application", "industry_application", "industry", "application",
        "行业/应用", "行业应用", "应用场景", "業界/用途", "산업/응용",
        "industria/aplicación", "industria/aplicacion", "industria", "aplicación", "aplicacion",
        "setor/aplicação", "setor/aplicacao", "secteur/application", "settore/applicazione",
        "branche/anwendung", "отрасль/применение",
    ],
}

# 用问题标题关键词做二次判断。实际 Meta field_data 的 name 有时会是完整问题文案，
# 例如 “What type of cooperation are you interested in?”。
CANONICAL_FIELD_KEYWORDS = {
    "type_of_cooperation": [
        "cooperation", "collaboration", "partner", "distributor", "dealer",
        "合作", "協力", "협력", "cooperación", "cooperacion", "cooperação", "cooperacao",
        "coopération", "cooperation", "collaborazione", "zusammenarbeit", "сотруднич",
    ],
    "industry_/_application": [
        "industry", "application", "scenario", "use case", "segment",
        "行业", "應用", "应用", "用途", "業界", "산업", "응용",
        "industria", "aplicación", "aplicacion", "setor", "secteur", "settore", "branche", "примен", "отрасл",
    ],
}
INVALID_SHEET_CHARS = r"[\\/*?:\[\]]"


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def log_line(message: str, log_path: Path | None = None) -> None:
    text = f"[{now_text()}] {message}"
    print(text)
    if log_path:
        with log_path.open("a", encoding="utf-8") as f:
            f.write(text + "\n")


def load_token() -> str:
    token = os.environ.get("META_ACCESS_TOKEN") or os.environ.get("FB_ACCESS_TOKEN")
    if token:
        return token.strip()
    if TOKEN_FILE.exists():
        token = TOKEN_FILE.read_text(encoding="utf-8-sig").strip()
        if token:
            return token
    raise RuntimeError("未找到 token。请在 ADS/token.txt 填入 Meta Access Token，或设置 META_ACCESS_TOKEN 环境变量。")


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8-sig") as f:
        return json.load(f)


def save_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def normalize_key(value: str) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch not in [" ", "-", "_"])


def parse_products(value: str | None) -> set[str]:
    if not value:
        return set()
    parts = []
    for chunk in str(value).replace("；", ",").replace(";", ",").replace("/", ",").split(","):
        item = chunk.strip()
        if item:
            parts.append(normalize_key(item))
    return set(parts)


def iter_forms(forms_library: Dict[str, Any], selected_products: set[str] | None = None) -> Iterable[Tuple[str, str, str, Dict[str, Any]]]:
    selected_products = selected_products or set()
    for page_id, products in forms_library.items():
        if not isinstance(products, dict):
            continue
        for product, langs in products.items():
            product_text = str(product)
            if selected_products and normalize_key(product_text) not in selected_products:
                continue
            if not isinstance(langs, dict):
                continue
            for lang, info in langs.items():
                if isinstance(info, dict) and info.get("form_id"):
                    yield str(page_id), product_text, str(lang).upper(), info


def available_products(forms_library: Dict[str, Any]) -> List[str]:
    found: Dict[str, str] = {}
    for _page_id, products in forms_library.items():
        if not isinstance(products, dict):
            continue
        for product in products.keys():
            text = str(product)
            found.setdefault(normalize_key(text), text)
    return sorted(found.values(), key=lambda x: x.lower())


def parse_meta_time(value: str | None) -> datetime | None:
    if not value:
        return None
    text = str(value).strip()
    try:
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        if len(text) >= 5 and (text[-5] in ["+", "-"]) and text[-3] != ":":
            text = text[:-2] + ":" + text[-2:]
        dt = datetime.fromisoformat(text)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        pass
    for fmt in ["%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d %H:%M:%S"]:
        try:
            dt = datetime.strptime(str(value), fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except Exception:
            continue
    return None


def parse_date_start(value: str | None) -> datetime | None:
    """Parse YYYY-MM-DD as local start of day, then convert to UTC."""
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ["%Y-%m-%d", "%Y/%m/%d"]:
        try:
            local_tz = datetime.now().astimezone().tzinfo or timezone.utc
            dt = datetime.strptime(text, fmt).replace(tzinfo=local_tz)
            return dt.astimezone(timezone.utc)
        except Exception:
            continue
    raise ValueError(f"日期格式不正确：{value}。请使用 YYYY-MM-DD，例如 2026-07-13。")


def parse_date_end_exclusive(value: str | None) -> datetime | None:
    """Parse YYYY-MM-DD as inclusive local date and return next day 00:00 UTC."""
    start = parse_date_start(value)
    if start is None:
        return None
    return start + timedelta(days=1)


def iso_z(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def unix_seconds(dt: datetime) -> int:
    return int(dt.astimezone(timezone.utc).timestamp())


def redact_token(value: Any, token: str = "") -> str:
    """Remove access tokens from errors before they are written to logs."""
    text = str(value)
    if token:
        text = text.replace(token, "[REDACTED]")
    return re.sub(r"(?i)(access_token=)[^&\s]+", r"\1[REDACTED]", text)


def graph_get(url: str, params: Dict[str, Any], timeout: int = 60) -> Dict[str, Any]:
    token = str(params.get("access_token", ""))
    resp = None
    for attempt in range(4):
        try:
            resp = requests.get(url, params=params, timeout=timeout)
            break
        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError, requests.exceptions.Timeout) as exc:
            if attempt == 3:
                raise RuntimeError(
                    "连接 Meta 失败，已自动重试 4 次：" + redact_token(exc, token)
                ) from exc
            time.sleep(2 ** (attempt + 1))
    assert resp is not None
    try:
        data = resp.json()
    except Exception:
        resp.raise_for_status()
        raise
    if resp.status_code >= 400 or "error" in data:
        err = data.get("error", {})
        message = err.get("message") or resp.text
        code = err.get("code", resp.status_code)
        raise RuntimeError(f"Graph API error {code}: {message}")
    return data


def fetch_form_leads(form_id: str, token: str, *, since_dt: datetime | None = None, until_dt: datetime | None = None, limit: int = 100, max_pages: int = 100) -> Tuple[List[Dict[str, Any]], bool]:
    fields = ",".join([
        "id",
        "created_time",
        "ad_id",
        "ad_name",
        "adset_id",
        "adset_name",
        "campaign_id",
        "campaign_name",
        "is_organic",
        "platform",
        "field_data",
    ])
    url = f"{GRAPH_BASE}/{form_id}/leads"
    params: Dict[str, Any] = {
        "access_token": token,
        "fields": fields,
        "limit": limit,
    }
    filtering = []
    if since_dt:
        params["since"] = unix_seconds(since_dt)
        filtering.append({"field": "time_created", "operator": "GREATER_THAN", "value": unix_seconds(since_dt)})
    if until_dt:
        params["until"] = unix_seconds(until_dt)
        filtering.append({"field": "time_created", "operator": "LESS_THAN", "value": unix_seconds(until_dt)})
    # 有些 leadgen endpoint 对 since/until 不稳定，同时加 filtering 做兼容。
    if filtering:
        params["filtering"] = json.dumps(filtering)

    all_rows: List[Dict[str, Any]] = []
    page = 0
    stopped_by_time = False

    while url and page < max_pages:
        page += 1
        data = graph_get(url, params)
        rows = data.get("data", [])
        if not isinstance(rows, list):
            rows = []

        for row in rows:
            created_dt = parse_meta_time(row.get("created_time"))
            if since_dt and created_dt and created_dt < since_dt:
                stopped_by_time = True
                continue
            if until_dt and created_dt and created_dt >= until_dt:
                continue
            all_rows.append(row)

        if stopped_by_time:
            break

        next_url = data.get("paging", {}).get("next")
        url = next_url or ""
        params = {}
        if url:
            time.sleep(0.2)

    return all_rows, stopped_by_time


def flatten_field_data(field_data: Any) -> Dict[str, str]:
    flat: Dict[str, str] = {}
    if not isinstance(field_data, list):
        return flat
    for item in field_data:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        values = item.get("values")
        if isinstance(values, list):
            value = ", ".join(str(v) for v in values if v is not None)
        else:
            value = "" if values is None else str(values)
        if name:
            flat[name] = value
    return flat


def pick_field(flat: Dict[str, str], aliases: List[str]) -> str:
    lower_map = {str(k).lower(): v for k, v in flat.items()}
    normalized_map = {normalize_field_name(k): v for k, v in flat.items()}
    for alias in aliases:
        alias_l = str(alias).lower()
        if alias_l in lower_map:
            return lower_map[alias_l]
        alias_n = normalize_field_name(alias)
        if alias_n in normalized_map:
            return normalized_map[alias_n]
    for key, value in flat.items():
        key_l = str(key).lower()
        key_n = normalize_field_name(key)
        for alias in aliases:
            alias_l = str(alias).lower()
            alias_n = normalize_field_name(alias)
            if alias_l and alias_l in key_l:
                return value
            if alias_n and alias_n in key_n:
                return value
    return ""


def normalize_field_name(value: str) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("＆", "&").replace("／", "/")
    text = re.sub(r"[\s\-]+", "_", text)
    text = re.sub(r"[^0-9a-zA-Z_/&\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af\u0400-\u04ffÀ-ÿ]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def canonical_field_name(field_name: str) -> str | None:
    name = normalize_field_name(field_name)
    raw = str(field_name or "").strip().lower()
    for canonical, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            alias_n = normalize_field_name(alias)
            alias_l = str(alias).lower()
            if name == alias_n or alias_n in name or alias_l in raw:
                return canonical
    for canonical, keywords in CANONICAL_FIELD_KEYWORDS.items():
        for keyword in keywords:
            key_l = str(keyword).lower()
            key_n = normalize_field_name(keyword)
            if key_l and key_l in raw:
                return canonical
            if key_n and key_n in name:
                return canonical
    return None


def append_extra_field(row: Dict[str, Any], key: str, value: str) -> None:
    # 不能识别的自定义字段仍然保留在表格末尾，但用固定前缀，避免丢数据。
    col = "extra_" + normalize_field_name(key)
    if not col or col == "extra_":
        col = "extra_field"
    if col in DEFAULT_HEADERS:
        col = "extra_" + col
    if col not in row:
        row[col] = value
    elif value and str(value) not in str(row[col]):
        row[col] = str(row[col]) + "; " + str(value)


def safe_sheet_name(name: str, existing: set[str]) -> str:
    text = re.sub(INVALID_SHEET_CHARS, "_", str(name)).strip() or "Sheet"
    text = text[:31]
    base = text
    idx = 2
    while text in existing:
        suffix = f"_{idx}"
        text = (base[: 31 - len(suffix)] + suffix)[:31]
        idx += 1
    existing.add(text)
    return text


def product_sheet_name(product: str, wb) -> str:
    # 对已存在的 sheet，直接复用同名产品 sheet。
    raw = re.sub(INVALID_SHEET_CHARS, "_", str(product)).strip() or "Unknown"
    raw = raw[:31]
    if raw in wb.sheetnames:
        return raw
    existing = set(wb.sheetnames)
    return safe_sheet_name(raw, existing)


def ensure_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)


def ensure_headers(ws, required_headers: List[str]) -> List[str]:
    headers = [cell.value for cell in ws[1] if cell.value] if ws.max_row >= 1 else []
    if not headers:
        headers = []
    for h in required_headers:
        if h not in headers:
            headers.append(h)
            ws.cell(row=1, column=len(headers), value=h)
    return headers


def auto_width(ws) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 10
        for cell in col[:80]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = min(len(value), 45)
        ws.column_dimensions[letter].width = max_len + 2


def new_output_path() -> Path:
    return SCRIPT_DIR / f"{OUTPUT_XLSX_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def create_workbook():
    wb = Workbook()
    default = wb.active
    default.title = ALL_SHEET_NAME
    return wb


def append_rows(path: Path, rows: List[Dict[str, Any]], summary_rows: List[Dict[str, Any]] | None = None) -> None:
    # 每次运行新建一个 Excel 文件，不再读旧 leads.xlsx，也不按产品拆 sheet。
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = create_workbook()

    dynamic_headers: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in DEFAULT_HEADERS and key not in dynamic_headers:
                dynamic_headers.append(key)
    headers = DEFAULT_HEADERS + dynamic_headers

    all_ws = ensure_sheet(wb, ALL_SHEET_NAME)
    all_headers = ensure_headers(all_ws, headers)
    for row in rows:
        all_ws.append([row.get(h, "") for h in all_headers])

    if summary_rows is not None:
        summary_ws = ensure_sheet(wb, SUMMARY_SHEET_NAME)
        summary_headers = [
            "run_time",
            "mode",
            "product_filter",
            "total_forms",
            "candidate_leads",
            "skipped_seen",
            "new_rows",
            "failed_forms",
            "date_range",
            "output_file",
        ]
        s_headers = ensure_headers(summary_ws, summary_headers)
        for row in summary_rows:
            summary_ws.append([row.get(h, "") for h in s_headers])

    if ALL_SHEET_NAME in wb.sheetnames:
        wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index(ALL_SHEET_NAME)))

    for ws in wb.worksheets:
        if ws.max_row > 0:
            ws.freeze_panes = "A2"
            auto_width(ws)

    wb.save(path)

def build_output_row(*, page_id: str, product: str, language: str, form_info: Dict[str, Any], lead: Dict[str, Any]) -> Dict[str, Any]:
    flat = flatten_field_data(lead.get("field_data"))

    row: Dict[str, Any] = {
        "id": lead.get("id", ""),
        "created_time": lead.get("created_time", ""),
        "ad_id": lead.get("ad_id", ""),
        "ad_name": lead.get("ad_name", ""),
        "adset_id": lead.get("adset_id", ""),
        "adset_name": lead.get("adset_name", ""),
        "campaign_id": lead.get("campaign_id", ""),
        "campaign_name": lead.get("campaign_name", ""),
        "form_id": form_info.get("form_id", ""),
        "form_name": form_info.get("name", ""),
        "is_organic": str(lead.get("is_organic", "false")).lower(),
        "platform": lead.get("platform", ""),
        "type_of_cooperation": "",
        "industry_/_application": "",
        "full_name": pick_field(flat, FIELD_ALIASES["full_name"]),
        "country": pick_field(flat, FIELD_ALIASES["country"]),
        "phone_number": pick_field(flat, FIELD_ALIASES["phone_number"]),
        "email": pick_field(flat, FIELD_ALIASES["email"]),
        "lead_status": "complete",
    }

    # 多语言自定义问题统一归到固定列，避免出现“字段_xxx/字段_西语问题/字段_葡语问题”一堆列。
    for key, value in flat.items():
        canonical = canonical_field_name(key)
        if canonical in row:
            if value and not row.get(canonical):
                row[canonical] = value
            elif value and str(value) not in str(row.get(canonical, "")):
                row[canonical] = str(row.get(canonical, "")) + "; " + str(value)
        elif canonical in ["company", "job_title"]:
            # 这两个不是当前目标格式里的核心列，先不展开，保持 Meta/Odoo 格式干净。
            continue
        else:
            append_extra_field(row, key, value)

    # 兜底：如果问题标题识别不到，但某些表单使用统一字段 key，也再按 alias 取一次。
    if not row["type_of_cooperation"]:
        row["type_of_cooperation"] = pick_field(flat, FIELD_ALIASES["type_of_cooperation"])
    if not row["industry_/_application"]:
        row["industry_/_application"] = pick_field(flat, FIELD_ALIASES["industry_/_application"])

    return row


def form_state_key(page_id: str, product: str, language: str, form_id: str) -> str:
    return f"{page_id}|{product}|{language}|{form_id}"


def get_since_dt(*, state: Dict[str, Any], state_key: str, initial_days: int, full: bool) -> datetime | None:
    if full:
        return None
    item = state.get(state_key, {}) if isinstance(state, dict) else {}
    last_sync = ""
    if isinstance(item, dict):
        last_sync = str(item.get("last_sync_time") or item.get("last_seen_created_time") or "")
    dt = parse_meta_time(last_sync)
    if dt:
        return dt - timedelta(seconds=SYNC_BUFFER_SECONDS)
    return utc_now() - timedelta(days=initial_days)


def update_form_state(*, state: Dict[str, Any], state_key: str, page_id: str, product: str, language: str, form_id: str, form_name: str, max_created_dt: datetime | None, returned_count: int, new_count: int) -> None:
    old = state.get(state_key, {}) if isinstance(state.get(state_key, {}), dict) else {}
    old_dt = parse_meta_time(str(old.get("last_sync_time") or old.get("last_seen_created_time") or ""))
    final_dt = max_created_dt or old_dt or utc_now()
    if old_dt and final_dt < old_dt:
        final_dt = old_dt
    state[state_key] = {
        "page_id": page_id,
        "product": product,
        "language": language,
        "form_id": form_id,
        "form_name": form_name,
        "last_sync_time": iso_z(final_dt),
        "last_run_time": iso_z(utc_now()),
        "last_returned_count": returned_count,
        "last_new_count": new_count,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="同步 Meta Lead Ads 表单线索到本次新建的 Excel")
    parser.add_argument("--full", action="store_true", help="全量拉取。慎用：会增加 API 请求；仍会按 lead_id 去重。")
    parser.add_argument("--products", default="", help="只同步指定产品，多个用英文逗号分隔，例如：EG10,P2,V4e。留空表示同步全部产品。")
    parser.add_argument("--all", action="store_true", help="兼容旧参数，等同于 --full。")
    parser.add_argument("--reset-seen", action="store_true", help="配合 --full 使用时清空去重记录，可能导致 leads.xlsx 追加重复线索。")
    parser.add_argument("--initial-days", type=int, default=DEFAULT_INITIAL_DAYS, help="首次增量同步回看天数，默认 30。")
    parser.add_argument("--start-date", default="", help="按时间段导出的开始日期，格式 YYYY-MM-DD。")
    parser.add_argument("--end-date", default="", help="按时间段导出的结束日期，格式 YYYY-MM-DD，包含当天。")
    parser.add_argument("--limit", type=int, default=100, help="每页拉取数量，默认 100。")
    parser.add_argument("--max-pages", type=int, default=100, help="每个表单最多翻页数量，默认 100。")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    full = bool(args.full or args.all)
    range_start_dt = parse_date_start(args.start_date)
    range_end_dt = parse_date_end_exclusive(args.end_date)
    date_range_mode = bool(range_start_dt or range_end_dt)
    if date_range_mode and range_start_dt and range_end_dt and range_start_dt >= range_end_dt:
        print("时间段不正确：开始日期必须早于或等于结束日期。")
        return 1

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_path = LOG_DIR / f"sync_leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    try:
        if date_range_mode:
            start_label = args.start_date.strip() or "不限开始"
            end_label = args.end_date.strip() or "不限结束"
            mode_text = f"时间段导出 {start_label} 至 {end_label}"
        else:
            mode_text = "全量同步" if full else "增量同步"
        log_line(f"开始同步 Meta Lead Ads 表单线索：{mode_text}...", log_path)
        token = load_token()
        forms_library = load_json(FORMS_LIBRARY, {})
        if not forms_library:
            raise RuntimeError(f"未找到表单库：{FORMS_LIBRARY}")

        products_filter = parse_products(args.products)
        all_products = available_products(forms_library)
        if products_filter:
            available_keys = {normalize_key(x) for x in all_products}
            missing_products = sorted(products_filter - available_keys)
            if missing_products:
                raise RuntimeError("产品未在表单库中找到：" + ", ".join(missing_products) + "。可用产品：" + ", ".join(all_products))
            log_line("本次只同步产品：" + str(args.products), log_path)
        else:
            log_line("本次同步全部产品。可用产品：" + ", ".join(all_products), log_path)

        if args.reset_seen and not full:
            raise RuntimeError("--reset-seen 只能和 --full 一起使用，避免误删去重记录。")

        # 时间段导出是临时导出报表：不受 leads_seen 去重影响，也不推进 sync_state。
        seen = set()
        if not args.reset_seen and not date_range_mode:
            seen_data = load_json(SEEN_FILE, [])
            if isinstance(seen_data, list):
                seen = {str(x) for x in seen_data}
            elif isinstance(seen_data, dict):
                seen = {str(x) for x in seen_data.keys()}

        state = load_json(STATE_FILE, {})
        if not isinstance(state, dict):
            state = {}

        new_rows: List[Dict[str, Any]] = []
        total_forms = 0
        total_leads = 0
        skipped_seen = 0
        failed_forms: List[str] = []

        for page_id, product, language, form_info in iter_forms(forms_library, products_filter):
            total_forms += 1
            form_id = str(form_info.get("form_id"))
            form_name = str(form_info.get("name", ""))
            state_key = form_state_key(page_id, product, language, form_id)
            if date_range_mode:
                since_dt = range_start_dt
                until_dt = range_end_dt
            else:
                since_dt = get_since_dt(state=state, state_key=state_key, initial_days=max(1, int(args.initial_days)), full=full)
                until_dt = None

            label = f"{product}/{language}/{form_id}"
            since_text = "ALL" if since_dt is None else iso_z(since_dt)
            until_text = "ALL" if until_dt is None else iso_z(until_dt)
            log_line(f"读取表单：{label} {form_name} | since={since_text} | until={until_text}", log_path)

            try:
                leads, stopped_by_time = fetch_form_leads(form_id, token, since_dt=since_dt, until_dt=until_dt, limit=args.limit, max_pages=args.max_pages)
            except Exception as exc:
                failed_forms.append(f"{label}: {exc}")
                log_line(f"失败：{label} -> {exc}", log_path)
                continue

            total_leads += len(leads)
            form_new_count = 0
            max_created_dt: datetime | None = None

            for lead in leads:
                lead_id = str(lead.get("id") or "")
                if not lead_id:
                    continue
                created_dt = parse_meta_time(lead.get("created_time"))
                if created_dt and (max_created_dt is None or created_dt > max_created_dt):
                    max_created_dt = created_dt
                if not date_range_mode and lead_id in seen:
                    skipped_seen += 1
                    continue
                new_rows.append(build_output_row(page_id=page_id, product=product, language=language, form_info=form_info, lead=lead))
                form_new_count += 1
                if not date_range_mode:
                    seen.add(lead_id)

            if not date_range_mode:
                update_form_state(state=state, state_key=state_key, page_id=page_id, product=product, language=language, form_id=form_id, form_name=form_name, max_created_dt=max_created_dt, returned_count=len(leads), new_count=form_new_count)

            if stopped_by_time:
                log_line(f"已遇到旧数据，提前停止翻页：{label}", log_path)
            log_line(f"表单完成：{label} 返回 {len(leads)} 条，新增 {form_new_count} 条", log_path)

        if total_forms == 0:
            raise RuntimeError("本次没有匹配到任何表单，请检查 --products 参数。")

        output_xlsx = new_output_path()
        summary_rows = [{
            "run_time": now_text(),
            "mode": mode_text,
            "product_filter": args.products or "ALL",
            "date_range": (f"{args.start_date or '不限开始'} 至 {args.end_date or '不限结束'}" if date_range_mode else ""),
            "total_forms": total_forms,
            "candidate_leads": total_leads,
            "skipped_seen": skipped_seen,
            "new_rows": len(new_rows),
            "failed_forms": len(failed_forms),
            "output_file": str(output_xlsx),
        }]

        append_rows(output_xlsx, new_rows, summary_rows)
        if not date_range_mode:
            save_json(SEEN_FILE, sorted(seen))
            save_json(STATE_FILE, state)

        log_line("同步完成。", log_path)
        log_line(f"模式：{mode_text}", log_path)
        log_line(f"表单数：{total_forms}", log_path)
        log_line(f"接口返回候选线索数：{total_leads}", log_path)
        log_line(f"已存在跳过：{skipped_seen}", log_path)
        log_line(f"本次写入：{len(new_rows)}", log_path)
        log_line(f"输出文件：{output_xlsx}", log_path)
        log_line("Excel 已新建为本次同步文件；客户留言在一个 sheet 中，并展开表单字段。", log_path)
        if failed_forms:
            log_line(f"失败表单数：{len(failed_forms)}", log_path)
            for item in failed_forms:
                log_line(f"  - {item}", log_path)
            return 1
        return 0
    except Exception as exc:
        log_line(f"同步失败：{exc}", log_path)
        return 1


if __name__ == "__main__":
    sys.exit(main())
