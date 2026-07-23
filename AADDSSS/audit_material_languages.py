# -*- coding: utf-8 -*-
"""Audit active Meta ads against quarterly-budget material languages.

Default is read-only. Pass --apply to pause ad sets with explicit hash/path mismatch.
"""
import argparse
import hashlib
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import app


BASE = Path(__file__).resolve().parent
OUT_DIR = BASE / "导出"
LANG_CODES = {"AR", "CHT", "CN", "DE", "EN", "ES", "FR", "ID", "IT", "JP", "KR", "MS", "MY", "NL", "PL", "PT", "TH", "TR", "TW", "VN", "ZH"}
# Meta 返回这批图片名时中文前缀会乱码；已于 2026-07-17 直接查看画面，五张均为西班牙语。
VISUALLY_VERIFIED_LANGUAGES = {
    "086ca8defe328130a35027d23ca579d8": "ES",
    "59e91c10f76f6bf3940e491f86e1014e": "ES",
    "b767b3bd10284bd2d4656ce5f8f5e006": "ES",
    "d066925d2b0a4e1dc7d6e9442e3b2a03": "ES",
    "e079563aa52a521cf493eb257fff46b7": "ES",
}


def file_md5(path):
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def compact(value):
    return re.sub(r"[^A-Za-z0-9]+", "", str(value or "")).upper()


def graph_all(url, params):
    rows = []
    while url:
        response = requests.get(url, params=params, timeout=120)
        data = response.json()
        if data.get("error"):
            raise RuntimeError(data["error"].get("message", "Meta API error"))
        rows.extend(data.get("data") or [])
        url = (data.get("paging") or {}).get("next")
        params = None
    return rows


def load_budget():
    data = json.loads((BASE / "config" / "quarterly_budget.json").read_text(encoding="utf-8"))
    rows = data.get("rows") or []
    products = sorted({str(row.get("product") or "").strip() for row in rows if row.get("product")}, key=lambda x: len(compact(x)), reverse=True)
    return rows, products, data.get("source_name", "")


def infer_product(products, *names):
    haystack = compact(" ".join(str(name or "") for name in names))
    for product in products:
        if compact(product) and compact(product) in haystack:
            return product
    return ""


def infer_selling_point(adset_name, product):
    parts = str(adset_name or "").split("_")
    if len(parts) >= 3 and parts[0].upper() == "FJD" and compact(parts[1]) == compact(product):
        return parts[2].strip()
    return ""


def expected_languages(budget_rows, product, countries):
    country_set = {str(x).upper() for x in countries}
    exact, partial = set(), set()
    for row in budget_rows:
        if compact(row.get("product")) != compact(product):
            continue
        row_countries = {str(x).upper() for x in (row.get("countries") or [])}
        lang = str(row.get("asset_language") or "").strip().upper()
        if not lang:
            continue
        if row_countries == country_set:
            exact.add(lang)
        elif row_countries & country_set:
            partial.add(lang)
    return exact or partial


def build_asset_hash_index():
    library_path = BASE / "_cache" / "image_library.json"
    library = json.loads(library_path.read_text(encoding="utf-8")) if library_path.exists() else {"by_md5": {}}
    by_md5 = library.get("by_md5") or {}
    index = defaultdict(list)
    root = BASE / "素材包"
    for product_dir in root.iterdir() if root.exists() else []:
        if not product_dir.is_dir():
            continue
        for point_dir in product_dir.iterdir():
            if not point_dir.is_dir():
                continue
            for lang_dir in point_dir.iterdir():
                if not lang_dir.is_dir() or lang_dir.name.upper() not in LANG_CODES:
                    continue
                for image in lang_dir.rglob("*"):
                    if not image.is_file() or image.suffix.lower() not in app.IMAGE_EXT:
                        continue
                    md5 = file_md5(image)
                    item = by_md5.get(md5) or by_md5.get(md5.upper()) or {}
                    image_hash = str(item.get("hash") or item.get("image_hash") or "").strip()
                    if image_hash:
                        index[image_hash].append({
                            "product": product_dir.name,
                            "selling_point": point_dir.name,
                            "language": lang_dir.name.upper(),
                            "file": image.name,
                            "path": str(image),
                        })
    return index


def is_active(item):
    return str(item.get("status") or "").upper() == "ACTIVE" and str(item.get("effective_status") or "").upper() == "ACTIVE"


def active_context(ad):
    adset = ad.get("adset") or {}
    campaign = adset.get("campaign") or {}
    return is_active(ad) and is_active(adset) and is_active(campaign)


def creative_hashes(ad):
    feed = (ad.get("creative") or {}).get("asset_feed_spec") or {}
    return [str(item.get("hash") or "").strip() for item in (feed.get("images") or []) if item.get("hash")]


def infer_meta_image_record(name):
    text = str(name or "")
    if "西班牙语" in text:
        language = "ES"
    else:
        match = re.search(r"(?:^|_)FJD_([^_]+)_([A-Z]{2,3})(?:_|$)", text, re.I)
        language = match.group(2).upper() if match else ""
    product_match = re.search(r"(?:^|_)FJD_([^_]+)_", text, re.I)
    product = product_match.group(1) if product_match else ""
    if not product or not language:
        return None
    return {"product": product, "selling_point": "", "language": language, "file": text, "path": "Meta 后台素材库", "source": "meta_name"}


def fetch_meta_image_records(account, token, hashes):
    if not hashes:
        return {}
    result = defaultdict(list)
    hashes = sorted(set(hashes))
    for start in range(0, len(hashes), 100):
        batch = hashes[start:start + 100]
        params = {"access_token": token, "fields": "hash,name,created_time", "hashes": json.dumps(batch), "limit": 200}
        items = graph_all(f"https://graph.facebook.com/{app.API_VERSION}/{account}/adimages", params)
        for item in items:
            record = infer_meta_image_record(item.get("name"))
            verified_language = VISUALLY_VERIFIED_LANGUAGES.get(str(item.get("hash") or ""))
            if verified_language:
                record = record or {"product": "EC100", "selling_point": "", "file": str(item.get("name") or ""), "path": "Meta 后台素材库", "source": "meta_name"}
                record["product"] = "EC100"
                record["language"] = verified_language
                record["source"] = "visual_verified"
            if record:
                result[str(item.get("hash") or "")].append(record)
    return result


def audit(apply_changes=False):
    budget_rows, products, budget_source = load_budget()
    hash_index = build_asset_hash_index()
    token = app.get_token()
    account = "act_" + app.clean_id(app.get_defaults().get("ad_account_id"))
    fields = "id,name,status,effective_status,adset_id,adset{id,name,status,effective_status,targeting,campaign_id,campaign{id,name,status,effective_status}},creative{id,name,asset_feed_spec,image_hash,object_story_spec}"
    ads = graph_all(f"https://graph.facebook.com/{app.API_VERSION}/{account}/ads", {"access_token": token, "fields": fields, "limit": 500})
    groups = defaultdict(list)
    for ad in ads:
        if active_context(ad):
            groups[str(ad.get("adset_id") or "")].append(ad)
    all_hashes = {h for group_ads in groups.values() for ad in group_ads for h in creative_hashes(ad)}
    meta_image_records = fetch_meta_image_records(account, token, [h for h in all_hashes if h not in hash_index])

    report_rows = []
    paused = set()
    for adset_id, group_ads in groups.items():
        first = group_ads[0]
        adset = first.get("adset") or {}
        campaign = adset.get("campaign") or {}
        countries = [str(x).upper() for x in (((adset.get("targeting") or {}).get("geo_locations") or {}).get("countries") or [])]
        product = infer_product(products, adset.get("name"), campaign.get("name"), *(ad.get("name") for ad in group_ads))
        selling_point = infer_selling_point(adset.get("name"), product)
        expected = expected_languages(budget_rows, product, countries) if product and countries else set()
        hashes = sorted({h for ad in group_ads for h in creative_hashes(ad)})
        actual_languages, asset_files, mismatched_hashes, unknown_hashes = set(), [], [], []

        for image_hash in hashes:
            records = hash_index.get(image_hash) or meta_image_records.get(image_hash) or []
            if not records:
                unknown_hashes.append(image_hash)
                continue
            actual_languages.update(record["language"] for record in records)
            asset_files.extend(f"{record['language']} | {record['product']} | {record['selling_point'] or '后台历史素材'} | {record['file']}" for record in records)
            matches = [record for record in records if compact(record["product"]) == compact(product) and record["language"] in expected and (record.get("source") in {"meta_name", "visual_verified"} or not selling_point or record["selling_point"] == selling_point)]
            if not matches:
                mismatched_hashes.append(image_hash)

        if not product:
            decision, reason = "待人工确认", "无法从广告组/广告系列名称识别产品"
        elif not countries:
            decision, reason = "待人工确认", "广告组没有可读取的国家定向"
        elif not expected:
            decision, reason = "待人工确认", "季度预算表未匹配到该产品和国家"
        elif not hashes:
            decision, reason = "待人工确认", "Creative 没有返回图片 Hash"
        elif mismatched_hashes:
            decision, reason = "应关闭", "图片 Hash 对应的产品/卖点/语种与预算表不一致"
        elif unknown_hashes:
            decision, reason = "待人工确认", "部分图片 Hash 无法映射到本地素材库"
        else:
            decision, reason = "匹配", "全部图片 Hash 均匹配产品、卖点和预算语种"

        action_result = "未操作"
        if decision == "应关闭" and apply_changes:
            response = requests.post(f"https://graph.facebook.com/{app.API_VERSION}/{adset_id}", data={"access_token": token, "status": "PAUSED"}, timeout=60)
            data = response.json()
            if data.get("error"):
                action_result = "关闭失败：" + str(data["error"].get("message") or data["error"])
            elif data.get("success", True):
                action_result = "已关闭"
                paused.add(adset_id)
            else:
                action_result = "关闭失败：Meta 未返回成功"

        report_rows.append({
            "判断": decision,
            "执行结果": action_result,
            "广告组ID": adset_id,
            "广告组名称": str(adset.get("name") or ""),
            "广告系列": str(campaign.get("name") or ""),
            "产品": product,
            "卖点": selling_point,
            "国家": ",".join(countries),
            "预算表语种": ",".join(sorted(expected)),
            "素材实际语种": ",".join(sorted(actual_languages)),
            "活动广告数": len(group_ads),
            "Creative ID": ",".join(sorted({str((ad.get("creative") or {}).get("id") or "") for ad in group_ads})),
            "广告名称": "\n".join(str(ad.get("name") or "") for ad in group_ads),
            "图片Hash": "\n".join(hashes),
            "不匹配Hash": "\n".join(mismatched_hashes),
            "未知Hash": "\n".join(unknown_hashes),
            "素材文件": "\n".join(sorted(set(asset_files))),
            "原因": reason,
            "检查时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    priority = {"应关闭": 0, "待人工确认": 1, "匹配": 2}
    report_rows.sort(key=lambda row: (priority.get(row["判断"], 9), row["产品"], row["国家"], row["广告组名称"]))
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    mode = "已执行关停" if apply_changes else "只读审计"
    output = OUT_DIR / f"广告素材语种核对_{stamp}.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    counts = defaultdict(int)
    for row in report_rows:
        counts[row["判断"]] += 1
    summary_rows = [
        ("检查模式", mode), ("预算表来源", budget_source), ("Meta 广告账户", account),
        ("ACTIVE 广告组", len(report_rows)), ("匹配", counts["匹配"]),
        ("明确不匹配", counts["应关闭"]), ("待人工确认", counts["待人工确认"]),
        ("本次已关闭", len(paused)), ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for key, value in summary_rows:
        summary.append([key, value])
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 70
    summary["A1"].font = Font(bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="1E3A5F")

    detail = workbook.create_sheet("广告组明细")
    headers = list(report_rows[0].keys()) if report_rows else ["判断", "执行结果", "广告组ID", "广告组名称", "原因"]
    detail.append(headers)
    for row in report_rows:
        detail.append([row.get(header, "") for header in headers])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")
    widths = {"A": 14, "B": 18, "C": 20, "D": 48, "E": 45, "F": 14, "G": 20, "H": 16, "I": 16, "J": 18, "K": 12, "L": 25, "M": 55, "N": 42, "O": 42, "P": 42, "Q": 75, "R": 48, "S": 20}
    for column, width in widths.items():
        detail.column_dimensions[column].width = width
    for row in detail.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].value == "应关闭":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FEE2E2")
        elif row[0].value == "待人工确认":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FEF3C7")

    workbook.save(output)
    json_output = output.with_suffix(".json")
    json_output.write_text(json.dumps({"mode": mode, "summary": dict(counts), "paused": sorted(paused), "rows": report_rows}, ensure_ascii=False, indent=2), encoding="utf-8")
    return output, report_rows, paused


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Pause explicitly mismatched ad sets")
    args = parser.parse_args()
    path, rows, paused = audit(args.apply)
    counts = defaultdict(int)
    for row in rows:
        counts[row["判断"]] += 1
    print(json.dumps({"file": str(path), "active_adsets": len(rows), "matched": counts["匹配"], "mismatched": counts["应关闭"], "manual_review": counts["待人工确认"], "paused": len(paused)}, ensure_ascii=False))
