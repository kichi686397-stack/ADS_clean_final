# -*- coding: utf-8 -*-
"""
ADS 本地前端 v4
功能：素材包扫描、素材预览、Lead Forms 同步、投放前检查、生成 plan.xlsx、运行 run.py。
运行：py app.py，然后打开 http://127.0.0.1:8765
"""
import base64
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import webbrowser
import traceback
from copy import copy
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO, StringIO
from pathlib import Path
from urllib.parse import urlparse, parse_qs, unquote

import requests
from openpyxl import load_workbook, Workbook

BASE_DIR = Path(__file__).resolve().parent
CONFIG_DIR = BASE_DIR / "config"
PLAN_FILE = BASE_DIR / "plan.xlsx"
ASSET_ROOT = BASE_DIR / "素材包"
CACHE_DIR = BASE_DIR / "_cache"
FORMS_LIBRARY_FILE = CACHE_DIR / "forms_library.json"
PRODUCT_URLS_FILE = CONFIG_DIR / "product_urls.json"
PRODUCT_SETTINGS_FILE = CONFIG_DIR / "product_settings.json"
AUTOMATION_RULES_FILE = CONFIG_DIR / "automation_rules.json"
AUTOMATION_PREVIEW_FILE = CACHE_DIR / "automation_last_preview.json"
IMAGE_LIBRARY_FILE = CACHE_DIR / "image_library.json"
INSIGHTS_CACHE_FILE = CACHE_DIR / "insights_data.json"
QUARTER_BUDGET_FILE = CONFIG_DIR / "quarterly_budget.json"
QUARTER_BUDGET_SOURCE_FILE = CACHE_DIR / "quarterly_budget_source.xlsx"
QUARTER_SPEND_CACHE_FILE = CACHE_DIR / "quarter_spend.json"
ACTIVE_COVERAGE_CACHE_FILE = CACHE_DIR / "active_ad_coverage.json"
RUN_SCRIPT = BASE_DIR / "run.py"
TOKEN_FILE = BASE_DIR / "token.txt"
LOG_DIR = BASE_DIR / "日志"
AUTOMATION_LOG_FILE = LOG_DIR / "automation_actions.json"
AUTOMATION_ACTION_STATE_FILE = CONFIG_DIR / "automation_action_state.json"
AUTOMATION_ADSET_STATUS_CACHE_FILE = CACHE_DIR / "automation_adset_status.json"
LEAD_SYNC_DIR = BASE_DIR / "lead_sync"
LEAD_SYNC_SCRIPT = LEAD_SYNC_DIR / "sync_leads.py"
API_VERSION = "v23.0"
HOST = "127.0.0.1"
PORT = 8765
INSIGHTS_CACHE_TTL_SECONDS = 60 * 60
INSIGHTS_CACHE_SCHEMA_VERSION = 2
QUARTER_SPEND_CACHE_TTL_SECONDS = 60 * 60
AUTOMATION_STATUS_CACHE_TTL_SECONDS = 15 * 60
META_MIN_DAILY_BUDGET = 1.01
BLOCKED_COUNTRIES = {"TW", "SG"}

PLAN_HEADERS = [
    "enabled", "product", "selling_point", "asset_language", "countries", "daily_budget",
    "page_id", "lead_form_id", "website_url", "cta_type", "language_mode", "custom_language",
    "dry_run", "delay_max_seconds", "ad_account_id", "campaign_id", "url_tags"
]

COUNTRY_OPTIONS = [
    ("US","美国"),("CA","加拿大"),("MX","墨西哥"),("BR","巴西"),("CL","智利"),("CO","哥伦比亚"),("PE","秘鲁"),("AR","阿根廷"),
    ("GB","英国"),("IE","爱尔兰"),("FR","法国"),("DE","德国"),("IT","意大利"),("ES","西班牙"),("PT","葡萄牙"),("NL","荷兰"),("BE","比利时"),("CH","瑞士"),("AT","奥地利"),
    ("PL","波兰"),("CZ","捷克"),("SK","斯洛伐克"),("HU","匈牙利"),("RO","罗马尼亚"),("BG","保加利亚"),("HR","克罗地亚"),("SI","斯洛文尼亚"),("RS","塞尔维亚"),
    ("EE","爱沙尼亚"),("LV","拉脱维亚"),("LT","立陶宛"),("SE","瑞典"),("NO","挪威"),("FI","芬兰"),("DK","丹麦"),("GR","希腊"),("TR","土耳其"),
    ("SA","沙特"),("AE","阿联酋"),("QA","卡塔尔"),("KW","科威特"),("OM","阿曼"),("BH","巴林"),("JO","约旦"),("IL","以色列"),
    ("EG","埃及"),("MA","摩洛哥"),("DZ","阿尔及利亚"),("TN","突尼斯"),("ZA","南非"),("NG","尼日利亚"),("KE","肯尼亚"),("TZ","坦桑尼亚"),("GH","加纳"),
    ("IN","印度"),("PK","巴基斯坦"),("BD","孟加拉"),("LK","斯里兰卡"),("JP","日本"),("KR","韩国"),("ID","印尼"),("MY","马来西亚"),("SG","新加坡"),("TH","泰国"),("VN","越南"),("PH","菲律宾"),
    ("AU","澳大利亚"),("NZ","新西兰")
]

COUNTRY_TO_FORM_LANG = {
    "KR":"KR","JP":"JP","ES":"ES","MX":"ES","CO":"ES","CL":"ES","PE":"ES","AR":"ES","BR":"PT","PT":"PT","FR":"FR","DE":"DE","IT":"IT","TR":"TR","PL":"PL","ID":"ID","TH":"TH","VN":"VN",
    "US":"EN","GB":"EN","IE":"EN","CA":"EN","AU":"EN","NZ":"EN","SG":"EN","MY":"EN","AE":"EN","SA":"EN","QA":"EN","KW":"EN","OM":"EN","BH":"EN","ZA":"EN","NG":"EN","KE":"EN","PH":"EN","IN":"EN"
}

IMAGE_EXT = {".jpg", ".jpeg", ".png", ".webp"}
TEXT_FILES = ["文案.txt", "copy.txt", "ad_copy.txt", "text.txt"]
IMAGE_DIRS = ["图片", "images", "image"]
TEXT_SUFFIXES = {".txt"}
last_run = {"running": False, "started_at": "", "finished_at": "", "returncode": None, "output": "", "run_id": "", "paused": False, "run_error": None, "summary": {"planned": 0, "success": 0, "failed": 0, "skipped": 0, "dry_run": 0}}
last_lead_sync = {"running": False, "started_at": "", "finished_at": "", "returncode": None, "output": "", "output_file": ""}
insights_cache = {"key": "", "loaded_at": None, "data": None}
insights_cache_lock = threading.Lock()
quarter_spend_lock = threading.Lock()
active_coverage_lock = threading.Lock()
automation_state = {"running": False, "last_started": "", "last_finished": "", "last_error": "", "last_result": None}


def norm(v): return str(v or "").strip()

def load_json(path, default):
    try:
        p = Path(path)
        return json.loads(p.read_text(encoding="utf-8")) if p.exists() else default
    except Exception:
        return default

def save_json(path, data):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def get_token():
    tok = os.getenv("META_ACCESS_TOKEN", "").strip()
    if tok: return tok
    if TOKEN_FILE.exists():
        tok = TOKEN_FILE.read_text(encoding="utf-8").strip()
        if tok and tok != "PASTE_TOKEN_HERE": return tok
    raise RuntimeError("没有找到 token：请把 Meta Access Token 放到 token.txt。")

def jresp(h, data, status=200):
    # 统一 JSON 返回。发生错误时尽量把请求方法、路径、提示都带回前端，
    # 避免浏览器只显示 Failed to fetch / Not found。
    if isinstance(data, dict):
        data.setdefault("ok", status < 400)
        if status >= 400:
            data.setdefault("method", getattr(h, "command", ""))
            data.setdefault("path", getattr(h, "path", ""))
    b = json.dumps(data, ensure_ascii=False).encode("utf-8")
    h.send_response(status); h.send_header("Content-Type", "application/json; charset=utf-8"); h.send_header("Content-Length", str(len(b))); h.end_headers(); h.wfile.write(b)


def not_found_response(h, path):
    return jresp(h, {
        "error": "接口不存在 / Not found",
        "detail": f"当前请求没有匹配到后端接口：{getattr(h, 'command', '')} {path}",
        "hint": "请确认：1）浏览器地址是 http://127.0.0.1:8765；2）frontend.bat 黑色窗口没有关闭；3）当前页面和 app.py 来自同一个最新版 ADS 文件夹；4）不要混用旧版页面缓存，按 Ctrl+F5 强制刷新。",
        "available_get": ["/", "/index.html", "/api/ping", "/api/options", "/api/asset_preview", "/api/status", "/api/check_packs", "/api/logs", "/api/image_hash_status", "/api/leads/options", "/api/leads/status", "/api/leads/files", "/api/leads/download", "/api/insights", "/api/insights/export", "/api/report/export", "/api/budget/status", "/api/budget/template", "/api/automation/settings", "/api/automation/status", "/api/automation/effectiveness", "/asset"],
        "available_post": ["/api/sync_forms", "/api/preflight", "/api/generate_plan", "/api/create_pack", "/api/sync_image_hashes", "/api/run", "/api/leads/start", "/api/budget/import", "/api/automation/settings", "/api/automation/preview", "/api/automation/run", "/api/automation/apply", "/api/automation/rollback"],
    }, 404)

def tresp(h, text, status=200, ctype="text/html; charset=utf-8"):
    b = text.encode("utf-8")
    h.send_response(status); h.send_header("Content-Type", ctype); h.send_header("Content-Length", str(len(b))); h.end_headers(); h.wfile.write(b)

def read_body(h):
    n = int(h.headers.get("Content-Length", "0") or 0)
    return json.loads(h.rfile.read(n).decode("utf-8") or "{}") if n else {}

def read_text_file(path):
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try: return Path(path).read_text(encoding=enc)
        except UnicodeDecodeError: pass
    return Path(path).read_text(encoding="utf-8", errors="ignore")

def _copy_key_from_filename(path):
    stem = Path(path).stem.lower().replace("-", "_").replace(" ", "_")
    if any(word in stem for word in ("headline", "title", "标题")):
        return "headline"
    if any(word in stem for word in ("description", "desc", "描述")):
        return "description"
    return "primary_text"


def _plain_text_variants(text):
    normalized = str(text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not normalized:
        return []
    paragraphs = [re.sub(r"\s*\n\s*", " ", block).strip() for block in re.split(r"\n\s*\n+", normalized) if block.strip()]
    if len(paragraphs) > 1:
        return paragraphs
    return [line.strip() for line in normalized.split("\n") if line.strip()]


def _merge_copy_value(target, key, value):
    existing = [line.strip() for line in str(target.get(key, "")).split("\n") if line.strip()]
    incoming = [line.strip() for line in str(value or "").split("\n") if line.strip()]
    target[key] = "\n".join(existing + [line for line in incoming if line not in existing])


def parse_copy_text(text, default_key="primary_text"):
    """解析素材包里的文案文件。

    支持：
    正文1/正文2/正文3/正文4/正文5
    标题1/标题2/标题3/标题4/标题5
    描述1/描述2/描述3/描述4/描述5
    也支持 primary_text/headline/description。多条会用换行合并，run.py 会作为动态素材多条文案。
    """
    key_map = {
        "primary_text":"primary_text","body":"primary_text","text":"primary_text","正文":"primary_text","主文案":"primary_text",
        "headline":"headline","title":"headline","标题":"headline",
        "description":"description","desc":"description","描述":"description",
        "cta":"cta_type","cta_type":"cta_type","行动号召":"cta_type",
        "url_tags":"url_tags","utm":"url_tags","link":"website_url","链接":"website_url","website_url":"website_url"
    }
    res, cur = {}, None
    for raw in text.replace("\r\n","\n").replace("\r","\n").split("\n"):
        s = raw.strip()
        if not s:
            continue
        section = re.match(r"^(?:#{1,6}\s*)?[\[【]?\s*(正文|主文案|primary_text|body|text|标题|headline|title|描述|description|desc)\s*(?:\d{0,2})\s*[\]】]?$", s, re.I)
        if section:
            raw_label = section.group(1)
            label = raw_label.lower().replace(" ", "_")
            cur = key_map.get(label) or key_map.get(raw_label)
            if cur:
                res.setdefault(cur, "")
                continue
        m = re.match(r"^([A-Za-z_ ]+|[\u4e00-\u9fff]+)\s*(\d{0,2})\s*[:：]\s*(.*)$", s)
        if m:
            raw_label = m.group(1).strip()
            label = raw_label.lower().replace(" ", "_")
            label = re.sub(r"\d+$", "", label)
            key = key_map.get(label) or key_map.get(re.sub(r"\d+$", "", raw_label))
            if key:
                cur = key
                val = m.group(3).strip()
                if val:
                    res[key] = (res.get(key, "") + ("\n" if res.get(key) else "") + val).strip()
                else:
                    res.setdefault(key, "")
                continue
        if cur:
            res[cur] = (res.get(cur, "") + ("\n" if res.get(cur) else "") + s).strip()
    if not any(str(value or "").strip() for value in res.values()) and default_key:
        variants = _plain_text_variants(text)
        if variants:
            res = {default_key: "\n".join(variants)}
    return res

def find_pack(product, selling_point, asset_language=""):
    base_candidates = [ASSET_ROOT / product / selling_point, ASSET_ROOT / selling_point, Path(selling_point)]
    # 语种版结构：素材包/产品/卖点/EN/文案.txt + 图片/
    if asset_language:
        for b in base_candidates:
            if not b.exists() or not b.is_dir():
                continue
            for child in b.iterdir():
                if child.is_dir() and child.name.casefold() == str(asset_language).casefold():
                    return child
        # 预算表指定语种后严格匹配，缺目录就交给预检拦截，不再静默回退。
        return ASSET_ROOT / product / selling_point / asset_language
    for p in base_candidates:
        if p.exists() and p.is_dir(): return p
    return ASSET_ROOT / product / selling_point


def detect_language_variants(product, selling_point):
    base = ASSET_ROOT / product / selling_point
    langs = []
    if base.exists() and base.is_dir():
        # 根目录本身也可以作为 DEFAULT，兼容老素材包。
        root_info = inspect_pack(product, selling_point, "") if not getattr(detect_language_variants, "_in_call", False) else {"images":[],"copy_exists":False}
        if root_info.get("images") or root_info.get("copy_exists"):
            langs.append("DEFAULT")
        for d in sorted([x for x in base.iterdir() if x.is_dir()], key=lambda x:x.name.lower()):
            if d.name in IMAGE_DIRS:
                continue
            has_images = any(x.is_file() and x.suffix.lower() in IMAGE_EXT for x in d.iterdir())
            has_text = any((d / n).exists() for n in TEXT_FILES) or any(x.is_file() and x.suffix.lower() in TEXT_SUFFIXES for x in d.iterdir())
            has_image_dir = any((d / img).exists() and (d / img).is_dir() for img in IMAGE_DIRS)
            if has_text or has_images or has_image_dir:
                langs.append(d.name)
    return langs or ["DEFAULT"]


def inspect_pack(product, selling_point, asset_language=""):
    lang = "" if str(asset_language or "").upper() in {"", "DEFAULT", "AUTO"} else str(asset_language).strip()
    folder = find_pack(product, selling_point, lang)
    info = {"exists": folder.exists(), "folder": str(folder), "asset_language": lang or "DEFAULT", "images": [], "copy": {}, "copy_exists": False, "copy_counts": {}, "copy_sources": []}
    if not folder.exists(): return info
    images = []
    for d in IMAGE_DIRS:
        p = folder / d
        if p.exists() and p.is_dir(): images += sorted([x for x in p.iterdir() if x.is_file() and x.suffix.lower() in IMAGE_EXT])
    if not images: images = sorted([x for x in folder.iterdir() if x.is_file() and x.suffix.lower() in IMAGE_EXT])
    info["images"] = [{"name": x.name, "rel": str(x.relative_to(BASE_DIR)).replace("\\", "/")} for x in images]
    copy, copy_sources = {}, []
    text_files = sorted([x for x in folder.iterdir() if x.is_file() and x.suffix.lower() in TEXT_SUFFIXES], key=lambda x: x.name.lower())
    priority = {name.lower(): index for index, name in enumerate(["primary_text.txt", "body.txt", "正文.txt", "headline.txt", "title.txt", "标题.txt", "description.txt", "desc.txt", "描述.txt", *TEXT_FILES])}
    text_files.sort(key=lambda x: (priority.get(x.name.lower(), 999), x.name.lower()))
    for path in text_files:
        parsed = parse_copy_text(read_text_file(path), default_key=_copy_key_from_filename(path))
        recognized = False
        for key, value in parsed.items():
            if value:
                _merge_copy_value(copy, key, value)
                recognized = True
        if recognized:
            copy_sources.append(path.name)
    info["copy"] = copy
    info["copy_sources"] = copy_sources
    info["copy_exists"] = bool(copy.get("primary_text") or copy.get("headline") or copy.get("description"))
    for k in ["primary_text", "headline", "description"]:
        info["copy_counts"][k] = len([x for x in str(copy.get(k, "")).split("\n") if x.strip()])
    return info


def scan_asset_packs():
    ASSET_ROOT.mkdir(parents=True, exist_ok=True)
    result = {}
    for pd in sorted([p for p in ASSET_ROOT.iterdir() if p.is_dir()], key=lambda p:p.name.lower()):
        pts=[]
        for pt in sorted([p for p in pd.iterdir() if p.is_dir()], key=lambda p:p.name.lower()):
            langs = detect_language_variants(pd.name, pt.name)
            total_images = 0
            any_copy = False
            for lang in langs:
                data = inspect_pack(pd.name, pt.name, lang)
                total_images += len(data["images"])
                any_copy = any_copy or data["copy_exists"]
            pts.append({"name": pt.name, "image_count": total_images, "copy_exists": any_copy, "languages": langs})
        result[pd.name] = pts
    return result

def check_all_packs():
    """扫描所有素材包，输出每个产品/卖点/语种是否缺图片或文案。"""
    rows = []
    assets = scan_asset_packs()
    for product, points in assets.items():
        for point in points:
            for lang in point.get("languages", ["DEFAULT"]):
                info = inspect_pack(product, point["name"], lang)
                problems = []
                if not info.get("exists"):
                    problems.append("素材包不存在")
                if not info.get("images"):
                    problems.append("缺少图片")
                if not info.get("copy_exists"):
                    problems.append("缺少文案")
                rows.append({
                    "product": product,
                    "selling_point": point["name"],
                    "asset_language": lang,
                    "folder": info.get("folder", ""),
                    "images": len(info.get("images", [])),
                    "copy_exists": bool(info.get("copy_exists")),
                    "copy_counts": info.get("copy_counts", {}),
                    "status": "OK" if not problems else "WARN",
                    "problems": problems,
                })
    return {"count": len(rows), "items": rows}

def create_asset_pack(product, selling_point, asset_language="EN"):
    product = norm(product)
    selling_point = norm(selling_point)
    asset_language = norm(asset_language) or "EN"
    if not product or not selling_point:
        raise RuntimeError("产品和卖点名称不能为空。")
    if re.search(r"[\\/:*?\"<>|]", product + selling_point + asset_language):
        raise RuntimeError("产品/卖点/语种名称不能包含这些字符：\\ / : * ? \" < > |")
    folder = ASSET_ROOT / product / selling_point / asset_language
    image_dir = folder / "图片"
    image_dir.mkdir(parents=True, exist_ok=True)
    copy_file = folder / "文案.txt"
    created_copy = False
    if not copy_file.exists():
        copy_file.write_text("正文1：\n\n正文2：\n\n正文3：\n\n正文4：\n\n正文5：\n\n标题1：\n\n标题2：\n\n标题3：\n\n标题4：\n\n标题5：\n\n描述1：\n\n描述2：\n\n描述3：\n\n描述4：\n\n描述5：\n\nCTA：\nGET_QUOTE\n", encoding="utf-8")
        created_copy = True
    return {"folder": str(folder), "image_dir": str(image_dir), "copy_file": str(copy_file), "asset_language": asset_language, "created_copy": created_copy}



def file_md5(path):
    import hashlib
    h = hashlib.md5()
    with Path(path).open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def upload_image_to_meta(ad_account_id, token, image_path):
    account = clean_id(ad_account_id)
    if not account:
        raise RuntimeError("Ad Account ID 为空，无法上传图片 Hash。")
    if not str(account).startswith("act_"):
        account = "act_" + account
    url = f"https://graph.facebook.com/{API_VERSION}/{account}/adimages"
    res = None
    for attempt in range(4):
        try:
            with Path(image_path).open("rb") as f:
                res = requests.post(url, data={"access_token": token}, files={"filename": f}, timeout=120)
            break
        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError, requests.exceptions.Timeout) as exc:
            if attempt == 3:
                safe_error = re.sub(r"(?i)(access_token=)[^&\s]+", r"\1[REDACTED]", str(exc)).replace(token, "[REDACTED]")
                raise RuntimeError(f"上传图片时连接 Meta 失败，已自动重试 4 次：{safe_error}") from exc
            time.sleep(2 ** (attempt + 1))
    assert res is not None
    try:
        data = res.json()
    except Exception:
        raise RuntimeError("Meta 上传图片返回非 JSON：" + res.text[:800])
    if "error" in data:
        raise RuntimeError(json.dumps(data, ensure_ascii=False, indent=2))
    images = data.get("images", {})
    if not images:
        raise RuntimeError(f"上传后没有返回 image hash：{Path(image_path).name}")
    first = next(iter(images.values()))
    h = first.get("hash")
    if not h:
        raise RuntimeError(f"上传后没有找到 hash：{Path(image_path).name}")
    return h


def load_image_library():
    return load_json(IMAGE_LIBRARY_FILE, {"by_md5": {}, "by_filename": {}, "updated_at": ""})


def save_image_library(lib):
    lib["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_json(IMAGE_LIBRARY_FILE, lib)


def image_hash_status():
    """扫描素材包，检查每张图是否已经有 Meta image hash。"""
    lib = load_image_library()
    rows = []
    total = ready = missing = 0
    assets = scan_asset_packs()
    for product, points in assets.items():
        for point in points:
            for lang in point.get("languages", ["DEFAULT"]):
                info = inspect_pack(product, point["name"], lang)
                for img in info.get("images", []):
                    p = (BASE_DIR / img["rel"]).resolve()
                    if not p.exists():
                        continue
                    md5 = file_md5(p)
                    item = (lib.get("by_md5", {}) or {}).get(md5) or (lib.get("by_filename", {}) or {}).get(p.name) or {}
                    h = item.get("hash") or item.get("image_hash") or ""
                    total += 1
                    if h:
                        ready += 1
                        status = "READY"
                    else:
                        missing += 1
                        status = "MISSING"
                    rows.append({
                        "status": status,
                        "product": product,
                        "selling_point": point["name"],
                        "asset_language": lang,
                        "file_name": p.name,
                        "rel": img["rel"],
                        "md5": md5,
                        "hash": h,
                    })
    return {"total": total, "ready": ready, "missing": missing, "items": rows, "library": str(IMAGE_LIBRARY_FILE)}


def sync_image_hashes(ad_account_id, limit=0):
    """把素材包中缺 hash 的图片上传到 Meta，写入 _cache/image_library.json。"""
    token = get_token()
    lib = load_image_library()
    status = image_hash_status()
    uploaded = []
    skipped = 0
    errors = []
    count = 0
    for row in status.get("items", []):
        if row.get("status") == "READY":
            skipped += 1
            continue
        if limit and count >= int(limit):
            break
        p = (BASE_DIR / row["rel"]).resolve()
        try:
            h = upload_image_to_meta(ad_account_id, token, p)
            item = {
                "filename": p.name,
                "path": str(p),
                "rel": row["rel"],
                "md5": row["md5"],
                "hash": h,
                "product": row["product"],
                "selling_point": row["selling_point"],
                "asset_language": row["asset_language"],
                "ad_account_id": clean_id(ad_account_id),
                "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source": "frontend_asset_library_upload",
            }
            lib.setdefault("by_md5", {})[row["md5"]] = item
            lib.setdefault("by_filename", {})[p.name] = item
            uploaded.append({**row, "hash": h})
            count += 1
        except Exception as e:
            errors.append({**row, "error": str(e)[:1000]})
    save_image_library(lib)
    after = image_hash_status()
    return {"uploaded": len(uploaded), "skipped_ready": skipped, "errors": errors, "after": after, "library": str(IMAGE_LIBRARY_FILE)}

def list_logs():
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    files = []
    for p in sorted(LOG_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True)[:20]:
        st = p.stat()
        files.append({"name": p.name, "path": str(p), "size_kb": round(st.st_size/1024, 1), "modified": datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S")})
    return {"count": len(files), "items": files}

def clean_id(v):
    s = norm(v)
    return s.split(":")[-1] if ":" in s else s



def product_alias_key(value):
    return re.sub(r"[^A-Za-z0-9+]+", "", str(value or "")).upper()

def load_product_urls():
    return load_json(PRODUCT_URLS_FILE, {"default_url":"", "products":{}, "aliases":{}})

def resolve_product_url(product):
    data = load_product_urls()
    products = data.get("products", {})
    aliases = data.get("aliases", {})
    p = norm(product)
    if p in products:
        return products[p]
    compact = product_alias_key(p)
    for k, v in products.items():
        if product_alias_key(k) == compact:
            return v
    alias_target = aliases.get(compact) or aliases.get(p.upper())
    if alias_target and alias_target in products:
        return products[alias_target]
    return data.get("default_url") or ""


def load_product_settings():
    return load_json(PRODUCT_SETTINGS_FILE, {"default_page_id":"1076514822219251", "product_lines":{}})


def normalize_product_name(product):
    text = str(product or "").strip()
    compact = re.sub(r"[^A-Za-z0-9]+", "", text).upper()
    aliases = {
        "V4EPRO":"V4e Pro", "V4EMINI":"V4e Pro", "V10A":"V10a", "V10I":"V10i",
        "V10L":"V10L", "V1T":"V1t", "V1T5W":"V1t 5W", "EG10":"EG10", "EC100":"EC100",
        "GC100":"GC100", "DC100":"DC100", "TERRAMIND":"TerraMind", "P1":"P1", "P2":"P2",
        "P2VISION":"P2 Vision+", "P2VISIONPLUS":"P2 Vision+", "P2VISION+":"P2 Vision+",
        "S1":"S1", "S2":"S2", "V4E":"V4e", "MODELWEB":"Model Web"
    }
    return aliases.get(compact, text)


def product_line_for_product(product):
    product_key = normalize_product_name(product)
    settings = load_product_settings()
    for line_name, info in (settings.get("product_lines") or {}).items():
        products = [normalize_product_name(p) for p in info.get("products", [])]
        if product_key in products:
            return line_name
    return "MC"


def page_id_for_product(product):
    settings = load_product_settings()
    line = product_line_for_product(product)
    info = (settings.get("product_lines") or {}).get(line, {})
    return str(info.get("page_id") or settings.get("default_page_id") or "").strip()

def get_campaign_map():
    mapping={}
    if not PLAN_FILE.exists(): return mapping
    try:
        wb=load_workbook(PLAN_FILE, read_only=True, data_only=True)
        if "产品Campaign对照" not in wb.sheetnames: return mapping
        ws=wb["产品Campaign对照"]; headers=[str(c.value or "").strip().lower() for c in ws[1]]
        if "product" not in headers or "campaign_id" not in headers: return mapping
        pc,cc=headers.index("product")+1,headers.index("campaign_id")+1
        for r in range(2, ws.max_row+1):
            p=norm(ws.cell(r,pc).value); cid=clean_id(ws.cell(r,cc).value)
            if p and cid: mapping[p]=cid
    except Exception: pass
    return mapping


def quarter_bounds(year=None, quarter=None):
    today = datetime.now().date()
    year = int(year or today.year)
    quarter = int(quarter or ((today.month - 1) // 3 + 1))
    if quarter not in {1, 2, 3, 4}:
        raise RuntimeError("季度必须是 1、2、3 或 4。")
    start_month = (quarter - 1) * 3 + 1
    start = datetime(year, start_month, 1).date()
    next_start = datetime(year + 1, 1, 1).date() if quarter == 4 else datetime(year, start_month + 3, 1).date()
    return start, next_start - timedelta(days=1), year, quarter


def _budget_number(value):
    text = re.sub(r"[^0-9.\-]", "", str(value or "").replace(",", ""))
    try:
        number = float(text)
    except (TypeError, ValueError):
        return 0.0
    return round(number, 2)


def _budget_header(value):
    return re.sub(r"[\s_()（）/\-]+", "", str(value or "").strip().lower())


def _parse_budget_countries(value):
    raw = str(value or "").strip()
    if not raw:
        return []
    aliases = {}
    for code, name in COUNTRY_OPTIONS:
        aliases[code.upper()] = code
        aliases[str(name).strip().upper()] = code
    parts = [part.strip() for part in re.split(r"[|,，;；/\s]+", raw) if part.strip()]
    countries = []
    for part in parts:
        code = aliases.get(part.upper(), part.upper() if re.fullmatch(r"[A-Za-z]{2}", part) else "")
        if code and code not in countries:
            countries.append(code)
    return countries


def _parse_budget_period(year_value, quarter_value):
    today = datetime.now().date()
    combined = f"{year_value or ''} {quarter_value or ''}"
    year_match = re.search(r"(20\d{2})", combined)
    year = int(year_match.group(1)) if year_match else today.year
    quarter_text = str(quarter_value or "").strip().upper()
    quarter_match = re.search(r"Q?\s*([1-4])", quarter_text)
    if quarter_match:
        quarter = int(quarter_match.group(1))
    else:
        chinese = {"一": 1, "二": 2, "三": 3, "四": 4}
        quarter = next((value for key, value in chinese.items() if key in quarter_text), (today.month - 1) // 3 + 1)
    return year, quarter


def import_quarter_budget(payload):
    name = Path(str(payload.get("name") or "季度预算.xlsx")).name
    encoded = str(payload.get("data") or "")
    if not encoded:
        raise RuntimeError("没有收到预算文件。")
    if "," in encoded:
        encoded = encoded.split(",", 1)[1]
    try:
        raw = base64.b64decode(encoded, validate=True)
    except Exception:
        raise RuntimeError("预算文件内容无法读取，请重新选择文件。")
    suffix = Path(name).suffix.lower()
    if suffix == ".csv":
        text = None
        for encoding in ("utf-8-sig", "gbk", "utf-8"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise RuntimeError("CSV 编码无法识别，请另存为 UTF-8 CSV 或 XLSX。")
        matrix = list(csv.reader(StringIO(text)))
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        matrix = []
        for sheet in workbook.worksheets:
            candidate = [list(row) for row in sheet.iter_rows(values_only=True)]
            if candidate:
                matrix = candidate
                if any("预算" in str(cell or "") or "budget" in str(cell or "").lower() for row in candidate[:10] for cell in row):
                    break
    else:
        raise RuntimeError("预算表仅支持 .xlsx、.xlsm 或 .csv。")
    if not matrix:
        raise RuntimeError("预算表是空的。")

    aliases = {
        "product": {"产品", "产品名", "产品名称", "产品型号", "product", "productname", "model"},
        "campaign_id": {"campaignid", "广告系列id", "广告系列编号", "campaign编号"},
        "campaign_name": {"campaign", "campaignname", "广告系列", "广告系列名称"},
        "countries": {"国家", "国家/地区", "国家地区", "国家代码", "投放国家", "地区", "投放地区", "市场", "country", "countries", "region", "market"},
        "asset_language": {"语种", "语种可选", "素材语种", "固定语种", "语言", "language", "lang", "assetlanguage"},
        "budget": {"预算", "预算金额", "金额", "初始预算", "季度预算", "预算usd", "预算美元", "初始预算usd", "初始预算美元", "季度预算usd", "季度预算美元", "budget", "budgetusd", "initialbudget", "quarterbudget", "totalbudget"},
        "year": {"年", "年份", "年度", "year"},
        "quarter": {"季度", "quarter", "q"},
    }
    aliases = {key: {_budget_header(value) for value in values} for key, values in aliases.items()}
    header_row, columns = None, {}
    for row_index, row in enumerate(matrix[:20]):
        normalized = [_budget_header(cell) for cell in row]
        found = {}
        for key, names in aliases.items():
            for column_index, value in enumerate(normalized):
                if value in names:
                    found[key] = column_index
                    break
        if "budget" in found and ("product" in found or "campaign_id" in found or "campaign_name" in found):
            header_row, columns = row_index, found
            break
    if header_row is None:
        raise RuntimeError("没有识别到表头。至少需要：产品（或 Campaign ID）和初始预算。")

    campaign_map = get_campaign_map()
    reverse_campaign_map = {clean_id(value): key for key, value in campaign_map.items()}
    rows, errors = [], []
    for row_number, row in enumerate(matrix[header_row + 1:], start=header_row + 2):
        def value(key):
            index = columns.get(key)
            return row[index] if index is not None and index < len(row) else ""
        product = norm(value("product"))
        campaign_id = clean_id(value("campaign_id"))
        campaign_name = norm(value("campaign_name"))
        countries = _parse_budget_countries(value("countries"))
        asset_language = norm(value("asset_language")).upper()
        initial_budget = _budget_number(value("budget"))
        if not any(norm(cell) for cell in row):
            continue
        if not product and campaign_name:
            product = _infer_product(campaign_name)
        if not product and campaign_id:
            product = reverse_campaign_map.get(campaign_id, "")
        if not campaign_id and product:
            campaign_id = clean_id(campaign_map.get(product))
        if not product and not campaign_id:
            errors.append(f"第 {row_number} 行缺少产品或 Campaign ID")
            continue
        if initial_budget <= 0:
            errors.append(f"第 {row_number} 行初始预算必须大于 0")
            continue
        year, quarter = _parse_budget_period(value("year"), value("quarter"))
        rows.append({
            "year": year, "quarter": quarter, "product": product, "campaign_id": campaign_id,
            "campaign_name": campaign_name, "countries": countries, "asset_language": asset_language, "initial_budget": initial_budget,
        })
    if not rows:
        raise RuntimeError("预算表没有可用数据。" + ("；" + "；".join(errors[:5]) if errors else ""))
    deduplicated = {}
    for row in rows:
        identity = row["campaign_id"] or product_key(row["product"])
        deduplicated[(row["year"], row["quarter"], identity, "|".join(row.get("countries") or []))] = row
    rows = list(deduplicated.values())
    config = {"updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "source_name": name, "rows": rows}
    save_json(QUARTER_BUDGET_FILE, config)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    source_suffix = suffix if suffix in {".xlsx", ".xlsm", ".csv"} else ".xlsx"
    (CACHE_DIR / f"quarterly_budget_source{source_suffix}").write_bytes(raw)
    return {"imported": len(rows), "errors": errors, "config": config}


def fetch_quarter_spend(year=None, quarter=None, force=False):
    start, end, year, quarter = quarter_bounds(year, quarter)
    today = datetime.now().date()
    until = min(today, end)
    if until < start:
        return {"year": year, "quarter": quarter, "since": start.isoformat(), "until": until.isoformat(), "campaigns": {}, "cached_at": datetime.now().isoformat(timespec="seconds"), "source": "future"}
    account_id = clean_id(get_defaults().get("ad_account_id"))
    if not account_id:
        raise RuntimeError("缺少 Ad Account ID，无法读取季度已花费。")
    key = f"{account_id}:{year}Q{quarter}:country-v2"
    with quarter_spend_lock:
        store = load_json(QUARTER_SPEND_CACHE_FILE, {"entries": {}})
        entry = (store.get("entries") or {}).get(key) or {}
        try:
            cached_at = datetime.fromisoformat(str(entry.get("cached_at") or ""))
        except (TypeError, ValueError):
            cached_at = None
        if not force and entry.get("data") and cached_at and (datetime.now() - cached_at).total_seconds() < QUARTER_SPEND_CACHE_TTL_SECONDS:
            result = dict(entry["data"])
            result.update({"source": "disk", "cached_at": cached_at.isoformat(timespec="seconds")})
            return result
        account = account_id if account_id.startswith("act_") else "act_" + account_id
        params = {
            "access_token": get_token(), "level": "campaign", "fields": "campaign_id,campaign_name,spend,impressions,reach,clicks,actions",
            "time_range": json.dumps({"since": start.isoformat(), "until": until.isoformat()}), "breakdowns": "country", "limit": 500,
        }
        try:
            raw_rows = _graph_get_all(f"https://graph.facebook.com/{API_VERSION}/{account}/insights", params)
            campaigns = {}
            for row in raw_rows:
                campaign_id = clean_id(row.get("campaign_id"))
                if not campaign_id:
                    continue
                item = campaigns.setdefault(campaign_id, {"campaign_id": campaign_id, "campaign_name": str(row.get("campaign_name") or ""), "spend": 0.0, "impressions": 0, "reach": 0, "clicks": 0, "leads": 0.0, "countries": {}})
                country = str(row.get("country") or "UNKNOWN").upper()
                country_spend = _number(row.get("spend"))
                country_leads = _action_value(row.get("actions"), ["onsite_conversion.lead_grouped", "lead", "onsite_conversion.lead", "offsite_conversion.fb_pixel_lead"])
                item["spend"] += country_spend
                item["impressions"] += int(_number(row.get("impressions")))
                item["reach"] += int(_number(row.get("reach")))
                item["clicks"] += int(_number(row.get("clicks")))
                item["leads"] += country_leads
                country_item = item["countries"].setdefault(country, {"spend": 0.0, "impressions": 0, "reach": 0, "clicks": 0, "leads": 0.0})
                country_item["spend"] += country_spend
                country_item["impressions"] += int(_number(row.get("impressions")))
                country_item["reach"] += int(_number(row.get("reach")))
                country_item["clicks"] += int(_number(row.get("clicks")))
                country_item["leads"] += country_leads
            for item in campaigns.values():
                item["spend"] = round(item["spend"], 2)
                item["leads"] = round(item["leads"], 2)
                for country_item in item["countries"].values():
                    country_item["spend"] = round(country_item["spend"], 2)
                    country_item["leads"] = round(country_item["leads"], 2)
            data = {"year": year, "quarter": quarter, "since": start.isoformat(), "until": until.isoformat(), "campaigns": campaigns}
            cached_at = datetime.now()
            store.setdefault("entries", {})[key] = {"cached_at": cached_at.isoformat(timespec="seconds"), "data": data}
            save_json(QUARTER_SPEND_CACHE_FILE, store)
            data.update({"source": "meta_api", "cached_at": cached_at.isoformat(timespec="seconds")})
            return data
        except Exception:
            if entry.get("data") and cached_at:
                result = dict(entry["data"])
                result.update({"source": "stale_disk", "cached_at": cached_at.isoformat(timespec="seconds"), "stale": True})
                return result
            raise


def fetch_active_ad_coverage(force=False):
    account_id = clean_id(get_defaults().get("ad_account_id"))
    if not account_id:
        raise RuntimeError("缺少 Ad Account ID，无法检测地区投放状态。")
    key = f"{account_id}:active-v1"
    with active_coverage_lock:
        store = load_json(ACTIVE_COVERAGE_CACHE_FILE, {"entries": {}})
        entry = (store.get("entries") or {}).get(key) or {}
        try:
            cached_at = datetime.fromisoformat(str(entry.get("cached_at") or ""))
        except (TypeError, ValueError):
            cached_at = None
        if not force and entry.get("data") and cached_at and (datetime.now() - cached_at).total_seconds() < QUARTER_SPEND_CACHE_TTL_SECONDS:
            result = dict(entry["data"])
            result.update({"source": "disk", "cached_at": cached_at.isoformat(timespec="seconds")})
            return result
        account = account_id if account_id.startswith("act_") else "act_" + account_id
        params = {
            "access_token": get_token(),
            "fields": "id,name,status,effective_status,campaign_id,targeting,campaign{name}",
            "limit": 500,
        }
        try:
            adsets = _graph_get_all(f"https://graph.facebook.com/{API_VERSION}/{account}/adsets", params)
            campaigns = {}
            active_count = 0
            for adset in adsets:
                if str(adset.get("effective_status") or adset.get("status") or "").upper() != "ACTIVE":
                    continue
                active_count += 1
                campaign_id = clean_id(adset.get("campaign_id"))
                campaign = adset.get("campaign") or {}
                item = campaigns.setdefault(campaign_id, {"campaign_id": campaign_id, "campaign_name": str(campaign.get("name") or ""), "countries": set(), "adsets": []})
                countries = ((adset.get("targeting") or {}).get("geo_locations") or {}).get("countries") or []
                item["countries"].update(str(country).upper() for country in countries if country)
                item["adsets"].append({"id": str(adset.get("id") or ""), "name": str(adset.get("name") or ""), "countries": [str(country).upper() for country in countries if country]})
            for item in campaigns.values():
                item["countries"] = sorted(item["countries"])
            data = {"campaigns": campaigns, "active_adsets": active_count}
            cached_at = datetime.now()
            store.setdefault("entries", {})[key] = {"cached_at": cached_at.isoformat(timespec="seconds"), "data": data}
            save_json(ACTIVE_COVERAGE_CACHE_FILE, store)
            data.update({"source": "meta_api", "cached_at": cached_at.isoformat(timespec="seconds")})
            return data
        except Exception:
            if entry.get("data") and cached_at:
                result = dict(entry["data"])
                result.update({"source": "stale_disk", "cached_at": cached_at.isoformat(timespec="seconds"), "stale": True})
                return result
            raise


def _duplicate_name_key(value):
    return re.sub(r"[\W_]+", "", str(value or ""), flags=re.UNICODE).casefold()


def find_duplicate_active_ads(rows, coverage_data=None):
    """Find same Campaign + selling point + overlapping country before creating ads."""
    if not rows:
        return {}
    coverage_data = coverage_data or fetch_active_ad_coverage(force=False)
    campaigns = coverage_data.get("campaigns") or {}
    duplicates = {}
    planned = {}
    for index, row in enumerate(rows):
        campaign_id = clean_id(row.get("campaign_id"))
        point_key = _duplicate_name_key(row.get("selling_point"))
        countries = {str(country).upper() for country in parse_countries(row.get("countries"))}
        if not campaign_id or not point_key or not countries:
            continue
        hits = []
        for adset in (campaigns.get(campaign_id) or {}).get("adsets", []):
            adset_key = _duplicate_name_key(adset.get("name"))
            overlap = countries.intersection({str(country).upper() for country in (adset.get("countries") or [])})
            if point_key in adset_key and overlap:
                hits.append({"source": "active", "id": adset.get("id", ""), "name": adset.get("name", ""), "countries": sorted(overlap)})
        planned_key = (campaign_id, point_key)
        for earlier in planned.get(planned_key, []):
            overlap = countries.intersection(earlier["countries"])
            if overlap:
                hits.append({"source": "queue", "id": "", "name": earlier["name"], "countries": sorted(overlap)})
        planned.setdefault(planned_key, []).append({"name": f"队列第 {index + 1} 行", "countries": countries})
        if hits:
            duplicates[index] = hits
    return duplicates


def quarter_budget_status(force=False):
    config = load_json(QUARTER_BUDGET_FILE, {"rows": []})
    start, end, year, quarter = quarter_bounds()
    current_rows = [row for row in (config.get("rows") or []) if int(row.get("year") or 0) == year and int(row.get("quarter") or 0) == quarter]
    spend_data = fetch_quarter_spend(year, quarter, force=force) if current_rows else {"campaigns": {}, "source": "none", "cached_at": ""}
    coverage_data = fetch_active_ad_coverage(force=force) if current_rows else {"campaigns": {}, "source": "none", "cached_at": "", "active_adsets": 0}
    campaign_map = get_campaign_map()
    today = datetime.now().date()
    remaining_days = max(0, (end - today).days + 1)
    total_days = (end - start).days + 1
    elapsed_days = max(0, min(total_days, (min(today, end) - start).days + 1)) if today >= start else 0
    items = []
    for row in current_rows:
        product = norm(row.get("product"))
        campaign_id = clean_id(row.get("campaign_id")) or clean_id(campaign_map.get(product))
        spend_item = (spend_data.get("campaigns") or {}).get(campaign_id, {})
        coverage_item = (coverage_data.get("campaigns") or {}).get(campaign_id, {})
        countries = [str(country).upper() for country in (row.get("countries") or []) if country]
        initial = round(_number(row.get("initial_budget")), 2)
        if countries:
            country_metrics = [(spend_item.get("countries") or {}).get(country) or {} for country in countries]
            spent = round(sum(_number(metric.get("spend")) for metric in country_metrics), 2)
            leads = round(sum(_number(metric.get("leads")) for metric in country_metrics), 2)
            impressions = sum(int(_number(metric.get("impressions"))) for metric in country_metrics)
            reach = sum(int(_number(metric.get("reach"))) for metric in country_metrics)
            clicks = sum(int(_number(metric.get("clicks"))) for metric in country_metrics)
        else:
            spent = round(_number(spend_item.get("spend")), 2)
            leads = round(_number(spend_item.get("leads")), 2)
            impressions = int(_number(spend_item.get("impressions")))
            reach = int(_number(spend_item.get("reach")))
            clicks = int(_number(spend_item.get("clicks")))
        remaining = max(0, round(initial - spent, 2))
        calculated_daily = round(remaining / remaining_days, 2) if remaining_days else 0
        daily = calculated_daily
        waiting_for_min_budget = remaining > 0 and remaining_days > 0 and daily < META_MIN_DAILY_BUDGET
        eligible_remaining_days = int(remaining / META_MIN_DAILY_BUDGET) if waiting_for_min_budget else remaining_days
        budget_eligible_in_days = max(0, remaining_days - eligible_remaining_days) if waiting_for_min_budget else 0
        budget_eligible_date = (today + timedelta(days=budget_eligible_in_days)).isoformat() if waiting_for_min_budget else ""
        expected_spend = round(initial * elapsed_days / total_days, 2) if total_days else 0
        pace_percent = round(spent / expected_spend * 100, 1) if expected_spend else 0
        pace_status = "overspend" if expected_spend and spent > expected_spend * 1.1 else ("slow" if elapsed_days >= 7 and expected_spend and spent < expected_spend * 0.8 else "on_track")
        average_daily_spend = round(spent / elapsed_days, 2) if elapsed_days else 0
        remaining_after_today = max(0, (end - today).days)
        projected_spend = round(spent + average_daily_spend * remaining_after_today, 2)
        projected_completion = round(projected_spend / initial * 100, 1) if initial else 0
        days_to_exhaust = int((remaining / average_daily_spend) + 0.9999) if average_daily_spend > 0 and remaining > 0 else None
        exhaustion_date = (today + timedelta(days=days_to_exhaust)).isoformat() if days_to_exhaust is not None else ""
        cpl = round(spent / leads, 2) if leads else 0
        ctr = round(clicks / impressions * 100, 3) if impressions else 0
        frequency = round(impressions / reach, 2) if reach else 0
        active_countries = [str(country).upper() for country in (coverage_item.get("countries") or [])]
        blocked_countries = [country for country in countries if country in BLOCKED_COUNTRIES]
        missing_countries = [country for country in countries if country not in active_countries and country not in BLOCKED_COUNTRIES]
        has_active = bool(coverage_item.get("adsets"))
        items.append({
            **row, "countries": countries, "campaign_id": campaign_id, "campaign_name": row.get("campaign_name") or spend_item.get("campaign_name", "") or coverage_item.get("campaign_name", ""),
            "spent": spent, "remaining": remaining, "remaining_days": remaining_days, "daily_budget": daily,
            "calculated_daily_budget": calculated_daily, "waiting_for_min_budget": waiting_for_min_budget,
            "budget_eligible_in_days": budget_eligible_in_days, "budget_eligible_date": budget_eligible_date,
            "leads": leads, "cpl": cpl, "impressions": impressions, "reach": reach, "clicks": clicks, "ctr": ctr, "frequency": frequency,
            "expected_spend": expected_spend, "pace_percent": pace_percent, "pace_status": pace_status,
            "average_daily_spend": average_daily_spend, "projected_spend": projected_spend, "projected_completion": projected_completion,
            "days_to_exhaust": days_to_exhaust, "exhaustion_date": exhaustion_date,
            "matched": bool(campaign_id), "exhausted": remaining <= 0, "active_countries": active_countries,
            "missing_countries": missing_countries, "blocked_countries": blocked_countries, "has_active": has_active,
            "coverage_status": "blocked" if blocked_countries and not [country for country in countries if country not in BLOCKED_COUNTRIES] else ("missing" if missing_countries else ("active" if (countries or has_active) else "no_region")),
        })
    return {
        "year": year, "quarter": quarter, "period": f"{start.isoformat()} 至 {end.isoformat()}",
        "remaining_days": remaining_days, "elapsed_days": elapsed_days, "total_days": total_days, "updated_at": config.get("updated_at", ""), "source_name": config.get("source_name", ""),
        "spend_source": spend_data.get("source", "none"), "spend_cached_at": spend_data.get("cached_at", ""),
        "coverage_source": coverage_data.get("source", "none"), "coverage_cached_at": coverage_data.get("cached_at", ""), "active_adsets": coverage_data.get("active_adsets", 0),
        "items": items, "imported_total": len(config.get("rows") or []), "current_count": len(items),
    }


def match_quarter_budget(status, product, campaign_id, countries=None):
    campaign_id = clean_id(campaign_id)
    product_match = product_key(product)
    items = status.get("items") or []
    candidates = [item for item in items if campaign_id and clean_id(item.get("campaign_id")) == campaign_id]
    if not candidates:
        candidates = [item for item in items if product_match and product_key(item.get("product")) == product_match]
    if not candidates:
        return None
    requested = {str(country).upper() for country in (countries or []) if country}
    regional = [item for item in candidates if item.get("countries")]
    if requested and regional:
        matched = [item for item in regional if requested.intersection({str(country).upper() for country in item.get("countries") or []})]
        if matched:
            aggregate = dict(matched[0])
            for field in ("initial_budget", "spent", "remaining", "daily_budget"):
                total = 0.0
                for item in matched:
                    item_countries = {str(country).upper() for country in item.get("countries") or []}
                    ratio = len(requested.intersection(item_countries)) / max(1, len(item_countries))
                    total += _number(item.get(field)) * ratio
                aggregate[field] = round(total, 2)
            aggregate["countries"] = sorted(requested)
            aggregate["missing_countries"] = sorted(set().union(*(set(item.get("missing_countries") or []) for item in matched)).intersection(requested))
            return aggregate
        # 有按地区拆分的预算时，禁止回退到同产品的其他地区，否则会串用预算和素材语种。
        return next((item for item in candidates if not item.get("countries")), None)
    return next((item for item in candidates if not item.get("countries")), candidates[0] if not requested else None)


def send_budget_template(handler):
    _, _, year, quarter = quarter_bounds()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "季度预算"
    sheet.append(["年份", "季度", "产品", "Campaign ID", "国家/地区", "语种(可选)", "初始预算(USD)"])
    campaign_map = get_campaign_map()
    products = list(campaign_map.items()) or [(product, "") for product in scan_asset_packs().keys()]
    for product, campaign_id in products:
        sheet.append([year, f"Q{quarter}", product, campaign_id, "US", "", ""])
    sheet.freeze_panes = "A2"
    for column, width in zip("ABCDEFG", [12, 12, 20, 24, 20, 16, 20]):
        sheet.column_dimensions[column].width = width
    from openpyxl.styles import Font, PatternFill
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="4F46E5")
    buffer = BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()
    from urllib.parse import quote
    filename = f"quarter_budget_template_{year}Q{quarter}.xlsx"
    handler.send_response(200)
    handler.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    handler.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
    handler.send_header("Content-Length", str(len(payload)))
    handler.end_headers()
    handler.wfile.write(payload)

def get_defaults():
    d={"enabled":"Y","daily_budget":"100","page_id":"1076514822219251","lead_form_id":"","website_url":"","cta_type":"GET_QUOTE","language_mode":"auto","custom_language":"","dry_run":"no","delay_max_seconds":"60","ad_account_id":"1682488829626268","url_tags":"utm_source=facebook&utm_medium=socialad&utm_campaign=fjd-{产品名}&utm_id=5224"}
    if PLAN_FILE.exists():
        try:
            wb=load_workbook(PLAN_FILE, read_only=True, data_only=True); ws=wb["广告主表"] if "广告主表" in wb.sheetnames else wb[wb.sheetnames[0]]
            headers=[str(c.value or "").strip() for c in ws[1]]
            for r in range(2, ws.max_row+1):
                row={headers[i]:ws.cell(r,i+1).value for i in range(len(headers)) if headers[i]}
                if any(v not in (None,"") for v in row.values()):
                    for k in d:
                        # Dry Run 不再从旧 plan.xlsx 继承，避免旧表里 yes 导致用户以为已关闭但实际仍预检
                        if k in {"dry_run", "url_tags"}:
                            continue
                        if row.get(k) not in (None,""): d[k]=str(row.get(k)).strip()
                    break
        except Exception: pass
    d["delay_max_seconds"] = "60"
    return d

def backup_plan():
    if PLAN_FILE.exists():
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        p=CACHE_DIR/f"plan_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(PLAN_FILE,p); return str(p)
    return ""

def ensure_plan():
    if PLAN_FILE.exists(): wb=load_workbook(PLAN_FILE)
    else:
        wb=Workbook(); wb.active.title="广告主表"; wb.create_sheet("产品Campaign对照"); wb.create_sheet("字段说明")
    if "广告主表" not in wb.sheetnames:
        wb.create_sheet("广告主表", 0)
    return wb, wb["广告主表"]

def safe_save_workbook(wb, target_path):
    target_path = Path(target_path)
    try:
        wb.save(target_path)
        return str(target_path), ""
    except PermissionError:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        fallback = CACHE_DIR / f"{target_path.stem}_autosave_{datetime.now().strftime('%Y%m%d_%H%M%S')}{target_path.suffix}"
        wb.save(fallback)
        return str(fallback), f"plan.xlsx 可能正在被 Excel 打开，已另存为 {fallback}"

def write_plan(rows):
    backup=backup_plan(); wb,ws=ensure_plan(); ws.delete_rows(1, ws.max_row); ws.append(PLAN_HEADERS)
    for x in rows: ws.append([x.get(h,"") for h in PLAN_HEADERS])
    widths=[10,14,18,18,12,18,20,32,14,14,10,16,18,22,48]
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
    for c in ws[1]:
        f = copy(c.font)
        f.bold = True
        c.font = f
    saved_path, warning = safe_save_workbook(wb, PLAN_FILE)
    return {"plan":saved_path,"backup":backup,"warning":warning}

def parse_countries(raw):
    if isinstance(raw, list): parts=raw
    else: parts=re.split(r"[|,;\s]+", str(raw or ""))
    out=[]
    for x in parts:
        c=str(x).strip().upper()
        if c and c not in out: out.append(c)
    return out

def form_lang_for_countries(countries, merge=True):
    langs=[]
    for c in countries:
        lang=COUNTRY_TO_FORM_LANG.get(c, "EN")
        if lang not in langs: langs.append(lang)
    if merge and len(langs)>1: return "EN"
    return langs[0] if langs else "EN"

def parse_form_name(name):
    # EN_EG10_Leadform_20260611 / KR_V10L_Leadform_20260611 / EN_Model Web_Leadform_20260611
    m=re.match(r"^([A-Z]{2})_(.+?)_Leadform_(\d{8})", str(name or ""), re.I)
    if not m: return None
    return {"lang":m.group(1).upper(),"product":m.group(2).strip(),"date":m.group(3)}

def product_key(p):
    return re.sub(r"[^A-Za-z0-9]+","",str(p or "")).upper()

FORM_LANG_ALIASES={
    "EN":"EN","ENGLISH":"EN","US":"EN","GB":"EN",
    "FR":"FR","FRENCH":"FR","ES":"ES","SPANISH":"ES","PT":"PT","PORTUGUESE":"PT",
    "IT":"IT","ITALIAN":"IT","JP":"JP","JA":"JP","JAPANESE":"JP","KR":"KR","KO":"KR","KOREAN":"KR",
    "RU":"RU","RUSSIAN":"RU","TR":"TR","TURKISH":"TR","DE":"DE","GERMAN":"DE",
    "NL":"NL","DUTCH":"NL","PL":"PL","POLISH":"PL","ID":"ID","INDONESIAN":"ID",
}

def custom_form_lang(value):
    raw=str(value or "").strip()
    if not raw: return None
    first=re.split(r"[,+/;|\s]+", raw)[0].strip()
    key=re.sub(r"[^A-Za-z]+","",first).upper()
    return FORM_LANG_ALIASES.get(key, key[:2] if len(key)>=2 else None)

def latest_form(forms):
    return sorted(forms or [], key=lambda x: str(x.get("date") or x.get("created_time") or ""), reverse=True)[0] if forms else None

def _form_item(item, lang=None, page_id=None):
    if not isinstance(item, dict): return None
    fid=str(item.get("id") or item.get("form_id") or "").strip()
    if not fid: return None
    out=dict(item)
    out["id"]=fid
    out["form_id"]=fid
    if lang: out["lang"]=lang
    if page_id: out["page_id"]=str(page_id)
    if "date" not in out:
        m=re.search(r"(20\d{6})", str(out.get("name") or ""))
        if m: out["date"]=m.group(1)
    return out

def normalize_forms_library(lib):
    # 兼容两种格式：
    # 1) 新格式 {by_page:{page:{by_product:{P2:{EN:[...]}}}}}
    # 2) 你手动导入的直接格式 {page:{P2:{EN:{form_id:...}}}}
    out={"by_product":{}, "by_page":{}, "raw":lib.get("raw",[]) if isinstance(lib,dict) else [], "updated_at":lib.get("updated_at","") if isinstance(lib,dict) else ""}
    def add(page_id, product, lang, item):
        it=_form_item(item, lang, page_id)
        if not it: return
        pkey=product_key(product)
        out["by_product"].setdefault(pkey,{}).setdefault(lang,[]).append(it)
        if page_id:
            out["by_page"].setdefault(str(page_id), {"by_product":{}})["by_product"].setdefault(pkey,{}).setdefault(lang,[]).append(it)
    if not isinstance(lib, dict):
        return out
    # 已有 by_product / by_page
    for product, langs in (lib.get("by_product") or {}).items():
        for lang, items in (langs or {}).items():
            if isinstance(items, dict): items=[items]
            for item in items or []: add(None, product, str(lang).upper(), item)
    for page_id, pdata in (lib.get("by_page") or {}).items():
        byp=(pdata or {}).get("by_product", {}) if isinstance(pdata, dict) else {}
        for product, langs in byp.items():
            for lang, items in (langs or {}).items():
                if isinstance(items, dict): items=[items]
                for item in items or []: add(page_id, product, str(lang).upper(), item)
    # 直接 page_id -> product -> lang -> item 格式
    for page_id, pdata in lib.items():
        if str(page_id) in {"by_product","by_page","raw","updated_at"}: continue
        if not isinstance(pdata, dict): continue
        for product, langs in pdata.items():
            if not isinstance(langs, dict): continue
            for lang, item in langs.items():
                if isinstance(item, list):
                    for one in item: add(page_id, product, str(lang).upper(), one)
                else:
                    add(page_id, product, str(lang).upper(), item)
    return out

def load_forms_library():
    return normalize_forms_library(load_json(FORMS_LIBRARY_FILE, {"by_product":{},"raw":[],"updated_at":""}))

def match_form(product, countries, merge=True, page_id=None, preferred_lang=None):
    lib=load_forms_library(); pkey=product_key(product); lang=(preferred_lang or form_lang_for_countries(countries, merge) or "EN").upper()
    by_page=lib.get("by_page",{}) or {}
    by=None
    if page_id:
        by=(by_page.get(clean_id(page_id),{}) or {}).get("by_product")
    if not by:
        by=lib.get("by_product",{})
    plist=by.get(pkey, {})
    item=latest_form(plist.get(lang, [])) or latest_form(plist.get("EN", []))
    return item, lang
def build_rows(payload):
    defaults=payload.get("defaults") or {}; items=payload.get("items") or []; split=payload.get("country_mode")=="split"; cmap=get_campaign_map(); rows=[]
    default_budget = norm(defaults.get("daily_budget"))
    needs_auto_budget = any(not norm(item.get("daily_budget")) and not default_budget for item in items)
    needs_fixed_language = any(norm(item.get("asset_language")).upper() in {"", "DEFAULT", "AUTO"} for item in items)
    budget_status = quarter_budget_status() if needs_auto_budget or needs_fixed_language else None
    for item in items:
        product=norm(item.get("product")); point=norm(item.get("selling_point")); countries=parse_countries(item.get("countries"))
        if not product or not point or not countries: raise RuntimeError("请检查产品、卖点、国家是否都已选择。")
        groups=[[c] for c in countries] if split else [countries]
        for group in groups:
            manual_budget = norm(item.get("daily_budget")) or default_budget
            row=dict(defaults); row.update({"enabled":"Y","product":product,"selling_point":point,"asset_language":norm(item.get("asset_language") or "AUTO").upper(),"countries":"|".join(group),"daily_budget":manual_budget,"dry_run":norm(item.get("dry_run") or defaults.get("dry_run") or "no"),"delay_max_seconds":"60","campaign_id":norm(item.get("campaign_id")) or norm(defaults.get("campaign_id")) or cmap.get(product,"")})
            budget_info = match_quarter_budget(budget_status or {}, product, row.get("campaign_id"), group)
            fixed_language = norm((budget_info or {}).get("asset_language")).upper()
            if fixed_language and norm(row.get("asset_language")).upper() in {"", "DEFAULT", "AUTO"}:
                row["asset_language"] = fixed_language
            if manual_budget:
                row["budget_source"] = "manual"
            else:
                if budget_info:
                    row["daily_budget"] = f"{_number(budget_info.get('daily_budget')):.2f}"
                    row["budget_source"] = "quarter"
                    row["budget_info"] = budget_info
                    row["waiting_for_min_budget"] = bool(budget_info.get("waiting_for_min_budget"))
                    row["calculated_daily_budget"] = budget_info.get("calculated_daily_budget")
                    row["budget_eligible_in_days"] = budget_info.get("budget_eligible_in_days")
                    row["budget_eligible_date"] = budget_info.get("budget_eligible_date")
                else:
                    row["budget_source"] = "missing"
            row["page_id"] = page_id_for_product(product) or row.get("page_id")
            # 官网链接自动按产品匹配：用户不用手填。只有手动填了非首页链接才保留。
            manual_url = norm(row.get("website_url"))
            row["website_url"] = resolve_product_url(product)
            if not norm(row.get("lead_form_id")):
                form, _ = match_form(product, group, merge=not split, page_id=row.get("page_id"), preferred_lang=custom_form_lang(row.get("custom_language")) if str(row.get("language_mode") or "").lower()=="custom" else None)
                if form: row["lead_form_id"] = form.get("id", "")
            current_tags = norm(row.get("url_tags"))
            auto_tags = f"utm_source=facebook&utm_medium=socialad&utm_campaign=fjd-{product.lower()}&utm_id=5224"
            if (not current_tags) or "{产品名}" in current_tags:
                row["url_tags"] = auto_tags
            else:
                row["url_tags"] = re.sub(r"(utm_campaign=)fjd-[^&]+", rf"\1fjd-{product.lower()}", current_tags, flags=re.I)
            rows.append(row)
    return rows

def preflight(payload):
    rows=build_rows(payload); split=payload.get("country_mode")=="split"; out=[]
    duplicate_map = find_duplicate_active_ads(rows)
    for row_index, row in enumerate(rows):
        countries=parse_countries(row.get("countries")); blocked_countries=[country for country in countries if country in BLOCKED_COUNTRIES]; pack=inspect_pack(row.get("product"), row.get("selling_point"), row.get("asset_language")); form, lang=match_form(row.get("product"), countries, merge=not split, page_id=row.get("page_id") or page_id_for_product(row.get("product")), preferred_lang=custom_form_lang(row.get("custom_language")) if str(row.get("language_mode") or "").lower()=="custom" else None)
        problems=[]
        if blocked_countries: problems.append(f"本季度已禁投地区：{','.join(blocked_countries)}；请下季度从预算表移除")
        if str(row.get("language_mode") or "").lower()=="custom" and not norm(row.get("custom_language")):
            problems.append("语言模式为 custom，但没有填写自定义语言")
        if norm(row.get("asset_language")).upper() in {"", "DEFAULT", "AUTO"}:
            problems.append("预算表没有匹配到素材语种；请先在季度预算表填写该产品和地区的语种")
        if not pack["exists"]: problems.append(f"缺少预算表指定语种 {row.get('asset_language') or 'DEFAULT'} 的素材目录，已禁止回退到其他语种")
        if not pack["images"]: problems.append("没有图片")
        if not pack["copy_exists"]: problems.append("没有文案")
        if not row.get("campaign_id"): problems.append("缺少 Campaign ID")
        if row.get("budget_source") == "missing": problems.append("未填写手动预算，且当前季度预算表没有匹配该产品")
        elif row.get("budget_source") == "quarter" and _number(row.get("daily_budget")) <= 0: problems.append("该产品季度预算已经用完，自动日预算为 0")
        elif row.get("budget_source") == "quarter" and row.get("waiting_for_min_budget"): problems.append(f"自动日预算 ${_number(row.get('daily_budget')):.2f} 低于 Meta 最低 ${META_MIN_DAILY_BUDGET:.2f}，暂不投放；预计 {row.get('budget_eligible_date') or '后续日期'} 可重新检测")
        elif row.get("budget_source") == "manual" and 0 < _number(row.get("daily_budget")) < META_MIN_DAILY_BUDGET: problems.append(f"手动日预算必须至少为 ${META_MIN_DAILY_BUDGET:.2f}")
        elif not norm(row.get("daily_budget")): problems.append("缺少每日预算")
        if not (row.get("lead_form_id") or form): problems.append("没有匹配表单")
        row["page_id"] = row.get("page_id") or page_id_for_product(row.get("product"))
        if not row.get("page_id"): problems.append("缺少 Page ID")
        for duplicate in duplicate_map.get(row_index, []):
            source = "ACTIVE 广告组" if duplicate.get("source") == "active" else "当前投放队列"
            countries_text = ",".join(duplicate.get("countries") or [])
            problems.append(f"重复投放：{source}“{duplicate.get('name') or duplicate.get('id')}”已覆盖 {countries_text}，且使用相同卖点")
        out.append({"product":row.get("product"),"selling_point":row.get("selling_point"),"asset_language":row.get("asset_language"),"asset_folder":pack.get("folder",""),"countries":countries,"blocked_countries":blocked_countries,"budget":row.get("daily_budget"),"budget_source":row.get("budget_source"),"budget_info":row.get("budget_info"),"waiting_for_min_budget":row.get("waiting_for_min_budget",False),"calculated_daily_budget":row.get("calculated_daily_budget"),"budget_eligible_in_days":row.get("budget_eligible_in_days",0),"budget_eligible_date":row.get("budget_eligible_date",""),"images":len(pack["images"]),"copy":pack["copy"],"form_lang":lang,"form_name":(form or {}).get("name",""),"form_id":row.get("lead_form_id") or (form or {}).get("id",""),"campaign_id":row.get("campaign_id"),"website_url":row.get("website_url"),"status":"OK" if not problems else "WARN","problems":problems})
    return out

def sync_forms(page_id):
    page_id=clean_id(page_id)
    if not page_id: raise RuntimeError("Page ID 为空。")
    token=get_token(); endpoint=f"https://graph.facebook.com/{API_VERSION}/{page_id}/leadgen_forms"
    params={"access_token":token,"fields":"id,name,status,locale,created_time","limit":100}
    raw=[]
    while True:
        r=requests.get(endpoint, params=params, timeout=120); data=r.json()
        if "error" in data: raise RuntimeError(json.dumps(data, ensure_ascii=False, indent=2))
        raw += data.get("data", [])
        nxt=data.get("paging",{}).get("next")
        if not nxt: break
        endpoint=nxt; params={}
    by={}
    for item in raw:
        parsed=parse_form_name(item.get("name"))
        if not parsed: continue
        pkey=product_key(parsed["product"]); lang=parsed["lang"]
        obj={"id":item.get("id"),"name":item.get("name"),"status":item.get("status"),"locale":item.get("locale"),"created_time":item.get("created_time"),"date":parsed["date"],"page_id":page_id}
        by.setdefault(pkey,{}).setdefault(lang,[]).append(obj)
    old_lib=load_forms_library()
    by_page=old_lib.get("by_page",{}) or {}
    by_page[page_id]={"updated_at":datetime.now().strftime("%Y-%m-%d %H:%M:%S"),"page_id":page_id,"raw":raw,"by_product":by}
    lib={"updated_at":datetime.now().strftime("%Y-%m-%d %H:%M:%S"),"page_id":page_id,"raw":raw,"by_product":by,"by_page":by_page}
    save_json(FORMS_LIBRARY_FILE, lib)
    return {"count":len(raw),"matched":sum(len(vv) for p in by.values() for vv in p.values()),"library":lib}

def token_ready_for_run():
    token_file = BASE_DIR / "token.txt"
    if token_file.exists():
        token = token_file.read_text(encoding="utf-8", errors="ignore").strip()
        if token and token not in {"PASTE_TOKEN_HERE", "把你的Meta Token粘贴到这里"}:
            return True, ""
    if os.getenv("META_ACCESS_TOKEN"):
        return True, ""
    return False, "缺少 Meta Access Token：请打开 token.txt，删除占位文字并粘贴有效 token 后保存。"

def run_ads_script():
    global last_run
    ok, msg = token_ready_for_run()
    if not ok:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        run_id = datetime.now().strftime("%Y%m%d%H%M%S%f")
        last_run = {"running": False, "started_at": now, "finished_at": now, "returncode": 1, "ok": False, "output": msg, "log_path": "", "run_id": run_id, "paused": True, "run_error": {"message": msg, "skipped_count": 0}, "summary": {"planned": 0, "success": 0, "failed": 1, "skipped": 0, "dry_run": 0}}
        return False
    if last_run.get("running"):
        return False

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    started = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    run_id = datetime.now().strftime("%Y%m%d%H%M%S%f")
    log_path = LOG_DIR / f"run_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    last_run = {
        "running": True,
        "started_at": started,
        "finished_at": "",
        "returncode": None,
        "output": f"开始运行 run.py\n日志文件：{log_path}\n",
        "log_path": str(log_path),
        "run_id": run_id,
        "paused": False,
        "run_error": None,
        "summary": {"planned": 0, "success": 0, "failed": 0, "skipped": 0, "dry_run": 0},
    }

    def append_output(text):
        global last_run
        old = last_run.get("output", "")
        last_run["output"] = (old + text)[-50000:]

    def target():
        global last_run
        try:
            cmd = [sys.executable, "-u", str(RUN_SCRIPT)]
            with log_path.open("w", encoding="utf-8", errors="replace") as log_file:
                log_file.write(f"Command: {' '.join(cmd)}\n")
                log_file.write(f"Working directory: {BASE_DIR}\n\n")
                log_file.flush()
                env = os.environ.copy()
                env["PYTHONIOENCODING"] = "utf-8"
                env["PYTHONUTF8"] = "1"
                proc = subprocess.Popen(
                    cmd,
                    cwd=str(BASE_DIR),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    bufsize=1,
                    env=env,
                )
                for line in proc.stdout or []:
                    log_file.write(line)
                    log_file.flush()
                    if line.startswith("__ADS_RUN_START__"):
                        try:
                            start_payload = json.loads(line[len("__ADS_RUN_START__"):].strip())
                            last_run["summary"]["planned"] = int(start_payload.get("planned") or 0)
                        except Exception:
                            pass
                    elif line.startswith("__ADS_RUN_ITEM__"):
                        try:
                            item_payload = json.loads(line[len("__ADS_RUN_ITEM__"):].strip())
                            item_status = str(item_payload.get("status") or "")
                            if item_status in last_run["summary"]:
                                last_run["summary"][item_status] += 1
                        except Exception:
                            pass
                    elif line.startswith("__ADS_RUN_ERROR__"):
                        try:
                            error_payload = json.loads(line[len("__ADS_RUN_ERROR__"):].strip())
                        except Exception:
                            error_payload = {"message": line.strip(), "skipped_count": 0}
                        last_run.update({"paused": True, "run_error": error_payload})
                        last_run["summary"]["skipped"] = int(error_payload.get("skipped_count") or 0)
                    else:
                        append_output(line)
                rc = proc.wait()
            if rc != 0 and not last_run.get("run_error"):
                output_lines = [line.strip() for line in str(last_run.get("output") or "").splitlines() if line.strip()]
                last_run.update({"paused": True, "run_error": {"message": output_lines[-1] if output_lines else "投放脚本运行失败，请查看日志。", "skipped_count": 0}})
            last_run.update({
                "running": False,
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "returncode": rc,
                "ok": rc == 0,
                "log_path": str(log_path),
            })
        except Exception:
            err = traceback.format_exc()
            try:
                log_path.write_text(err, encoding="utf-8")
            except Exception:
                pass
            last_run.update({
                "running": False,
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "returncode": -1,
                "ok": False,
                "paused": True,
                "run_error": {"message": str(err).strip().splitlines()[-1] if str(err).strip() else "投放脚本启动失败。", "skipped_count": 0},
                "output": (last_run.get("output", "") + "\n" + err)[-50000:],
                "log_path": str(log_path),
            })

    threading.Thread(target=target, daemon=True).start()
    return True


def lead_products():
    """Return the product names understood by lead_sync.py's forms library."""
    data = load_json(FORMS_LIBRARY_FILE, {})
    found = {}
    if isinstance(data, dict):
        for products in data.values():
            if not isinstance(products, dict):
                continue
            for product in products:
                name = norm(product)
                if name:
                    found.setdefault(re.sub(r"[\s_-]+", "", name).lower(), name)
    return sorted(found.values(), key=str.lower)


def list_lead_files():
    LEAD_SYNC_DIR.mkdir(parents=True, exist_ok=True)
    items = []
    for path in sorted(LEAD_SYNC_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
        stat = path.stat()
        items.append({
            "name": path.name,
            "size_kb": round(stat.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            "download_url": "/api/leads/download?name=" + path.name,
        })
    return {"count": len(items), "items": items[:50]}


def start_lead_sync(payload):
    global last_lead_sync
    if last_lead_sync.get("running"):
        return False
    ok, msg = token_ready_for_run()
    if not ok:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        last_lead_sync = {"running": False, "started_at": now, "finished_at": now, "returncode": 1, "ok": False, "output": msg, "output_file": ""}
        return False
    if not LEAD_SYNC_SCRIPT.exists():
        raise RuntimeError(f"找不到客户同步脚本：{LEAD_SYNC_SCRIPT}")

    available = lead_products()
    available_by_key = {re.sub(r"[\s_-]+", "", x).lower(): x for x in available}
    requested = payload.get("products") or []
    if not isinstance(requested, list):
        raise RuntimeError("products 必须是产品名称列表。")
    selected = []
    for raw in requested:
        key = re.sub(r"[\s_-]+", "", norm(raw)).lower()
        if key not in available_by_key:
            raise RuntimeError(f"产品未在表单库中找到：{raw}")
        if available_by_key[key] not in selected:
            selected.append(available_by_key[key])

    start_date = norm(payload.get("start_date"))
    end_date = norm(payload.get("end_date"))
    date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    for label, value in (("开始日期", start_date), ("结束日期", end_date)):
        if value and not date_pattern.match(value):
            raise RuntimeError(f"{label}格式必须是 YYYY-MM-DD。")
        if value:
            datetime.strptime(value, "%Y-%m-%d")
    if start_date and end_date and start_date > end_date:
        raise RuntimeError("开始日期不能晚于结束日期。")

    cmd = [sys.executable, "-u", str(LEAD_SYNC_SCRIPT)]
    if selected:
        cmd += ["--products", ",".join(selected)]
    if start_date:
        cmd += ["--start-date", start_date]
    if end_date:
        cmd += ["--end-date", end_date]
    if bool(payload.get("full")):
        cmd.append("--full")

    before = {p.name for p in LEAD_SYNC_DIR.glob("*.xlsx")}
    started = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    last_lead_sync = {
        "running": True, "started_at": started, "finished_at": "", "returncode": None,
        "ok": None, "output": "开始同步客户信息...\n", "output_file": "",
        "mode": "指定日期导出" if (start_date or end_date) else ("全量同步" if payload.get("full") else "增量同步"),
    }

    def append_output(text):
        global last_lead_sync
        last_lead_sync["output"] = (last_lead_sync.get("output", "") + text)[-50000:]

    def target():
        global last_lead_sync
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"] = "1"
            proc = subprocess.Popen(
                cmd, cwd=str(BASE_DIR), stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, encoding="utf-8", errors="replace", bufsize=1, env=env,
            )
            for line in proc.stdout or []:
                append_output(line)
            rc = proc.wait()
            new_files = [p for p in LEAD_SYNC_DIR.glob("*.xlsx") if p.name not in before]
            newest = max(new_files, key=lambda p: p.stat().st_mtime) if new_files else None
            last_lead_sync.update({
                "running": False, "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "returncode": rc, "ok": rc == 0, "output_file": newest.name if newest else "",
            })
        except Exception:
            err = traceback.format_exc()
            last_lead_sync.update({
                "running": False, "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "returncode": -1, "ok": False, "output": (last_lead_sync.get("output", "") + "\n" + err)[-50000:],
            })

    threading.Thread(target=target, daemon=True).start()
    return True


def send_lead_file(handler, name):
    safe_name = Path(unquote(name or "")).name
    path = (LEAD_SYNC_DIR / safe_name).resolve()
    if not safe_name or path.parent != LEAD_SYNC_DIR.resolve() or path.suffix.lower() != ".xlsx" or not path.is_file():
        return jresp(handler, {"error": "客户信息文件不存在。"}, 404)
    data = path.read_bytes()
    from urllib.parse import quote
    handler.send_response(200)
    handler.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    handler.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(path.name)}")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def _number(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _action_value(actions, preferred_types):
    values = {}
    for item in actions or []:
        if isinstance(item, dict) and item.get("action_type"):
            values[str(item["action_type"])] = _number(item.get("value"))
    for action_type in preferred_types:
        if action_type in values:
            return values[action_type]
    return 0.0


def _summarize_insight_period(rows, start_date, days):
    daily_by_date = {}
    for i in range(days):
        date_text = (start_date + timedelta(days=i)).isoformat()
        daily_by_date[date_text] = {"date": date_text, "spend": 0.0, "impressions": 0, "reach": 0, "clicks": 0, "leads": 0.0}
    for row in rows:
        day = daily_by_date.setdefault(row["date"], {"date": row["date"], "spend": 0.0, "impressions": 0, "reach": 0, "clicks": 0, "leads": 0.0})
        for metric in ("spend", "impressions", "reach", "clicks", "leads"):
            day[metric] += row[metric]
    daily = []
    for day in daily_by_date.values():
        day["spend"], day["leads"] = round(day["spend"], 2), round(day["leads"], 2)
        day["ctr"] = round(day["clicks"] / day["impressions"] * 100, 3) if day["impressions"] else 0
        day["cpc"] = round(day["spend"] / day["clicks"], 3) if day["clicks"] else 0
        day["cpm"] = round(day["spend"] / day["impressions"] * 1000, 3) if day["impressions"] else 0
        day["cost_per_lead"] = round(day["spend"] / day["leads"], 2) if day["leads"] else 0
        daily.append(day)
    daily.sort(key=lambda x: x["date"])
    totals = {
        "spend": round(sum(x["spend"] for x in rows), 2), "impressions": sum(x["impressions"] for x in rows),
        "clicks": sum(x["clicks"] for x in rows), "leads": round(sum(x["leads"] for x in rows), 2),
    }
    totals["ctr"] = round(totals["clicks"] / totals["impressions"] * 100, 3) if totals["impressions"] else 0
    totals["cost_per_lead"] = round(totals["spend"] / totals["leads"], 2) if totals["leads"] else 0
    return daily, totals


def _insights_with_cache_meta(data, cached_at, source):
    """Attach cache information without polluting the saved insight payload."""
    now = datetime.now()
    age_seconds = max(0, int((now - cached_at).total_seconds()))
    result = dict(data)
    result["cache"] = {
        "source": source,
        "cached_at": cached_at.isoformat(timespec="seconds"),
        "age_minutes": age_seconds // 60,
        "ttl_minutes": INSIGHTS_CACHE_TTL_SECONDS // 60,
        "expires_at": (cached_at + timedelta(seconds=INSIGHTS_CACHE_TTL_SECONDS)).isoformat(timespec="seconds"),
    }
    return result


def fetch_adset_insights(days=7, force=False):
    # Prevent two tabs from refreshing the same Meta dataset at the same time.
    with insights_cache_lock:
        try:
            return _fetch_adset_insights(days=days, force=force)
        except Exception:
            # If Meta is temporarily unavailable, keep the dashboard usable with
            # the last disk snapshot instead of spending retries or showing no data.
            account_id = clean_id(get_defaults().get("ad_account_id"))
            key = f"{account_id}:{max(1, min(int(days or 7), 90))}:country-v1"
            disk_cache = load_json(INSIGHTS_CACHE_FILE, {})
            entry = (disk_cache.get("entries") or {}).get(key) or {}
            if entry.get("data") and entry.get("cached_at"):
                try:
                    cached_at = datetime.fromisoformat(str(entry["cached_at"]))
                    result = _insights_with_cache_meta(entry["data"], cached_at, "stale_disk")
                    result["cache"]["stale"] = True
                    return result
                except (TypeError, ValueError):
                    pass
            raise


def _fetch_adset_insights(days=7, force=False):
    """Fetch daily ad-set performance, reusing a one-hour memory/disk cache."""
    global insights_cache
    days = max(1, min(int(days or 7), 90))
    account_id = clean_id(get_defaults().get("ad_account_id"))
    if not account_id:
        raise RuntimeError("Ad Account ID 为空，请先在高级设置或 plan.xlsx 中填写。")
    key = f"{account_id}:{days}:country-v1"
    now = datetime.now()
    loaded_at = insights_cache.get("loaded_at")
    if not force and insights_cache.get("key") == key and insights_cache.get("data") and loaded_at and (now - loaded_at).total_seconds() < INSIGHTS_CACHE_TTL_SECONDS:
        return _insights_with_cache_meta(insights_cache["data"], loaded_at, "memory")

    disk_cache = load_json(INSIGHTS_CACHE_FILE, {})
    if disk_cache.get("schema_version") != INSIGHTS_CACHE_SCHEMA_VERSION:
        disk_cache = {"schema_version": INSIGHTS_CACHE_SCHEMA_VERSION, "entries": {}}
    entry = (disk_cache.get("entries") or {}).get(key) or {}
    try:
        disk_cached_at = datetime.fromisoformat(str(entry.get("cached_at") or ""))
    except (TypeError, ValueError):
        disk_cached_at = None
    if not force and entry.get("data") and disk_cached_at and (now - disk_cached_at).total_seconds() < INSIGHTS_CACHE_TTL_SECONDS:
        insights_cache = {"key": key, "loaded_at": disk_cached_at, "data": entry["data"]}
        return _insights_with_cache_meta(entry["data"], disk_cached_at, "disk")

    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=days - 1)
    previous_end_date = start_date - timedelta(days=1)
    previous_start_date = previous_end_date - timedelta(days=days - 1)
    account = account_id if account_id.startswith("act_") else "act_" + account_id
    fields = [
        "date_start", "date_stop", "campaign_id", "campaign_name", "adset_id", "adset_name",
        "spend", "impressions", "reach", "clicks", "ctr", "cpc", "cpm", "actions",
    ]
    params = {
        "access_token": get_token(), "level": "adset", "fields": ",".join(fields), "time_increment": 1,
        "time_range": json.dumps({"since": previous_start_date.isoformat(), "until": end_date.isoformat()}),
        "breakdowns": "country", "limit": 500,
    }
    url = f"https://graph.facebook.com/{API_VERSION}/{account}/insights"
    raw_rows, page_count = [], 0
    while url and page_count < 100:
        response = requests.get(url, params=params if page_count == 0 else None, timeout=120)
        try:
            payload = response.json()
        except Exception:
            raise RuntimeError("Meta Insights 返回的不是 JSON：" + response.text[:500])
        if payload.get("error"):
            err = payload["error"]
            message = err.get("message") if isinstance(err, dict) else str(err)
            raise RuntimeError("读取 Meta 广告数据失败：" + str(message))
        raw_rows.extend(payload.get("data") or [])
        url = (payload.get("paging") or {}).get("next")
        params = None
        page_count += 1

    all_rows = []
    for raw in raw_rows:
        spend = _number(raw.get("spend"))
        leads = _action_value(raw.get("actions"), [
            "onsite_conversion.lead_grouped", "lead", "onsite_conversion.lead", "offsite_conversion.fb_pixel_lead",
        ])
        all_rows.append({
            "date": str(raw.get("date_start") or ""), "country": str(raw.get("country") or "UNKNOWN"),
            "campaign_id": str(raw.get("campaign_id") or ""),
            "campaign_name": str(raw.get("campaign_name") or "未命名 Campaign"), "adset_id": str(raw.get("adset_id") or ""),
            "adset_name": str(raw.get("adset_name") or "未命名广告组"), "spend": round(spend, 2),
            "impressions": int(_number(raw.get("impressions"))), "reach": int(_number(raw.get("reach"))),
            "clicks": int(_number(raw.get("clicks"))), "ctr": round(_number(raw.get("ctr")), 3),
            "cpc": round(_number(raw.get("cpc")), 3), "cpm": round(_number(raw.get("cpm")), 3),
            "leads": round(leads, 2), "cost_per_lead": round(spend / leads, 2) if leads else 0,
        })
    all_rows.sort(key=lambda x: (x["date"], x["campaign_name"].lower(), x["adset_name"].lower()))
    rows = [row for row in all_rows if row["date"] >= start_date.isoformat()]
    previous_rows = [row for row in all_rows if previous_start_date.isoformat() <= row["date"] <= previous_end_date.isoformat()]
    daily, totals = _summarize_insight_period(rows, start_date, days)
    previous_daily, previous_totals = _summarize_insight_period(previous_rows, previous_start_date, days)
    result = {
        "account_id": account_id, "days": days, "since": start_date.isoformat(), "until": end_date.isoformat(),
        "previous_since": previous_start_date.isoformat(), "previous_until": previous_end_date.isoformat(),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "rows": rows, "daily": daily,
        "previous_rows": previous_rows, "previous_daily": previous_daily, "previous_totals": previous_totals,
        "totals": totals, "row_count": len(rows), "previous_row_count": len(previous_rows),
    }
    cached_at = datetime.now()
    insights_cache = {"key": key, "loaded_at": cached_at, "data": result}
    entries = disk_cache.setdefault("entries", {})
    entries[key] = {"cached_at": cached_at.isoformat(timespec="seconds"), "data": result}
    save_json(INSIGHTS_CACHE_FILE, disk_cache)
    return _insights_with_cache_meta(result, cached_at, "meta_api")


def load_automation_rules():
    defaults = {
        "enabled": False, "check_interval_hours": 6,
        "warning_over_baseline_pct": 25, "close_over_baseline_pct": 50,
        "no_lead_days": 3, "current_window_days": 7,
        "spend_spike_enabled": True, "spend_spike_multiplier": 1.5, "spend_spike_min_spend": 20,
        "scale_enabled": True, "scale_min_leads": 5, "scale_cpl_better_pct": 20,
        "pacing_enabled": True, "budget_adjust_pct": 10, "budget_cooldown_hours": 24,
        "no_delivery_enabled": True, "no_delivery_hours": 24,
        "config_checks_enabled": True, "execution_mode": "approval",
        "new_ad_protection_hours": 72, "decision_delay_days": 1,
    }
    data = load_json(AUTOMATION_RULES_FILE, {})
    if isinstance(data, dict):
        defaults.update({k: data[k] for k in defaults if k in data})
    return defaults


def save_automation_rules(payload):
    rules = load_automation_rules()
    if "enabled" in payload:
        rules["enabled"] = bool(payload.get("enabled"))
    for key, low, high in (("check_interval_hours", 1, 24), ("warning_over_baseline_pct", 1, 500), ("close_over_baseline_pct", 1, 500), ("no_lead_days", 1, 30), ("current_window_days", 1, 30)):
        if key in payload:
            rules[key] = max(low, min(int(payload[key]), high))
    for key in ("spend_spike_enabled", "scale_enabled", "pacing_enabled", "no_delivery_enabled", "config_checks_enabled"):
        if key in payload:
            rules[key] = bool(payload.get(key))
    if "execution_mode" in payload:
        mode = str(payload.get("execution_mode") or "approval").lower()
        if mode not in {"notify", "approval", "auto"}:
            raise RuntimeError("自动化模式必须是仅提醒、待确认或全自动。")
        rules["execution_mode"] = mode
    for key, low, high in (("spend_spike_multiplier", 1.1, 5), ("spend_spike_min_spend", 1, 10000), ("scale_min_leads", 1, 1000), ("scale_cpl_better_pct", 1, 80), ("budget_adjust_pct", 1, 10), ("budget_cooldown_hours", 1, 168), ("no_delivery_hours", 1, 168), ("new_ad_protection_hours", 24, 168), ("decision_delay_days", 1, 3)):
        if key in payload:
            rules[key] = max(low, min(float(payload[key]), high))
    if rules["close_over_baseline_pct"] <= rules["warning_over_baseline_pct"]:
        raise RuntimeError("自动关闭阈值必须高于更换素材提醒阈值。")
    save_json(AUTOMATION_RULES_FILE, rules)
    return rules


def _graph_get_all(url, params):
    rows, page = [], 0
    while url and page < 100:
        response = requests.get(url, params=params if page == 0 else None, timeout=120)
        try:
            data = response.json()
        except Exception:
            raise RuntimeError("Meta 返回的不是 JSON：" + response.text[:500])
        if data.get("error"):
            error = data["error"]
            raise RuntimeError("Meta API 错误：" + str(error.get("message") if isinstance(error, dict) else error))
        rows.extend(data.get("data") or [])
        url = (data.get("paging") or {}).get("next")
        params = None
        page += 1
    return rows


def _adset_status_is_active(item):
    status = str(item.get("status") or item.get("effective_status") or "").upper()
    effective = str(item.get("effective_status") or status).upper()
    if status != "ACTIVE" or effective != "ACTIVE":
        return False
    campaign = item.get("campaign") or {}
    if isinstance(campaign, dict):
        campaign_status = str(campaign.get("status") or campaign.get("effective_status") or "ACTIVE").upper()
        campaign_effective = str(campaign.get("effective_status") or campaign_status).upper()
        if campaign_status != "ACTIVE" or campaign_effective != "ACTIVE":
            return False
    return True


def _save_automation_adset_status_cache(adsets, source="meta_api"):
    items = {}
    for item in adsets or []:
        adset_id = str(item.get("id") or "")
        if not adset_id:
            continue
        campaign = item.get("campaign") or {}
        items[adset_id] = {
            "id": adset_id, "status": item.get("status", ""), "effective_status": item.get("effective_status", ""),
            "campaign": campaign if isinstance(campaign, dict) else {}, "active": _adset_status_is_active(item),
        }
    payload = {"cached_at": datetime.now().isoformat(timespec="seconds"), "source": source, "items": items}
    save_json(AUTOMATION_ADSET_STATUS_CACHE_FILE, payload)
    return payload


def _automation_adset_status_map(force=False):
    cached = load_json(AUTOMATION_ADSET_STATUS_CACHE_FILE, {})
    cached_at = None
    try:
        cached_at = datetime.fromisoformat(str(cached.get("cached_at") or ""))
    except Exception:
        pass
    if not force and cached.get("items") and cached_at and (datetime.now() - cached_at).total_seconds() < AUTOMATION_STATUS_CACHE_TTL_SECONDS:
        return cached
    try:
        account_id = clean_id(get_defaults().get("ad_account_id"))
        account = account_id if account_id.startswith("act_") else "act_" + account_id
        adsets = _graph_get_all(f"https://graph.facebook.com/{API_VERSION}/{account}/adsets", {
            "access_token": get_token(), "fields": "id,status,effective_status,campaign{id,name,status,effective_status}", "limit": 500,
        })
        return _save_automation_adset_status_cache(adsets)
    except Exception:
        if cached.get("items"):
            cached["source"] = "stale_cache"
            return cached
        return {"cached_at": "", "source": "unavailable", "items": {}}


def _update_cached_adset_status(adset_id, status):
    cached = load_json(AUTOMATION_ADSET_STATUS_CACHE_FILE, {"items": {}})
    if not isinstance(cached, dict):
        cached = {"items": {}}
    items = cached.setdefault("items", {})
    item = items.setdefault(str(adset_id), {"id": str(adset_id), "campaign": {}})
    item["status"] = str(status).upper()
    item["effective_status"] = str(status).upper()
    item["active"] = str(status).upper() == "ACTIVE" and _adset_status_is_active(item)
    cached["cached_at"] = datetime.now().isoformat(timespec="seconds")
    cached["source"] = "local_action"
    save_json(AUTOMATION_ADSET_STATUS_CACHE_FILE, cached)


def _insight_rows(account, since, until, level, breakdowns="", time_increment=None):
    params = {
        "access_token": get_token(), "level": level,
        "fields": "adset_id,adset_name,campaign_id,campaign_name,spend,actions",
        "time_range": json.dumps({"since": since.isoformat(), "until": until.isoformat()}), "limit": 500,
    }
    if breakdowns:
        params["breakdowns"] = breakdowns
    if time_increment:
        params["time_increment"] = time_increment
    return _graph_get_all(f"https://graph.facebook.com/{API_VERSION}/{account}/insights", params)


def _infer_product(*names):
    haystack = re.sub(r"[^A-Za-z0-9]+", "", " ".join(str(x or "") for x in names)).upper()
    products = sorted(scan_asset_packs().keys(), key=lambda x: len(re.sub(r"[^A-Za-z0-9]+", "", x)), reverse=True)
    for product in products:
        key = re.sub(r"[^A-Za-z0-9]+", "", product).upper()
        if key and key in haystack:
            return product
    return ""


def _infer_selling_point(product, *names, assets=None):
    catalog = assets if isinstance(assets, dict) else scan_asset_packs()
    points = [str(item.get("name") or "").strip() for item in catalog.get(str(product or ""), [])]
    haystack = " ".join(str(value or "") for value in names).casefold()
    compact_haystack = re.sub(r"[\s_\-]+", "", haystack)
    for point in sorted([point for point in points if point], key=len, reverse=True):
        if point.casefold() in haystack or re.sub(r"[\s_\-]+", "", point.casefold()) in compact_haystack:
            return point
    return ""


def _set_adset_status(adset_id, status):
    response = requests.post(f"https://graph.facebook.com/{API_VERSION}/{clean_id(adset_id)}", data={"access_token": get_token(), "status": str(status).upper()}, timeout=60)
    data = response.json()
    if data.get("error"):
        error = data["error"]
        raise RuntimeError(str(error.get("message") if isinstance(error, dict) else error))
    success = bool(data.get("success", True))
    if success:
        _update_cached_adset_status(adset_id, status)
    return success


def _pause_adset(adset_id):
    return _set_adset_status(adset_id, "PAUSED")


def _update_adset_budget(adset_id, daily_budget):
    cents = max(1, int(round(float(daily_budget) * 100)))
    response = requests.post(
        f"https://graph.facebook.com/{API_VERSION}/{clean_id(adset_id)}",
        data={"access_token": get_token(), "daily_budget": cents}, timeout=60,
    )
    data = response.json()
    if data.get("error"):
        error = data["error"]
        raise RuntimeError(str(error.get("message") if isinstance(error, dict) else error))
    return bool(data.get("success", True))


def _automation_action_state():
    data = load_json(AUTOMATION_ACTION_STATE_FILE, {})
    return data if isinstance(data, dict) else {}


def _budget_cooldown(adset_id, hours):
    item = _automation_action_state().get(str(adset_id)) or {}
    text = str(item.get("last_budget_change") or "")
    try:
        changed = datetime.fromisoformat(text)
        remaining = float(hours) * 3600 - (datetime.now() - changed).total_seconds()
        return max(0, remaining)
    except Exception:
        return 0


def _record_budget_change(adset_id, action, old_budget, new_budget):
    state = _automation_action_state()
    state[str(adset_id)] = {
        "last_budget_change": datetime.now().isoformat(timespec="seconds"),
        "action": action, "old_budget": old_budget, "new_budget": new_budget,
    }
    save_json(AUTOMATION_ACTION_STATE_FILE, state)


def _adset_age_hours(created_time):
    text = str(created_time or "").strip()
    if not text:
        return 999999
    try:
        created = datetime.fromisoformat(text.replace("Z", "+00:00"))
        now = datetime.now(created.tzinfo) if created.tzinfo else datetime.now()
        return max(0, (now - created).total_seconds() / 3600)
    except Exception:
        return 999999


def _landing_page_health(url):
    if not url:
        return False, "未配置落地页"
    try:
        response = requests.head(url, allow_redirects=True, timeout=8, headers={"User-Agent": "ADS-Automation-Healthcheck/1.0"})
        if response.status_code in {403, 405}:
            response = requests.get(url, allow_redirects=True, timeout=8, stream=True, headers={"User-Agent": "ADS-Automation-Healthcheck/1.0"})
        if response.status_code >= 400:
            return False, f"落地页返回 HTTP {response.status_code}"
        return True, ""
    except Exception as exc:
        return False, f"落地页无法访问：{str(exc)[:120]}"


def _execute_automation_item(item, rules):
    action = str(item.get("action") or "")
    item["executed"] = False
    item["error"] = ""
    try:
        if action.startswith("CLOSE_"):
            item["previous_status"] = "ACTIVE"
            item["paused"] = _pause_adset(item["adset_id"])
            item["executed"] = bool(item["paused"])
        elif action in {"SCALE_UP", "PACE_UP", "PACE_DOWN"}:
            remaining = _budget_cooldown(item["adset_id"], rules.get("budget_cooldown_hours", 24))
            if remaining > 0:
                item["error"] = f"预算调整仍在冷却期，约 {remaining / 3600:.1f} 小时后可再次执行"
            else:
                item["budget_changed"] = _update_adset_budget(item["adset_id"], item["proposed_budget"])
                item["executed"] = bool(item["budget_changed"])
                if item["budget_changed"]:
                    _record_budget_change(item["adset_id"], action, item.get("daily_budget", 0), item["proposed_budget"])
        if item.get("executed"):
            item["executed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            item["action_id"] = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}-{item.get('adset_id')}-{action}"
            item["rollback_available"] = True
    except Exception as exc:
        item["error"] = str(exc)
    return item


def _safe_external_error(exc):
    message = str(exc or "未知错误")
    message = re.sub(r"(access_token=)[^&\s]+", r"\1[已隐藏]", message, flags=re.I)
    return message[:800]


def evaluate_automation_rules(execute=False):
    global automation_state
    if automation_state.get("running"):
        raise RuntimeError("规则检查正在运行，请稍后再试。")
    automation_state.update({"running": True, "last_started": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "last_error": ""})
    try:
        rules = load_automation_rules()
        today = datetime.now().date()
        decision_until = today - timedelta(days=int(rules.get("decision_delay_days", 1)))
        current_start = decision_until - timedelta(days=int(rules["current_window_days"]) - 1)
        no_lead_start = decision_until - timedelta(days=int(rules["no_lead_days"]) - 1)
        daily_start = today - timedelta(days=7)
        month_start = today.replace(day=1)
        previous_end = month_start - timedelta(days=1)
        previous_start = previous_end.replace(day=1)
        account_id = clean_id(get_defaults().get("ad_account_id"))
        account = account_id if account_id.startswith("act_") else "act_" + account_id
        token = get_token()
        adsets = _graph_get_all(f"https://graph.facebook.com/{API_VERSION}/{account}/adsets", {
            "access_token": token, "fields": "id,name,status,effective_status,campaign_id,daily_budget,lifetime_budget,targeting,created_time,campaign{id,name,status,effective_status}", "limit": 500,
        })
        _save_automation_adset_status_cache(adsets)
        active = {str(x.get("id")): x for x in adsets if _adset_status_is_active(x)}
        baseline_raw = _insight_rows(account, previous_start, previous_end, "account", "country")
        current_raw = _insight_rows(account, current_start, decision_until, "adset", "country")
        recent_raw = _insight_rows(account, no_lead_start, decision_until, "adset")
        daily_raw = _insight_rows(account, daily_start, today, "adset", time_increment=1)
        baseline = {}
        for row in baseline_raw:
            country = str(row.get("country") or "UNKNOWN")
            spend, leads = _number(row.get("spend")), _action_value(row.get("actions"), ["onsite_conversion.lead_grouped", "lead", "onsite_conversion.lead", "offsite_conversion.fb_pixel_lead"])
            if leads:
                baseline[country] = {"spend": spend, "leads": leads, "cpl": spend / leads}
        current_by_adset = {}
        for row in current_raw:
            adset_id = str(row.get("adset_id") or "")
            if adset_id not in active:
                continue
            spend, leads = _number(row.get("spend")), _action_value(row.get("actions"), ["onsite_conversion.lead_grouped", "lead", "onsite_conversion.lead", "offsite_conversion.fb_pixel_lead"])
            current_by_adset.setdefault(adset_id, []).append({"country": str(row.get("country") or "UNKNOWN"), "spend": spend, "leads": leads, "cpl": spend / leads if leads else 0})
        recent = {}
        for row in recent_raw:
            adset_id = str(row.get("adset_id") or "")
            recent[adset_id] = {"spend": _number(row.get("spend")), "leads": _action_value(row.get("actions"), ["onsite_conversion.lead_grouped", "lead", "onsite_conversion.lead", "offsite_conversion.fb_pixel_lead"])}
        daily = {}
        for row in daily_raw:
            adset_id = str(row.get("adset_id") or "")
            if adset_id not in active:
                continue
            day = str(row.get("date_start") or row.get("date_stop") or "")[:10]
            daily.setdefault(adset_id, {})[day] = {
                "spend": _number(row.get("spend")),
                "leads": _action_value(row.get("actions"), ["onsite_conversion.lead_grouped", "lead", "onsite_conversion.lead", "offsite_conversion.fb_pixel_lead"]),
            }
        try:
            budget_status = quarter_budget_status(force=False)
        except Exception:
            budget_status = {"items": []}
        assets = scan_asset_packs()
        landing_health = {}
        items, closed, budget_changed = [], [], []
        execution_halted, halt_reason = False, ""
        for adset_id, adset in active.items():
            targeting = adset.get("targeting") or {}
            countries = ((targeting.get("geo_locations") or {}).get("countries") or [])
            campaign = adset.get("campaign") or {}
            campaign_name = str(campaign.get("name") or "") if isinstance(campaign, dict) else ""
            regions = current_by_adset.get(adset_id, [])
            total_spend = sum(x["spend"] for x in regions)
            total_leads = sum(x["leads"] for x in regions)
            worst = None
            for region in regions:
                base = baseline.get(region["country"])
                if base and region["leads"] and base["cpl"]:
                    over = (region["cpl"] / base["cpl"] - 1) * 100
                    candidate = {**region, "baseline_cpl": round(base["cpl"], 2), "over_pct": round(over, 1)}
                    if worst is None or candidate["over_pct"] > worst["over_pct"]:
                        worst = candidate
            recent_value = recent.get(adset_id, {"spend": 0, "leads": 0})
            age_hours = _adset_age_hours(adset.get("created_time"))
            mature = age_hours >= float(rules["no_lead_days"]) * 24
            protected = age_hours < float(rules.get("new_ad_protection_hours", 72))
            product = _infer_product(campaign_name, adset.get("name"))
            campaign_id = str(adset.get("campaign_id") or campaign.get("id") or "")
            budget_raw = adset.get("daily_budget") or 0
            daily_budget = round(_number(budget_raw) / 100, 2) if budget_raw else 0
            base_values = [baseline[c]["cpl"] for c in countries if c in baseline and baseline[c].get("cpl")]
            baseline_cpl = sum(base_values) / len(base_values) if base_values else 0
            day_values = daily.get(adset_id, {})
            today_value = day_values.get(today.isoformat(), {"spend": 0, "leads": 0})
            previous_spend = [day_values.get((today - timedelta(days=offset)).isoformat(), {}).get("spend", 0) for offset in range(1, 8)]
            average_7d_spend = sum(previous_spend) / 7
            last_2d_spend = sum(day_values.get((today - timedelta(days=offset)).isoformat(), {}).get("spend", 0) for offset in range(0, 2))
            action, reason = "OK", "表现未触发规则"
            proposed_budget = 0
            cooldown_seconds = _budget_cooldown(adset_id, rules.get("budget_cooldown_hours", 24))
            config_warnings = []
            budget_info = match_quarter_budget(budget_status, product, campaign_id, countries) if product else None
            if rules.get("config_checks_enabled"):
                if not product:
                    config_warnings.append("未从广告组或 Campaign 名称识别出产品")
                else:
                    url = resolve_product_url(product)
                    if product not in landing_health:
                        landing_health[product] = _landing_page_health(url)
                    if not landing_health[product][0]:
                        config_warnings.append(landing_health[product][1])
                    try:
                        form, form_lang = match_form(product, countries, merge=True, page_id=page_id_for_product(product), preferred_lang=(budget_info or {}).get("asset_language"))
                        if not form:
                            config_warnings.append(f"没有匹配的 {form_lang} 留资表单")
                    except Exception as exc:
                        config_warnings.append(f"表单配置检查失败：{str(exc)[:100]}")
                    fixed_language = str((budget_info or {}).get("asset_language") or "").upper()
                    if fixed_language:
                        available_languages = {str(lang).upper() for point in assets.get(product, []) for lang in point.get("languages", [])}
                        if fixed_language not in available_languages:
                            config_warnings.append(f"季度预算固定 {fixed_language}，但素材包没有该语种")
            if protected:
                action, reason = "INFO_PROTECTED", f"新广告保护中：已运行 {age_hours:.0f} 小时，满 {float(rules.get('new_ad_protection_hours', 72)):.0f} 小时后再参与停投、扩量和预算调速"
                if age_hours >= float(rules.get("no_delivery_hours", 24)) and last_2d_spend <= 0:
                    reason += "；当前仍无消耗，可先人工检查审核、受众和支付状态"
                if config_warnings:
                    reason += "；配置提醒：" + "；".join(config_warnings)
            elif mature and recent_value["spend"] > 0 and recent_value["leads"] == 0:
                action, reason = "CLOSE_NO_LEADS", f"连续 {rules['no_lead_days']} 天有花费但 0 Leads"
            elif worst and worst["over_pct"] > float(rules["close_over_baseline_pct"]):
                action, reason = "CLOSE_HIGH_CPL", f"{worst['country']} CPL 比上月地区基准高 {worst['over_pct']:.1f}%"
            elif rules.get("spend_spike_enabled") and today_value["leads"] == 0 and today_value["spend"] >= float(rules["spend_spike_min_spend"]) and average_7d_spend > 0 and today_value["spend"] > average_7d_spend * float(rules["spend_spike_multiplier"]):
                action, reason = "CLOSE_SPEND_SPIKE", f"今日花费 ${today_value['spend']:.2f}，超过前 7 天日均 ${average_7d_spend:.2f} 的 {float(rules['spend_spike_multiplier']):.1f} 倍且 0 Leads"
            elif worst and worst["over_pct"] > float(rules["warning_over_baseline_pct"]):
                action, reason = "WARN_CREATIVE", f"{worst['country']} CPL 比上月地区基准高 {worst['over_pct']:.1f}%，建议更换素材"
            elif rules.get("no_delivery_enabled") and age_hours >= float(rules["no_delivery_hours"]) and last_2d_spend <= 0:
                action, reason = "WARN_NO_DELIVERY", f"广告组已 ACTIVE {age_hours / 24:.1f} 天，但最近 24 小时没有消耗，请检查审核、受众、出价和支付状态"
            elif config_warnings:
                action, reason = "WARN_CONFIG", "；".join(config_warnings)
            elif rules.get("pacing_enabled") and daily_budget > 0 and budget_info and cooldown_seconds <= 0 and budget_info.get("pace_status") == "overspend":
                action, reason = "PACE_DOWN", f"季度预算进度偏快（节奏 {float(budget_info.get('pace_percent') or 0):.1f}%），建议下调日预算"
                proposed_budget = round(daily_budget * (1 - float(rules["budget_adjust_pct"]) / 100), 2)
            elif rules.get("pacing_enabled") and daily_budget > 0 and budget_info and cooldown_seconds <= 0 and budget_info.get("pace_status") == "slow" and recent_value["leads"] > 0 and baseline_cpl > 0 and recent_value["spend"] / recent_value["leads"] <= baseline_cpl:
                action, reason = "PACE_UP", f"季度预算进度偏慢且近 {rules['no_lead_days']} 天 CPL 健康，建议补速"
                proposed_budget = round(daily_budget * (1 + float(rules["budget_adjust_pct"]) / 100), 2)
            elif rules.get("scale_enabled") and daily_budget > 0 and cooldown_seconds <= 0 and recent_value["leads"] >= float(rules["scale_min_leads"]) and baseline_cpl > 0 and recent_value["spend"] / recent_value["leads"] <= baseline_cpl * (1 - float(rules["scale_cpl_better_pct"]) / 100):
                action, reason = "SCALE_UP", f"近 {rules['no_lead_days']} 天 {recent_value['leads']:.0f} Leads，CPL 比地区上月基准低至少 {float(rules['scale_cpl_better_pct']):.0f}%"
                proposed_budget = round(daily_budget * (1 + float(rules["budget_adjust_pct"]) / 100), 2)
            if action == "OK":
                continue
            previous_selling_point = _infer_selling_point(product, adset.get("name"), campaign_name, assets=assets)
            item = {
                "adset_id": adset_id, "adset_name": str(adset.get("name") or ""), "campaign_id": campaign_id,
                "campaign_name": campaign_name, "product": product, "previous_selling_point": previous_selling_point, "countries": countries,
                "daily_budget": daily_budget, "proposed_budget": proposed_budget, "current_spend": round(total_spend, 2),
                "current_leads": round(total_leads, 2), "current_cpl": round(total_spend / total_leads, 2) if total_leads else 0,
                "last_days_spend": round(recent_value["spend"], 2), "last_days_leads": round(recent_value["leads"], 2),
                "today_spend": round(today_value["spend"], 2), "average_7d_spend": round(average_7d_spend, 2),
                "baseline_cpl": round(baseline_cpl, 2), "budget_pace": (budget_info or {}).get("pace_status", ""),
                "cooldown_hours": round(cooldown_seconds / 3600, 1), "config_warnings": config_warnings,
                "country_detail": worst, "action": action, "reason": reason, "paused": False, "budget_changed": False, "executed": False, "error": "",
            }
            if execute:
                _execute_automation_item(item, rules)
                if item.get("paused"):
                    closed.append(adset_id)
                if item.get("budget_changed"):
                    budget_changed.append(adset_id)
            items.append(item)
            if execute and item.get("error"):
                execution_halted = True
                halt_reason = _safe_external_error(item.get("error"))
                break
        result = {"ok": True, "execute": bool(execute), "checked_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "active_adsets": len(active), "items": items, "closed": closed, "budget_changed": budget_changed, "rules": rules, "baseline_period": f"{previous_start} 至 {previous_end}", "current_period": f"{current_start} 至 {decision_until}", "data_complete_through": decision_until.isoformat()}
        result["halted"] = execution_halted
        result["halt_reason"] = halt_reason
        if execute:
            history = load_json(AUTOMATION_LOG_FILE, [])
            if not isinstance(history, list): history = []
            history.append(result)
            save_json(AUTOMATION_LOG_FILE, history[-100:])
        save_json(AUTOMATION_PREVIEW_FILE, result)
        automation_state.update({"last_finished": result["checked_at"], "last_result": result})
        return result
    except Exception as exc:
        automation_state.update({"last_error": _safe_external_error(exc), "last_finished": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        raise
    finally:
        automation_state["running"] = False


def apply_automation_items(keys):
    requested = {str(key) for key in (keys or []) if key}
    if not requested:
        raise RuntimeError("请先选择要执行的自动化项目。")
    # Use the last successful full preview. Re-running every Meta insight query
    # immediately before writes wastes the request quota.
    preview = load_json(AUTOMATION_PREVIEW_FILE, {})
    if not isinstance(preview, dict) or not preview.get("checked_at"):
        raise RuntimeError("没有可用的自动化预览快照；请等待 Meta 额度恢复后先做一次只读预览。")
    try:
        preview_age = (datetime.now() - datetime.strptime(str(preview.get("checked_at")), "%Y-%m-%d %H:%M:%S")).total_seconds()
    except Exception:
        preview_age = 999999
    if preview_age > 60 * 60:
        raise RuntimeError("自动化预览已超过 60 分钟，为避免执行过期规则已停止。请等待 Meta 额度恢复后重新预览。")
    rules = preview.get("rules") or load_automation_rules()
    executable = {"CLOSE_NO_LEADS", "CLOSE_HIGH_CPL", "CLOSE_SPEND_SPIKE", "SCALE_UP", "PACE_UP", "PACE_DOWN"}
    candidates = []
    for source_item in preview.get("items") or []:
        item = dict(source_item)
        key = f"{item.get('adset_id')}|{item.get('action')}"
        if key not in requested or item.get("action") not in executable:
            continue
        candidates.append(item)
    if not candidates:
        raise RuntimeError("所选项目不在最近一次成功预览中，已停止执行。请刷新自动化中心后重新选择。")
    selected, closed, budget_changed = [], [], []
    halted, halt_reason = False, ""
    for index, item in enumerate(candidates):
        _execute_automation_item(item, rules)
        selected.append(item)
        if item.get("paused"):
            closed.append(item.get("adset_id"))
        if item.get("budget_changed"):
            budget_changed.append(item.get("adset_id"))
        if item.get("error"):
            halted = True
            halt_reason = _safe_external_error(item.get("error"))
            for pending in candidates[index + 1:]:
                pending["error"] = "前一项执行失败，已自动暂停后续操作，未向 Meta 提交"
                pending["executed"] = False
                selected.append(pending)
            break
    result = {
        **preview, "execute": True, "items": selected, "closed": closed,
        "budget_changed": budget_changed, "requested_count": len(requested),
        "used_cached_preview": True, "preview_age_minutes": round(preview_age / 60, 1),
        "halted": halted, "halt_reason": halt_reason,
        "checked_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    history = load_json(AUTOMATION_LOG_FILE, [])
    if not isinstance(history, list):
        history = []
    history.append(result)
    save_json(AUTOMATION_LOG_FILE, history[-100:])
    automation_state.update({"last_finished": result["checked_at"], "last_result": result})
    return result


def rollback_automation_action(action_id):
    action_id = str(action_id or "").strip()
    if not action_id:
        raise RuntimeError("缺少可撤销的操作编号。")
    history = load_json(AUTOMATION_LOG_FILE, [])
    if not isinstance(history, list):
        history = []
    target = None
    for run in reversed(history):
        for item in reversed(run.get("items") or []):
            if str(item.get("action_id") or "") == action_id:
                target = item
                break
        if target:
            break
    if not target or not target.get("executed"):
        raise RuntimeError("没有找到这条已执行操作，可能是旧版本记录。")
    if target.get("rolled_back_at"):
        raise RuntimeError("这条操作已经撤销过了。")
    action = str(target.get("action") or "")
    if action.startswith("CLOSE_"):
        success = _set_adset_status(target.get("adset_id"), target.get("previous_status") or "ACTIVE")
        rollback_text = "已恢复为 ACTIVE"
    elif action in {"SCALE_UP", "PACE_UP", "PACE_DOWN"}:
        old_budget = float(target.get("daily_budget") or 0)
        if old_budget <= 0:
            raise RuntimeError("原日预算缺失，无法自动恢复。")
        success = _update_adset_budget(target.get("adset_id"), old_budget)
        rollback_text = f"已恢复日预算 ${old_budget:.2f}"
        if success:
            _record_budget_change(target.get("adset_id"), "ROLLBACK", target.get("proposed_budget", 0), old_budget)
    else:
        raise RuntimeError("该提醒没有修改 Meta 广告，不需要撤销。")
    if not success:
        raise RuntimeError("Meta 没有确认撤销成功。")
    target["rolled_back_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    target["rollback_available"] = False
    target["rollback_result"] = rollback_text
    save_json(AUTOMATION_LOG_FILE, history[-100:])
    last_result = automation_state.get("last_result") or {}
    for item in last_result.get("items") or []:
        if str(item.get("action_id") or "") == action_id:
            item.update({"rolled_back_at": target["rolled_back_at"], "rollback_available": False, "rollback_result": rollback_text})
    return {"ok": True, "action_id": action_id, "item": target, "message": rollback_text}


def _metric_window(rows, adset_id, start, end):
    selected = [row for row in rows if str(row.get("adset_id") or "") == str(adset_id) and start <= str(row.get("date") or "")[:10] <= end]
    spend = sum(_number(row.get("spend")) for row in selected)
    leads = sum(_number(row.get("leads")) for row in selected)
    return {"spend": round(spend, 2), "leads": round(leads, 2), "cpl": round(spend / leads, 2) if leads else 0}


def automation_effectiveness():
    data = fetch_adset_insights(days=30, force=False)
    rows = data.get("rows") or []
    history = load_json(AUTOMATION_LOG_FILE, [])
    if not isinstance(history, list):
        history = []
    today = datetime.now().date()
    items, seen = [], set()
    for run in reversed(history):
        for action in reversed(run.get("items") or []):
            action_id = str(action.get("action_id") or "")
            if not action_id or action_id in seen or not action.get("executed"):
                continue
            seen.add(action_id)
            try:
                action_date = datetime.strptime(str(action.get("executed_at") or run.get("checked_at"))[:10], "%Y-%m-%d").date()
            except Exception:
                continue
            before = _metric_window(rows, action.get("adset_id"), (action_date - timedelta(days=3)).isoformat(), (action_date - timedelta(days=1)).isoformat())
            after3 = _metric_window(rows, action.get("adset_id"), action_date.isoformat(), min(today, action_date + timedelta(days=2)).isoformat())
            after7 = _metric_window(rows, action.get("adset_id"), action_date.isoformat(), min(today, action_date + timedelta(days=6)).isoformat())
            elapsed = max(0, (today - action_date).days + 1)
            action_type = str(action.get("action") or "")
            estimated_saved = round(float(action.get("daily_budget") or 0) * min(elapsed, 7), 2) if action_type.startswith("CLOSE_") and not action.get("rolled_back_at") else 0
            comparison = after7 if elapsed >= 7 else after3
            if action.get("rolled_back_at"):
                outcome = "已撤销"
            elif action_type.startswith("CLOSE_"):
                outcome = "已止损"
            elif elapsed < 3:
                outcome = "观察中"
            elif before["cpl"] and comparison["cpl"]:
                change = (comparison["cpl"] / before["cpl"] - 1) * 100
                outcome = "改善" if change <= -10 else ("恶化" if change >= 10 else "基本稳定")
            else:
                outcome = "数据不足"
            items.append({
                "action_id": action_id, "executed_at": action.get("executed_at") or run.get("checked_at"), "action": action_type,
                "adset_id": action.get("adset_id"), "adset_name": action.get("adset_name"), "product": action.get("product"),
                "countries": action.get("countries") or [], "before3": before, "after3": after3, "after7": after7,
                "elapsed_days": elapsed, "outcome": outcome, "estimated_saved": estimated_saved,
                "rolled_back_at": action.get("rolled_back_at") or "", "rollback_available": bool(action.get("rollback_available") and not action.get("rolled_back_at")),
            })
    summary = {
        "actions": len(items), "improved": len([x for x in items if x["outcome"] == "改善"]),
        "worsened": len([x for x in items if x["outcome"] == "恶化"]),
        "observing": len([x for x in items if x["outcome"] in {"观察中", "数据不足"}]),
        "estimated_saved": round(sum(x["estimated_saved"] for x in items), 2),
    }
    return {"ok": True, "cached_at": data.get("cached_at", ""), "source": data.get("source", ""), "summary": summary, "items": items[:100]}


def automation_scheduler():
    while True:
        threading.Event().wait(60)
        rules = load_automation_rules()
        if not rules.get("enabled") or automation_state.get("running"):
            continue
        last_text = automation_state.get("last_finished") or ""
        try:
            last_dt = datetime.strptime(last_text, "%Y-%m-%d %H:%M:%S") if last_text else None
        except Exception:
            last_dt = None
        if last_dt and (datetime.now() - last_dt).total_seconds() < float(rules["check_interval_hours"]) * 3600:
            continue
        try:
            evaluate_automation_rules(execute=rules.get("execution_mode") == "auto")
        except Exception:
            pass


def automation_status_payload():
    history = load_json(AUTOMATION_LOG_FILE, [])
    if not isinstance(history, list):
        history = []
    state = dict(automation_state)
    if not state.get("last_result"):
        cached_preview = load_json(AUTOMATION_PREVIEW_FILE, {})
        if isinstance(cached_preview, dict) and cached_preview.get("checked_at"):
            state["last_result"] = cached_preview
            state["last_finished"] = cached_preview.get("checked_at", "")
    if not state.get("last_result") and history:
        state["last_result"] = history[-1]
        state["last_finished"] = history[-1].get("checked_at", "")
    last_result = state.get("last_result") or {}
    status_cache = {"cached_at": "", "source": "not_needed", "items": {}}
    ignored_inactive = 0
    if last_result.get("items"):
        status_cache = _automation_adset_status_map(force=False)
        status_items = status_cache.get("items") or {}
        asset_catalog = scan_asset_packs()
        filtered = []
        for item in last_result.get("items") or []:
            status_item = status_items.get(str(item.get("adset_id") or ""))
            if status_item and status_item.get("active") is False:
                ignored_inactive += 1
                continue
            visible_item = dict(item)
            if not visible_item.get("previous_selling_point"):
                visible_item["previous_selling_point"] = _infer_selling_point(
                    visible_item.get("product"), visible_item.get("adset_name"), visible_item.get("campaign_name"), assets=asset_catalog
                )
            filtered.append(visible_item)
        clean_result = dict(last_result)
        clean_result["items"] = filtered
        clean_result["ignored_inactive"] = ignored_inactive
        try:
            clean_result["active_adsets"] = max(0, int(clean_result.get("active_adsets") or 0) - ignored_inactive)
        except Exception:
            pass
        clean_result["status_checked_at"] = status_cache.get("cached_at", "")
        clean_result["status_source"] = status_cache.get("source", "")
        state["last_result"] = clean_result
    return {"state": state, "rules": load_automation_rules(), "history": history[-20:], "ignored_inactive": ignored_inactive, "status_checked_at": status_cache.get("cached_at", ""), "status_source": status_cache.get("source", "")}


def send_insights_excel(handler, days=7, campaign_id="", country=""):
    data = fetch_adset_insights(days=days, force=False)
    campaign_id = clean_id(campaign_id)
    country = str(country or "").strip().upper()
    detail_rows = [row for row in data["rows"] if (not campaign_id or row["campaign_id"] == campaign_id) and (not country or row.get("country") == country)]
    previous_detail_rows = [row for row in data.get("previous_rows", []) if (not campaign_id or row["campaign_id"] == campaign_id) and (not country or row.get("country") == country)]
    daily_rows = data["daily"]
    if campaign_id or country:
        grouped = {day["date"]: {"date": day["date"], "spend": 0.0, "impressions": 0, "reach": 0, "clicks": 0, "leads": 0.0} for day in data["daily"]}
        for row in detail_rows:
            day = grouped[row["date"]]
            for metric in ("spend", "impressions", "reach", "clicks", "leads"):
                day[metric] += row[metric]
        daily_rows = []
        for day in grouped.values():
            day["ctr"] = round(day["clicks"] / day["impressions"] * 100, 3) if day["impressions"] else 0
            day["cpc"] = round(day["spend"] / day["clicks"], 3) if day["clicks"] else 0
            day["cpm"] = round(day["spend"] / day["impressions"] * 1000, 3) if day["impressions"] else 0
            day["cost_per_lead"] = round(day["spend"] / day["leads"], 2) if day["leads"] else 0
            daily_rows.append(day)
    workbook = Workbook()
    detail = workbook.active
    detail.title = "广告组明细"
    detail.append(["日期", "国家", "Campaign", "Campaign ID", "广告组", "广告组 ID", "花费(USD)", "展示", "触达", "点击", "CTR(%)", "CPC", "CPM", "Leads", "单条线索成本"])
    for row in detail_rows:
        detail.append([row["date"], row.get("country", "UNKNOWN"), row["campaign_name"], row["campaign_id"], row["adset_name"], row["adset_id"], row["spend"], row["impressions"], row["reach"], row["clicks"], row["ctr"], row["cpc"], row["cpm"], row["leads"], row["cost_per_lead"]])
    summary = workbook.create_sheet("每日汇总")
    summary.append(["日期", "花费(USD)", "展示", "触达", "点击", "CTR(%)", "CPC", "CPM", "Leads", "单条线索成本"])
    for row in daily_rows:
        summary.append([row["date"], row["spend"], row["impressions"], row["reach"], row["clicks"], row["ctr"], row["cpc"], row["cpm"], row["leads"], row["cost_per_lead"]])
    campaign_sheet = workbook.create_sheet("广告系列汇总")
    campaign_sheet.append(["Campaign", "Campaign ID", "广告组数", "花费(USD)", "展示", "点击", "CTR(%)", "Leads", "单条线索成本"])
    campaign_groups = {}
    for row in detail_rows:
        group = campaign_groups.setdefault(row["campaign_id"], {"name": row["campaign_name"], "adsets": set(), "spend": 0.0, "impressions": 0, "clicks": 0, "leads": 0.0})
        group["adsets"].add(row["adset_id"])
        for metric in ("spend", "impressions", "clicks", "leads"):
            group[metric] += row[metric]
    for campaign_id_value, group in sorted(campaign_groups.items(), key=lambda item: item[1]["spend"], reverse=True):
        ctr = group["clicks"] / group["impressions"] * 100 if group["impressions"] else 0
        cpl = group["spend"] / group["leads"] if group["leads"] else 0
        campaign_sheet.append([group["name"], campaign_id_value, len(group["adsets"]), round(group["spend"], 2), group["impressions"], group["clicks"], round(ctr, 3), round(group["leads"], 2), round(cpl, 2)])
    compare_sheet = workbook.create_sheet("环比汇总")
    compare_sheet.append(["指标", f"当前 {data['since']} 至 {data['until']}", f"上期 {data['previous_since']} 至 {data['previous_until']}", "环比"])
    _, current_totals = _summarize_insight_period(detail_rows, datetime.strptime(data["since"], "%Y-%m-%d").date(), int(data["days"]))
    _, previous_totals = _summarize_insight_period(previous_detail_rows, datetime.strptime(data["previous_since"], "%Y-%m-%d").date(), int(data["days"]))
    for label, metric in (("花费(USD)", "spend"), ("展示", "impressions"), ("点击", "clicks"), ("CTR(%)", "ctr"), ("Leads", "leads"), ("单条线索成本", "cost_per_lead")):
        current_value, previous_value = current_totals.get(metric, 0), previous_totals.get(metric, 0)
        if previous_value:
            change = f"{((current_value - previous_value) / previous_value * 100):+.1f}%"
        else:
            change = "新增" if current_value else "0.0%"
        compare_sheet.append([label, current_value, previous_value, change])
    from openpyxl.styles import Alignment, Font, PatternFill
    header_fill = PatternFill("solid", fgColor="312E81")
    for sheet in (detail, summary, campaign_sheet, compare_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill, cell.font, cell.alignment = header_fill, Font(color="FFFFFF", bold=True), Alignment(horizontal="center")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = max(min(max(len(str(cell.value or "")) for cell in column) + 2, 42), 11)
    buffer = BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()
    suffix = (f"_{campaign_id}" if campaign_id else "") + (f"_{country}" if country else "")
    filename = f"adset_insights_{data['since']}_{data['until']}{suffix}.xlsx"
    from urllib.parse import quote
    handler.send_response(200)
    handler.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    handler.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
    handler.send_header("Content-Length", str(len(payload)))
    handler.end_headers()
    handler.wfile.write(payload)


def calculate_creative_fatigue(rows):
    groups = {}
    for row in rows or []:
        key = (str(row.get("adset_id") or ""), str(row.get("country") or "UNKNOWN").upper())
        group = groups.setdefault(key, {
            "adset_id": key[0], "country": key[1], "adset_name": str(row.get("adset_name") or ""),
            "campaign_id": str(row.get("campaign_id") or ""), "campaign_name": str(row.get("campaign_name") or ""),
            "daily": {},
        })
        day = group["daily"].setdefault(str(row.get("date") or ""), {"spend": 0.0, "leads": 0.0})
        day["spend"] += _number(row.get("spend"))
        day["leads"] += _number(row.get("leads"))
    alerts = []
    for group in groups.values():
        daily_map = group.pop("daily")
        dates = sorted(date for date in daily_map if date)
        latest_dates, previous_dates = dates[-3:], dates[-6:-3]
        latest_spend = sum(daily_map[date]["spend"] for date in latest_dates)
        latest_leads = sum(daily_map[date]["leads"] for date in latest_dates)
        previous_spend = sum(daily_map[date]["spend"] for date in previous_dates)
        previous_leads = sum(daily_map[date]["leads"] for date in previous_dates)
        previous_cpl = previous_spend / previous_leads if previous_leads else 0
        latest_cpl = latest_spend / latest_leads if latest_leads else 0
        minimum_spend = max(20.0, previous_cpl)
        if not previous_dates or previous_cpl <= 0 or latest_spend < minimum_spend:
            continue
        comparison_cpl = latest_cpl if latest_leads else latest_spend
        cpl_change = (comparison_cpl / previous_cpl - 1) * 100
        if cpl_change < 25:
            continue
        severe = cpl_change >= 50
        reason = (f"近3日 CPL 比前3日高 {cpl_change:.1f}%" if latest_leads else
                  f"近3日已花 ${latest_spend:.2f} 仍为 0 Leads，已超过前3日 CPL {cpl_change:.1f}%")
        alerts.append({
            **group, "latest_spend": round(latest_spend, 2), "previous_spend": round(previous_spend, 2),
            "latest_leads": round(latest_leads, 2), "previous_leads": round(previous_leads, 2),
            "latest_cpl": round(latest_cpl, 2), "previous_cpl": round(previous_cpl, 2),
            "cpl_change": round(cpl_change, 1), "minimum_spend": round(minimum_spend, 2),
            "severity": "severe" if severe else "warning", "reason": reason,
            "suggestion": "建议立即更换素材" if severe else "建议准备新素材并持续观察",
        })
    return sorted(alerts, key=lambda item: (item["severity"] != "severe", -item["cpl_change"], -item["latest_spend"]))


def _report_change(current, previous):
    if previous:
        return f"{((current - previous) / previous * 100):+.1f}%"
    return "新增" if current else "0.0%"


def send_operating_report(handler, period="week"):
    period = "month" if str(period).lower() == "month" else "week"
    days = 30 if period == "month" else 7
    data = fetch_adset_insights(days=days, force=False)
    budget = quarter_budget_status(force=False)
    fatigue = calculate_creative_fatigue(data.get("rows") or [])
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "产品经营看板"
    dashboard.append(["产品", "Campaign", "国家/地区", "固定语种", "季度预算", "已花费", "Leads", "CPL", "剩余预算", "自动日预算", "当前应花", "预算节奏", "季末预计花费", "预计完成率", "预计耗尽日", "投放覆盖"])
    pace_labels = {"overspend": "超进度", "slow": "进度偏慢", "on_track": "正常"}
    for item in budget.get("items") or []:
        dashboard.append([
            item.get("product"), item.get("campaign_name") or item.get("campaign_id"), ",".join(item.get("countries") or []) or "全部", item.get("asset_language") or "自动",
            item.get("initial_budget"), item.get("spent"), item.get("leads"), item.get("cpl"), item.get("remaining"), item.get("daily_budget"),
            item.get("expected_spend"), pace_labels.get(item.get("pace_status"), item.get("pace_status")), item.get("projected_spend"),
            item.get("projected_completion"), item.get("exhaustion_date"), "缺投放：" + ",".join(item.get("missing_countries") or []) if item.get("missing_countries") else "已覆盖",
        ])
    campaign_sheet = workbook.create_sheet("广告系列汇总")
    campaign_sheet.append(["广告系列（产品）", "Campaign ID", "广告组数", "本期花费", "上期花费", "花费环比", "展示", "点击", "CTR(%)", "Leads", "上期 Leads", "Leads 环比", "CPL", "上期 CPL", "CPL 环比"])
    current_groups, previous_groups = {}, {}
    for target, rows in ((current_groups, data.get("rows") or []), (previous_groups, data.get("previous_rows") or [])):
        for row in rows:
            group = target.setdefault(row.get("campaign_id"), {"name": row.get("campaign_name"), "adsets": set(), "spend": 0.0, "impressions": 0, "clicks": 0, "leads": 0.0})
            group["adsets"].add(row.get("adset_id"))
            for metric in ("spend", "impressions", "clicks", "leads"):
                group[metric] += _number(row.get(metric))
    for campaign_id, current in sorted(current_groups.items(), key=lambda pair: pair[1]["spend"], reverse=True):
        previous = previous_groups.get(campaign_id) or {"spend": 0, "leads": 0}
        ctr = current["clicks"] / current["impressions"] * 100 if current["impressions"] else 0
        cpl = current["spend"] / current["leads"] if current["leads"] else 0
        previous_cpl = previous["spend"] / previous["leads"] if previous.get("leads") else 0
        campaign_sheet.append([current["name"], campaign_id, len(current["adsets"]), round(current["spend"], 2), round(previous.get("spend", 0), 2), _report_change(current["spend"], previous.get("spend", 0)), int(current["impressions"]), int(current["clicks"]), round(ctr, 3), round(current["leads"], 2), round(previous.get("leads", 0), 2), _report_change(current["leads"], previous.get("leads", 0)), round(cpl, 2), round(previous_cpl, 2), _report_change(cpl, previous_cpl)])
    daily_sheet = workbook.create_sheet("每日趋势")
    daily_sheet.append(["日期", "花费", "展示", "触达", "点击", "CTR(%)", "Leads", "CPL"])
    for row in data.get("daily") or []:
        daily_sheet.append([row.get("date"), row.get("spend"), row.get("impressions"), row.get("reach"), row.get("clicks"), row.get("ctr"), row.get("leads"), row.get("cost_per_lead")])
    fatigue_sheet = workbook.create_sheet("素材疲劳")
    fatigue_sheet.append(["级别", "广告系列", "国家", "广告组", "近3日花费", "前3日花费", "近3日 Leads", "近3日 CPL", "前3日 CPL", "CPL 变化", "原因", "建议"])
    for item in fatigue:
        fatigue_sheet.append(["高风险" if item["severity"] == "severe" else "关注", item["campaign_name"], item["country"], item["adset_name"], item["latest_spend"], item["previous_spend"], item["latest_leads"], item["latest_cpl"] or "—", item["previous_cpl"], f"{item['cpl_change']:+.1f}%", item["reason"], item["suggestion"]])
    automation_sheet = workbook.create_sheet("自动化记录")
    automation_sheet.append(["检查时间", "执行模式", "动作", "执行结果", "广告组", "产品", "国家", "原日预算", "建议/新日预算", "原因", "错误"])
    action_labels = {"CLOSE_NO_LEADS": "3天0留资停投", "CLOSE_HIGH_CPL": "高CPL停投", "CLOSE_SPEND_SPIKE": "花费突增止损", "WARN_CREATIVE": "更换素材", "SCALE_UP": "优质扩量", "PACE_UP": "预算补速", "PACE_DOWN": "预算降速", "WARN_NO_DELIVERY": "ACTIVE不消耗", "WARN_CONFIG": "配置异常"}
    history = load_json(AUTOMATION_LOG_FILE, [])
    report_since = str(data.get("since") or "")
    for run in history if isinstance(history, list) else []:
        if str(run.get("checked_at") or "")[:10] < report_since:
            continue
        for item in run.get("items") or []:
            result_text = "已撤销" if item.get("rolled_back_at") else ("已暂停" if item.get("paused") else ("已调预算" if item.get("budget_changed") else ("执行失败" if item.get("error") else "只读提醒")))
            automation_sheet.append([run.get("checked_at"), "执行" if run.get("execute") else "预览", action_labels.get(item.get("action"), item.get("action")), result_text, item.get("adset_name"), item.get("product"), ",".join(item.get("countries") or []), item.get("daily_budget"), item.get("proposed_budget") or "", item.get("reason"), item.get("error")])
    detail_sheet = workbook.create_sheet("广告组明细")
    detail_sheet.append(["日期", "国家", "广告系列", "广告组", "花费", "展示", "触达", "点击", "CTR(%)", "Leads", "CPL"])
    for row in data.get("rows") or []:
        detail_sheet.append([row.get("date"), row.get("country"), row.get("campaign_name"), row.get("adset_name"), row.get("spend"), row.get("impressions"), row.get("reach"), row.get("clicks"), row.get("ctr"), row.get("leads"), row.get("cost_per_lead")])
    from openpyxl.styles import Alignment, Font, PatternFill
    header_fill = PatternFill("solid", fgColor="312E81")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill, cell.font, cell.alignment = header_fill, Font(color="FFFFFF", bold=True), Alignment(horizontal="center")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = max(min(max(len(str(cell.value or "")) for cell in column) + 2, 42), 11)
    buffer = BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()
    filename = f"{'monthly' if period == 'month' else 'weekly'}_operating_report_{data['since']}_{data['until']}.xlsx"
    from urllib.parse import quote
    handler.send_response(200)
    handler.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    handler.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
    handler.send_header("Content-Length", str(len(payload)))
    handler.end_headers()
    handler.wfile.write(payload)

HTML = r'''<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>ADS 工作台｜投放、客户与广告数据</title>
<style>
:root{
  --bg:#f4f6fb;
  --paper:#ffffff;
  --paper-2:#f8f9fd;
  --ink:#111827;
  --text:#1f2937;
  --muted:#667085;
  --muted-2:#98a2b3;
  --line:#e6e9f0;
  --line-2:#d8deea;
  --purple:#5432eb;
  --purple-dark:#3b22b8;
  --purple-soft:#f1efff;
  --yellow:#ffad00;
  --green:#087f5b;
  --orange:#b45309;
  --red:#b42318;
  --shadow:0 18px 50px rgba(17,24,39,.08);
  --shadow-sm:0 8px 22px rgba(17,24,39,.055);
  --radius:18px;
  --radius-sm:12px;
}
*{box-sizing:border-box}
html{background:var(--bg)}
body{
  margin:0;
  color:var(--text);
  font-family:Inter,"Segoe UI",Arial,"Microsoft YaHei",sans-serif;
  font-size:13px;
  line-height:1.48;
  background:
    radial-gradient(circle at 14% -10%,rgba(84,50,235,.12),transparent 30%),
    radial-gradient(circle at 92% 0%,rgba(255,173,0,.11),transparent 28%),
    linear-gradient(180deg,#fbfcff 0,#f4f6fb 260px,#f4f6fb 100%);
}
header{
  position:sticky;top:0;z-index:50;
  min-height:72px;
  padding:14px 24px;
  background:rgba(255,255,255,.88);
  backdrop-filter:blur(18px);
  border-bottom:1px solid rgba(216,222,234,.86);
  display:flex;justify-content:space-between;gap:18px;align-items:center;
  box-shadow:0 10px 30px rgba(17,24,39,.055);
}
h1{margin:0;font-size:20px;letter-spacing:-.02em;font-weight:860;color:var(--ink);display:flex;gap:11px;align-items:center}
h1:before{content:"";width:34px;height:34px;border-radius:12px;background:linear-gradient(135deg,var(--purple),var(--purple-dark));box-shadow:inset 0 0 0 1px rgba(255,255,255,.22),0 10px 20px rgba(84,50,235,.20)}
header p{margin:4px 0 0;color:var(--muted);font-size:12px}
main{max-width:1580px;margin:0 auto;padding:18px 18px 28px}
.dashboard{display:grid;grid-template-columns:330px minmax(620px,1fr) 460px;gap:16px;align-items:start}
.card{
  background:rgba(255,255,255,.96);
  border:1px solid rgba(230,233,240,.96);
  border-radius:var(--radius);
  padding:16px;
  box-shadow:var(--shadow-sm);
  margin-bottom:16px;
  position:relative;
  overflow:hidden;
}
.card:before{content:"";position:absolute;left:0;right:0;top:0;height:1px;background:linear-gradient(90deg,rgba(84,50,235,.6),rgba(255,173,0,.3),transparent)}
.card.compact{padding:16px}
h2{font-size:15px;margin:0;font-weight:850;letter-spacing:-.01em;color:var(--ink)}
.topline{display:flex;justify-content:space-between;gap:10px;align-items:center;margin:-2px 0 14px;padding-bottom:10px;border-bottom:1px solid #eef1f6}
.muted-chip{display:inline-flex;align-items:center;border:1px solid #e5e8f0;border-radius:999px;padding:4px 9px;color:#667085;background:#fafbff;font-size:12px;font-weight:740}
.row{display:grid;grid-template-columns:1fr 1fr;gap:12px}.row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px}.full{grid-column:1/-1}
.field{margin-bottom:11px}
label{font-size:12px;color:#344054;font-weight:780;display:block;margin-bottom:6px}
input,select,textarea{
  width:100%;padding:10px 11px;border:1px solid #dfe4ee;border-radius:13px;background:#fff;
  font-size:13px;outline:none;color:#111827;transition:border-color .16s ease,box-shadow .16s ease,background .16s ease;
}
input:hover,select:hover,textarea:hover{border-color:#cfd6e4}
input:focus,select:focus,textarea:focus{border-color:rgba(84,50,235,.65);box-shadow:0 0 0 4px rgba(84,50,235,.10);background:#fff}
input[readonly]{background:#f8fafc;color:#667085}textarea{min-height:68px;resize:vertical}
button{
  border:0;border-radius:13px;padding:9px 13px;font-size:13px;font-weight:790;cursor:pointer;
  transition:transform .12s ease,box-shadow .12s ease,background .12s ease,border-color .12s ease;white-space:nowrap;
}
button:hover{transform:translateY(-1px);box-shadow:0 10px 20px rgba(17,24,39,.10)}button:active{transform:translateY(0);box-shadow:none}
.primary{background:linear-gradient(135deg,var(--purple),var(--purple-dark));color:#fff;box-shadow:0 12px 24px rgba(84,50,235,.22)}
.secondary{background:var(--purple-soft);color:#3820a6;border:1px solid rgba(84,50,235,.15)}
.dark{background:#111827;color:#fff;box-shadow:0 12px 24px rgba(17,24,39,.16)}
.ghost{background:#fff;color:#344054;border:1px solid #e1e6ef}.danger{background:#fff4f3;color:var(--red);border:1px solid #ffd5d1;padding:6px 9px}
.actions{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px}header .actions{margin-top:0;justify-content:flex-end}header .ghost{background:#fff;color:#344054;border:1px solid #e4e7ef}
.hint{font-size:12px;color:var(--muted);line-height:1.55}.warn{color:var(--orange)}.ok{color:var(--green)}.bad{color:var(--red)}.mini{font-size:12px;color:var(--muted);margin-left:6px}
.system-strip{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:16px}
.sys-pill{border:1px solid #e5e8f0;border-radius:16px;background:rgba(255,255,255,.92);padding:10px 12px;box-shadow:var(--shadow-sm);position:relative;overflow:hidden}
.sys-pill:before{content:"";position:absolute;left:0;top:0;bottom:0;width:3px;background:var(--purple)}.sys-pill b{display:block;font-size:12px;color:var(--ink)}.sys-pill span{font-size:11px;color:var(--muted)}
.preview{display:grid;grid-template-columns:1.12fr .88fr;gap:12px;margin-top:2px}.copybox{white-space:pre-wrap;border:1px solid var(--line);border-radius:15px;padding:11px;background:#fbfcff;font-size:12px;max-height:154px;overflow:auto;color:#344054}.thumbs{display:flex;gap:8px;flex-wrap:wrap;max-height:154px;overflow:auto;border:1px solid var(--line);border-radius:15px;padding:9px;background:#fbfcff;min-height:78px}.thumb{width:82px;height:56px;object-fit:cover;border:1px solid #dfe4ee;border-radius:12px;background:#f3f4f6;box-shadow:0 6px 14px rgba(17,24,39,.08)}
.country-wrap{margin-top:13px;padding:13px;border:1px solid #e5e8f0;background:linear-gradient(180deg,#fbfcff,#f8f9fd);border-radius:17px}.manual-country{margin-top:9px}#country_dropdown{margin:8px 0 0}
.queue,.checktable{width:100%;border-collapse:separate;border-spacing:0;font-size:12px}.queue th,.queue td,.checktable th,.checktable td{border-bottom:1px solid var(--line);padding:9px 8px;text-align:left;vertical-align:top}.queue th,.checktable th{position:sticky;top:0;background:#f8fafc;color:#475467;font-weight:820;z-index:2}.queue tr:hover td,.checktable tr:hover td{background:#fbfcff}.queue-wrap,.table-scroll{max-height:282px;overflow:auto;border:1px solid var(--line);border-radius:15px;background:#fff}.tag{display:inline-flex;align-items:center;padding:3px 8px;border-radius:999px;background:#eef2ff;color:#3730a3;margin:1px;font-size:11px;font-weight:780}
.status{white-space:pre-wrap;background:#0b1020;color:#d1fae5;border-radius:16px;padding:15px;min-height:152px;max-height:280px;overflow:auto;font-family:Consolas,"SFMono-Regular",Menlo,monospace;font-size:12px;border:1px solid rgba(255,255,255,.08);box-shadow:inset 0 0 0 1px rgba(255,255,255,.04)}
details{border:1px solid var(--line);border-radius:16px;background:#fbfcff;margin-top:10px;padding:0;overflow:hidden}summary{list-style:none;cursor:pointer;padding:11px 13px;font-weight:850;font-size:13px;color:#273142;display:flex;justify-content:space-between;align-items:center;background:linear-gradient(180deg,#fff,#f8fafc)}summary::-webkit-details-marker{display:none}summary:after{content:'展开';font-size:12px;color:#667085;font-weight:760;background:#eef1f7;border-radius:999px;padding:2px 8px}details[open] summary:after{content:'收起'}details .inside{padding:13px}
.split-actions{display:grid;grid-template-columns:1fr 1fr;gap:9px}.sticky-actions{position:sticky;top:92px;z-index:10}.preflight-box{padding:0!important}#preflight_result{min-height:80px;border:1px dashed #d8deea;border-radius:15px;padding:11px;background:#fbfcff}#batch_text{min-height:140px;font-family:Consolas,"SFMono-Regular",Menlo,monospace;font-size:12px}.smallbtn{padding:5px 8px;font-size:12px;border-radius:9px;margin:1px}.batch-preview{margin-top:10px}.kpi-row{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:10px}.kpi{padding:10px;border:1px solid #e6eaf2;border-radius:15px;background:#fbfcff}.kpi b{display:block;font-size:14px;color:#111827}.kpi span{font-size:11px;color:#667085}
@media(max-width:1260px){.system-strip{grid-template-columns:1fr 1fr}.dashboard{grid-template-columns:1fr}header{height:auto;align-items:flex-start;flex-direction:column}.preview{grid-template-columns:1fr}.sticky-actions{position:relative;top:auto}.row3{grid-template-columns:1fr}.row{grid-template-columns:1fr}}
@media(max-height:820px){main{padding:12px}.card{padding:13px;margin-bottom:12px}.field{margin-bottom:8px}input,select,textarea{padding:8px 10px}.copybox,.thumbs{max-height:118px}.queue-wrap{max-height:210px}.status{max-height:200px;min-height:120px}}
</style>
<style>
:root{
  --bg:#eef2f7;
  --ink:#0f172a;
  --text:#1e293b;
  --muted:#64748b;
  --line:#dbe3ef;
  --accent:#2563eb;
  --accent-2:#0f766e;
  --accent-soft:#e0f2fe;
  --warning:#b45309;
  --success:#15803d;
  --danger:#b42318;
  --shadow:0 28px 60px rgba(15,23,42,.10);
  --shadow-sm:0 10px 24px rgba(15,23,42,.06);
  --radius:22px;
}
html{background:var(--bg);scroll-behavior:smooth}
body{
  background:
    radial-gradient(circle at 14% 0%,rgba(37,99,235,.12),transparent 26%),
    radial-gradient(circle at 86% 4%,rgba(15,118,110,.11),transparent 24%),
    radial-gradient(circle at 50% -8%,rgba(245,158,11,.08),transparent 20%),
    linear-gradient(180deg,#fbfdff 0,#f3f7fc 220px,#eef2f7 100%) !important;
  color:var(--text);
  font-family:Inter,"Segoe UI","PingFang SC","Microsoft YaHei",Arial,sans-serif;
  min-height:100vh;
  position:relative;
}
body:before{
  content:"";
  position:fixed;
  inset:-10% -10% auto -10%;
  height:420px;
  background:
    radial-gradient(circle at 20% 30%,rgba(37,99,235,.12),transparent 28%),
    radial-gradient(circle at 78% 28%,rgba(15,118,110,.10),transparent 26%);
  pointer-events:none;
  z-index:-1;
}
header{
  padding:16px 24px !important;
  background:rgba(255,255,255,.82) !important;
  backdrop-filter:blur(20px) saturate(120%);
  border-bottom:1px solid rgba(219,227,239,.9) !important;
  box-shadow:0 12px 34px rgba(15,23,42,.06) !important;
}
h1{
  font-size:21px !important;
  letter-spacing:-.03em !important;
}
h1:before{
  width:36px !important;
  height:36px !important;
  border-radius:13px !important;
  background:linear-gradient(135deg,var(--accent),var(--accent-2)) !important;
  box-shadow:inset 0 0 0 1px rgba(255,255,255,.20),0 14px 24px rgba(37,99,235,.22) !important;
}
main{max-width:1600px !important;padding:20px 18px 30px !important}
.dashboard{grid-template-columns:minmax(320px,340px) minmax(640px,1fr) minmax(420px,460px) !important;gap:18px !important}
.card{
  background:linear-gradient(180deg,rgba(255,255,255,.98),rgba(249,251,255,.98)) !important;
  border:1px solid rgba(219,227,239,.95) !important;
  border-radius:var(--radius) !important;
  padding:17px !important;
  box-shadow:var(--shadow-sm) !important;
  transition:transform .18s ease,box-shadow .18s ease,border-color .18s ease;
}
.card:hover{transform:translateY(-1px);box-shadow:0 20px 40px rgba(15,23,42,.08);border-color:#cfd9e8}
.card:before{
  background:linear-gradient(90deg,rgba(37,99,235,.55),rgba(15,118,110,.42),rgba(245,158,11,.22),transparent) !important;
}
.topline{border-bottom:1px solid #edf2f7 !important}
.muted-chip{
  border:1px solid #dbe3ef !important;
  border-radius:999px !important;
  padding:4px 10px !important;
  color:#52627a !important;
  background:linear-gradient(180deg,#fff,#f8fbff) !important;
  box-shadow:0 4px 10px rgba(15,23,42,.04) !important;
}
label{color:#334155 !important;font-weight:800 !important;letter-spacing:.01em}
input,select,textarea{
  border:1px solid #d6deea !important;
  border-radius:14px !important;
  background:linear-gradient(180deg,#fff,#fbfcff) !important;
  color:#0f172a !important;
}
input:hover,select:hover,textarea:hover{border-color:#bfd0e4 !important}
input:focus,select:focus,textarea:focus{
  border-color:rgba(37,99,235,.62) !important;
  box-shadow:0 0 0 4px rgba(37,99,235,.10) !important;
}
input[readonly]{background:#f8fbfe !important;color:#64748b !important}
button{font-weight:800 !important;letter-spacing:.01em}
button:hover{box-shadow:0 12px 22px rgba(15,23,42,.10) !important}
button:focus-visible{outline:2px solid rgba(37,99,235,.35);outline-offset:2px}
.primary{background:linear-gradient(135deg,var(--accent),#1d4ed8) !important;color:#fff !important;box-shadow:0 12px 24px rgba(37,99,235,.22) !important}
.secondary{background:linear-gradient(180deg,#eaf2ff,#dfeaff) !important;color:#1d4ed8 !important;border:1px solid rgba(37,99,235,.14) !important}
.dark{background:linear-gradient(135deg,#0f172a,#1e293b) !important;color:#fff !important;box-shadow:0 12px 24px rgba(15,23,42,.18) !important}
.ghost{background:#fff !important;color:#334155 !important;border:1px solid #d7e0ec !important}
.danger{background:#fff4f2 !important;color:var(--danger) !important;border:1px solid #ffd5cf !important}
.hint{color:var(--muted) !important}
.warn{color:var(--warning) !important}.ok{color:var(--success) !important}.bad{color:var(--danger) !important}
.sys-pill{
  border:1px solid #dbe3ef !important;
  border-radius:18px !important;
  background:linear-gradient(180deg,rgba(255,255,255,.96),rgba(248,251,255,.96)) !important;
  box-shadow:var(--shadow-sm) !important;
}
.sys-pill:before{width:4px !important;background:linear-gradient(180deg,var(--accent),var(--accent-2)) !important}
.copybox,.thumbs{
  border:1px solid var(--line) !important;
  border-radius:16px !important;
  background:linear-gradient(180deg,#fbfcff,#f7faff) !important;
}
.copybox{color:#334155 !important;box-shadow:inset 0 1px 0 rgba(255,255,255,.75)}
.thumbs{box-shadow:inset 0 1px 0 rgba(255,255,255,.75)}
.thumb{border:1px solid #d7e0ec !important;box-shadow:0 8px 16px rgba(15,23,42,.08) !important}
.country-wrap{border:1px solid #dbe3ef !important;background:linear-gradient(180deg,#fbfcff,#f5f9ff) !important}
.saved-country-sets{
  display:flex;
  flex-wrap:wrap;
  gap:8px;
  margin:6px 0 10px;
}
.saved-country-sets .saved-set{
  display:inline-flex;
  align-items:center;
  gap:6px;
  padding:6px 10px;
  border-radius:999px;
  background:#eff6ff;
  border:1px solid #dbeafe;
  color:#1d4ed8;
  font-size:12px;
  font-weight:700;
  line-height:1;
}
.saved-country-sets .saved-set button{
  padding:0;
  border:0;
  background:transparent !important;
  box-shadow:none !important;
  color:inherit !important;
  font-size:12px;
  font-weight:800;
}
.saved-country-sets .saved-set .remove{
  color:#64748b !important;
  opacity:.85;
}
.queue-wrap,.table-scroll{
  border:1px solid var(--line) !important;
  border-radius:15px !important;
  box-shadow:inset 0 1px 0 rgba(255,255,255,.55) !important;
}
.queue th,.checktable th{
  background:linear-gradient(180deg,#f8fbff,#eff4fa) !important;
  color:#334155 !important;
}
.queue tr:nth-child(even) td,.checktable tr:nth-child(even) td{background:#fbfdff !important}
.queue tr:hover td,.checktable tr:hover td{background:#f2f7ff !important}
.tag{background:#e8f2ff !important;color:#1d4ed8 !important;font-weight:800 !important}
.status{
  background:linear-gradient(180deg,#0b1220,#111827 45%,#0f172a) !important;
  box-shadow:inset 0 0 0 1px rgba(255,255,255,.04),0 16px 30px rgba(15,23,42,.18) !important;
}
details{border:1px solid var(--line) !important;background:#fbfdff !important}
summary{background:linear-gradient(180deg,#fff,#f4f8fc) !important;color:#273142 !important}
summary:after{background:#edf2f7 !important;color:#64748b !important}
#preflight_result{border-color:#d7e0ec !important;background:linear-gradient(180deg,#fbfdff,#f7fbff) !important}
.kpi{border:1px solid #e6edf5 !important;background:linear-gradient(180deg,#fff,#f8fbff) !important}
@media(max-width:1260px){.dashboard{grid-template-columns:1fr !important}header{flex-direction:column !important;align-items:flex-start !important}.preview{grid-template-columns:1fr !important}.row3,.row{grid-template-columns:1fr !important}.sticky-actions{position:relative !important;top:auto !important}}
@media(max-width:900px){main{padding:14px 12px 20px !important}header{padding:14px 14px 16px !important}.card{padding:15px !important}}
</style>
<style>
.skip-nav{position:absolute;left:-999px;top:10px;padding:10px 14px;background:#fff;color:#0f172a;border:1px solid #dbe3ef;border-radius:999px;box-shadow:0 10px 24px rgba(15,23,42,.12);z-index:200}
.skip-nav:focus{left:14px;top:14px}
.workspace{display:grid;grid-template-columns:minmax(360px,1.1fr) minmax(320px,.9fr) minmax(330px,.95fr);gap:16px;align-items:start}
.panel{min-height:0}
.panel-right{position:sticky;top:92px}
.ui-dialog{border:0;padding:0;background:transparent;width:min(980px,calc(100vw - 28px))}
.ui-dialog::backdrop{background:rgba(15,23,42,.58);backdrop-filter:blur(4px)}
.dialog-card{background:#fff;border:1px solid #dbe3ef;border-radius:22px;box-shadow:0 28px 60px rgba(15,23,42,.20);overflow:hidden}
.dialog-card-wide{width:min(1080px,calc(100vw - 28px))}
.dialog-card-wide{
  width:100% !important;
  max-width:100% !important;
  box-sizing:border-box !important;
}
.dialog-head{display:flex;justify-content:space-between;gap:16px;align-items:flex-start;padding:18px 18px 12px;border-bottom:1px solid #edf2f7;background:linear-gradient(180deg,#fff,#f8fbff)}
.dialog-head h2{margin:0;font-size:18px}
.dialog-head p{margin:4px 0 0;color:#64748b;font-size:12px}
.dialog-body{padding:16px 18px 18px}
@media(max-width:1260px){.workspace{grid-template-columns:1fr}.panel-right{position:relative;top:auto}}
</style>
<style>
/* High-end admin skin */
body{
  background:
    radial-gradient(circle at 8% 4%,rgba(37,99,235,.10),transparent 18%),
    radial-gradient(circle at 90% 8%,rgba(245,158,11,.08),transparent 18%),
    linear-gradient(180deg,#f8fafc 0,#eef2f7 240px,#e9eef5 100%) !important;
}
header{
  border-bottom:1px solid rgba(203,213,225,.95) !important;
  box-shadow:0 14px 36px rgba(15,23,42,.06) !important;
}
h1{
  font-size:22px !important;
}
h1:before{
  width:38px !important;
  height:38px !important;
  border-radius:12px !important;
  background:linear-gradient(135deg,#0f172a,#2563eb) !important;
}
header p{
  font-size:12px !important;
  color:#64748b !important;
  letter-spacing:.01em;
}
.system-strip{
  gap:12px !important;
  margin-bottom:18px !important;
}
.sys-pill{
  padding:12px 13px !important;
  border-radius:18px !important;
  background:linear-gradient(180deg,rgba(255,255,255,.94),rgba(248,250,252,.98)) !important;
  border:1px solid rgba(203,213,225,.96) !important;
}
.sys-pill b{
  font-size:11.5px !important;
  letter-spacing:.02em;
  text-transform:none;
}
.sys-pill span{
  color:#64748b !important;
}
.workspace{
  grid-template-columns:minmax(360px,1.02fr) minmax(360px,.98fr) minmax(360px,1fr) !important;
  gap:14px !important;
}
.panel-left,.panel-mid,.panel-right{
  position:relative;
}
.panel-left:after,.panel-mid:after,.panel-right:after{
  content:"";
  position:absolute;
  left:16px;
  right:16px;
  top:0;
  height:3px;
  border-radius:999px;
  opacity:.95;
}
.panel-left:after{background:linear-gradient(90deg,#2563eb,rgba(37,99,235,.25))}
.panel-mid:after{background:linear-gradient(90deg,#0f766e,rgba(15,118,110,.25))}
.panel-right:after{background:linear-gradient(90deg,#f59e0b,rgba(245,158,11,.25))}
.card{
  box-shadow:0 18px 40px rgba(15,23,42,.08) !important;
}
.card:hover{
  transform:translateY(-1px) !important;
  box-shadow:0 24px 50px rgba(15,23,42,.10) !important;
}
.topline{
  margin-bottom:16px !important;
}
h2{
  font-size:16px !important;
  letter-spacing:-.02em !important;
}
.muted-chip{
  background:linear-gradient(180deg,#fff,#f8fafc) !important;
  border-color:#d5dce7 !important;
}
.panel-mid .copybox,
.panel-mid .thumbs,
.panel-right .queue-wrap,
.panel-right .table-scroll,
.panel-right #preflight_result{
  background:linear-gradient(180deg,#fbfdff,#f7f9fc) !important;
}
.copybox{
  color:#334155 !important;
}
.thumb{
  border-color:#cfd8e4 !important;
}
.queue th,.checktable th{
  background:linear-gradient(180deg,#f8fafc,#edf2f7) !important;
  color:#475569 !important;
}
.queue tr:hover td,.checktable tr:hover td{
  background:#eef4ff !important;
}
.tag{
  background:#e8efff !important;
  color:#1d4ed8 !important;
}
.status{
  background:linear-gradient(180deg,#0b1220,#111827 60%,#0f172a) !important;
  color:#d1fae5 !important;
  box-shadow:inset 0 1px 0 rgba(255,255,255,.04),0 18px 36px rgba(15,23,42,.18) !important;
}
.split-actions button{
  min-height:42px;
}
.ui-dialog{
  width:min(960px,calc(100vw - 20px)) !important;
  max-width:calc(100vw - 20px) !important;
  margin:12px auto !important;
  overflow:hidden !important;
}
.dialog-card{
  background:linear-gradient(180deg,#ffffff,#f8fbff) !important;
  border-color:rgba(203,213,225,.98) !important;
  box-shadow:0 38px 80px rgba(15,23,42,.26) !important;
  width:100% !important;
  max-width:100% !important;
  box-sizing:border-box !important;
}
.dialog-head{
  background:linear-gradient(180deg,#0f172a,#111827) !important;
  border-bottom:1px solid rgba(255,255,255,.06) !important;
}
.dialog-head h2,
.dialog-head p{
  color:#fff !important;
}
.dialog-head p{
  color:#94a3b8 !important;
}
.dialog-head .ghost{
  background:rgba(255,255,255,.08) !important;
  border-color:rgba(255,255,255,.12) !important;
  color:#fff !important;
}
.dialog-body{
  padding:18px !important;
  max-height:calc(100vh - 128px) !important;
  overflow:auto !important;
}
.skip-nav{
  background:#fff !important;
  border-color:#d5dce7 !important;
}
@media(max-width:1260px){
  .workspace{grid-template-columns:1fr !important}
  .panel-right{position:relative !important;top:auto !important}
}
@media(max-width:900px){
  .system-strip{grid-template-columns:1fr 1fr !important}
  .panel-left:after,.panel-mid:after,.panel-right:after{left:12px;right:12px}
  .dialog-head{padding:16px 16px 10px !important}
}
</style>
<style>
/* Compact one-screen mode */
main{
  padding:12px 14px 16px !important;
  max-width:1680px !important;
}
.system-strip{
  margin-bottom:12px !important;
}
.workspace{
  grid-template-columns:minmax(330px,1.03fr) minmax(320px,.95fr) minmax(320px,.92fr) !important;
  gap:12px !important;
  align-items:stretch !important;
}
.card{
  margin-bottom:12px !important;
}
.card.compact{
  padding:12px !important;
}
.panel-left,.panel-mid,.panel-right{
  max-height:calc(100vh - 192px) !important;
  overflow:auto !important;
  min-height:0 !important;
}
.panel-right{
  position:relative !important;
  top:auto !important;
}
.topline{
  margin-bottom:10px !important;
  padding-bottom:8px !important;
}
h2{
  font-size:15px !important;
}
.row,.row3{
  gap:8px !important;
}
.field{
  margin-bottom:8px !important;
}
label{
  margin-bottom:5px !important;
}
input,select,textarea{
  padding:8px 10px !important;
  border-radius:12px !important;
}
.preview{
  gap:10px !important;
}
.panel-mid .preview{
  grid-template-columns:minmax(0,1fr) minmax(0,.92fr) !important;
}
.copybox{
  max-height:92px !important;
}
.thumbs{
  max-height:92px !important;
  min-height:72px !important;
}
.thumb{
  width:68px !important;
  height:48px !important;
}
.country-wrap{
  padding:10px !important;
  margin-top:10px !important;
}
.saved-country-sets{
  margin:4px 0 8px !important;
  gap:6px !important;
}
.saved-country-sets .saved-set{
  padding:5px 9px !important;
}
.queue-wrap,.table-scroll{
  max-height:190px !important;
}
.status{
  min-height:84px !important;
  max-height:140px !important;
}
.dialog-card{
  max-height:calc(100vh - 40px);
  overflow:auto;
}
@media(max-width:1260px){
  .panel-left,.panel-mid,.panel-right{
    max-height:none !important;
    overflow:visible !important;
  }
}
.app-nav{display:flex;gap:8px;margin-top:10px}
.nav-btn{padding:7px 12px!important;border:1px solid #dbe3ef!important;background:#fff!important;color:#475569!important}
.nav-btn.active{background:linear-gradient(135deg,var(--accent),var(--accent-2))!important;color:#fff!important;border-color:transparent!important}
.leads-layout{display:grid;grid-template-columns:minmax(360px,.85fr) minmax(480px,1.15fr);gap:18px;align-items:start}
.lead-products{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:8px;max-height:310px;overflow:auto;padding:10px;border:1px solid var(--line);border-radius:15px;background:#fbfcff}
.product-check{display:flex;align-items:center;gap:8px;padding:8px 10px;background:#fff;border:1px solid #e5eaf2;border-radius:11px;cursor:pointer}
.product-check input{width:auto!important;margin:0;accent-color:var(--accent)}
.lead-toolbar{display:flex;gap:8px;align-items:center;margin:10px 0;flex-wrap:wrap}
.lead-file-list{max-height:310px;overflow:auto;border:1px solid var(--line);border-radius:15px;background:#fff}
.lead-file{display:grid;grid-template-columns:1fr auto;gap:12px;align-items:center;padding:11px 13px;border-bottom:1px solid var(--line)}
.lead-file:last-child{border-bottom:0}.lead-file b{display:block;color:var(--ink)}
.lead-log{min-height:230px!important;max-height:390px!important}
section[hidden],div[hidden]{display:none!important}
@media(max-width:980px){.leads-layout{grid-template-columns:1fr}.lead-products{grid-template-columns:repeat(2,minmax(0,1fr))}}

/* Final design layer: keep behavior stable while giving the workspace one visual system. */
:root{
  --shell:#0b1220;
  --shell-2:#111c32;
  --surface:#ffffff;
  --canvas:#f3f6fb;
  --accent:#5b5cf0;
  --accent-2:#18a99a;
  --accent-soft:#eeefff;
  --ink:#111827;
  --text:#334155;
  --muted:#738197;
  --line:#e3e9f2;
  --shadow-sm:0 12px 34px rgba(15,23,42,.07);
  --shadow-lg:0 24px 70px rgba(15,23,42,.12);
}
html{background:var(--canvas)!important}
body{
  background:
    radial-gradient(circle at 12% 0%,rgba(91,92,240,.10),transparent 25%),
    radial-gradient(circle at 88% 2%,rgba(24,169,154,.09),transparent 23%),
    linear-gradient(180deg,#f9fbff 0,#f3f6fb 260px,#f3f6fb 100%)!important;
  color:var(--text)!important;
}
body:before{display:none!important}
header{
  min-height:84px!important;
  padding:12px 24px!important;
  background:linear-gradient(115deg,var(--shell) 0,var(--shell-2) 62%,#10243b 100%)!important;
  border-bottom:1px solid rgba(255,255,255,.08)!important;
  box-shadow:0 14px 38px rgba(15,23,42,.20)!important;
}
.brand-area{display:grid;grid-template-columns:auto auto;grid-template-rows:auto auto;align-items:center;column-gap:26px}
.brand-area h1{grid-column:1;grid-row:1;color:#fff!important;font-size:20px!important}
.brand-area h1:before{
  width:38px!important;height:38px!important;border-radius:13px!important;
  background:linear-gradient(145deg,#7475ff,#42d5c4)!important;
  box-shadow:0 10px 28px rgba(79,70,229,.42),inset 0 0 0 1px rgba(255,255,255,.24)!important;
}
.brand-area p{grid-column:1;grid-row:2;color:#94a3b8!important;margin-top:1px!important}
.app-nav{grid-column:2;grid-row:1/3;margin:0!important;padding:4px;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.10);border-radius:14px}
.nav-btn{padding:8px 14px!important;border:0!important;background:transparent!important;color:#aebbd0!important;box-shadow:none!important}
.nav-btn:hover{color:#fff!important;background:rgba(255,255,255,.07)!important}
.nav-btn.active{background:#fff!important;color:#172033!important;box-shadow:0 7px 18px rgba(0,0,0,.18)!important}
header .actions{align-items:center!important}
header .actions button{background:rgba(255,255,255,.07)!important;color:#dbe7f7!important;border:1px solid rgba(255,255,255,.11)!important;box-shadow:none!important}
header .actions button:hover{background:rgba(255,255,255,.14)!important;color:#fff!important}
header .actions .secondary{background:rgba(91,92,240,.30)!important;border-color:rgba(135,137,255,.38)!important}
header .actions .dark{background:#fff!important;color:#172033!important}
main{max-width:1580px!important;padding:20px 20px 34px!important}
.system-strip{gap:12px!important;margin-bottom:14px!important}
.sys-pill{
  min-height:64px;padding:11px 14px 10px 17px!important;border:1px solid rgba(215,224,237,.92)!important;
  border-radius:16px!important;background:rgba(255,255,255,.88)!important;box-shadow:0 8px 24px rgba(15,23,42,.055)!important;
}
.sys-pill:before{width:4px!important;background:linear-gradient(180deg,var(--accent),var(--accent-2))!important}
.sys-pill b{font-size:12px!important;color:#172033!important;margin-bottom:2px}.sys-pill span{color:#8290a5!important}
.page-intro{
  display:flex;align-items:center;justify-content:space-between;gap:24px;
  min-height:88px;margin-bottom:16px;padding:15px 18px 15px 20px;
  color:#fff;border-radius:20px;overflow:hidden;position:relative;
  background:
    radial-gradient(circle at 88% 10%,rgba(70,219,202,.24),transparent 28%),
    linear-gradient(120deg,#4f46e5,#5b5cf0 48%,#138c83);
  box-shadow:0 18px 42px rgba(79,70,229,.20);
}
.page-intro:after{content:"";position:absolute;width:190px;height:190px;border:1px solid rgba(255,255,255,.14);border-radius:50%;right:-46px;top:-104px}
.page-intro h2{font-size:22px!important;color:#fff!important;margin:1px 0 2px!important;letter-spacing:-.02em!important}
.page-intro p{margin:0;color:rgba(255,255,255,.76);font-size:12px}.eyebrow{font-size:10px;letter-spacing:.16em;font-weight:900;color:#c9fffa}
.flow-steps{display:flex;align-items:center;gap:9px;position:relative;z-index:1;color:rgba(255,255,255,.78);font-size:12px;font-weight:800}
.flow-steps span{display:flex;align-items:center;gap:6px;padding:7px 10px;border:1px solid rgba(255,255,255,.16);background:rgba(255,255,255,.09);border-radius:999px;white-space:nowrap}
.flow-steps i{display:grid;place-items:center;width:19px;height:19px;border-radius:50%;font-style:normal;background:#fff;color:#4f46e5;font-size:10px}.flow-steps b{opacity:.5}
.intro-badge{display:flex;align-items:center;gap:8px;position:relative;z-index:1;padding:9px 13px;border:1px solid rgba(255,255,255,.18);background:rgba(4,24,36,.16);border-radius:999px;font-size:12px;font-weight:850;white-space:nowrap}
.intro-badge span{width:8px;height:8px;border-radius:50%;background:#70f0bd;box-shadow:0 0 0 5px rgba(112,240,189,.14)}
.leads-intro{grid-column:1/-1;margin-bottom:0!important}
.workspace{gap:16px!important;grid-template-columns:minmax(300px,.86fr) minmax(560px,1.35fr) minmax(390px,1.05fr)!important}
.leads-layout{gap:16px!important;grid-template-columns:minmax(390px,.88fr) minmax(520px,1.12fr)!important}
.card{
  background:rgba(255,255,255,.96)!important;border:1px solid rgba(220,227,238,.98)!important;
  border-radius:19px!important;box-shadow:var(--shadow-sm)!important;padding:17px!important;
}
.card:hover{transform:none!important;border-color:#d5deeb!important;box-shadow:0 16px 42px rgba(15,23,42,.085)!important}
.card:before{height:0!important}.topline{margin:-1px 0 15px!important;padding-bottom:11px!important;border-bottom:1px solid #edf1f6!important}
.topline h2,.card>h2{font-size:15px!important;color:#172033!important}.muted-chip{background:#f4f6fb!important;border-color:#e3e8f1!important;color:#64748b!important;box-shadow:none!important}
label{color:#445168!important;font-size:11px!important;letter-spacing:.02em!important}
input,select,textarea{
  min-height:38px!important;background:#fbfcff!important;border:1px solid #dce3ed!important;border-radius:11px!important;
  color:#172033!important;box-shadow:inset 0 1px 2px rgba(15,23,42,.025)!important;
}
input:focus,select:focus,textarea:focus{background:#fff!important;border-color:#7778f2!important;box-shadow:0 0 0 4px rgba(91,92,240,.11)!important}
button{border-radius:11px!important;transition:transform .14s ease,box-shadow .14s ease,background .14s ease!important}
button:hover{transform:translateY(-1px)!important}.primary{background:linear-gradient(135deg,#5b5cf0,#4849d9)!important;box-shadow:0 10px 24px rgba(91,92,240,.24)!important}
.secondary{background:#eeefff!important;color:#4546c7!important;border-color:#dfe0ff!important}.dark{background:#172033!important;color:#fff!important}.ghost{background:#f8fafc!important}
.preview>div{min-width:0}.copybox,.thumbs{background:#f8faff!important;border-color:#e4eaf3!important;border-radius:13px!important}.thumb{border-radius:9px!important}
.country-wrap{background:#f8faff!important;border-color:#e4eaf3!important;border-radius:14px!important}
.queue-wrap,.table-scroll,.lead-file-list{border-color:#e3e9f2!important;border-radius:13px!important}.queue th,.checktable th{background:#f5f7fb!important;color:#59677d!important}.queue td,.checktable td{border-color:#edf1f6!important}
.status{background:linear-gradient(145deg,#0b1220,#111a2d)!important;color:#c7f9e9!important;border-radius:14px!important;box-shadow:inset 0 0 0 1px rgba(255,255,255,.04)!important}
.lead-products{background:#f7f9fd!important;border-color:#e3e9f2!important;border-radius:14px!important;padding:9px!important;gap:7px!important}
.product-check{background:#fff!important;border-color:#e3e9f2!important;border-radius:10px!important;box-shadow:0 3px 9px rgba(15,23,42,.025);transition:border-color .14s ease,box-shadow .14s ease,transform .14s ease}
.product-check:hover{border-color:#bfc7ff!important;box-shadow:0 6px 16px rgba(91,92,240,.08);transform:translateY(-1px)}
.product-check:has(input:checked){border-color:#9fa1ff!important;background:#f4f4ff!important;color:#4143c5!important}
.product-check input{accent-color:#5b5cf0!important}.lead-file{padding:12px 13px!important}.lead-file:hover{background:#f8faff}.lead-file b{font-size:12px;margin-bottom:3px}
.lead-file .secondary{display:inline-flex;align-items:center;justify-content:center;font-weight:800}
.hint{color:#758399!important}.ok{color:#07806b!important}.warn{color:#aa6508!important}.bad{color:#b42318!important}
*{scrollbar-width:thin;scrollbar-color:#c4cfdd transparent}*::-webkit-scrollbar{width:8px;height:8px}*::-webkit-scrollbar-thumb{background:#c4cfdd;border-radius:999px;border:2px solid transparent;background-clip:padding-box}
#ads_page,#leads_page,#ads_intro,#insights_page,#budget_page{animation:workspaceIn .22s ease-out}@keyframes workspaceIn{from{opacity:.35;transform:translateY(5px)}to{opacity:1;transform:none}}
.insights-page{display:grid;gap:14px}.insights-intro{margin-bottom:0!important}.dark-eyebrow{color:#6264d9!important}.insights-toolbar{display:flex;align-items:center;justify-content:space-between;gap:20px;margin:0!important;padding:14px 17px!important}.insights-toolbar h2{margin:2px 0 1px!important}.insights-toolbar p{margin:0}.toolbar-actions{display:flex;align-items:center;gap:9px}.toolbar-actions select{width:135px!important}.button-link{display:inline-flex;align-items:center;justify-content:center;min-height:38px;padding:8px 14px;border-radius:11px;text-decoration:none;font-size:13px;font-weight:800}.button-link.dark{background:#172033;color:#fff;box-shadow:0 9px 20px rgba(15,23,42,.15)}
.insight-kpis{display:grid;grid-template-columns:repeat(5,1fr);gap:12px}.metric-card{position:relative;overflow:hidden;width:100%;min-height:116px;padding:15px 16px;border:1px solid #e0e7f0;border-radius:17px;background:rgba(255,255,255,.96);box-shadow:var(--shadow-sm);color:inherit;text-align:left;cursor:pointer}.metric-card:hover{transform:translateY(-2px);border-color:#bfc5ff;box-shadow:0 13px 28px rgba(57,61,135,.12)}.metric-card.active{border-color:#6667ed;box-shadow:0 0 0 3px rgba(91,92,240,.12),0 13px 28px rgba(57,61,135,.12)}.metric-card:focus-visible{outline:3px solid rgba(91,92,240,.24);outline-offset:2px}.metric-card:after{content:"";position:absolute;width:64px;height:64px;border-radius:50%;right:-24px;top:-25px;background:rgba(91,92,240,.07)}.metric-card span{display:block;color:#758399;font-size:11px;font-weight:800}.metric-card strong{display:inline-block;margin:7px 5px 5px 0;color:#172033;font-size:25px;line-height:1;letter-spacing:-.04em}.metric-card small{color:#93a0b3;font-size:10px}.compare-pill{display:block;width:max-content;margin-top:5px;padding:3px 7px;border-radius:999px;background:#f0f3f8;color:#708096;font-size:9px;font-style:normal;font-weight:850}.compare-pill.up{background:#e9f9f3;color:#07805f}.compare-pill.down{background:#fff1ef;color:#b04437}.compare-pill.neutral{background:#eef1ff;color:#5557cd}.accent-metric{color:#fff;border-color:transparent;background:linear-gradient(135deg,#4f46e5,#168f87);box-shadow:0 14px 34px rgba(79,70,229,.20)}.accent-metric:hover,.accent-metric.active{border-color:#fff;box-shadow:0 0 0 3px rgba(79,70,229,.16),0 16px 34px rgba(79,70,229,.24)}.accent-metric span,.accent-metric strong,.accent-metric small{color:#fff!important}.accent-metric small{opacity:.72}.accent-metric .compare-pill{background:rgba(255,255,255,.15);color:#fff}.accent-metric:after{background:rgba(255,255,255,.12)}
.insights-main{margin:0!important;min-height:470px}.insights-head{display:flex;align-items:center;justify-content:space-between;gap:16px;padding-bottom:13px;border-bottom:1px solid #edf1f6}.insights-head h2{margin:0 0 3px!important}.insights-head p{margin:0}.insights-controls{display:flex;align-items:center;gap:9px}.insights-controls select{width:145px!important}.insights-controls .campaign-filter{width:220px!important}.insights-controls .country-filter{width:155px!important}.view-switch{display:flex;padding:3px;border:1px solid #e0e6ef;border-radius:12px;background:#f5f7fb}.view-switch button{padding:7px 12px!important;background:transparent!important;color:#77859a!important;box-shadow:none!important}.view-switch button.active{background:#fff!important;color:#3e40c2!important;box-shadow:0 5px 13px rgba(15,23,42,.09)!important}
.insights-empty{display:flex;align-items:center;justify-content:center;gap:9px;min-height:350px;color:#7b899d}.loader-dot{width:9px;height:9px;border-radius:50%;background:#7778f2;box-shadow:0 0 0 6px rgba(91,92,240,.11)}.insights-empty.loading .loader-dot{animation:pulseData 1s infinite}@keyframes pulseData{50%{transform:scale(.6);opacity:.45}}
.chart-wrap{margin-top:15px;padding:12px 14px 6px;border:1px solid #e5eaf2;border-radius:15px;background:linear-gradient(180deg,#fbfcff,#f7f9fd)}#insights_chart{display:block;width:100%;height:350px}.chart-legend{display:flex;justify-content:center;gap:18px;padding:4px 0 7px;color:#657389;font-size:11px}.chart-legend span{display:flex;align-items:center;gap:6px}.chart-legend i{width:17px;height:3px;border-radius:999px;background:#5b5cf0}.chart-legend i.previous{height:0;border-top:2px dashed #94a3b8;background:transparent}.table-tools{display:flex;align-items:center;justify-content:space-between;gap:12px;margin:14px 0 9px}.table-tools input{max-width:320px}.insights-table-scroll{max-height:390px;overflow:auto;border:1px solid #e3e9f2;border-radius:13px}.insights-table{min-width:1120px}.insights-table td:nth-child(2),.insights-table td:nth-child(3){max-width:280px}.insights-note{margin:11px 1px 0;color:#8a96a8;font-size:10px}.number-cell{font-variant-numeric:tabular-nums;white-space:nowrap}.detail-compare{display:grid;gap:2px;min-width:112px;font-size:10px;line-height:1.35}.detail-compare span{display:flex;justify-content:space-between;gap:8px}.detail-compare b{color:#7c899b;font-weight:750}.empty-row{text-align:center!important;color:#8491a4!important;padding:36px!important}.campaign-summary-head{display:flex;align-items:center;justify-content:space-between;gap:14px;margin:14px 1px 10px}.campaign-summary-head h3{margin:0 0 3px;color:#172033;font-size:14px}.campaign-summary-head p{margin:0}.campaign-table{min-width:1240px}.campaign-link{padding:0!important;background:transparent!important;color:#4546c7!important;text-align:left!important;white-space:normal!important;box-shadow:none!important}.campaign-link:hover{text-decoration:underline;transform:none!important}.delta-up{color:#07805f!important}.delta-down{color:#b04437!important}.delta-neutral{color:#69778c!important}
.automation-card{margin:0!important}.automation-head{display:flex;align-items:center;justify-content:space-between;gap:18px;padding-bottom:13px;border-bottom:1px solid #edf1f6}.automation-head h2{margin:2px 0 3px!important}.automation-head p{margin:0}.automation-actions{display:flex;align-items:center;justify-content:flex-end;gap:8px;flex-wrap:wrap}.automation-mode-select{width:auto;min-width:126px;background:#fff}.rule-strip{display:grid;grid-template-columns:repeat(4,1fr);gap:9px;margin:13px 0}.rule-strip span{padding:10px 12px;border:1px solid #e4e9f1;border-radius:12px;background:#f8faff;color:#64748b;font-size:11px}.rule-strip b{display:block;margin-bottom:2px;color:#3f42c4;font-size:13px}.automation-summary{padding:10px 12px;border-radius:11px;background:#f5f7fb;margin-bottom:10px}.automation-table-wrap{max-height:430px;overflow:auto;border:1px solid #e3e9f2;border-radius:13px}.automation-table{min-width:1180px}.rule-status{display:inline-flex;padding:4px 8px;border-radius:999px;font-size:10px;font-weight:900;white-space:nowrap}.rule-status.warn{background:#fff5dd;color:#9a5d06}.rule-status.close{background:#ffefed;color:#b23d32}.rule-status.paused{background:#e9f9f3;color:#08795d}.rule-status.scale{background:#e9f8ef;color:#087646}.rule-status.info{background:#edf2ff;color:#3846b5}.automation-recreate{white-space:nowrap}.automation-row-actions{display:flex;gap:6px;flex-wrap:wrap;min-width:190px}.automation-batch-toolbar .actions{flex-wrap:wrap}.effectiveness-panel{margin-top:14px;padding-top:14px;border-top:1px solid #e5eaf2}.effectiveness-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:9px}.effectiveness-head h3{margin:0}.effectiveness-table-wrap{max-height:320px;overflow:auto;border:1px solid #e3e9f2;border-radius:13px}.effectiveness-table{min-width:1040px}.effectiveness-kpis{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:9px}.effectiveness-kpis span{padding:7px 10px;border-radius:999px;background:#f1f4fa;color:#526075;font-size:11px}.effectiveness-kpis b{color:#23266f}
.automation-batch-toolbar{margin:0 0 10px!important;padding:9px 10px;border:1px solid #e3e8f1;border-radius:12px;background:#f8faff}.automation-batch-toolbar .automation-product-filter{width:auto;min-width:170px;background:#fff}.automation-filter-count{white-space:nowrap}.automation-batch-table{min-width:900px}.automation-batch-note{margin:0 0 10px}.automation-select-cell{width:48px;text-align:center!important}.automation-select-cell input{width:auto}
.fatigue-card{margin:0!important}.fatigue-head{display:flex;align-items:center;justify-content:space-between;gap:18px;margin-bottom:12px}.fatigue-head h2{margin:2px 0 3px!important}.fatigue-head p{margin:0}.fatigue-table-wrap{max-height:330px;overflow:auto;border:1px solid #e3e9f2;border-radius:13px}.fatigue-table{min-width:1220px}.fatigue-status{display:inline-flex;padding:4px 8px;border-radius:999px;font-size:10px;font-weight:900}.fatigue-status.severe{background:#ffefed;color:#b23d32}.fatigue-status.warning{background:#fff5dd;color:#9a5d06}
.insight-section-tabs{display:flex;gap:8px;padding:7px;border:1px solid #dfe6f0;border-radius:16px;background:#f3f6fb;position:sticky;top:8px;z-index:12;box-shadow:0 8px 24px rgba(31,42,68,.08)}.insight-section-tabs button{flex:1;min-height:44px;border:0;border-radius:11px;background:transparent;color:#64748b;font-weight:900;cursor:pointer}.insight-section-tabs button.active{background:#fff;color:#3538ad;box-shadow:0 5px 15px rgba(48,54,120,.13)}.winner-card{margin:0!important}.winner-table-wrap{max-height:430px;overflow:auto;border:1px solid #e3e9f2;border-radius:13px}.winner-table{min-width:1080px}.winner-rank{display:inline-flex;align-items:center;justify-content:center;min-width:32px;height:28px;border-radius:9px;background:#eef1ff;color:#4548bd;font-weight:950}.winner-rank.top{background:linear-gradient(135deg,#fff0bd,#ffd269);color:#7b4d00}.winner-confidence{display:inline-flex;padding:4px 8px;border-radius:999px;background:#e8f8f1;color:#08795d;font-size:10px;font-weight:900}.winner-confidence.candidate{background:#eef1ff;color:#5557cd}
.budget-import-bar{display:grid;grid-template-columns:minmax(280px,1fr) auto;align-items:end;gap:18px;margin-bottom:12px}.budget-import-bar .actions{justify-content:flex-end}.budget-kpis{display:grid;grid-template-columns:repeat(5,1fr);gap:9px;margin:10px 0}.budget-kpi{padding:11px 12px;border:1px solid #e3e8f1;border-radius:13px;background:linear-gradient(180deg,#fff,#f8faff)}.budget-kpi span{display:block;color:#758399;font-size:10px;font-weight:800}.budget-kpi b{display:block;margin-top:4px;color:#172033;font-size:17px}.budget-table{min-width:1480px}.budget-table td:first-child{max-width:300px}.budget-source{display:inline-flex;padding:3px 7px;border-radius:999px;background:#eef1ff;color:#4d4fc5;font-size:9px;font-weight:850}.budget-source.manual{background:#f0f3f8;color:#68778c}.budget-source.exhausted{background:#ffefed;color:#b23d32}.pace-pill{display:inline-flex;padding:4px 8px;border-radius:999px;font-size:10px;font-weight:900}.pace-pill.on_track{background:#e9f9f3;color:#08795d}.pace-pill.slow{background:#fff5dd;color:#9a5d06}.pace-pill.overspend{background:#ffefed;color:#b23d32}.coverage-ok{color:#08795d;font-weight:850}.coverage-missing{color:#b23d32;font-weight:850}.quick-select-toolbar{display:flex;align-items:center;justify-content:space-between;gap:12px;margin:10px 0}.quick-product-controls{display:grid;grid-template-columns:repeat(auto-fit,minmax(min(100%,360px),1fr));gap:9px;margin-bottom:11px}.quick-product-card{display:grid;min-width:0;grid-template-columns:minmax(0,1fr) minmax(0,1.2fr) minmax(92px,.72fr) auto;align-items:center;gap:8px;padding:9px 10px;border:1px solid #e3e8f1;border-radius:12px;background:#f8faff}.quick-product-card>*{min-width:0}.quick-product-card b{color:#3033b5;overflow-wrap:anywhere}.quick-product-card select{width:100%!important;max-width:100%}.quick-product-card button{white-space:normal}.quick-launch-table{min-width:900px}.quick-launch-table td{overflow-wrap:anywhere}.quick-launch-table select{width:100%;min-width:145px;max-width:220px}.quick-launch-dialog{width:min(1180px,calc(100vw - 24px))!important}.quick-launch-dialog .dialog-card{display:flex;max-height:calc(100vh - 24px);flex-direction:column;overflow:hidden!important}.quick-launch-dialog .dialog-body{display:flex;min-height:0;flex:1;flex-direction:column;overflow:hidden!important;padding-bottom:0!important}.quick-launch-content{min-height:0;flex:1;overflow:auto;padding-right:3px}.quick-launch-content .insights-table-scroll{max-height:none}.quick-launch-footer{display:flex;flex:0 0 auto;align-items:end;justify-content:space-between;gap:18px;margin:0 -18px;padding:13px 18px;border-top:1px solid #e0e7f0;background:#fff;box-shadow:0 -10px 24px rgba(15,23,42,.06)}.quick-launch-footer .field{max-width:240px;margin:0}.quick-launch-footer .actions{justify-content:flex-end;flex-wrap:wrap}.quick-launch-confirm{min-width:150px;font-weight:900}.quick-launch-warning{flex:0 0 auto;margin:0 -18px;padding:8px 18px 11px;background:#fff}
.run-error-dialog{width:min(640px,calc(100vw - 24px))!important}.run-error-head{background:linear-gradient(135deg,#7f1d1d,#b42318)!important}.run-error-banner{display:flex;align-items:center;justify-content:space-between;gap:12px;padding:11px 13px;border:1px solid #fecaca;border-radius:12px;background:#fff1f2;color:#991b1b}.run-error-banner span{color:#b45353;font-size:11px}.run-error-message{min-height:90px;max-height:260px;margin:12px 0;padding:13px;overflow:auto;border:1px solid #e2e8f0;border-radius:12px;background:#0f172a;color:#fecaca;white-space:pre-wrap;overflow-wrap:anywhere;font:12px/1.55 ui-monospace,SFMono-Regular,Consolas,monospace}.quick-launch-product-card{grid-template-columns:minmax(145px,1fr) minmax(180px,1.35fr) auto}.quick-product-meta{display:grid;gap:3px}.quick-product-meta span{color:#7b8798;font-size:10px;overflow-wrap:anywhere}.quick-launch-table{width:100%;min-width:760px;table-layout:fixed}.quick-launch-table th:nth-child(1){width:58px}.quick-launch-table th:nth-child(2){width:22%}.quick-launch-table th:nth-child(3){width:22%}.quick-launch-table th:nth-child(4){width:21%}.quick-launch-table td{overflow-wrap:anywhere;vertical-align:top}.quick-launch-table select{width:100%;min-width:0;max-width:none}.budget-language-note{display:inline-flex;margin-top:4px;padding:2px 6px;border-radius:999px;background:#eef1ff;color:#5557c8;font-size:9px;font-weight:850}.budget-language-note.bad{background:#ffefed;color:#b23d32}.quick-material-error{margin-top:4px;color:#b23d32;font-size:10px}.quick-launch-dialog{width:min(1120px,calc(100vw - 24px))!important}.quick-launch-content .insights-table-scroll{max-height:340px}.budget-page{display:grid;min-width:0;gap:14px}.budget-intro{margin-bottom:0!important}.report-download-grid{display:grid;min-width:0;grid-template-columns:1fr 1fr;gap:14px}.report-download-card{position:relative;display:flex;min-width:0;min-height:150px;overflow:hidden;flex-direction:column;justify-content:center;padding:22px 24px;border-radius:19px;color:#fff;text-decoration:none;box-shadow:0 16px 36px rgba(15,23,42,.16);transition:transform .15s ease,box-shadow .15s ease}.report-download-card:hover{transform:translateY(-2px);box-shadow:0 20px 42px rgba(15,23,42,.22)}.report-download-card:after{content:"";position:absolute;width:180px;height:180px;right:-55px;top:-75px;border-radius:50%;background:rgba(255,255,255,.10)}.report-download-card.weekly{background:linear-gradient(135deg,#4f46e5,#168f87)}.report-download-card.monthly{background:linear-gradient(135deg,#172033,#334155 58%,#a16207)}.report-download-card span{font-size:10px;font-weight:900;letter-spacing:.16em;opacity:.72}.report-download-card strong{margin:7px 0 4px;font-size:21px;letter-spacing:-.02em}.report-download-card small{max-width:560px;color:rgba(255,255,255,.75);font-size:11px}.report-download-card b{margin-top:12px;font-size:12px}.budget-workspace-card{min-width:0;max-width:100%;margin:0!important;padding:20px!important;overflow:hidden}.budget-page-head{display:flex;min-width:0;align-items:center;justify-content:space-between;gap:18px;margin-bottom:14px;padding-bottom:14px;border-bottom:1px solid #e8edf4}.budget-page-head h2{margin:2px 0 3px!important}.budget-page-head p{margin:0}.budget-page-head .actions{justify-content:flex-end}.budget-import-panel{display:grid;min-width:0;grid-template-columns:minmax(360px,1fr) auto;align-items:end;gap:14px;margin-bottom:12px;padding:14px 15px;border:1px solid #dde5f0;border-radius:15px;background:linear-gradient(180deg,#fbfdff,#f6f8fc)}.budget-import-panel .field{min-width:0;margin:0}.budget-import-button{min-height:42px}.budget-table-scroll{width:100%;max-width:100%;overflow:auto;border:1px solid #e0e7f0;border-radius:15px;background:#fff}.budget-table-scroll .budget-table{margin:0}.budget-table-scroll th{top:0}.budget-footnote{margin:12px 2px 0}
.run-panel-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-top:18px;margin-bottom:8px}.run-panel-head h2{margin:0!important}.run-panel-head p{margin:3px 0 0}.run-state-badge{display:inline-flex;align-items:center;min-height:30px;padding:6px 11px;border-radius:999px;font-size:11px;font-weight:900}.run-state-badge.idle{background:#eef2f7;color:#64748b}.run-state-badge.running{background:#e8efff;color:#3748bf}.run-state-badge.success{background:#e8f8f1;color:#08795d}.run-state-badge.error{background:#ffefed;color:#b23d32}.run-status-panel{overflow:hidden;border:2px solid #dce4f0;border-radius:16px;background:#f8faff;box-shadow:0 10px 24px rgba(15,23,42,.06)}.run-status-panel details{border:0!important;border-radius:0!important}.run-status-panel summary{min-height:46px;padding:14px 16px!important;font-weight:900}.run-status-panel .inside{border-top:1px solid #e4e9f1}.run-help{border-top:1px solid #e1e7f0!important}.run-help-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;padding:0 14px 14px}.run-help-grid div{display:grid;gap:3px;padding:10px 11px;border:1px solid #e2e8f0;border-radius:11px;background:#fff}.run-help-grid b{color:#3033b5;font-size:11px}.run-help-grid span{color:#68778c;font-size:10px;line-height:1.5}.refresh-time-chip{grid-column:1/-1;justify-self:end;padding:6px 10px;border-radius:999px;background:#eef2ff;color:#4c51bf;font-size:10px;font-weight:850}.product-line-filter{min-width:150px!important}.run-summary-dialog{width:min(650px,calc(100vw - 24px))!important}.run-summary-head{background:linear-gradient(135deg,#173b69,#3157c8)!important}.run-summary-kpis{display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin-bottom:12px}.run-summary-kpis div{padding:12px 9px;border:1px solid #e2e8f0;border-radius:12px;background:#f8faff;text-align:center}.run-summary-kpis span{display:block;color:#778398;font-size:10px}.run-summary-kpis b{display:block;margin-top:4px;color:#1e2b44;font-size:21px}.automation-select-cell{min-width:64px!important;height:50px;cursor:pointer}.automation-select-cell input{width:24px!important;height:24px!important;margin:0;cursor:pointer;accent-color:#4f46e5}.automation-table tbody .automation-select-cell:hover{background:#eef1ff}.automation-card{padding:24px!important}.automation-head{padding:4px 4px 17px}.automation-actions button,.automation-actions select{min-height:42px}.automation-summary{min-height:44px;padding:13px 15px}
@media(max-width:1260px){
  header{align-items:stretch!important}.brand-area{grid-template-columns:1fr auto}.workspace,.leads-layout{grid-template-columns:1fr!important}.page-intro{min-height:80px}.insight-kpis{grid-template-columns:repeat(3,1fr)}.automation-head{align-items:flex-start;flex-direction:column}.automation-actions{justify-content:flex-start}.rule-strip{grid-template-columns:1fr 1fr}
}
@media(max-width:760px){
  header{padding:12px 14px!important}.brand-area{display:flex;flex-wrap:wrap;gap:5px 12px}.brand-area p{width:100%}.app-nav{order:3;width:100%;margin-top:6px!important}.nav-btn{flex:1}
  main{padding:14px 10px 26px!important}.system-strip{grid-template-columns:1fr 1fr!important}.page-intro{align-items:flex-start;flex-direction:column;gap:10px}.flow-steps{width:100%;justify-content:space-between}.flow-steps span{padding:6px 8px}
  .lead-products{grid-template-columns:1fr!important}.insights-toolbar,.insights-head,.fatigue-head,.budget-page-head{align-items:flex-start;flex-direction:column}.toolbar-actions,.insights-controls{width:100%;flex-wrap:wrap}.toolbar-actions>*{flex:1}.insight-kpis{grid-template-columns:1fr 1fr}.metric-card{min-height:94px}.insights-controls select,.insights-controls .campaign-filter,.insights-controls .country-filter{flex:1;width:auto!important}.view-switch{flex:1}.view-switch button{flex:1}.chart-wrap{padding:5px}#insights_chart{height:290px}.rule-strip{grid-template-columns:1fr}.automation-actions>*{flex:1}.report-download-grid{grid-template-columns:1fr}.budget-import-panel{grid-template-columns:1fr}.budget-import-button{width:100%}.budget-page-head .actions{justify-content:flex-start}.budget-kpis{grid-template-columns:1fr 1fr}.quick-product-controls{grid-template-columns:1fr}.quick-product-card{grid-template-columns:1fr}.quick-launch-footer{align-items:stretch;flex-direction:column}.quick-launch-footer .field{max-width:none;width:100%}.quick-launch-footer .actions{display:grid;width:100%;grid-template-columns:1fr 1fr}.quick-launch-footer .actions button{width:100%}
}
</style>
</head>
<body>
<a class="skip-nav" href="#main">跳到主内容</a>
<header>
  <div class="brand-area">
    <h1>ADS 工作台</h1>
    <p id="header_subtitle">批量投放、客户信息同步集中在一个网页里。</p>
    <div class="app-nav">
      <button id="nav_ads" class="nav-btn active" onclick="switchPage('ads')">批量投放</button>
      <button id="nav_leads" class="nav-btn" onclick="switchPage('leads')">客户信息</button>
      <button id="nav_insights" class="nav-btn" onclick="switchPage('insights')">广告数据</button>
      <button id="nav_budget" class="nav-btn" onclick="switchPage('budget')">预算中心</button>
    </div>
  </div>
  <div class="actions" id="ads_header_actions">
    <button class="ghost" onclick="refreshOptions()">刷新素材包</button>
    <button class="ghost" onclick="openDialog('dlg_batch')">批量导入</button>
    <button class="ghost" onclick="openDialog('dlg_pack')">素材工具</button>
    <button class="ghost" onclick="openDialog('dlg_adv')">高级设置</button>
    <button class="secondary" onclick="preflight()">投放前检查</button>
    <button class="dark" onclick="checkStatus()">刷新状态</button>
  </div>
</header>
<main id="main">
  <div class="system-strip full">
    <div class="sys-pill"><b id="sys_token">Token：检测中</b><span>token.txt / 环境变量</span></div>
    <div class="sys-pill"><b id="sys_assets">素材包：检测中</b><span>ADS/素材包</span></div>
    <div class="sys-pill"><b id="sys_forms">表单库：检测中</b><span>_cache/forms_library.json</span></div>
    <div class="sys-pill"><b id="sys_mode">模式：-</b><span>Dry Run / LIVE</span></div>
    <div class="sys-pill"><b id="sys_ads">广告状态：ACTIVE</b><span>默认投放后打开</span></div>
  </div>
  <div id="ads_intro" class="page-intro">
    <div><span class="eyebrow">CAMPAIGN STUDIO</span><h2>批量投放</h2><p>从素材检查到广告创建，把重复步骤压缩成一条清晰流程。</p></div>
    <div class="flow-steps"><span><i>1</i>配置</span><b>→</b><span><i>2</i>检查</span><b>→</b><span><i>3</i>执行</span></div>
  </div>
  <div class="workspace" id="ads_page">
    <section class="card panel panel-left">
      <div class="topline"><h2>快速配置</h2><span class="muted-chip">主流程留在一屏</span></div>
      <div class="row">
        <div class="field"><label>每日预算（USD，可留空）</label><input id="daily_budget" placeholder="留空则按季度剩余预算自动计算" oninput="updateAutoBudgetHint()"><p id="auto_budget_hint" class="hint">自动预算 =（季度初始预算 − 已花费）÷ 季度剩余天数</p></div>
        <div class="field"><label>Dry Run</label><select id="dry_run"><option value="no">正式创建（会投到 Meta）</option><option value="yes">只预检 Dry Run（不创建）</option></select></div>
      </div>
      <div class="row">
        <div class="field"><label>CTA</label><select id="cta_type"><option>GET_QUOTE</option><option>LEARN_MORE</option><option>SIGN_UP</option><option>CONTACT_US</option></select></div>
        <div class="field"><label>语言模式</label><select id="language_mode" onchange="toggleCustomLanguage()"><option value="auto">auto 按国家官方语言</option><option value="all">all 不限制语言</option><option value="custom">custom 手动指定语言</option></select></div>
      </div>
      <div class="field" id="custom_language_wrap" style="display:none"><label>自定义语言</label><input id="custom_language" placeholder="例如 EN,FR 或 English,French"><p class="hint">仅在 language_mode=custom 生效，支持逗号或加号分隔多个语言。</p></div>
      <div class="field"><label>Website URL</label><input id="website_url" readonly placeholder="选择产品后自动匹配官网产品页"></div>
      <div class="field country-wrap" id="country_wrap">
        <label>国家 <span id="country_count" class="mini">已选 0</span></label>
        <div class="row">
          <div class="field"><input id="country_set_name" placeholder="给这组国家起个名字，例如 北美 / 欧洲5国"></div>
          <div class="field"><button class="secondary" style="width:100%" onclick="saveCountrySet()">保存国家组合</button></div>
        </div>
        <div id="country_presets" class="saved-country-sets"></div>
        <select id="country_dropdown" onchange="addCountryFromDropdown()"><option value="">从下拉列表选择国家，选中后自动加入输入框</option></select>
        <div class="manual-country"><label>国家代码输入框</label><input id="custom_countries" placeholder="例如 US,CA,FR，也可用空格/逗号分隔" oninput="updateCountryCount()"></div>
        <div class="actions" style="margin-top:10px">
          <button class="ghost" onclick="clearCountries()">清空</button>
        </div>
      </div>
      <div class="row">
        <div class="field"><label>多国家怎么建</label><select id="country_mode"><option value="merge">合并 1 个广告组，共用预算</option><option value="split">按国家拆成多个广告组</option></select></div>
        <div class="field"><label>&nbsp;</label><button class="primary" style="width:100%" onclick="addItem()">加入队列</button></div>
      </div>
      <p class="hint warn">正式创建前建议先 Dry Run。多国家合并时若表单语言不同，默认使用 EN 表单。</p>
    </section>
    <section class="card panel panel-mid">
      <div class="topline"><h2>素材预览</h2><span class="hint">左边是文案，右边是图片</span></div>
      <div class="field">
        <label>产品与素材</label>
        <div class="row3">
          <div class="field"><select id="product" onchange="updateSellingPoints()"></select></div>
          <div class="field"><select id="selling_point" onchange="updateAssetLanguages()"></select><div id="asset_note" class="hint ok"></div></div>
          <div class="field"><select id="asset_language" onchange="loadPreview()"></select></div>
        </div>
      </div>
      <div class="preview">
        <div><label>文案预览</label><div id="copy_preview" class="copybox">请选择素材包</div></div>
        <div><label>图片预览</label><div id="image_preview" class="thumbs"></div></div>
      </div>
      <div class="field" style="margin-top:12px">
        <label>当前素材说明</label>
        <div class="copybox" style="max-height:none;min-height:100px" id="asset_help">这里会显示当前产品、卖点和语种是否齐全。你也可以通过右上角“素材工具”去检查素材包、Hash 和运行日志。</div>
      </div>
      <div class="actions">
        <button class="secondary" onclick="openDialog('dlg_pack')">打开素材工具</button>
        <button class="ghost" onclick="refreshOptions()">重新读取素材</button>
      </div>
    </section>
    <section class="card panel panel-right sticky-actions">
      <div class="topline"><h2>队列与执行</h2><span class="hint" id="queue_hint">等待添加</span></div>
      <div class="split-actions">
        <button class="secondary" onclick="preflight()">投放前检查</button>
        <button class="secondary" onclick="generatePlan(false)">生成 plan.xlsx</button>
        <button class="dark" onclick="generatePlan(true)">生成并运行</button>
        <button class="ghost" onclick="checkStatus()">刷新状态</button>
      </div>
      <h2 style="margin-top:14px;margin-bottom:8px">当前队列</h2>
      <div class="queue-wrap"><table class="queue"><thead><tr><th>#</th><th>产品</th><th>卖点</th><th>语种</th><th>国家</th><th>预算</th><th>模式</th><th>操作</th></tr></thead><tbody id="queue"></tbody></table></div>
      <h2 style="margin-top:14px;margin-bottom:8px">投放前检查结果</h2>
      <div class="inside preflight-box"><div id="preflight_result" class="hint">还没有检查。</div></div>
      <div class="run-panel-head"><div><h2>运行状态</h2><p id="run_state_meta" class="hint">尚未开始投放</p></div><span id="run_state_badge" class="run-state-badge idle">等待运行</span></div>
      <div class="run-status-panel">
        <details open><summary>实时运行日志</summary><div class="inside"><div id="status" class="status">等待操作...</div></div></details>
        <details class="run-help" open><summary>常见报错与解决方法</summary><div class="run-help-grid"><div><b>预算过低</b><span>日均低于 $1.01 会暂缓投放，等待预算达到门槛。</span></div><div><b>地区已禁投</b><span>台湾 TW、新加坡 SG 本季度不能生成计划。</span></div><div><b>素材或表单缺失</b><span>补齐对应产品、卖点、语种的图片、TXT 文案与 Lead Form。</span></div><div><b>Invalid parameter</b><span>查看弹窗中的中文具体原因；修复后重新做投放前检查。</span></div></div></details>
      </div>
    </section>
  </div>
  <section id="leads_page" class="leads-layout" hidden>
    <div class="page-intro leads-intro">
      <div><span class="eyebrow">LEAD OPERATIONS</span><h2>客户信息</h2><p>按产品和时间范围同步 Meta 表单，结果自动整理成可下载的 Excel。</p></div>
      <div class="intro-badge"><span></span>增量同步已启用</div>
    </div>
    <div>
      <div class="card">
        <div class="topline"><h2>同步客户信息</h2><span class="muted-chip" id="lead_mode_badge">默认增量同步</span></div>
        <p class="hint">选择产品和时间范围。产品全部不选时，会同步表单库中的全部产品。</p>
        <div class="lead-toolbar">
          <input id="lead_search" style="max-width:260px" placeholder="搜索产品" oninput="renderLeadProducts()">
          <button class="ghost" onclick="setAllLeadProducts(true)">全选</button>
          <button class="ghost" onclick="setAllLeadProducts(false)">清空</button>
          <span class="hint" id="lead_selected_count">已选择 0 个</span>
        </div>
        <div id="lead_products" class="lead-products"><span class="hint">正在读取产品...</span></div>
      </div>
      <div class="card">
        <div class="topline"><h2>同步设置</h2><span class="hint">日期均包含当天</span></div>
        <div class="row">
          <div class="field"><label>开始日期（可选）</label><input id="lead_start_date" type="date"></div>
          <div class="field"><label>结束日期（可选）</label><input id="lead_end_date" type="date"></div>
        </div>
        <label class="product-check"><input id="lead_full" type="checkbox">全量重拉（会增加 Meta API 请求，仍按客户 ID 去重）</label>
        <p class="hint warn" style="margin-top:10px">填写日期时属于临时报表导出，不影响增量同步状态；不填日期默认只拉新增客户。</p>
        <div class="actions">
          <button id="lead_start_btn" class="primary" onclick="startLeadSync()">开始同步</button>
          <button class="ghost" onclick="refreshLeadStatus()">刷新状态</button>
        </div>
      </div>
    </div>
    <div>
      <div class="card">
        <div class="topline"><h2>运行状态</h2><span id="lead_status_text" class="hint">等待操作</span></div>
        <div id="lead_log" class="status lead-log">尚未运行客户信息同步。</div>
      </div>
      <div class="card">
        <div class="topline"><h2>客户信息文件</h2><button class="ghost" onclick="refreshLeadFiles()">刷新列表</button></div>
        <p class="hint">每次同步都会生成一个新的 Excel。点击即可下载到电脑。</p>
        <div id="lead_files" class="lead-file-list"><div class="lead-file"><span class="hint">正在读取文件...</span></div></div>
      </div>
    </div>
  </section>
  <section id="insights_page" class="insights-page" hidden>
    <div class="page-intro insights-intro">
      <div><span class="eyebrow">PERFORMANCE ANALYTICS</span><h2>广告组数据</h2><p>每小时自动更新一次，其余时间优先读取本地缓存，避免重复消耗 Meta API。</p></div>
      <div class="intro-badge"><span></span>60 分钟缓存 · 广告组层级</div>
    </div>
    <div class="insights-toolbar card">
      <div>
        <span class="eyebrow dark-eyebrow">DATE RANGE</span>
        <h2>数据范围</h2>
        <p class="hint" id="insights_range">默认最近 7 天（包含今天）</p>
      </div>
      <div class="toolbar-actions">
        <select id="insights_days" onchange="resetInsights()"><option value="7" selected>最近 7 天</option><option value="14">最近 14 天</option><option value="30">最近 30 天</option></select>
        <button id="insights_refresh_btn" class="primary" title="跳过一小时缓存并立即请求 Meta" onclick="loadInsights(true)">强制刷新</button>
        <a id="insights_download" class="button-link dark" href="/api/insights/export?days=7">下载 Excel</a>
        <a class="button-link secondary" href="/api/report/export?period=week">下载周报</a>
        <a class="button-link secondary" href="/api/report/export?period=month">下载月报</a>
      </div>
      <div id="insights_last_refresh" class="refresh-time-chip">上次刷新：尚未读取</div>
    </div>
    <div class="insight-section-tabs" role="tablist" aria-label="数据面板分类">
      <button id="insight_section_overview" class="active" onclick="setInsightsSection('overview')">数据总览</button>
      <button id="insight_section_creative" onclick="setInsightsSection('creative')">素材表现</button>
      <button id="insight_section_automation" onclick="setInsightsSection('automation')">自动化规则</button>
    </div>
    <div class="insight-kpis" data-insight-section="overview">
      <button type="button" class="metric-card active" data-metric="spend" onclick="selectMetricFromCard('spend')"><span>花费</span><strong id="metric_spend">—</strong><small>USD</small><em id="metric_spend_compare" class="compare-pill">环比 —</em></button>
      <button type="button" class="metric-card" data-metric="impressions" onclick="selectMetricFromCard('impressions')"><span>展示</span><strong id="metric_impressions">—</strong><small>Impressions</small><em id="metric_impressions_compare" class="compare-pill">环比 —</em></button>
      <button type="button" class="metric-card" data-metric="clicks" onclick="selectMetricFromCard('clicks')"><span>点击</span><strong id="metric_clicks">—</strong><small>Clicks</small><em id="metric_clicks_compare" class="compare-pill">环比 —</em></button>
      <button type="button" class="metric-card" data-metric="leads" onclick="selectMetricFromCard('leads')"><span>Leads</span><strong id="metric_leads">—</strong><small>Meta Leads</small><em id="metric_leads_compare" class="compare-pill">环比 —</em></button>
      <button type="button" class="metric-card accent-metric" data-metric="cost_per_lead" onclick="selectMetricFromCard('cost_per_lead')"><span>单条线索成本</span><strong id="metric_cpl">—</strong><small>USD / Lead</small><em id="metric_cpl_compare" class="compare-pill">环比 —</em></button>
    </div>
    <div class="card insights-main" data-insight-section="overview">
      <div class="insights-head">
        <div><h2>广告表现</h2><p class="hint" id="insights_updated">点击“读取最新数据”开始</p></div>
        <div class="insights-controls">
          <select id="insights_product_filter" class="product-line-filter" onchange="applyInsightsProduct()"><option value="">全部产品线</option></select>
          <select id="insights_campaign_filter" class="campaign-filter" onchange="applyInsightsCampaign()"><option value="">全部广告系列</option></select>
          <select id="insights_country_filter" class="country-filter" onchange="applyInsightsCountry()"><option value="">全部国家</option></select>
          <div class="view-switch"><button id="view_chart_btn" class="active" onclick="setInsightsView('chart')">折线图</button><button id="view_campaign_btn" onclick="setInsightsView('campaign')">广告系列</button><button id="view_table_btn" onclick="setInsightsView('table')">明细表</button></div>
          <select id="chart_metric" onchange="metricSelectChanged()"><option value="spend">花费</option><option value="leads">Leads</option><option value="cost_per_lead">单条线索成本</option><option value="clicks">点击</option><option value="impressions">展示</option><option value="ctr">CTR</option></select>
        </div>
      </div>
      <div id="insights_loading" class="insights-empty"><span class="loader-dot"></span>等待读取 Meta 广告数据</div>
      <div id="insights_chart_wrap" class="chart-wrap" hidden>
        <svg id="insights_chart" viewBox="0 0 1000 360" preserveAspectRatio="xMidYMid meet" role="img" aria-label="广告指标趋势折线图"></svg>
        <div id="chart_legend" class="chart-legend"></div>
      </div>
      <div id="insights_table_wrap" hidden>
        <div class="table-tools"><input id="insights_search" placeholder="搜索国家或广告组" oninput="renderInsightsTable()"><span id="insights_row_count" class="hint"></span></div>
        <div class="insights-table-scroll"><table class="checktable insights-table"><thead><tr><th>日期</th><th>国家</th><th>广告组</th><th>花费</th><th>展示</th><th>点击</th><th>CTR</th><th>Leads</th><th>单条线索成本</th><th>环比</th></tr></thead><tbody id="insights_rows"></tbody></table></div>
      </div>
      <div id="insights_campaign_wrap" hidden>
        <div class="campaign-summary-head"><div><h3>按广告系列拆分</h3><p class="hint">每个广告系列作为一个产品分组，点击名称可查看该系列趋势。</p></div><span id="campaign_count" class="muted-chip"></span></div>
        <div class="insights-table-scroll"><table class="checktable campaign-table"><thead><tr><th>广告系列（产品）</th><th>广告组数</th><th>花费</th><th>花费环比</th><th>展示</th><th>点击</th><th>CTR</th><th>Leads</th><th>Leads 环比</th><th>单条线索成本</th><th>成本环比</th></tr></thead><tbody id="campaign_rows"></tbody></table></div>
      </div>
      <p class="insights-note">Leads 优先采用 Meta 的 grouped lead 指标；不同账户归因设置可能导致数字与 Ads Manager 略有差异。</p>
    </div>
    <div class="card fatigue-card" data-insight-section="creative" hidden>
      <div class="fatigue-head"><div><span class="eyebrow dark-eyebrow">CREATIVE HEALTH</span><h2>素材疲劳检测</h2><p class="hint">近 3 日花费达到有效观察量后，对比前后 3 日 CPL；恶化 25% 提醒，恶化 50% 标记高风险。</p></div><span id="fatigue_badge" class="muted-chip">等待数据</span></div>
      <div class="fatigue-table-wrap"><table class="checktable fatigue-table"><thead><tr><th>风险</th><th>广告系列</th><th>国家</th><th>广告组</th><th>近3日花费</th><th>前3日花费</th><th>近3日 Leads</th><th>近3日 CPL</th><th>前3日 CPL</th><th>变化 / 建议</th></tr></thead><tbody id="fatigue_rows"><tr><td colspan="10" class="empty-row">读取广告数据后自动检测</td></tr></tbody></table></div>
    </div>
    <div class="card winner-card" data-insight-section="creative" hidden>
      <div class="fatigue-head"><div><span class="eyebrow dark-eyebrow">WINNING CREATIVE</span><h2>优胜素材</h2><p class="hint">按当前筛选范围汇总卖点与地区；优先展示有稳定留资、花费达到观察量且 CPL 较低的素材。</p></div><span id="winner_badge" class="muted-chip">等待数据</span></div>
      <div class="winner-table-wrap"><table class="checktable winner-table"><thead><tr><th>排名</th><th>产品 / 卖点</th><th>国家</th><th>广告组数</th><th>花费</th><th>Leads</th><th>CPL</th><th>CTR</th><th>可信度</th><th>素材</th></tr></thead><tbody id="winner_rows"><tr><td colspan="10" class="empty-row">读取广告数据后自动排名</td></tr></tbody></table></div>
    </div>
    <div class="card automation-card" data-insight-section="automation" hidden>
      <div class="automation-head">
        <div><span class="eyebrow dark-eyebrow">AUTOMATION CONTROL</span><h2>自动化规则中心</h2><p class="hint">自动停投、预算调速与配置体检统一管理；只检查有效状态为 ACTIVE 的广告组。</p></div>
        <div class="automation-actions"><span id="automation_badge" class="muted-chip">读取设置中</span><select id="automation_mode" class="automation-mode-select" onchange="changeAutomationMode()"><option value="notify">仅提醒</option><option value="approval" selected>待人工确认</option><option value="auto">全自动执行</option></select><button id="automation_toggle" class="secondary" onclick="toggleAutomation()">启用自动化</button><button class="ghost" onclick="previewAutomation()">只读预览</button><button class="dark" onclick="runAutomationNow()">执行全部命中项</button></div>
      </div>
      <div class="rule-strip"><span><b>72 小时保护</b> 新广告只观察，不自动操作</span><span><b>数据截至昨天</b> 避开 Leads 回传延迟</span><span><b>CPL +25% / +50%</b> 提醒换素材 / 自动暂停</span><span><b>3 天 0 Leads</b> 有花费则自动暂停</span><span><b>花费突增 1.5×</b> 当天 0 Leads 自动止损</span><span><b>优质组自动扩量</b> 5 Leads 且 CPL 低 20%</span><span><b>季度预算调速</b> 单次 10%，冷却 24 小时</span><span><b>异常只提醒</b> 不消耗 / 落地页 / 表单 / 语种</span></div>
      <div id="automation_summary" class="hint automation-summary">点击“只读预览”可在不修改广告的情况下查看可能触发的规则。</div>
      <div class="quick-select-toolbar automation-batch-toolbar"><div class="actions"><select id="automation_product_filter" class="automation-product-filter" onchange="changeAutomationProductFilter()"><option value="">全部产品</option></select><button class="secondary" onclick="setAutomationSelection(true)">全选当前产品</button><button class="ghost" onclick="setAutomationSelection(false)">清空选择</button><button id="automation_execute_button" class="primary" onclick="applyAutomationSelected()" disabled>执行所选规则</button><button id="automation_batch_button" class="dark" onclick="openAutomationBatch()" disabled>批量换卖点重新创建</button></div><div class="actions"><span id="automation_filter_count" class="hint automation-filter-count">显示全部</span><span id="automation_selected_count" class="muted-chip">已选择 0 项</span></div></div>
      <div class="automation-table-wrap"><table class="checktable automation-table"><thead><tr><th class="automation-select-cell">选择</th><th>处理</th><th>广告组 / Campaign</th><th>产品 / 地区</th><th>近7天</th><th>触发原因</th><th>操作</th></tr></thead><tbody id="automation_rows"><tr><td colspan="7" class="empty-row">尚未检查</td></tr></tbody></table></div>
      <div class="effectiveness-panel"><div class="effectiveness-head"><div><h3>自动化效果复盘</h3><p class="hint">对比执行前 3 天与执行后 3/7 天；沿用广告数据的一小时缓存。</p></div><button class="secondary" onclick="loadAutomationEffectiveness()">刷新复盘</button></div><div id="effectiveness_summary" class="effectiveness-kpis"><span>点击“刷新复盘”查看执行效果</span></div><div class="effectiveness-table-wrap"><table class="checktable effectiveness-table"><thead><tr><th>结果</th><th>执行时间 / 动作</th><th>广告组</th><th>执行前3天</th><th>执行后3天</th><th>执行后7天</th><th>估算节省</th><th>操作</th></tr></thead><tbody id="effectiveness_rows"><tr><td colspan="8" class="empty-row">尚未读取复盘数据</td></tr></tbody></table></div></div>
    </div>
  </section>
  <section id="budget_page" class="budget-page" hidden>
    <div class="page-intro budget-intro">
      <div><span class="eyebrow">BUDGET OPERATIONS</span><h2>季度预算中心</h2><p>整页查看产品和地区预算、投放覆盖、预算节奏及耗尽预测。</p></div>
      <div class="intro-badge"><span></span>60 分钟缓存 · 当前季度</div>
    </div>
    <div class="report-download-grid">
      <a class="report-download-card weekly" href="/api/report/export?period=week"><span>WEEKLY REPORT</span><strong>一键生成并下载周报</strong><small>最近 7 天 · 产品经营、环比、趋势、素材疲劳、广告组明细</small><b>生成周报 Excel →</b></a>
      <a class="report-download-card monthly" href="/api/report/export?period=month"><span>MONTHLY REPORT</span><strong>一键生成并下载月报</strong><small>最近 30 天 · 产品经营、环比、趋势、素材疲劳、广告组明细</small><b>生成月报 Excel →</b></a>
    </div>
    <div class="card budget-workspace-card">
      <div class="budget-page-head"><div><span class="eyebrow dark-eyebrow">QUARTER PLAN</span><h2>季度预算明细</h2><p class="hint">导入初始预算后，自动扣除已花费并计算剩余每日预算。</p></div><div class="actions"><a class="button-link dark" href="/api/budget/template">下载预算模板</a><button class="secondary" onclick="loadBudgetStatus(true)">刷新预算与投放检测</button><button id="quick_launch_open" class="primary" onclick="openQuickLaunch()">一键投放缺失地区</button></div></div>
      <div class="budget-import-panel">
        <div class="field"><label>导入初始季度预算表</label><input id="quarter_budget_file" type="file" accept=".xlsx,.xlsm,.csv"><p class="hint">至少包含“产品（或 Campaign ID）”“国家/地区”和“初始预算”；“语种”可选，填写后会自动带入该地区的一键投放。</p></div>
        <button class="primary budget-import-button" onclick="importQuarterBudget()">导入并更新预算</button>
      </div>
      <div id="budget_summary" class="automation-summary">尚未读取季度预算表。</div>
      <div id="budget_kpis" class="budget-kpis"><div class="budget-kpi"><span>季度初始预算</span><b>—</b></div><div class="budget-kpi"><span>已花费</span><b>—</b></div><div class="budget-kpi"><span>Leads</span><b>—</b></div><div class="budget-kpi"><span>剩余预算</span><b>—</b></div><div class="budget-kpi"><span>季末预计完成率</span><b>—</b></div></div>
      <div class="budget-table-scroll"><table class="checktable budget-table"><thead><tr><th>产品 / Campaign</th><th>国家 / 地区</th><th>固定语种</th><th>季度预算</th><th>已花 / Leads / CPL</th><th>剩余 / 自动日预算</th><th>预算节奏</th><th>季末与耗尽预测</th><th>投放检测</th><th>预算状态</th></tr></thead><tbody id="budget_rows"><tr><td colspan="10" class="empty-row">尚未读取预算</td></tr></tbody></table></div>
      <p class="hint budget-footnote">投放检测仅统计 ACTIVE 广告组；手动填写每日预算时，手动值优先。</p>
    </div>
  </section>
</main>
<dialog id="dlg_adv" class="ui-dialog">
  <div class="dialog-card">
    <div class="dialog-head">
      <div><h2>高级设置</h2><p>Page / Form / Account / UTM</p></div>
      <button class="ghost" type="button" onclick="closeDialog('dlg_adv')">关闭</button>
    </div>
    <div class="dialog-body">
      <div class="row">
        <div class="field"><label>Page ID</label><input id="page_id"></div>
        <div class="field"><label>Lead Form ID（可空）</label><input id="lead_form_id"></div>
      </div>
      <div class="row">
        <div class="field"><label>投递随机间隔</label><input id="delay_max_seconds" value="60" readonly><p class="hint">已锁定为每个广告组之间随机等待 1–60 秒。</p></div>
        <div class="field"><label>Ad Account ID</label><input id="ad_account_id"></div>
      </div>
      <div class="field"><label>URL Tags / UTM</label><textarea id="url_tags"></textarea><p class="hint">默认会把 {产品名} 替换成当前产品小写。</p></div>
      <p class="hint">Token 放在 token.txt；前端不会显示 token。</p>
    </div>
  </div>
</dialog>
<dialog id="dlg_pack" class="ui-dialog">
  <div class="dialog-card">
    <div class="dialog-head">
      <div><h2>素材工具</h2><p>检查素材包、Hash 和日志</p></div>
      <button class="ghost" type="button" onclick="closeDialog('dlg_pack')">关闭</button>
    </div>
    <div class="dialog-body">
      <div class="row">
        <div class="field"><label>新建产品</label><input id="new_product" placeholder="例如 EG10"></div>
        <div class="field"><label>新建卖点</label><input id="new_selling_point" placeholder="例如 安装便捷"></div>
      </div>
      <div class="field"><label>新建素材语种</label><input id="new_asset_language" placeholder="EN / KR / FR" value="EN"></div>
      <div class="actions">
        <button class="secondary" onclick="createPack()">新建模板</button>
        <button class="ghost" onclick="checkPacks()">检查素材包</button>
        <button class="ghost" onclick="showLogs()">查看日志</button>
        <button class="ghost" onclick="checkHashLibrary()">Hash 状态</button>
        <button class="secondary" onclick="syncImageHashes()">同步图片 Hash</button>
      </div>
      <div id="pack_result" class="hint" style="margin-top:10px">TXT 文案会自动识别：文件名可以自定义；一个文件里用“正文 / 标题 / 描述”分段，或分别放正文.txt、标题.txt、描述.txt。</div>
    </div>
  </div>
</dialog>
<dialog id="dlg_run_error" class="ui-dialog run-error-dialog">
  <div class="dialog-card">
    <div class="dialog-head run-error-head"><div><h2>投放运行已暂停</h2><p>检测到错误，后续计划已停止执行</p></div><button class="ghost" onclick="closeDialog('dlg_run_error')">关闭</button></div>
    <div class="dialog-body">
      <div class="run-error-banner"><b id="run_error_title">运行错误</b><span id="run_error_meta"></span></div>
      <pre id="run_error_message" class="run-error-message"></pre>
      <p class="hint">已经成功完成的广告组不会回滚；未执行的项目仍保留在计划中。修复问题后请先重新做投放前检查，避免重复创建。</p>
      <div class="actions"><button class="primary" onclick="focusRunErrorLog()">查看完整运行日志</button><button class="ghost" onclick="closeDialog('dlg_run_error')">知道了</button></div>
    </div>
  </div>
</dialog>
<dialog id="dlg_run_summary" class="ui-dialog run-summary-dialog">
  <div class="dialog-card">
    <div class="dialog-head run-summary-head"><div><h2>队列投放完成</h2><p id="run_summary_time">本次运行已结束</p></div><button class="ghost" onclick="closeDialog('dlg_run_summary')">关闭</button></div>
    <div class="dialog-body">
      <div id="run_summary_kpis" class="run-summary-kpis"></div>
      <div id="run_summary_message" class="automation-summary"></div>
      <div class="actions"><button class="primary" onclick="focusRunSummaryLog()">查看运行日志</button><button class="secondary" onclick="closeDialog('dlg_run_summary')">完成</button></div>
    </div>
  </div>
</dialog>
<dialog id="dlg_quick_launch" class="ui-dialog quick-launch-dialog">
  <div class="dialog-card dialog-card-wide">
    <div class="dialog-head"><div><h2>缺失地区一键投放</h2><p>为有预算但没有 ACTIVE 广告组的地区快速补投</p></div><button class="ghost" onclick="closeDialog('dlg_quick_launch')">关闭</button></div>
    <div class="dialog-body">
      <div class="quick-launch-content">
        <div class="automation-summary" id="quick_launch_summary">正在整理缺失地区...</div>
        <div class="quick-select-toolbar"><div class="actions"><button class="secondary" onclick="setQuickSelection(true)">全选</button><button class="ghost" onclick="setQuickSelection(false)">清空</button></div><span id="quick_selected_count" class="muted-chip">已选择 0 项</span></div>
        <p class="hint">投放语种直接使用预算表中的固定语种；这里只需要选择卖点。没有对应语种素材的预算项会自动取消勾选并标记。</p>
        <div id="quick_product_controls" class="quick-product-controls"></div>
        <div class="insights-table-scroll"><table class="checktable quick-launch-table"><thead><tr><th>选择</th><th>产品</th><th>地区</th><th>自动日预算</th><th>卖点素材</th></tr></thead><tbody id="quick_launch_rows"></tbody></table></div>
      </div>
      <div class="quick-launch-footer">
        <div class="field"><label>执行模式</label><select id="quick_launch_mode"><option value="yes">先 Dry Run（推荐）</option><option value="no">LIVE 正式投放</option></select></div>
        <div class="actions"><button class="secondary" onclick="addMissingToQueue()">加入队列检查</button><button id="quick_launch_confirm" class="primary quick-launch-confirm" onclick="launchMissingNow()">确定投放</button></div>
      </div>
      <p class="hint warn quick-launch-warning">LIVE 会生成 plan.xlsx 并运行正式投放。系统仍会执行素材、表单、Campaign 和预算检查。</p>
    </div>
  </div>
</dialog>
<dialog id="dlg_batch" class="ui-dialog">
  <div class="dialog-card dialog-card-wide">
    <div class="dialog-head">
      <div><h2>批量导入</h2><p>直接粘贴 Excel 表格内容</p></div>
      <button class="ghost" type="button" onclick="closeDialog('dlg_batch')">关闭</button>
    </div>
    <div class="dialog-body">
      <div class="field"><label>批量投放表</label><textarea id="batch_text" placeholder="可从 Excel 直接粘贴：产品 / 卖点 / 国家 / 预算（可空）"></textarea><p class="hint">语种无需填写，统一按照季度预算表中该产品和地区的语种。无表头顺序：产品 / 卖点 / 国家 / 预算；预算也可留空。</p></div>
      <div class="actions"><button class="secondary" onclick="previewBatchItems()">解析预览</button><button class="primary" onclick="confirmBatchItems()">确认加入队列</button><button class="ghost" onclick="setv('batch_text','');document.getElementById('batch_result').innerHTML=''">清空</button></div>
      <div id="batch_result" class="hint batch-preview"></div>
    </div>
  </div>
</dialog>
<dialog id="dlg_recreate" class="ui-dialog">
  <div class="dialog-card">
    <div class="dialog-head"><div><h2>更换卖点重新创建</h2><p>保留 Campaign、国家和预算，只替换素材卖点。</p></div><button class="ghost" onclick="closeDialog('dlg_recreate')">关闭</button></div>
    <div class="dialog-body">
      <div class="field"><label>原广告组</label><input id="recreate_adset" readonly></div>
      <div class="row"><div class="field"><label>上一次卖点</label><input id="recreate_previous_point" readonly></div><div class="field"><label>产品</label><input id="recreate_product" readonly></div></div>
      <div class="row"><div class="field"><label>新卖点</label><select id="recreate_point" onchange="updateRecreateLanguages()"></select></div><div class="field"><label>素材语种</label><select id="recreate_language"></select></div></div>
      <div class="field"><label>保留配置</label><input id="recreate_config" readonly></div>
      <div class="actions"><button class="primary" onclick="confirmRecreate()">加入投放队列</button><button class="ghost" onclick="closeDialog('dlg_recreate')">取消</button></div>
      <p class="hint warn">加入队列后请先做“投放前检查”，确认无误再生成并运行。</p>
    </div>
  </div>
</dialog>
<dialog id="dlg_automation_batch" class="ui-dialog">
  <div class="dialog-card dialog-card-wide">
    <div class="dialog-head"><div><h2>批量换卖点重新创建</h2><p>按产品统一选择新卖点；语种读取预算表，原 Campaign、国家和预算不变。</p></div><button class="ghost" onclick="closeDialog('dlg_automation_batch')">关闭</button></div>
    <div class="dialog-body">
      <div id="automation_batch_summary" class="automation-summary">正在整理所选广告组...</div>
      <p class="hint automation-batch-note">同一产品只需选择一次卖点；素材语种会按每个地区的季度预算配置自动匹配。</p>
      <div id="automation_batch_product_controls" class="quick-product-controls"></div>
      <div class="insights-table-scroll"><table class="checktable automation-batch-table"><thead><tr><th>原广告组</th><th>产品</th><th>上一次卖点</th><th>地区</th><th>保留预算</th><th>Campaign</th></tr></thead><tbody id="automation_batch_rows"></tbody></table></div>
      <div class="actions"><button class="primary" onclick="confirmAutomationBatch()">全部加入投放队列</button><button class="ghost" onclick="closeDialog('dlg_automation_batch')">取消</button></div>
      <p class="hint warn">加入队列后仍需执行“投放前检查”。如果新卖点在同一地区已有 ACTIVE 广告，重复投放检查会阻止创建。</p>
    </div>
  </div>
</dialog>
<script>
let OPTIONS={assets:{},countries:[],defaults:{},forms:{}};let QUEUE=[];let LEAD_PRODUCTS=[];let LEAD_SELECTED=new Set();let leadPollTimer=null;let INSIGHTS_DATA=null;let INSIGHTS_VIEW='chart';let INSIGHTS_SECTION='overview';let AUTOMATION_ITEMS=[];let AUTOMATION_SETTINGS=null;let AUTOMATION_SELECTED=new Set();let AUTOMATION_BATCH_GROUPS=[];let RECREATE_ITEM=null;let BUDGET_STATUS=null;let QUICK_LAUNCH_ITEMS=[];let QUICK_PRODUCT_GROUPS=[];let LAST_RUN_ERROR_POPUP='';let LAST_RUN_SUMMARY_POPUP='';function openDialog(id){const d=document.getElementById(id);if(d&&d.showModal)d.showModal()}function closeDialog(id){const d=document.getElementById(id);if(d&&d.open)d.close()}document.addEventListener("click",e=>{if(e.target&&e.target.tagName==="DIALOG"&&e.target.open)e.target.close()});async function api(path,body){const opt=body?{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}:{};let r;try{r=await fetch(path,opt)}catch(e){throw new Error('本地前端服务没有连接上。请确认 frontend.bat 的黑色窗口还开着，并且浏览器地址是 http://127.0.0.1:8765，不要直接打开 html 文件。原始错误：'+e.message)}let text=await r.text();let d={};try{d=text?JSON.parse(text):{raw:text}}catch(e){throw new Error('服务返回的不是 JSON，可能 app.py 报错退出了。请求：'+(opt.method||'GET')+' '+path+'。返回内容：'+text.slice(0,800))}if(!r.ok){let msg=[];msg.push('请求：'+(opt.method||'GET')+' '+path);msg.push('HTTP：'+r.status);if(d.error)msg.push('错误：'+d.error);if(d.detail)msg.push('详情：'+d.detail);if(d.hint)msg.push('建议：'+d.hint);if(d.traceback)msg.push('Python traceback：\n'+d.traceback);if(d.available_get)msg.push('可用 GET 接口：'+d.available_get.join(', '));if(d.available_post)msg.push('可用 POST 接口：'+d.available_post.join(', '));if(!d.error&&!d.detail)msg.push('原始返回：'+text.slice(0,800));throw new Error(msg.join('\n'))}return d}function v(id){return document.getElementById(id).value.trim()}function setv(id,val){document.getElementById(id).value=val||''}function log(t){document.getElementById('status').textContent=t}
function insightsNeedsRefresh(){const cachedAt=INSIGHTS_DATA?.cache?.cached_at;if(!INSIGHTS_DATA||!cachedAt)return true;const parsed=new Date(cachedAt);return Number.isNaN(parsed.getTime())||Date.now()-parsed.getTime()>=60*60*1000}
function switchPage(page){const leads=page==='leads',insights=page==='insights',budget=page==='budget',ads=page==='ads';document.getElementById('ads_intro').hidden=!ads;document.getElementById('ads_page').hidden=!ads;document.getElementById('leads_page').hidden=!leads;document.getElementById('insights_page').hidden=!insights;document.getElementById('budget_page').hidden=!budget;document.getElementById('ads_header_actions').hidden=!ads;document.getElementById('nav_ads').classList.toggle('active',ads);document.getElementById('nav_leads').classList.toggle('active',leads);document.getElementById('nav_insights').classList.toggle('active',insights);document.getElementById('nav_budget').classList.toggle('active',budget);document.getElementById('header_subtitle').textContent=leads?'从 Meta 表单同步客户信息，并直接下载 Excel。':insights?'查看广告组趋势、明细和自动规则。':budget?'查看季度预算、地区覆盖，并一键生成经营周报和月报。':'批量投放、素材预览和投放检查。';if(leads)refreshLeadDashboard();if(insights){loadAutomationSettings();if(insightsNeedsRefresh())loadInsights(false)}if(budget)loadBudgetStatus(false)}
function setInsightsSection(section){INSIGHTS_SECTION=['overview','creative','automation'].includes(section)?section:'overview';document.querySelectorAll('[data-insight-section]').forEach(el=>el.hidden=el.dataset.insightSection!==INSIGHTS_SECTION);['overview','creative','automation'].forEach(name=>document.getElementById('insight_section_'+name)?.classList.toggle('active',name===INSIGHTS_SECTION));if(INSIGHTS_SECTION==='creative'&&INSIGHTS_DATA){renderCreativeFatigue();renderCreativeWinners()}if(INSIGHTS_SECTION==='automation')loadAutomationSettings()}
setInterval(()=>{const page=document.getElementById('insights_page'),btn=document.getElementById('insights_refresh_btn');if(INSIGHTS_DATA&&page&&!page.hidden&&btn&&!btn.disabled&&insightsNeedsRefresh())loadInsights(false)},60*1000)
async function loadLeadOptions(){try{const d=await api('/api/leads/options');LEAD_PRODUCTS=d.products||[];renderLeadProducts()}catch(e){document.getElementById('lead_products').innerHTML='<span class="bad">'+escapeHtml(e.message)+'</span>'}}
function renderLeadProducts(){const q=v('lead_search').toLowerCase();const visible=LEAD_PRODUCTS.filter(x=>x.toLowerCase().includes(q));document.getElementById('lead_products').innerHTML=visible.map(x=>{const i=LEAD_PRODUCTS.indexOf(x);return `<label class="product-check"><input type="checkbox" ${LEAD_SELECTED.has(x)?'checked':''} onchange="toggleLeadProduct(LEAD_PRODUCTS[${i}],this.checked)"><span>${escapeHtml(x)}</span></label>`}).join('')||'<span class="hint">没有匹配的产品</span>';document.getElementById('lead_selected_count').textContent=`已选择 ${LEAD_SELECTED.size} 个`}
function toggleLeadProduct(name,checked){if(checked)LEAD_SELECTED.add(name);else LEAD_SELECTED.delete(name);document.getElementById('lead_selected_count').textContent=`已选择 ${LEAD_SELECTED.size} 个`}
function setAllLeadProducts(checked){const q=v('lead_search').toLowerCase();LEAD_PRODUCTS.filter(x=>x.toLowerCase().includes(q)).forEach(x=>checked?LEAD_SELECTED.add(x):LEAD_SELECTED.delete(x));renderLeadProducts()}
async function startLeadSync(){const start=v('lead_start_date'),end=v('lead_end_date'),full=document.getElementById('lead_full').checked;if(start&&end&&start>end){alert('开始日期不能晚于结束日期');return}let mode=(start||end)?'指定日期导出':(full?'全量同步':'增量同步');let scope=LEAD_SELECTED.size?Array.from(LEAD_SELECTED).join('、'):'全部产品';if(!confirm(`即将开始${mode}\n产品：${scope}\n\n确认继续？`))return;try{const d=await api('/api/leads/start',{products:Array.from(LEAD_SELECTED),start_date:start,end_date:end,full});document.getElementById('lead_log').textContent=d.status?.output||'同步已启动';await refreshLeadStatus()}catch(e){document.getElementById('lead_log').textContent=e.message}}
async function refreshLeadStatus(){try{const s=await api('/api/leads/status');const btn=document.getElementById('lead_start_btn');btn.disabled=!!s.running;btn.textContent=s.running?'同步中...':'开始同步';document.getElementById('lead_status_text').textContent=s.running?`${s.mode||''} · 运行中`:(s.finished_at?(s.ok?'同步完成':'同步失败'):'等待操作');document.getElementById('lead_mode_badge').textContent=s.mode||'默认增量同步';document.getElementById('lead_log').textContent=s.output||'尚未运行客户信息同步。';if(leadPollTimer){clearTimeout(leadPollTimer);leadPollTimer=null}if(s.running)leadPollTimer=setTimeout(refreshLeadStatus,1200);else if(s.finished_at)refreshLeadFiles()}catch(e){document.getElementById('lead_log').textContent='状态读取失败：'+e.message}}
async function refreshLeadFiles(){try{const d=await api('/api/leads/files');document.getElementById('lead_files').innerHTML=(d.items||[]).map(x=>`<div class="lead-file"><div><b>${escapeHtml(x.name)}</b><span class="hint">${escapeHtml(x.modified)} · ${x.size_kb} KB</span></div><a class="secondary" style="text-decoration:none;padding:8px 12px;border-radius:12px" href="/api/leads/download?name=${encodeURIComponent(x.name)}">下载 Excel</a></div>`).join('')||'<div class="lead-file"><span class="hint">还没有客户信息文件。</span></div>'}catch(e){document.getElementById('lead_files').innerHTML='<div class="lead-file"><span class="bad">'+escapeHtml(e.message)+'</span></div>'}}
async function refreshLeadDashboard(){await Promise.all([loadLeadOptions(),refreshLeadFiles(),refreshLeadStatus()])}
function insightFormat(metric,value){const n=Number(value||0);if(metric==='spend'||metric==='cost_per_lead'||metric==='cpc'||metric==='cpm')return '$'+n.toLocaleString('zh-CN',{minimumFractionDigits:2,maximumFractionDigits:2});if(metric==='ctr')return n.toFixed(2)+'%';return n.toLocaleString('zh-CN',{maximumFractionDigits:2})}
function compactNumber(value){const n=Number(value||0);if(Math.abs(n)>=1000000)return (n/1000000).toFixed(1)+'M';if(Math.abs(n)>=1000)return (n/1000).toFixed(1)+'K';return n.toLocaleString('zh-CN',{maximumFractionDigits:1})}
function insightsCacheLabel(){const c=INSIGHTS_DATA?.cache;if(!c)return '缓存状态未知';if(c.source==='meta_api')return '刚从 Meta 更新';if(c.stale)return `Meta 暂不可用 · 使用 ${Number(c.age_minutes||0)} 分钟前缓存`;return `本地缓存 · ${Number(c.age_minutes||0)} 分钟前`}
function selectedProductLine(){return v('insights_product_filter')}
function selectedCampaignId(){return v('insights_campaign_filter')}
function selectedCountry(){return v('insights_country_filter')}
function countryLabel(code){const item=(OPTIONS.countries||[]).find(x=>String(x[0]).toUpperCase()===String(code).toUpperCase());return item?`${item[0]} · ${item[1]}`:String(code||'未知地区')}
function insightProductLine(row){const raw=String(row?.campaign_name||'')+' '+String(row?.adset_name||''),text=raw.replace(/[^A-Za-z0-9]+/g,'').toUpperCase(),products=Object.keys(OPTIONS.assets||{}).sort((a,b)=>b.length-a.length),known=products.find(product=>text.includes(String(product).replace(/[^A-Za-z0-9]+/g,'').toUpperCase()));if(known)return known;const campaign=String(row?.campaign_name||'').match(/Leads_New_FJD_(.+?)_FJD BU/i),adset=String(row?.adset_name||'').match(/^FJD_([^_]+)/i);return String(campaign?.[1]||adset?.[1]||'未识别产品').trim()}
function filteredInsightRows(){const product=selectedProductLine(),id=selectedCampaignId(),country=selectedCountry();return (INSIGHTS_DATA?.rows||[]).filter(x=>(!product||insightProductLine(x)===product)&&(!id||x.campaign_id===id)&&(!country||x.country===country))}
function filteredPreviousRows(){const product=selectedProductLine(),id=selectedCampaignId(),country=selectedCountry();return (INSIGHTS_DATA?.previous_rows||[]).filter(x=>(!product||insightProductLine(x)===product)&&(!id||x.campaign_id===id)&&(!country||x.country===country))}
function aggregateInsightRows(rows){const t={spend:0,impressions:0,clicks:0,leads:0};rows.forEach(x=>{t.spend+=Number(x.spend||0);t.impressions+=Number(x.impressions||0);t.clicks+=Number(x.clicks||0);t.leads+=Number(x.leads||0)});t.spend=Math.round(t.spend*100)/100;t.leads=Math.round(t.leads*100)/100;t.ctr=t.impressions?Math.round(t.clicks/t.impressions*100000)/1000:0;t.cost_per_lead=t.leads?Math.round(t.spend/t.leads*100)/100:0;return t}
function comparisonRate(current,previous){const c=Number(current||0),p=Number(previous||0);if(!p)return c?'new':0;return (c-p)/p*100}
function comparisonText(rate){if(rate==='new')return '环比 新增';const n=Number(rate||0),arrow=n>0?'↑':n<0?'↓':'→';return `环比 ${arrow} ${Math.abs(n).toFixed(1)}%`}
function comparisonClass(metric,rate){if(rate==='new')return metric==='cost_per_lead'?'down':'up';const n=Number(rate||0);if(metric==='spend')return 'neutral';if(metric==='cost_per_lead')return n<=0?'up':'down';return n>=0?'up':'down'}
function deltaCell(current,previous,metric){const rate=comparisonRate(current,previous),cls=comparisonClass(metric,rate).replace('up','delta-up').replace('down','delta-down').replace('neutral','delta-neutral');return `<span class="${cls}">${escapeHtml(comparisonText(rate).replace('环比 ',''))}</span>`}
function populateCampaignFilter(){const select=document.getElementById('insights_campaign_filter'),old=select.value,campaigns=new Map();[...(INSIGHTS_DATA?.previous_rows||[]),...(INSIGHTS_DATA?.rows||[])].forEach(x=>campaigns.set(x.campaign_id,x.campaign_name));select.innerHTML='';select.add(new Option(`全部广告系列（${campaigns.size}）`,''));Array.from(campaigns.entries()).sort((a,b)=>a[1].localeCompare(b[1],'zh-CN')).forEach(([id,name])=>select.add(new Option(name,id)));if(Array.from(campaigns.keys()).includes(old))select.value=old}
function populateProductFilter(){const select=document.getElementById('insights_product_filter'),old=select.value,products=new Set();[...(INSIGHTS_DATA?.previous_rows||[]),...(INSIGHTS_DATA?.rows||[])].forEach(x=>products.add(insightProductLine(x)));const items=Array.from(products).sort((a,b)=>a.localeCompare(b,'zh-CN'));select.innerHTML='';select.add(new Option(`全部产品线（${items.length}）`,''));items.forEach(product=>select.add(new Option(product,product)));if(products.has(old))select.value=old}
function populateCountryFilter(){const select=document.getElementById('insights_country_filter'),old=select.value,countries=new Set();[...(INSIGHTS_DATA?.previous_rows||[]),...(INSIGHTS_DATA?.rows||[])].forEach(x=>countries.add(x.country||'UNKNOWN'));const items=Array.from(countries).sort((a,b)=>countryLabel(a).localeCompare(countryLabel(b),'zh-CN'));select.innerHTML='';select.add(new Option(`全部国家（${items.length}）`,''));items.forEach(code=>select.add(new Option(countryLabel(code),code)));if(countries.has(old))select.value=old}
function syncMetricCards(){const metric=v('chart_metric')||'spend';document.querySelectorAll('.metric-card[data-metric]').forEach(card=>card.classList.toggle('active',card.dataset.metric===metric))}
function selectMetricFromCard(metric){document.getElementById('chart_metric').value=metric;syncMetricCards();setInsightsView('chart')}
function metricSelectChanged(){syncMetricCards();renderInsightChart()}
function resetInsights(){INSIGHTS_DATA=null;document.getElementById('insights_product_filter').innerHTML='<option value="">全部产品线</option>';document.getElementById('insights_campaign_filter').innerHTML='<option value="">全部广告系列</option>';document.getElementById('insights_country_filter').innerHTML='<option value="">全部国家</option>';const days=v('insights_days');document.getElementById('insights_download').href='/api/insights/export?days='+encodeURIComponent(days);loadInsights(false)}
function creativeFatigueItems(rows){const groups=new Map();(rows||[]).forEach(x=>{const key=(x.adset_id||'')+'|'+(x.country||'UNKNOWN'),g=groups.get(key)||{adset_id:x.adset_id,adset_name:x.adset_name,campaign_name:x.campaign_name,country:x.country||'UNKNOWN',daily:new Map()};const d=g.daily.get(x.date)||{spend:0,leads:0};d.spend+=Number(x.spend||0);d.leads+=Number(x.leads||0);g.daily.set(x.date,d);groups.set(key,g)});return Array.from(groups.values()).map(g=>{const dates=Array.from(g.daily.keys()).sort(),latest=dates.slice(-3),previous=dates.slice(-6,-3),sum=(ds,k)=>ds.reduce((n,d)=>n+Number(g.daily.get(d)?.[k]||0),0),latestSpend=sum(latest,'spend'),latestLeads=sum(latest,'leads'),previousSpend=sum(previous,'spend'),previousLeads=sum(previous,'leads'),previousCpl=previousLeads?previousSpend/previousLeads:0,latestCpl=latestLeads?latestSpend/latestLeads:0,minimumSpend=Math.max(20,previousCpl),comparisonCpl=latestLeads?latestCpl:latestSpend,cplChange=previousCpl?(comparisonCpl/previousCpl-1)*100:0,qualified=previous.length>0&&previousCpl>0&&latestSpend>=minimumSpend&&cplChange>=25,severe=cplChange>=50;return {...g,latestSpend,latestLeads,previousSpend,previousLeads,latestCpl,previousCpl,minimumSpend,cplChange,qualified,severity:severe?'severe':'warning'}}).filter(x=>x.qualified).sort((a,b)=>(a.severity===b.severity?b.cplChange-a.cplChange:a.severity==='severe'?-1:1))}
function renderCreativeFatigue(){const items=creativeFatigueItems(filteredInsightRows()),badge=document.getElementById('fatigue_badge'),body=document.getElementById('fatigue_rows');if(!badge||!body)return;const severe=items.filter(x=>x.severity==='severe').length;badge.textContent=items.length?`${items.length} 个需关注 · ${severe} 个高风险`:'当前未发现疲劳';badge.className='muted-chip '+(severe?'bad':items.length?'warn':'ok');body.innerHTML=items.map(x=>{const reason=x.latestLeads?`CPL 上升 ${x.cplChange.toFixed(1)}%`:`已花 $${x.latestSpend.toFixed(2)}，仍为 0 Leads`;return `<tr><td><span class="fatigue-status ${x.severity}">${x.severity==='severe'?'高风险':'关注'}</span></td><td>${escapeHtml(x.campaign_name)}</td><td>${escapeHtml(countryLabel(x.country))}</td><td title="${escapeHtml(x.adset_id)}">${escapeHtml(x.adset_name)}</td><td class="number-cell">${insightFormat('spend',x.latestSpend)}</td><td class="number-cell">${insightFormat('spend',x.previousSpend)}</td><td class="number-cell">${compactNumber(x.latestLeads)}</td><td class="number-cell">${x.latestLeads?insightFormat('cost_per_lead',x.latestCpl):'—'}</td><td class="number-cell">${insightFormat('cost_per_lead',x.previousCpl)}</td><td><b class="${x.cplChange>=50?'bad':'warn'}">${escapeHtml(reason)}</b><div class="hint">${x.severity==='severe'?'建议立即更换素材':'建议准备新素材并持续观察'}</div></td></tr>`}).join('')||'<tr><td colspan="10" class="empty-row">当前筛选范围内未发现明显素材疲劳</td></tr>'}
function insightSellingPoint(product,row){const text=String(row?.adset_name||'')+' '+String(row?.campaign_name||''),points=OPTIONS.assets?.[product]||[];return points.find(point=>text.includes(point.name))?.name||''}
function winningCreativeItems(rows){const groups=new Map();(rows||[]).forEach(row=>{const product=insightProductLine(row),point=insightSellingPoint(product,row),label=point||row.adset_name||'未识别素材',country=row.country||'UNKNOWN',key=[product,label,country].join('|'),g=groups.get(key)||{product,point,label,country,adsets:new Set(),spend:0,leads:0,clicks:0,impressions:0};g.adsets.add(row.adset_id);['spend','leads','clicks','impressions'].forEach(k=>g[k]+=Number(row[k]||0));groups.set(key,g)});return Array.from(groups.values()).filter(x=>x.leads>0).map(x=>({...x,adset_count:x.adsets.size,cpl:x.spend/x.leads,ctr:x.impressions?x.clicks/x.impressions*100:0,stable:x.leads>=3&&(x.spend>=10||x.leads>=5)})).sort((a,b)=>(Number(b.stable)-Number(a.stable))||(a.cpl-b.cpl)||(b.leads-a.leads)).slice(0,20)}
function renderCreativeWinners(){const items=winningCreativeItems(filteredInsightRows()),badge=document.getElementById('winner_badge'),body=document.getElementById('winner_rows');if(!badge||!body)return;const stable=items.filter(x=>x.stable).length;badge.textContent=items.length?`${items.length} 组上榜 · ${stable} 组稳定优胜`:'暂无有留资素材';badge.className='muted-chip '+(stable?'ok':'');body.innerHTML=items.map((x,i)=>`<tr><td><span class="winner-rank ${i<3?'top':''}">${i+1}</span></td><td><b>${escapeHtml(x.product)}</b><br><span class="hint">${escapeHtml(x.label)}</span></td><td>${escapeHtml(countryLabel(x.country))}</td><td class="number-cell">${x.adset_count}</td><td class="number-cell">${insightFormat('spend',x.spend)}</td><td class="number-cell"><b>${compactNumber(x.leads)}</b></td><td class="number-cell"><b>${insightFormat('cost_per_lead',x.cpl)}</b></td><td class="number-cell">${Number(x.ctr).toFixed(2)}%</td><td><span class="winner-confidence ${x.stable?'':'candidate'}">${x.stable?'稳定优胜':'潜力素材'}</span></td><td>${x.point?`<button class="secondary smallbtn" onclick="openWinningPack('${escapeHtml(x.product)}','${escapeHtml(x.point)}')">查看素材包</button>`:'<span class="hint">未识别卖点</span>'}</td></tr>`).join('')||'<tr><td colspan="10" class="empty-row">当前筛选范围内还没有产生留资的素材</td></tr>'}
function openWinningPack(product,point){switchPage('ads');setv('product',product);updateSellingPoints();setv('selling_point',point);updateAssetLanguages();document.getElementById('asset_note')?.scrollIntoView({behavior:'smooth',block:'center'})}
async function loadInsights(force){const days=v('insights_days')||'7',btn=document.getElementById('insights_refresh_btn'),empty=document.getElementById('insights_loading');if(btn.disabled)return;btn.disabled=true;btn.textContent=force?'强制刷新中...':'读取缓存中...';empty.hidden=false;empty.classList.add('loading');empty.innerHTML=`<span class="loader-dot"></span>${force?'正在从 Meta 强制刷新数据...':'正在读取本地缓存；超过 60 分钟才会请求 Meta...'}`;document.getElementById('insights_chart_wrap').hidden=true;document.getElementById('insights_table_wrap').hidden=true;document.getElementById('insights_campaign_wrap').hidden=true;try{INSIGHTS_DATA=await api(`/api/insights?days=${encodeURIComponent(days)}${force?'&force=1':''}`);const refreshTime=INSIGHTS_DATA.cache?.cached_at||INSIGHTS_DATA.updated_at||'未知';document.getElementById('insights_range').textContent=`当前 ${INSIGHTS_DATA.since} 至 ${INSIGHTS_DATA.until} · 环比 ${INSIGHTS_DATA.previous_since} 至 ${INSIGHTS_DATA.previous_until}`;document.getElementById('insights_updated').textContent=`更新于 ${INSIGHTS_DATA.updated_at} · ${insightsCacheLabel()} · 当前 ${INSIGHTS_DATA.row_count} 条 / 上期 ${INSIGHTS_DATA.previous_row_count} 条`;document.getElementById('insights_last_refresh').textContent=`上次刷新：${String(refreshTime).replace('T',' ')} · ${insightsCacheLabel()}`;document.getElementById('insights_download').href='/api/insights/export?days='+encodeURIComponent(days);populateProductFilter();populateCampaignFilter();populateCountryFilter();renderInsightKpis();renderCreativeFatigue();renderCreativeWinners();syncMetricCards();empty.hidden=true;setInsightsView(INSIGHTS_VIEW)}catch(e){INSIGHTS_DATA=null;empty.hidden=false;empty.classList.remove('loading');empty.innerHTML='<span class="bad">读取失败：'+escapeHtml(e.message)+'</span>'}finally{btn.disabled=false;btn.textContent='强制刷新'}}
function renderInsightKpis(){const t=aggregateInsightRows(filteredInsightRows()),p=aggregateInsightRows(filteredPreviousRows());document.getElementById('metric_spend').textContent=insightFormat('spend',t.spend);document.getElementById('metric_impressions').textContent=compactNumber(t.impressions);document.getElementById('metric_clicks').textContent=compactNumber(t.clicks);document.getElementById('metric_leads').textContent=compactNumber(t.leads);document.getElementById('metric_cpl').textContent=insightFormat('cost_per_lead',t.cost_per_lead);[['spend','spend'],['impressions','impressions'],['clicks','clicks'],['leads','leads'],['cpl','cost_per_lead']].forEach(([id,metric])=>{const el=document.getElementById('metric_'+id+'_compare'),rate=comparisonRate(t[metric],p[metric]);el.textContent=comparisonText(rate);el.className='compare-pill '+comparisonClass(metric,rate)})}
function applyInsightsFilters(){if(!INSIGHTS_DATA)return;renderInsightKpis();renderCreativeFatigue();renderCreativeWinners();const productSelect=document.getElementById('insights_product_filter'),campaignSelect=document.getElementById('insights_campaign_filter'),countrySelect=document.getElementById('insights_country_filter'),productName=productSelect.options[productSelect.selectedIndex]?.text||'全部产品线',campaignName=campaignSelect.options[campaignSelect.selectedIndex]?.text||'全部广告系列',countryName=countrySelect.options[countrySelect.selectedIndex]?.text||'全部国家',campaign=campaignSelect.value,country=countrySelect.value;document.getElementById('insights_updated').textContent=`更新于 ${INSIGHTS_DATA.updated_at} · ${insightsCacheLabel()} · ${productName} · ${campaignName} · ${countryName}`;document.getElementById('insights_download').href=`/api/insights/export?days=${encodeURIComponent(v('insights_days')||'7')}${campaign?'&campaign_id='+encodeURIComponent(campaign):''}${country?'&country='+encodeURIComponent(country):''}`;if(INSIGHTS_VIEW==='chart')renderInsightChart();else if(INSIGHTS_VIEW==='campaign')renderCampaignSummary();else renderInsightsTable()}
function applyInsightsProduct(){applyInsightsFilters()}
function applyInsightsCampaign(){applyInsightsFilters()}
function applyInsightsCountry(){applyInsightsFilters()}
function setInsightsView(view){INSIGHTS_VIEW=view;const chart=view==='chart',campaign=view==='campaign',table=view==='table';document.getElementById('view_chart_btn').classList.toggle('active',chart);document.getElementById('view_campaign_btn').classList.toggle('active',campaign);document.getElementById('view_table_btn').classList.toggle('active',table);document.getElementById('chart_metric').hidden=!chart;if(!INSIGHTS_DATA)return;document.getElementById('insights_loading').hidden=true;document.getElementById('insights_chart_wrap').hidden=!chart;document.getElementById('insights_campaign_wrap').hidden=!campaign;document.getElementById('insights_table_wrap').hidden=!table;if(chart)renderInsightChart();else if(campaign)renderCampaignSummary();else renderInsightsTable()}
function aggregateDailyForRows(rows,template){const byDate={};(template||[]).forEach(x=>byDate[x.date]={date:x.date,spend:0,impressions:0,clicks:0,leads:0});rows.forEach(x=>{const d=byDate[x.date]||(byDate[x.date]={date:x.date,spend:0,impressions:0,clicks:0,leads:0});['spend','impressions','clicks','leads'].forEach(k=>d[k]+=Number(x[k]||0))});return Object.values(byDate).sort((a,b)=>a.date.localeCompare(b.date)).map(d=>({...d,ctr:d.impressions?d.clicks/d.impressions*100:0,cost_per_lead:d.leads?d.spend/d.leads:0}))}
function filteredInsightDaily(){return selectedProductLine()||selectedCampaignId()||selectedCountry()?aggregateDailyForRows(filteredInsightRows(),INSIGHTS_DATA?.daily):INSIGHTS_DATA?.daily||[]}
function filteredPreviousDaily(){return selectedProductLine()||selectedCampaignId()||selectedCountry()?aggregateDailyForRows(filteredPreviousRows(),INSIGHTS_DATA?.previous_daily):INSIGHTS_DATA?.previous_daily||[]}
function renderInsightChart(){if(!INSIGHTS_DATA)return;const metric=v('chart_metric')||'spend',labels={spend:'花费',leads:'Leads',cost_per_lead:'单条线索成本',clicks:'点击',impressions:'展示',ctr:'CTR'},data=filteredInsightDaily(),previous=filteredPreviousDaily(),svg=document.getElementById('insights_chart');const W=1000,H=360,L=72,R=28,T=28,B=52,IW=W-L-R,IH=H-T-B;const values=[...data,...previous].map(x=>Number(x[metric]||0)),maxValue=Math.max(...values,1),ceiling=maxValue*1.12;let parts=['<defs><linearGradient id="chartArea" x1="0" y1="0" x2="0" y2="1"><stop offset="0%" stop-color="#5b5cf0" stop-opacity=".25"/><stop offset="100%" stop-color="#5b5cf0" stop-opacity=".02"/></linearGradient></defs>'];for(let i=0;i<=4;i++){const y=T+IH*i/4,val=ceiling*(1-i/4);parts.push(`<line x1="${L}" y1="${y}" x2="${W-R}" y2="${y}" stroke="#e4e9f1" stroke-width="1"/><text x="${L-12}" y="${y+4}" text-anchor="end" fill="#8491a4" font-size="12">${escapeHtml(compactNumber(val))}</text>`)}const makePoints=series=>series.map((x,i)=>{const px=series.length===1?L+IW/2:L+IW*i/(series.length-1),py=T+IH-(Number(x[metric]||0)/ceiling)*IH;return {x:px,y:py,row:x}}),points=makePoints(data),previousPoints=makePoints(previous);if(previousPoints.length){const previousLine=previousPoints.map((p,i)=>(i?'L':'M')+p.x.toFixed(1)+' '+p.y.toFixed(1)).join(' ');parts.push(`<path d="${previousLine}" fill="none" stroke="#94a3b8" stroke-width="3" stroke-dasharray="9 8" stroke-linecap="round"/>`);previousPoints.forEach(p=>parts.push(`<circle cx="${p.x}" cy="${p.y}" r="3" fill="#fff" stroke="#94a3b8" stroke-width="2"><title>上一周期 ${escapeHtml(p.row.date)} · ${escapeHtml(labels[metric])}: ${escapeHtml(insightFormat(metric,p.row[metric]))}</title></circle>`))}if(points.length){const line=points.map((p,i)=>(i?'L':'M')+p.x.toFixed(1)+' '+p.y.toFixed(1)).join(' '),area=line+` L ${points[points.length-1].x.toFixed(1)} ${T+IH} L ${points[0].x.toFixed(1)} ${T+IH} Z`;parts.push(`<path d="${area}" fill="url(#chartArea)"/><path d="${line}" fill="none" stroke="#5b5cf0" stroke-width="4" stroke-linecap="round" stroke-linejoin="round"/>`);points.forEach((p,i)=>{parts.push(`<circle cx="${p.x}" cy="${p.y}" r="5" fill="#fff" stroke="#5b5cf0" stroke-width="3"><title>当前周期 ${escapeHtml(p.row.date)} · ${escapeHtml(labels[metric])}: ${escapeHtml(insightFormat(metric,p.row[metric]))}</title></circle>`);const show=data.length<=14||i%3===0||i===data.length-1;if(show)parts.push(`<text x="${p.x}" y="${H-20}" text-anchor="middle" fill="#718096" font-size="12">${escapeHtml(p.row.date.slice(5))}</text>`)});}svg.innerHTML=parts.join('');const campaignSelect=document.getElementById('insights_campaign_filter'),countrySelect=document.getElementById('insights_country_filter'),campaignName=campaignSelect.options[campaignSelect.selectedIndex]?.text||'全部广告系列',countryName=countrySelect.options[countrySelect.selectedIndex]?.text||'全部国家';document.getElementById('chart_legend').innerHTML=`<span><i></i>当前 ${escapeHtml(labels[metric])}</span><span><i class="previous"></i>上一周期</span><span>${escapeHtml(campaignName)}</span><span>${escapeHtml(countryName)}</span>`}
function previousDetailDate(date){const offset=Math.round((Date.parse(date+'T00:00:00Z')-Date.parse(INSIGHTS_DATA.since+'T00:00:00Z'))/86400000),previous=new Date(Date.parse(INSIGHTS_DATA.previous_since+'T00:00:00Z')+offset*86400000);return previous.toISOString().slice(0,10)}
function previousDetailTotals(row){const date=previousDetailDate(row.date),matches=filteredPreviousRows().filter(x=>x.date===date&&x.adset_id===row.adset_id&&x.country===row.country);return {date,...aggregateInsightRows(matches)}}
function detailComparison(row){const previous=previousDetailTotals(row),cpl=row.leads&&previous.leads?deltaCell(row.cost_per_lead,previous.cost_per_lead,'cost_per_lead'):'<span class="delta-neutral">—</span>';return `<div class="detail-compare" title="对比 ${escapeHtml(previous.date)}"><span><b>花费</b>${deltaCell(row.spend,previous.spend,'spend')}</span><span><b>Leads</b>${deltaCell(row.leads,previous.leads,'leads')}</span><span><b>CPL</b>${cpl}</span></div>`}
function renderInsightsTable(){if(!INSIGHTS_DATA)return;const q=v('insights_search').toLowerCase(),base=filteredInsightRows(),rows=base.filter(x=>!q||(x.adset_name+' '+x.country+' '+countryLabel(x.country)).toLowerCase().includes(q));document.getElementById('insights_row_count').textContent=`显示 ${rows.length} / ${base.length} 条`;document.getElementById('insights_rows').innerHTML=rows.map(x=>`<tr><td>${escapeHtml(x.date)}</td><td>${escapeHtml(countryLabel(x.country))}</td><td title="${escapeHtml(x.adset_id)}">${escapeHtml(x.adset_name)}</td><td class="number-cell">${insightFormat('spend',x.spend)}</td><td class="number-cell">${compactNumber(x.impressions)}</td><td class="number-cell">${compactNumber(x.clicks)}</td><td class="number-cell">${insightFormat('ctr',x.ctr)}</td><td class="number-cell">${compactNumber(x.leads)}</td><td class="number-cell">${x.leads?insightFormat('cost_per_lead',x.cost_per_lead):'—'}</td><td>${detailComparison(x)}</td></tr>`).join('')||'<tr><td colspan="10" class="empty-row">没有匹配的广告组数据</td></tr>'}
function renderCampaignSummary(){if(!INSIGHTS_DATA)return;const groups=new Map();[...filteredPreviousRows(),...filteredInsightRows()].forEach(x=>{if(!groups.has(x.campaign_id))groups.set(x.campaign_id,{id:x.campaign_id,name:x.campaign_name,current:[],previous:[],adsets:new Set()})});filteredInsightRows().forEach(x=>{const g=groups.get(x.campaign_id);g.current.push(x);g.adsets.add(x.adset_id)});filteredPreviousRows().forEach(x=>groups.get(x.campaign_id).previous.push(x));const items=Array.from(groups.values()).map(g=>({...g,totals:aggregateInsightRows(g.current),previousTotals:aggregateInsightRows(g.previous)})).sort((a,b)=>b.totals.spend-a.totals.spend);document.getElementById('campaign_count').textContent=`${items.length} 个广告系列`;document.getElementById('campaign_rows').innerHTML=items.map(g=>`<tr><td><button class="campaign-link" data-campaign-id="${escapeHtml(g.id)}" onclick="selectCampaignFromRow(this.dataset.campaignId)">${escapeHtml(g.name)}</button><div class="hint">${escapeHtml(g.id)}</div></td><td class="number-cell">${g.adsets.size}</td><td class="number-cell">${insightFormat('spend',g.totals.spend)}</td><td class="number-cell">${deltaCell(g.totals.spend,g.previousTotals.spend,'spend')}</td><td class="number-cell">${compactNumber(g.totals.impressions)}</td><td class="number-cell">${compactNumber(g.totals.clicks)}</td><td class="number-cell">${insightFormat('ctr',g.totals.ctr)}</td><td class="number-cell">${compactNumber(g.totals.leads)}</td><td class="number-cell">${deltaCell(g.totals.leads,g.previousTotals.leads,'leads')}</td><td class="number-cell">${g.totals.leads?insightFormat('cost_per_lead',g.totals.cost_per_lead):'—'}</td><td class="number-cell">${deltaCell(g.totals.cost_per_lead,g.previousTotals.cost_per_lead,'cost_per_lead')}</td></tr>`).join('')||'<tr><td colspan="11" class="empty-row">没有广告系列数据</td></tr>'}
function selectCampaignFromRow(id){document.getElementById('insights_campaign_filter').value=id;applyInsightsCampaign();setInsightsView('chart')}
async function loadAutomationSettings(){try{const d=await api('/api/automation/settings');AUTOMATION_SETTINGS=d.rules;const enabled=!!d.rules.enabled,mode=d.rules.execution_mode||'approval',modeLabels={notify:'仅提醒',approval:'待人工确认',auto:'全自动'};document.getElementById('automation_mode').value=mode;document.getElementById('automation_badge').textContent=enabled?`已启用 · ${modeLabels[mode]}`:'自动化：未启用';document.getElementById('automation_badge').className='muted-chip '+(enabled?'ok':'');document.getElementById('automation_toggle').textContent=enabled?'停用自动化':'启用自动化';if(d.state?.last_result){AUTOMATION_ITEMS=d.state.last_result.items||[];renderAutomationResults(d.state.last_result);const hidden=Number(d.state.last_result.ignored_inactive||0);if(!AUTOMATION_ITEMS.length&&hidden)document.getElementById('automation_rows').innerHTML=`<tr><td colspan="7" class="empty-row"><b>当前没有新的规则命中</b><br>${hidden} 个先前命中项已暂停或关闭，按设置不再重复提示。</td></tr>`}else{AUTOMATION_ITEMS=[];document.getElementById('automation_rows').innerHTML='<tr><td colspan="7" class="empty-row">还没有成功的规则检查记录，请点击“只读预览”。</td></tr>'}if(d.state?.last_error){document.getElementById('automation_summary').innerHTML=`<span class="bad">最近一次检查失败，未清空上次成功结果：${escapeHtml(d.state.last_error)}</span><br><span class="hint">请恢复 Meta 网络后再点“只读预览”。</span>`}}catch(e){document.getElementById('automation_badge').textContent='设置读取失败';document.getElementById('automation_summary').innerHTML='<span class="bad">自动化设置读取失败：'+escapeHtml(e.message)+'</span>'}}
async function changeAutomationMode(){const mode=v('automation_mode'),labels={notify:'仅提醒',approval:'待人工确认',auto:'全自动执行'};if(mode==='auto'&&!confirm('全自动模式会在每次定时检查后真实暂停广告组或调整预算。72 小时保护、截至昨天数据、单次 10% 和 24 小时冷却仍会生效。确定切换吗？')){document.getElementById('automation_mode').value=AUTOMATION_SETTINGS?.execution_mode||'approval';return}try{const d=await api('/api/automation/settings',{execution_mode:mode});AUTOMATION_SETTINGS=d.rules;await loadAutomationSettings();document.getElementById('automation_summary').textContent=`执行模式已切换为“${labels[mode]}”。`}catch(e){alert(e.message);document.getElementById('automation_mode').value=AUTOMATION_SETTINGS?.execution_mode||'approval'}}
async function toggleAutomation(){const enabled=!AUTOMATION_SETTINGS?.enabled,mode=v('automation_mode'),labels={notify:'仅提醒',approval:'待人工确认',auto:'全自动执行'};if(enabled&&!confirm(`启用后每 6 小时检查一次，当前模式为“${labels[mode]}”。\n\n新广告保护 72 小时，绩效判断使用截至昨天的完整数据；配置异常只提醒。确定启用吗？`))return;try{const d=await api('/api/automation/settings',{enabled,execution_mode:mode});AUTOMATION_SETTINGS=d.rules;await loadAutomationSettings();document.getElementById('automation_summary').textContent=enabled?'自动化定时检查已启用。建议先点击“只读预览”。':'自动化已停用，历史日志和复盘数据仍会保留。'}catch(e){alert(e.message)}}
async function previewAutomation(){await runAutomationRequest(false)}
async function runAutomationNow(){if(!confirm('将立即执行全部命中项：可能真实暂停广告组，也可能调整日预算（单次最多 10%）。配置异常和不消耗提醒不会自动修改广告。确定继续？'))return;await runAutomationRequest(true)}
async function runAutomationRequest(execute){const box=document.getElementById('automation_summary');box.textContent=execute?'正在检查并执行自动关闭...':'正在只读检查，不会修改广告...';try{const d=await api(execute?'/api/automation/run':'/api/automation/preview',{});AUTOMATION_ITEMS=d.items||[];renderAutomationResults(d);if(d.halted)alert('执行遇到错误，后续规则已自动暂停。\n\n'+(d.halt_reason||'请稍后重试。'))}catch(e){box.innerHTML='<span class="bad">规则检查失败：'+escapeHtml(e.message)+'</span>'}}
function automationItemSelectable(x){return !!(x?.product&&OPTIONS.assets?.[x.product]&&x?.countries?.length&&(x.action==='WARN_CREATIVE'||String(x.action||'').startsWith('CLOSE_')))}
function automationItemExecutable(x){return ['CLOSE_NO_LEADS','CLOSE_HIGH_CPL','CLOSE_SPEND_SPIKE','SCALE_UP','PACE_UP','PACE_DOWN'].includes(x?.action)&&!x?.executed}
function automationSelectionAllowed(x){return automationItemExecutable(x)||automationItemSelectable(x)}
function automationActionKey(x){return String(x?.adset_id||'')+'|'+String(x?.action||'')}
function updateAutomationSelectionUi(){const selected=Array.from(AUTOMATION_SELECTED).map(i=>AUTOMATION_ITEMS[i]).filter(Boolean),count=selected.length,executeCount=selected.filter(automationItemExecutable).length,recreateCount=selected.filter(automationItemSelectable).length,badge=document.getElementById('automation_selected_count'),executeButton=document.getElementById('automation_execute_button'),recreateButton=document.getElementById('automation_batch_button');if(badge)badge.textContent=`已选择 ${count} 项 · 可执行 ${executeCount} · 可重建 ${recreateCount}`;if(executeButton){executeButton.disabled=!executeCount;executeButton.textContent=executeCount?`执行所选规则（${executeCount}）`:'执行所选规则'}if(recreateButton){recreateButton.disabled=!recreateCount;recreateButton.textContent=recreateCount?`批量换卖点重新创建（${recreateCount}）`:'批量换卖点重新创建'}}
function toggleAutomationSelection(index,checked){if(checked&&automationSelectionAllowed(AUTOMATION_ITEMS[index]))AUTOMATION_SELECTED.add(index);else AUTOMATION_SELECTED.delete(index);updateAutomationSelectionUi()}
document.addEventListener('click',event=>{const cell=event.target.closest?.('.automation-table tbody .automation-select-cell');if(!cell||event.target.matches('input'))return;const input=cell.querySelector('input:not(:disabled)');if(input)input.click()})
function automationProductMatches(x){const product=document.getElementById('automation_product_filter')?.value||'';return !product||String(x?.product||'未识别产品')===product}
function applyAutomationProductVisibility(){let visible=0;document.querySelectorAll('#automation_rows .automation-item-check').forEach(input=>{const index=Number(input.dataset.index),show=automationProductMatches(AUTOMATION_ITEMS[index]),row=input.closest('tr');if(row)row.hidden=!show;if(show)visible++});const count=document.getElementById('automation_filter_count'),product=document.getElementById('automation_product_filter')?.value||'';if(count)count.textContent=product?`${product} · 显示 ${visible} 项`:`显示全部 ${visible} 项`}
function refreshAutomationProductFilter(){const select=document.getElementById('automation_product_filter');if(!select)return;const products=Array.from(new Set(AUTOMATION_ITEMS.map(x=>String(x?.product||'未识别产品')))).sort((a,b)=>a.localeCompare(b,'zh-CN')),signature=products.join('|');if(select.dataset.signature!==signature){const old=select.value;select.innerHTML='<option value="">全部产品</option>';products.forEach(product=>select.add(new Option(product,product)));select.dataset.signature=signature;if(products.includes(old))select.value=old}applyAutomationProductVisibility()}
function changeAutomationProductFilter(){AUTOMATION_SELECTED.clear();document.querySelectorAll('.automation-item-check').forEach(input=>input.checked=false);applyAutomationProductVisibility();updateAutomationSelectionUi()}
function setAutomationSelection(all){AUTOMATION_SELECTED.clear();if(all)AUTOMATION_ITEMS.forEach((x,i)=>{if(automationProductMatches(x)&&automationSelectionAllowed(x))AUTOMATION_SELECTED.add(i)});document.querySelectorAll('.automation-item-check').forEach(input=>{input.checked=AUTOMATION_SELECTED.has(Number(input.dataset.index))});updateAutomationSelectionUi()}
const automationRowsElement=document.getElementById('automation_rows');if(automationRowsElement)new MutationObserver(()=>refreshAutomationProductFilter()).observe(automationRowsElement,{childList:true});
async function applyAutomationKeys(keys){if(!keys.length){alert('所选项目里没有可直接执行的规则。');return}if(!confirm(`将执行 ${keys.length} 个自动化动作，可能暂停广告组或调整日预算。确定继续？`))return;const box=document.getElementById('automation_summary');box.textContent='正在使用最近一次成功预览执行所选规则，不重复消耗检查额度...';try{const d=await api('/api/automation/apply',{keys});AUTOMATION_ITEMS=d.items||[];renderAutomationResults(d);if(d.halted)alert('执行遇到错误，后续操作已自动暂停，未继续向 Meta 提交。\n\n'+(d.halt_reason||'请稍后重试。'))}catch(e){box.innerHTML='<span class="bad">执行失败：'+escapeHtml(e.message)+'</span>'}}
async function applyAutomationSelected(){const keys=Array.from(AUTOMATION_SELECTED).map(i=>AUTOMATION_ITEMS[i]).filter(automationItemExecutable).map(automationActionKey);await applyAutomationKeys(keys)}
function renderAutomationResults(data){const items=data.items||[],closed=(data.closed||[]).length,budgetChanged=(data.budget_changed||[]).length,warnings=items.filter(x=>x.action==='WARN_CREATIVE').length,closeCount=items.filter(x=>x.action.startsWith('CLOSE_')).length,scaleCount=items.filter(x=>['SCALE_UP','PACE_UP','PACE_DOWN'].includes(x.action)).length,inspectCount=items.filter(x=>['WARN_CONFIG','WARN_NO_DELIVERY','INFO_PROTECTED'].includes(x.action)).length,ignoredInactive=Number(data.ignored_inactive||0);AUTOMATION_SELECTED.clear();document.getElementById('automation_summary').innerHTML=`检查 ACTIVE 广告组 ${data.active_adsets} 个；换素材 ${warnings}；停投 ${closeCount}；预算动作 ${scaleCount}；保护/待排查 ${inspectCount}${data.execute?`；本次已暂停 ${closed}、已调预算 ${budgetChanged}`:'（只读预览，未修改广告）'}。<br><span class="hint">地区基准：${escapeHtml(data.baseline_period)} · 决策数据：${escapeHtml(data.current_period)} · 当天仅用于花费突增观察</span>${ignoredInactive?`<br><span class="ok">已过滤 ${ignoredInactive} 个已暂停、关闭或所属 Campaign 已关闭的广告组，不再重复提示。</span>`:''}`;document.getElementById('automation_rows').innerHTML=items.map((x,i)=>{const closedAction=x.action.startsWith('CLOSE_'),budgetAction=['SCALE_UP','PACE_UP','PACE_DOWN'].includes(x.action),inspectAction=['WARN_CONFIG','WARN_NO_DELIVERY','INFO_PROTECTED'].includes(x.action),status=x.rolled_back_at?'info':x.executed?'paused':closedAction?'close':budgetAction?'scale':inspectAction?'info':'warn',labels={CLOSE_NO_LEADS:'3天0留资',CLOSE_HIGH_CPL:'高CPL停投',CLOSE_SPEND_SPIKE:'花费突增',WARN_CREATIVE:'换素材',SCALE_UP:'优质扩量',PACE_UP:'预算补速',PACE_DOWN:'预算降速',WARN_NO_DELIVERY:'不消耗',WARN_CONFIG:'配置异常',INFO_PROTECTED:'72小时保护'},label=x.rolled_back_at?'已撤销':x.executed?(x.paused?'已暂停':'已执行'):(labels[x.action]||x.action),selectable=automationItemSelectable(x),executable=automationItemExecutable(x),allowed=automationSelectionAllowed(x),country=(x.country_detail?.country||x.countries?.join(',')||'-'),budgetNote=budgetAction?`<br><span class="hint">日预算 $${Number(x.daily_budget||0).toFixed(2)} → $${Number(x.proposed_budget||0).toFixed(2)}</span>`:'',todayNote=x.action==='CLOSE_SPEND_SPIKE'?`<br><span class="hint">今日 $${Number(x.today_spend||0).toFixed(2)} / 7日均 $${Number(x.average_7d_spend||0).toFixed(2)}</span>`:'';return `<tr><td class="automation-select-cell"><input class="automation-item-check" data-index="${i}" type="checkbox" ${allowed?'':'disabled title="该项仅供查看"'} onchange="toggleAutomationSelection(${i},this.checked)"></td><td><span class="rule-status ${status}">${escapeHtml(label)}</span>${x.error?'<div class="bad">'+escapeHtml(x.error)+'</div>':''}${x.rollback_result?'<div class="ok">'+escapeHtml(x.rollback_result)+'</div>':''}${budgetNote}</td><td><b>${escapeHtml(x.adset_name)}</b><div class="hint">${escapeHtml(x.campaign_name||x.campaign_id)}<br>${escapeHtml(x.adset_id)}</div></td><td>${escapeHtml(x.product||'未识别产品')}<br><span class="hint">${escapeHtml(country)}</span></td><td class="number-cell">花费 ${insightFormat('spend',x.current_spend)}<br>Leads ${compactNumber(x.current_leads)}<br>CPL ${x.current_leads?insightFormat('cost_per_lead',x.current_cpl):'—'}${todayNote}</td><td>${escapeHtml(x.reason)}</td><td><div class="automation-row-actions">${x.rollback_available&&!x.rolled_back_at?`<button class="ghost" onclick="rollbackAutomation('${escapeHtml(x.action_id)}')">一键撤销</button>`:''}${executable?`<button class="primary" onclick="applyAutomationKeys(['${escapeHtml(automationActionKey(x))}'])">执行此规则</button>`:''}${selectable?`<button class="secondary automation-recreate" onclick="openRecreate(${i})">换卖点重新创建</button>`:''}${!x.rollback_available&&!executable&&!selectable?'<span class="hint">仅提醒，不自动修改</span>':''}</div></td></tr>`}).join('')||'<tr><td colspan="7" class="empty-row">当前没有广告组触发规则</td></tr>';updateAutomationSelectionUi()}
async function rollbackAutomation(actionId){if(!confirm('将恢复这次操作之前的广告组状态或日预算。若之后有人手动调整过该广告组，撤销会覆盖当前值。确定继续？'))return;try{const d=await api('/api/automation/rollback',{action_id:actionId});alert(d.message||'撤销成功');await loadAutomationSettings();await loadAutomationEffectiveness()}catch(e){alert(e.message)}}
function effectivenessMetric(metric){if(!metric)return '—';return `花费 $${Number(metric.spend||0).toFixed(2)}<br>Leads ${compactNumber(metric.leads||0)}<br>CPL ${metric.leads?'$'+Number(metric.cpl||0).toFixed(2):'—'}`}
async function loadAutomationEffectiveness(){const summary=document.getElementById('effectiveness_summary'),rows=document.getElementById('effectiveness_rows');summary.innerHTML='<span>正在读取缓存并计算...</span>';try{const d=await api('/api/automation/effectiveness'),s=d.summary||{},labels={CLOSE_NO_LEADS:'3天0留资停投',CLOSE_HIGH_CPL:'高CPL停投',CLOSE_SPEND_SPIKE:'花费突增止损',SCALE_UP:'优质扩量',PACE_UP:'预算补速',PACE_DOWN:'预算降速'};summary.innerHTML=`<span>已复盘 <b>${s.actions||0}</b> 次</span><span>改善 <b>${s.improved||0}</b></span><span>恶化 <b>${s.worsened||0}</b></span><span>观察中 <b>${s.observing||0}</b></span><span>估算节省 <b>$${Number(s.estimated_saved||0).toFixed(2)}</b></span><span>数据缓存：${escapeHtml(d.cached_at||'刚刚')}</span>`;rows.innerHTML=(d.items||[]).map(x=>{const status=x.outcome==='改善'||x.outcome==='已止损'?'scale':x.outcome==='恶化'?'close':'info';return `<tr><td><span class="rule-status ${status}">${escapeHtml(x.outcome)}</span></td><td>${escapeHtml(x.executed_at)}<br><span class="hint">${escapeHtml(labels[x.action]||x.action)} · 已过 ${x.elapsed_days} 天</span></td><td><b>${escapeHtml(x.adset_name||x.adset_id)}</b><br><span class="hint">${escapeHtml(x.product||'')} · ${escapeHtml((x.countries||[]).join(','))}</span></td><td class="number-cell">${effectivenessMetric(x.before3)}</td><td class="number-cell">${effectivenessMetric(x.after3)}</td><td class="number-cell">${effectivenessMetric(x.after7)}</td><td class="number-cell">${x.estimated_saved?'$'+Number(x.estimated_saved).toFixed(2):'—'}</td><td>${x.rollback_available?`<button class="ghost" onclick="rollbackAutomation('${escapeHtml(x.action_id)}')">一键撤销</button>`:x.rolled_back_at?'<span class="hint">已撤销</span>':'—'}</td></tr>`}).join('')||'<tr><td colspan="8" class="empty-row">还没有已执行的自动化动作；预览记录不会计入复盘。</td></tr>'}catch(e){summary.innerHTML='<span class="bad">复盘读取失败：'+escapeHtml(e.message)+'</span>';rows.innerHTML='<tr><td colspan="8" class="empty-row">请稍后重试</td></tr>'}}
function openAutomationBatch(){const selected=Array.from(AUTOMATION_SELECTED).map(i=>({index:i,item:AUTOMATION_ITEMS[i]})).filter(x=>automationItemSelectable(x.item));if(!selected.length){alert('请先勾选可重建的广告组。');return}const grouped=new Map();selected.forEach(x=>{if(!grouped.has(x.item.product))grouped.set(x.item.product,[]);grouped.get(x.item.product).push(x)});AUTOMATION_BATCH_GROUPS=Array.from(grouped.entries()).map(([product,items])=>({product,items}));document.getElementById('automation_batch_summary').innerHTML=`已选择 <b>${selected.length}</b> 个广告组，涉及 <b>${AUTOMATION_BATCH_GROUPS.length}</b> 个产品。`;document.getElementById('automation_batch_product_controls').innerHTML=AUTOMATION_BATCH_GROUPS.map((g,i)=>{const points=OPTIONS.assets?.[g.product]||[],previous=new Set(g.items.map(x=>x.item.previous_selling_point).filter(Boolean));return `<div class="quick-product-card"><b>${escapeHtml(g.product)} · ${g.items.length} 项</b><select id="automation_batch_point_${i}" onchange="refreshAutomationBatchLanguage(${i})">${points.map(p=>`<option value="${escapeHtml(p.name)}">${escapeHtml(p.name)}${previous.has(p.name)?'（上次使用）':''}</option>`).join('')}</select><select id="automation_batch_language_${i}"></select><span class="hint">统一应用</span></div>`}).join('');AUTOMATION_BATCH_GROUPS.forEach((g,i)=>{const select=document.getElementById('automation_batch_point_'+i),previous=new Set(g.items.map(x=>x.item.previous_selling_point).filter(Boolean)),alternative=[...(select?.options||[])].find(option=>!previous.has(option.value));if(select&&alternative)select.value=alternative.value;refreshAutomationBatchLanguage(i)});document.getElementById('automation_batch_rows').innerHTML=selected.map(({item})=>`<tr><td><b>${escapeHtml(item.adset_name)}</b><div class="hint">${escapeHtml(item.adset_id)}</div></td><td>${escapeHtml(item.product)}</td><td><span class="muted-chip">${escapeHtml(item.previous_selling_point||'未识别')}</span></td><td>${(item.countries||[]).map(countryLabel).map(escapeHtml).join('、')}</td><td class="number-cell">$${Number(item.daily_budget||0).toFixed(2)}/天</td><td>${escapeHtml(item.campaign_name||item.campaign_id)}</td></tr>`).join('');openDialog('dlg_automation_batch')}
function refreshAutomationBatchLanguage(groupIndex){const select=document.getElementById('automation_batch_language_'+groupIndex);if(!select)return;select.innerHTML='<option value="AUTO">按预算表自动匹配</option>';select.disabled=true}
function confirmAutomationBatch(){if(!AUTOMATION_BATCH_GROUPS.length)return;let added=0;AUTOMATION_BATCH_GROUPS.forEach((group,groupIndex)=>{const point=v('automation_batch_point_'+groupIndex);group.items.forEach(({item})=>{pushQueueItem(item.product,point,'AUTO',item.countries,String(item.daily_budget||v('daily_budget')),item.campaign_id);added++})});closeDialog('dlg_automation_batch');setAutomationSelection(false);switchPage('ads');renderQueue();alert(`已将 ${added} 个广告组加入投放队列，语种将按预算表匹配。请先做投放前检查，再生成并运行。`)}
function openRecreate(index){const item=AUTOMATION_ITEMS[index];if(!item)return;if(!item.product||!OPTIONS.assets?.[item.product]){alert('无法从广告系列或广告组名称识别产品，请先确保名称中包含产品型号。');return}if(!item.countries?.length){alert('原广告组没有读取到国家配置，无法保持其余配置不变。');return}RECREATE_ITEM=item;setv('recreate_adset',item.adset_name);setv('recreate_previous_point',item.previous_selling_point||'未识别');setv('recreate_product',item.product);setv('recreate_config',`${item.countries.join(',')} · $${item.daily_budget||'-'}/天`);const point=document.getElementById('recreate_point');point.innerHTML='';(OPTIONS.assets[item.product]||[]).forEach(x=>point.add(new Option(`${x.name}${x.name===item.previous_selling_point?'（上次使用）':''}`,x.name)));const alternative=[...point.options].find(option=>option.value!==item.previous_selling_point);if(alternative)point.value=alternative.value;updateRecreateLanguages();openDialog('dlg_recreate')}
function updateRecreateLanguages(){if(!RECREATE_ITEM)return;const select=document.getElementById('recreate_language');select.innerHTML='<option value="AUTO">按预算表自动匹配</option>';select.disabled=true}
function confirmRecreate(){if(!RECREATE_ITEM||!v('recreate_point'))return;pushQueueItem(RECREATE_ITEM.product,v('recreate_point'),'AUTO',RECREATE_ITEM.countries,String(RECREATE_ITEM.daily_budget||v('daily_budget')),RECREATE_ITEM.campaign_id);closeDialog('dlg_recreate');switchPage('ads');renderQueue();alert('已加入投放队列，语种将按预算表匹配。请先做投放前检查，再生成并运行。')}
function fileAsDataUrl(file){return new Promise((resolve,reject)=>{const reader=new FileReader();reader.onload=()=>resolve(reader.result);reader.onerror=()=>reject(new Error('文件读取失败'));reader.readAsDataURL(file)})}
function openBudgetDialog(){switchPage('budget')}
async function importQuarterBudget(){const input=document.getElementById('quarter_budget_file'),file=input.files?.[0],summary=document.getElementById('budget_summary');if(!file){summary.innerHTML='<span class="warn">请先选择预算表。</span>';return}summary.textContent='正在导入预算表...';try{const data=await fileAsDataUrl(file),result=await api('/api/budget/import',{name:file.name,data});summary.innerHTML=`<span class="ok">已导入 ${result.imported} 条预算。</span>${result.errors?.length?' · 忽略 '+result.errors.length+' 条错误':''}`;await loadBudgetStatus(false);input.value=''}catch(e){summary.innerHTML='<span class="bad">导入失败：'+escapeHtml(e.message)+'</span>'}}
function budgetItemForProduct(product){const campaign=(OPTIONS.campaign_map||{})[product]||'',key=String(product||'').replace(/[^A-Za-z0-9]+/g,'').toUpperCase();return (BUDGET_STATUS?.items||[]).find(x=>(campaign&&String(x.campaign_id)===String(campaign))||String(x.product||'').replace(/[^A-Za-z0-9]+/g,'').toUpperCase()===key)}
function updateAutoBudgetHint(){const hint=document.getElementById('auto_budget_hint');if(!hint)return;const item=budgetItemForProduct(v('product'));if(v('daily_budget')){hint.textContent='当前使用手动日预算；清空后改用季度预算自动计算。';return}hint.textContent=item?(item.waiting_for_min_budget?`自动预算仅 $${Number(item.daily_budget||0).toFixed(2)}/天，低于 Meta 最低 $1.01，已暂缓投放；预计 ${item.budget_eligible_date||'后续日期'} 可重新检测。`:`自动预算：$${Number(item.daily_budget||0).toFixed(2)}/天（剩余 $${Number(item.remaining||0).toFixed(2)} ÷ ${item.remaining_days} 天）`):'自动预算 =（季度初始预算 − 已花费）÷ 季度剩余天数；请先导入预算表。'}
async function loadBudgetStatus(force){
  const summary=document.getElementById('budget_summary'),rows=document.getElementById('budget_rows'),kpis=document.getElementById('budget_kpis');
  summary.textContent=force?'正在刷新季度花费与 ACTIVE 地区...':'正在读取季度预算与投放覆盖...';
  try{
    BUDGET_STATUS=await api('/api/budget/status'+(force?'?force=1':''));
    const d=BUDGET_STATUS,items=d.items||[],source=d.spend_source==='meta_api'||d.coverage_source==='meta_api'?'刚从 Meta 更新':d.spend_source==='stale_disk'||d.coverage_source==='stale_disk'?'Meta 暂不可用，使用旧缓存':d.current_count?'一小时缓存':'尚无预算',missing=items.reduce((n,x)=>n+(x.missing_countries||[]).length,0),waiting=items.filter(x=>x.waiting_for_min_budget).length,openBtn=document.getElementById('quick_launch_open'),totals=items.reduce((t,x)=>{t.initial+=Number(x.initial_budget||0);t.spent+=Number(x.spent||0);t.leads+=Number(x.leads||0);t.remaining+=Number(x.remaining||0);t.projected+=Number(x.projected_spend||0);return t},{initial:0,spent:0,leads:0,remaining:0,projected:0});
    summary.innerHTML=`<b>${d.year} Q${d.quarter}</b> · ${escapeHtml(d.period)} · 已过 ${d.elapsed_days}/${d.total_days} 天 · 剩余 ${d.remaining_days} 天 · ACTIVE 广告组 ${d.active_adsets||0} 个 · <span class="${missing?'coverage-missing':'coverage-ok'}">缺失地区 ${missing} 个</span>${waiting?` · <span class="warn">等待预算 ${waiting} 组</span>`:''} · ${escapeHtml(source)}${d.source_name?' · '+escapeHtml(d.source_name):''}`;
    if(kpis)kpis.innerHTML=`<div class="budget-kpi"><span>季度初始预算</span><b>$${totals.initial.toFixed(2)}</b></div><div class="budget-kpi"><span>已花费</span><b>$${totals.spent.toFixed(2)}</b></div><div class="budget-kpi"><span>Leads</span><b>${compactNumber(totals.leads)}</b></div><div class="budget-kpi"><span>剩余预算</span><b>$${totals.remaining.toFixed(2)}</b></div><div class="budget-kpi"><span>季末预计完成率</span><b>${totals.initial?(totals.projected/totals.initial*100).toFixed(1):'0.0'}%</b></div>`;
    if(openBtn){openBtn.disabled=!missing;openBtn.textContent=missing?`一键投放缺失地区（${missing}）`:'没有缺失地区'}
    const paceLabels={on_track:'正常',slow:'进度偏慢',overspend:'超进度'};
    rows.innerHTML=items.map(x=>{
      const countries=(x.countries||[]).length?(x.countries||[]).map(countryLabel).join('、'):'未配置地区',coverage=(x.blocked_countries||[]).length?`<span class="bad">本季度禁投：${(x.blocked_countries||[]).map(countryLabel).join('、')}</span>${(x.missing_countries||[]).length?`<br><span class="coverage-missing">其余缺投放：${(x.missing_countries||[]).map(countryLabel).join('、')}</span>`:''}`:x.coverage_status==='missing'?`<span class="coverage-missing">缺投放：${(x.missing_countries||[]).map(countryLabel).join('、')}</span>`:x.coverage_status==='active'?'<span class="coverage-ok">已有 ACTIVE 广告</span>':'<span class="hint">预算表未配置地区</span>',forecast=Number(x.projected_completion||0)>=100&&x.exhaustion_date?`按当前速度预计 ${escapeHtml(x.exhaustion_date)} 用完`:`季末预计剩余 $${Math.max(0,Number(x.initial_budget||0)-Number(x.projected_spend||0)).toFixed(2)}`;
      return `<tr><td><b>${escapeHtml(x.product||'未填写产品')}</b><div class="hint">${escapeHtml(x.campaign_name||x.campaign_id||'未匹配 Campaign')}</div></td><td>${escapeHtml(countries)}</td><td><span class="budget-source ${x.asset_language?'':'manual'}">${escapeHtml(x.asset_language||'自动')}</span></td><td class="number-cell"><b>$${Number(x.initial_budget||0).toFixed(2)}</b></td><td class="number-cell"><b>$${Number(x.spent||0).toFixed(2)}</b><div class="hint">Leads ${compactNumber(x.leads)} · CPL ${Number(x.leads||0)?'$'+Number(x.cpl||0).toFixed(2):'—'}</div></td><td class="number-cell"><b>$${Number(x.remaining||0).toFixed(2)}</b><div class="hint">$${Number(x.daily_budget||0).toFixed(2)}/天 · ${x.remaining_days} 天</div>${x.waiting_for_min_budget?`<div class="warn">低于 Meta 最低 $1.01，暂不投放<br>预计 ${escapeHtml(x.budget_eligible_date||'后续日期')} 重新检测</div>`:''}</td><td><span class="pace-pill ${escapeHtml(x.pace_status||'on_track')}">${escapeHtml(paceLabels[x.pace_status]||'正常')}</span><div class="hint">已花 / 应花 ${Number(x.pace_percent||0).toFixed(1)}%<br>$${Number(x.spent||0).toFixed(2)} / $${Number(x.expected_spend||0).toFixed(2)}</div></td><td class="number-cell"><b>$${Number(x.projected_spend||0).toFixed(2)} · ${Number(x.projected_completion||0).toFixed(1)}%</b><div class="hint">${forecast}</div></td><td>${coverage}</td><td><span class="budget-source ${x.exhausted||x.waiting_for_min_budget||(x.blocked_countries||[]).length?'exhausted':''}">${!x.matched?'未匹配':(x.blocked_countries||[]).length?'本季度禁投':x.exhausted?'已用完':x.waiting_for_min_budget?'等待预算':'可投放'}</span></td></tr>`
    }).join('')||'<tr><td colspan="10" class="empty-row">当前季度没有预算数据，请导入预算表</td></tr>';
    updateAutoBudgetHint()
  }catch(e){
    summary.innerHTML='<span class="bad">预算读取失败：'+escapeHtml(e.message)+'</span>';
    rows.innerHTML='<tr><td colspan="10" class="empty-row">无法读取预算</td></tr>'
  }
}
function quickFixedLanguage(item){return String(item?.asset_language||'DEFAULT').trim().toUpperCase()||'DEFAULT'}
function quickCompatiblePoints(item){const fixed=quickFixedLanguage(item),points=OPTIONS.assets?.[item.product]||[];if(['DEFAULT','AUTO'].includes(fixed))return points;return points.filter(point=>(point.languages||['DEFAULT']).some(lang=>String(lang).toUpperCase()===fixed))}
function openQuickLaunch(){const allItems=BUDGET_STATUS?.items||[],waiting=allItems.filter(x=>(x.missing_countries||[]).length&&x.waiting_for_min_budget).length;QUICK_LAUNCH_ITEMS=allItems.filter(x=>(x.missing_countries||[]).length&&!x.exhausted&&!x.waiting_for_min_budget&&x.matched).map(x=>({product:x.product,campaign_id:x.campaign_id,countries:x.missing_countries,daily_budget:x.daily_budget,remaining:x.remaining,remaining_days:x.remaining_days,asset_language:x.asset_language||'DEFAULT'}));const body=document.getElementById('quick_launch_rows'),summary=document.getElementById('quick_launch_summary');if(!QUICK_LAUNCH_ITEMS.length){alert(waiting?`当前缺失地区均在等待预算（${waiting} 组），日均达到 $1.01 后才会进入一键投放。`:'当前没有可一键投放的缺失地区。');return}let unavailable=0;body.innerHTML=QUICK_LAUNCH_ITEMS.map((x,i)=>{const fixed=quickFixedLanguage(x),points=quickCompatiblePoints(x),ready=!!points.length,pointOptions=ready?points.map(p=>`<option value="${escapeHtml(p.name)}">${escapeHtml(p.name)} · ${escapeHtml((p.languages||['DEFAULT']).join('/'))}</option>`).join(''):`<option value="">无 ${escapeHtml(fixed)} 素材</option>`;if(!ready)unavailable++;return `<tr><td><input id="quick_check_${i}" type="checkbox" ${ready?'checked':'disabled'} onchange="updateQuickSelectionCount()"></td><td><b>${escapeHtml(x.product)}</b><div class="hint">${escapeHtml(x.campaign_id)}</div><span class="budget-language-note ${ready?'':'bad'}">预算语种：${escapeHtml(fixed)}</span></td><td>${x.countries.map(countryLabel).map(escapeHtml).join('、')}</td><td><b>$${Number(x.daily_budget||0).toFixed(2)}</b><div class="hint">剩余 $${Number(x.remaining||0).toFixed(2)} / ${x.remaining_days} 天</div></td><td><select id="quick_point_${i}" ${ready?'':'disabled'}>${pointOptions}</select>${ready?'':`<div class="quick-material-error">请先补充 ${escapeHtml(fixed)} 语种素材</div>`}</td></tr>`}).join('');summary.innerHTML=`发现 <b>${QUICK_LAUNCH_ITEMS.reduce((n,x)=>n+x.countries.length,0)}</b> 个可投缺失地区，涉及 <b>${QUICK_LAUNCH_ITEMS.length}</b> 组预算${waiting?`；<span class="warn">另有 ${waiting} 组日均不足 $1.01，已暂缓</span>`:''}。预算语种已自动带入${unavailable?`；<span class="bad">${unavailable} 组没有对应语种素材，已取消选择</span>`:'。'}`;renderQuickProductControls();updateQuickSelectionCount();openDialog('dlg_quick_launch')}
function setQuickSelection(checked){QUICK_LAUNCH_ITEMS.forEach((_,i)=>{const box=document.getElementById('quick_check_'+i);if(box&&!box.disabled)box.checked=checked});updateQuickSelectionCount()}
function updateQuickSelectionCount(){const selected=QUICK_LAUNCH_ITEMS.filter((_,i)=>{const box=document.getElementById('quick_check_'+i);return box?.checked&&!box.disabled}),countries=selected.reduce((n,x)=>n+x.countries.length,0),label=document.getElementById('quick_selected_count'),confirmButton=document.getElementById('quick_launch_confirm');if(label)label.textContent=`已选择 ${selected.length} 组 / ${countries} 个地区`;if(confirmButton){confirmButton.disabled=!selected.length;confirmButton.textContent=selected.length?`确定投放（${countries} 个地区）`:'确定投放'}}
function renderQuickProductControls(){const products=Array.from(new Set(QUICK_LAUNCH_ITEMS.map(x=>x.product)));QUICK_PRODUCT_GROUPS=products.map(product=>({product,indexes:QUICK_LAUNCH_ITEMS.map((x,i)=>x.product===product?i:-1).filter(i=>i>=0)}));document.getElementById('quick_product_controls').innerHTML=QUICK_PRODUCT_GROUPS.map((g,i)=>{const validIndexes=g.indexes.filter(index=>quickCompatiblePoints(QUICK_LAUNCH_ITEMS[index]).length),required=Array.from(new Set(validIndexes.map(index=>quickFixedLanguage(QUICK_LAUNCH_ITEMS[index])).filter(lang=>!['DEFAULT','AUTO'].includes(lang)))),points=(OPTIONS.assets?.[g.product]||[]).filter(point=>required.every(lang=>(point.languages||['DEFAULT']).some(itemLang=>String(itemLang).toUpperCase()===lang))),pointOptions=points.map(p=>`<option value="${escapeHtml(p.name)}">${escapeHtml(p.name)} · ${escapeHtml((p.languages||['DEFAULT']).join('/'))}</option>`).join('');return `<div class="quick-product-card quick-launch-product-card"><div class="quick-product-meta"><b>${escapeHtml(g.product)} · ${validIndexes.length} 组</b><span>预算语种：${escapeHtml(required.join('/')||'默认')}</span></div><select id="quick_group_point_${i}" ${points.length?'':'disabled'}>${pointOptions||'<option value="">没有覆盖全部预算语种的共同卖点</option>'}</select><button class="secondary" onclick="applyQuickProductSettings(${i})" ${points.length?'':'disabled'}>统一应用卖点</button></div>`}).join('')}
function applyQuickProductSettings(index){const group=QUICK_PRODUCT_GROUPS[index],point=v('quick_group_point_'+index);if(!point)return;group.indexes.forEach(i=>{const check=document.getElementById('quick_check_'+i),pointSelect=document.getElementById('quick_point_'+i);if(pointSelect&&[...pointSelect.options].some(option=>option.value===point)){pointSelect.value=point;if(check&&!check.disabled)check.checked=true}});updateQuickSelectionCount()}
function collectQuickLaunchItems(){return QUICK_LAUNCH_ITEMS.map((x,i)=>({item:x,index:i})).filter(x=>document.getElementById('quick_check_'+x.index)?.checked).map(({item,index})=>({product:item.product,selling_point:v('quick_point_'+index),asset_language:'AUTO',countries:item.countries,daily_budget:'',campaign_id:item.campaign_id,dry_run:v('quick_launch_mode')})).filter(x=>x.selling_point&&x.countries.length)}
function addMissingToQueue(){const items=collectQuickLaunchItems();if(!items.length){alert('请至少选择一项。');return}items.forEach(x=>pushQueueItem(x.product,x.selling_point,x.asset_language,x.countries,'',x.campaign_id));closeDialog('dlg_quick_launch');switchPage('ads');renderQueue();alert(`已将 ${items.length} 组缺失地区加入队列，请先做投放前检查。`)}
async function launchMissingNow(){const items=collectQuickLaunchItems(),box=document.getElementById('quick_launch_summary'),mode=v('quick_launch_mode');if(!items.length){alert('请至少选择一项。');return}const countryCount=items.reduce((n,x)=>n+x.countries.length,0),modeName=mode==='yes'?'DRY RUN（不会创建广告）':'LIVE 正式投放';if(!confirm(`即将对 ${countryCount} 个缺失地区执行 ${modeName}。\n系统会先检查预算、素材、表单和 Campaign。\n\n确认继续？`))return;box.textContent='正在执行投放前检查...';try{const payload={defaults:{...defaults(),daily_budget:'',dry_run:mode},items,country_mode:'merge'},check=await api('/api/preflight',payload),warnings=(check.items||[]).filter(x=>x.status!=='OK');if(warnings.length){box.innerHTML='<span class="bad">有 '+warnings.length+' 项未通过检查，请先加入队列查看详细问题。</span>';return}box.textContent='检查通过，正在生成计划...';const plan=await api('/api/generate_plan',payload),run=await api('/api/run',{});if(!run.started){showRunErrorPopup(run.status||{});throw new Error(run.status?.run_error?.message||'投放程序当前正在运行，未重复启动。')}closeDialog('dlg_quick_launch');switchPage('ads');log(`一键投放已启动（${modeName}）\n广告组：${plan.rows}\n计划：${plan.plan}`);setTimeout(pollStatus,1000)}catch(e){box.innerHTML='<span class="bad">一键投放失败：'+escapeHtml(e.message)+'</span>'}}
async function refreshOptions(){try{OPTIONS=await api('/api/options');const d=OPTIONS.defaults||{};['page_id','lead_form_id','website_url','cta_type','language_mode','custom_language','dry_run','delay_max_seconds','ad_account_id','url_tags'].forEach(k=>setv(k,d[k]||''));setv('daily_budget','');renderProducts();renderCountries();renderCountrySets();toggleCustomLanguage();loadPreview();updateAutoBudgetHint();loadBudgetStatus(false);log('素材包已刷新。')}catch(e){log('刷新失败：'+e.message)}}
function productUrlTags(product){const slug=String(product||'').trim().toLowerCase();return slug?`utm_source=facebook&utm_medium=socialad&utm_campaign=fjd-${slug}&utm_id=5224`:''}
function syncUrlTags(){const prod=v('product');const field=document.getElementById('url_tags');if(!field)return;const current=String(field.value||'').trim();if(!prod){if(!current)field.value='';return}const next=productUrlTags(prod);const m=current.match(/utm_campaign=fjd-([^&]+)/i);if(!current||current.includes('{产品名}')||(m&&m[1].toLowerCase()!==String(prod).trim().toLowerCase()))field.value=next}
function renderProducts(){const p=document.getElementById('product');const old=p.value;p.innerHTML='';Object.keys(OPTIONS.assets||{}).forEach(x=>p.add(new Option(x,x)));if(old)p.value=old;updateSellingPoints()}
function updateSellingPoints(){const prod=v('product');const s=document.getElementById('selling_point');const old=s.value;s.innerHTML='';(OPTIONS.assets[prod]||[]).forEach(x=>s.add(new Option(`${x.name} (${x.languages?.join('/')||'DEFAULT'} · ${x.image_count||0}图${x.copy_exists?' · 有文案':' · 缺文案'})`,x.name)));if(old)s.value=old;updateAssetLanguages();const u=(OPTIONS.product_urls?.products||{})[prod]||'';setv('website_url',u);syncUrlTags();updateAutoBudgetHint()}
function updateAssetLanguages(){const prod=v('product'), sp=v('selling_point');const langSel=document.getElementById('asset_language');const current=langSel.value;langSel.innerHTML='';let point=(OPTIONS.assets[prod]||[]).find(x=>x.name===sp);(point?.languages||['DEFAULT']).forEach(x=>langSel.add(new Option(x,x)));if(current)langSel.value=current;loadPreview()}async function loadPreview(){const prod=v('product'),sp=v('selling_point'),lang=v('asset_language');if(!prod||!sp)return;try{const info=await api(`/api/asset_preview?product=${encodeURIComponent(prod)}&selling_point=${encodeURIComponent(sp)}&asset_language=${encodeURIComponent(lang)}`),sources=(info.copy_sources||[]).join('、');document.getElementById('asset_note').textContent=`${info.asset_language||lang}：图片 ${info.images.length} 张，正文 ${info.copy_counts?.primary_text||0} 条，标题 ${info.copy_counts?.headline||0} 条，描述 ${info.copy_counts?.description||0} 条${sources?' · 已识别 '+sources:''}`;document.getElementById('copy_preview').textContent=['正文：\n'+(info.copy.primary_text||''),'标题：\n'+(info.copy.headline||''),'描述：\n'+(info.copy.description||'')].join('\n\n');document.getElementById('image_preview').innerHTML=info.images.map(x=>`<img class="thumb" src="/asset?path=${encodeURIComponent(x.rel)}" title="${x.name}">`).join('')||'<span class="hint warn">没有图片</span>'}catch(e){document.getElementById('copy_preview').textContent='预览失败：'+e.message}}
function escapeHtml(x){return String(x??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]))}
function countryAliasMap(){const m={};(OPTIONS.countries||[]).forEach(([code,name])=>{m[String(code).toUpperCase()]=String(code).toUpperCase();m[String(name).toUpperCase()]=String(code).toUpperCase()});const aliases={
'UNITED STATES':'US','USA':'US','AMERICA':'US','美国':'US','美國':'US','CANADA':'CA','加拿大':'CA','MEXICO':'MX','墨西哥':'MX','BRAZIL':'BR','巴西':'BR','CHILE':'CL','智利':'CL','COLOMBIA':'CO','哥伦比亚':'CO','PERU':'PE','秘鲁':'PE','ARGENTINA':'AR','阿根廷':'AR',
'UNITED KINGDOM':'GB','UK':'GB','BRITAIN':'GB','ENGLAND':'GB','英国':'GB','IRELAND':'IE','爱尔兰':'IE','FRANCE':'FR','法国':'FR','GERMANY':'DE','德国':'DE','ITALY':'IT','意大利':'IT','SPAIN':'ES','西班牙':'ES','PORTUGAL':'PT','葡萄牙':'PT','NETHERLANDS':'NL','HOLLAND':'NL','荷兰':'NL','BELGIUM':'BE','比利时':'BE','SWITZERLAND':'CH','瑞士':'CH','AUSTRIA':'AT','奥地利':'AT',
'POLAND':'PL','波兰':'PL','CZECH':'CZ','CZECH REPUBLIC':'CZ','捷克':'CZ','SLOVAKIA':'SK','斯洛伐克':'SK','HUNGARY':'HU','匈牙利':'HU','ROMANIA':'RO','罗马尼亚':'RO','BULGARIA':'BG','保加利亚':'BG','CROATIA':'HR','克罗地亚':'HR','SLOVENIA':'SI','斯洛文尼亚':'SI','SERBIA':'RS','塞尔维亚':'RS','ESTONIA':'EE','爱沙尼亚':'EE','LATVIA':'LV','拉脱维亚':'LV','LITHUANIA':'LT','立陶宛':'LT','SWEDEN':'SE','瑞典':'SE','NORWAY':'NO','挪威':'NO','FINLAND':'FI','芬兰':'FI','DENMARK':'DK','丹麦':'DK','GREECE':'GR','希腊':'GR','TURKEY':'TR','土耳其':'TR',
'SAUDI ARABIA':'SA','SAUDI':'SA','沙特':'SA','沙特阿拉伯':'SA','UAE':'AE','UNITED ARAB EMIRATES':'AE','阿联酋':'AE','QATAR':'QA','卡塔尔':'QA','KUWAIT':'KW','科威特':'KW','OMAN':'OM','阿曼':'OM','BAHRAIN':'BH','巴林':'BH','JORDAN':'JO','约旦':'JO','ISRAEL':'IL','以色列':'IL',
'EGYPT':'EG','埃及':'EG','MOROCCO':'MA','摩洛哥':'MA','ALGERIA':'DZ','阿尔及利亚':'DZ','TUNISIA':'TN','突尼斯':'TN','SOUTH AFRICA':'ZA','南非':'ZA','NIGERIA':'NG','尼日利亚':'NG','KENYA':'KE','肯尼亚':'KE','TANZANIA':'TZ','坦桑尼亚':'TZ','GHANA':'GH','加纳':'GH',
'INDIA':'IN','印度':'IN','PAKISTAN':'PK','巴基斯坦':'PK','BANGLADESH':'BD','孟加拉':'BD','SRI LANKA':'LK','斯里兰卡':'LK','JAPAN':'JP','日本':'JP','KOREA':'KR','SOUTH KOREA':'KR','韩国':'KR','INDONESIA':'ID','印尼':'ID','印度尼西亚':'ID','MALAYSIA':'MY','马来西亚':'MY','SINGAPORE':'SG','新加坡':'SG','THAILAND':'TH','泰国':'TH','VIETNAM':'VN','越南':'VN','PHILIPPINES':'PH','菲律宾':'PH','AUSTRALIA':'AU','澳大利亚':'AU','NEW ZEALAND':'NZ','新西兰':'NZ'};Object.entries(aliases).forEach(([k,v])=>m[k]=v);return m}
function parseCountriesText(raw){const text=String(raw||'').trim();if(!text)return[];const alias=countryAliasMap();const parts=text.split(/[|,;，、\n]+|\s{2,}/).map(x=>x.trim()).filter(Boolean);let out=[];for(const part of parts){const key=part.toUpperCase();let code=alias[key];if(!code&&/^[A-Z]{2}$/.test(key))code=key;if(!code){const compact=key.replace(/\./g,'').replace(/\s+/g,' ');code=alias[compact]}if(!code)throw new Error('国家无法识别：'+part+'。请填国家代码，例如 US,CA,FR');out.push(code)}return Array.from(new Set(out))}
function renderCountries(){const sel=document.getElementById('country_dropdown'),blocked=new Set(['TW','SG']);if(!sel)return;sel.innerHTML='<option value="">从下拉列表选择国家，选中后自动加入输入框</option>';OPTIONS.countries.forEach(([code,name])=>{const opt=document.createElement('option');opt.value=code;opt.disabled=blocked.has(code);opt.textContent=code+' '+name+(blocked.has(code)?'（本季度禁投）':'');sel.appendChild(opt)});updateTopStatus();updateCountryCount();renderCountrySets()}
function addCountryFromDropdown(){const sel=document.getElementById('country_dropdown');const code=(sel.value||'').toUpperCase();if(!code)return;const arr=Array.from(new Set([...parseCustom(),code]));setv('custom_countries',arr.join(','));sel.value='';updateCountryCount()}
function updateCountryCount(){let n=0;try{n=selectedCountries().length}catch(e){}document.getElementById('country_count').textContent='已选 '+n;updateTopStatus()}
function selectCountries(list){setv('custom_countries',Array.from(new Set(list)).join(','));updateCountryCount()}
function clearCountries(){setv('custom_countries','');const sel=document.getElementById('country_dropdown');if(sel)sel.value='';updateCountryCount()}
function countrySetsKey(){return 'ads_country_sets_v1'}
function loadCountrySets(){try{return JSON.parse(localStorage.getItem(countrySetsKey())||'[]')}catch(e){return[]}}
function saveCountrySets(items){localStorage.setItem(countrySetsKey(),JSON.stringify(items||[]))}
function renderCountrySets(){const wrap=document.getElementById('country_presets');if(!wrap)return;const sets=loadCountrySets();if(!sets.length){wrap.innerHTML='<span class="hint">还没有保存国家组合。先手动选一组，再命名保存。</span>';return}wrap.innerHTML=sets.map((x,i)=>`<span class="saved-set"><button type="button" onclick="loadSavedCountrySet(${i})">${escapeHtml(x.name)}</button><button type="button" class="remove" title="删除" onclick="removeCountrySet(${i})">×</button></span>`).join('')}
function saveCountrySet(){const name=(document.getElementById('country_set_name').value||'').trim();const countries=selectedCountries();if(!countries.length){alert('请先选择国家');return}if(!name){alert('请给这组国家起个名字');return}const sets=loadCountrySets();const idx=sets.findIndex(x=>String(x.name||'').toLowerCase()===name.toLowerCase());const item={name:name,countries:countries};if(idx>=0)sets[idx]=item;else sets.unshift(item);saveCountrySets(sets.slice(0,20));renderCountrySets();document.getElementById('country_set_name').value='';}
function loadSavedCountrySet(i){const sets=loadCountrySets();const item=sets[i];if(!item)return;setv('custom_countries',Array.from(new Set(item.countries||[])).join(','));updateCountryCount()}
function removeCountrySet(i){const sets=loadCountrySets();sets.splice(i,1);saveCountrySets(sets);renderCountrySets()}
function parseCustom(){return parseCountriesText(v('custom_countries'))}
function selectedCountries(){return parseCustom()}
function toggleCustomLanguage(){const w=document.getElementById('custom_language_wrap'); if(w) w.style.display=(v('language_mode')==='custom')?'block':'none';updateTopStatus()}
function updateTopStatus(){const assetCount=Object.keys(OPTIONS.assets||{}).length;const formPages=Object.keys(OPTIONS.forms||{}).length;const hasToken=(OPTIONS.status&&OPTIONS.status.token)||false;const set=(id,txt)=>{const el=document.getElementById(id);if(el)el.textContent=txt};set('sys_token','Token：'+(hasToken?'已检测':'未确认'));set('sys_assets','素材包：'+(assetCount?assetCount+' 个产品':'未检测'));set('sys_forms','表单库：'+(formPages?formPages+' 个 Page':'未加载'));set('sys_mode','模式：'+(v('dry_run')==='yes'?'DRY RUN':'LIVE'));set('sys_ads','广告状态：ACTIVE')}
function pushQueueItem(product,point,assetLang,countries,budget,campaignId=''){const runMode=v('dry_run');const delayNow=v('delay_max_seconds');const mode=v('country_mode'),language=String(assetLang||'AUTO').toUpperCase();if(mode==='split'){countries.forEach(c=>QUEUE.push({product:product,selling_point:point,asset_language:language,countries:[c],daily_budget:budget,campaign_id:campaignId,dry_run:runMode,delay_max_seconds:delayNow}))}else{QUEUE.push({product:product,selling_point:point,asset_language:language,countries:countries,daily_budget:budget,campaign_id:campaignId,dry_run:runMode,delay_max_seconds:delayNow})}}
function addItem(){let cs=[];try{cs=selectedCountries()}catch(e){alert(e.message);return}const budget=v('daily_budget');if(!v('product')||!v('selling_point')||!cs.length){alert('请选择产品、卖点和国家');return}pushQueueItem(v('product'),v('selling_point'),'AUTO',cs,budget);clearCountries();renderQueue()}
function splitBatchLine(line){if(line.includes('\t'))return line.split('\t').map(x=>x.trim());if(line.includes('|'))return line.split('|').map(x=>x.trim());return line.split(',').map(x=>x.trim())}
function headerIndex(cols){const names={product:['产品','product','prod'],point:['卖点','selling_point','selling point','素材包','point'],lang:['语种','素材语种','language','lang','asset_language','asset language'],countries:['国家','国家代码','countries','country','country_code','country code'],budget:['预算','budget','daily_budget','daily budget','日预算']};const lower=cols.map(x=>String(x||'').trim().toLowerCase());let map={};for(const [key,arr] of Object.entries(names)){for(const name of arr){const i=lower.indexOf(name.toLowerCase());if(i>=0){map[key]=i;break}}}return map}
function parseBatchItems(){const box=document.getElementById('batch_text');const text=(box.value||'').trim();if(!text)throw new Error('请先粘贴批量表');const lines=text.replace(/\r\n/g,'\n').replace(/\r/g,'\n').split('\n').map(x=>x.trim()).filter(Boolean);let map=null,start=0;let first=splitBatchLine(lines[0]);let detected=headerIndex(first);if(['product','point','countries'].every(k=>detected[k]!==undefined)){map=detected;start=1}else{map=first.length>=5?{product:0,point:1,lang:2,countries:3,budget:4}:{product:0,point:1,countries:2,budget:3}}
let items=[],errors=[];for(let idx=start;idx<lines.length;idx++){try{const cols=splitBatchLine(lines[idx]);const get=k=>map[k]===undefined?'':String(cols[map[k]]??'').trim();if(cols.length<3&&start===0)throw new Error('列数不足');const product=get('product');const point=get('point');const lang='AUTO';const countries=parseCountriesText(get('countries'));const budget=get('budget');if(!product||!point||!countries.length)throw new Error('产品/卖点/国家不能为空');if(!OPTIONS.assets[product])throw new Error('产品不存在：'+product);const p=(OPTIONS.assets[product]||[]).find(x=>x.name===point);if(!p)throw new Error('卖点素材包不存在：'+product+' / '+point);items.push({product:product,selling_point:point,asset_language:lang,countries:countries,daily_budget:budget})}catch(e){errors.push({line:idx+1,error:e.message,raw:lines[idx]})}}
return {items,errors,hasHeader:start===1}}
function renderBatchPreview(parsed){const result=document.getElementById('batch_result');let html='';if(parsed.items.length){html+=`<div class="table-scroll"><table class="checktable"><tr><th>#</th><th>产品</th><th>卖点</th><th>语种</th><th>国家</th><th>预算</th></tr>`;parsed.items.forEach((x,i)=>{html+=`<tr><td>${i+1}</td><td>${escapeHtml(x.product)}</td><td>${escapeHtml(x.selling_point)}</td><td><span class="budget-source">按预算表</span></td><td>${x.countries.map(c=>`<span class="tag">${c}</span>`).join('')}</td><td>${x.daily_budget?'$'+escapeHtml(x.daily_budget):'<span class="budget-source">自动计算</span>'}</td></tr>`});html+='</table></div>'}else{html+='<span class="warn">没有解析到可加入的行。</span>'}if(parsed.errors.length){html+=`<div class="bad" style="margin-top:8px">错误 ${parsed.errors.length} 行：<br>`+parsed.errors.map(e=>`第 ${e.line} 行：${escapeHtml(e.error)} <span class="hint">${escapeHtml(e.raw)}</span>`).join('<br>')+'</div>'}html+=`<p class="hint">${parsed.hasHeader?'已识别表头。':'未识别表头，按固定顺序解析。'} 可加入 ${parsed.items.length} 行；语种和空白预算均按季度预算表匹配。</p>`;result.innerHTML=html}
function previewBatchItems(){try{const parsed=parseBatchItems();window.__BATCH_PARSED=parsed;renderBatchPreview(parsed)}catch(e){document.getElementById('batch_result').innerHTML='<span class="bad">解析失败：'+escapeHtml(e.message)+'</span>'}}
function confirmBatchItems(){try{const parsed=parseBatchItems();window.__BATCH_PARSED=parsed;renderBatchPreview(parsed);if(parsed.errors.length&&!confirm('有 '+parsed.errors.length+' 行错误，只加入正确的 '+parsed.items.length+' 行吗？'))return;parsed.items.forEach(x=>pushQueueItem(x.product,x.selling_point,x.asset_language,x.countries,x.daily_budget));renderQueue();if(parsed.items.length){document.getElementById('batch_text').value='';document.getElementById('batch_result').innerHTML='<span class="ok">已加入 '+parsed.items.length+' 行到队列。</span>'}}catch(e){document.getElementById('batch_result').innerHTML='<span class="bad">加入失败：'+escapeHtml(e.message)+'</span>'}}
function renderQueue(){const tb=document.getElementById('queue');tb.innerHTML='';QUEUE.forEach((x,i)=>{const b=x.daily_budget||v('daily_budget')||'',autoLang=['','DEFAULT','AUTO'].includes(String(x.asset_language||'').toUpperCase());tb.innerHTML+=`<tr><td>${i+1}</td><td>${escapeHtml(x.product)}</td><td>${escapeHtml(x.selling_point)}</td><td>${autoLang?'<span class="budget-source">按预算表</span>':escapeHtml(x.asset_language)}</td><td>${x.countries.map(c=>`<span class="tag">${c}</span>`).join('')}</td><td>${b?('$'+escapeHtml(b)):'<span class="budget-source">自动计算</span>'}</td><td>${(x.dry_run||v('dry_run'))==='yes'?'预检':'正式'}</td><td><button class="ghost smallbtn" onclick="editItem(${i})">编辑</button><button class="secondary smallbtn" onclick="copyItem(${i})">复制</button><button class="danger smallbtn" onclick="removeItem(${i})">删除</button></td></tr>`});document.getElementById('queue_hint').textContent=QUEUE.length?`共 ${QUEUE.length} 个广告组 / ${v('daily_budget')?'手动预算 $'+v('daily_budget'):'季度预算自动计算'}`:'待添加'}
function removeItem(i){QUEUE.splice(i,1);renderQueue()}
function copyItem(i){const x=JSON.parse(JSON.stringify(QUEUE[i]));QUEUE.splice(i+1,0,x);renderQueue()}
function editItem(i){const x=QUEUE[i];const product=prompt('产品',x.product);if(product===null)return;const point=prompt('卖点',x.selling_point);if(point===null)return;const countries=prompt('国家，可填 US,CA 或国家名',x.countries.join(','));if(countries===null)return;const budget=prompt('预算 USD',x.daily_budget||v('daily_budget')||'');if(budget===null)return;try{x.product=product.trim();x.selling_point=point.trim();x.asset_language='AUTO';x.countries=parseCountriesText(countries);x.daily_budget=budget.trim();renderQueue()}catch(e){alert(e.message)}}
function defaults(){return{enabled:'Y',daily_budget:v('daily_budget'),page_id:v('page_id'),lead_form_id:v('lead_form_id'),website_url:v('website_url'),cta_type:v('cta_type'),language_mode:v('language_mode'),custom_language:v('custom_language'),dry_run:v('dry_run'),delay_max_seconds:v('delay_max_seconds'),ad_account_id:v('ad_account_id'),campaign_id:'',url_tags:v('url_tags'),asset_language:'AUTO'}}
async function createPack(){try{const p=v('new_product')||v('product');const sp=v('new_selling_point')||v('selling_point');const lang=v('new_asset_language')||'EN';const res=await api('/api/create_pack',{product:p,selling_point:sp,asset_language:lang});document.getElementById('pack_result').innerHTML='已创建：<br>'+res.folder+'<br>请把图片放到：'+res.image_dir+'<br>文案文件：'+res.copy_file;await refreshOptions()}catch(e){document.getElementById('pack_result').innerHTML='<span class="bad">创建失败：'+e.message+'</span>'}}
async function checkPacks(){try{const res=await api('/api/check_packs');let html='<div class="table-scroll">共 '+res.count+' 个素材包。<table class="checktable"><tr><th>状态</th><th>产品</th><th>卖点</th><th>语种</th><th>图片</th><th>文案</th><th>问题</th></tr>';res.items.forEach(x=>{html+=`<tr><td class="${x.status==='OK'?'ok':'warn'}">${x.status}</td><td>${x.product}</td><td>${x.selling_point}</td><td>${x.asset_language||'DEFAULT'}</td><td>${x.images}</td><td>${x.copy_exists?'有':'缺'}</td><td>${(x.problems||[]).join('<br>')||'-'}</td></tr>`});html+='</table></div>';document.getElementById('pack_result').innerHTML=html}catch(e){document.getElementById('pack_result').innerHTML='<span class="bad">检查失败：'+e.message+'</span>'}}

async function checkHashLibrary(){try{const res=await api('/api/image_hash_status');let html='<div class="table-scroll">图片 Hash 状态：共 '+res.total+' 张，已入库 '+res.ready+'，缺失 '+res.missing+'。<br><span class="hint">库文件：'+res.library+'</span><table class="checktable"><tr><th>状态</th><th>产品</th><th>卖点</th><th>语种</th><th>图片</th><th>Hash</th></tr>';res.items.slice(0,300).forEach(x=>{html+=`<tr><td class="${x.status==='READY'?'ok':'warn'}">${x.status}</td><td>${x.product}</td><td>${x.selling_point}</td><td>${x.asset_language||'DEFAULT'}</td><td>${x.file_name}</td><td><span class="hint">${x.hash||'未入库'}</span></td></tr>`});html+='</table></div>';document.getElementById('pack_result').innerHTML=html}catch(e){document.getElementById('pack_result').innerHTML='<span class="bad">Hash状态读取失败：'+e.message+'</span>'}}
async function syncImageHashes(){if(!confirm('会把素材包中缺少 Hash 的图片上传到 Meta 广告账户。确认继续吗？'))return;try{const res=await api('/api/sync_image_hashes',{ad_account_id:v('ad_account_id')});let msg=`图片 Hash 同步完成：上传 ${res.uploaded} 张，已存在跳过 ${res.skipped_ready} 张，剩余缺失 ${res.after?.missing||0} 张。`;if(res.errors&&res.errors.length){msg+='\n\n失败：\n'+res.errors.map(x=>`${x.file_name}: ${x.error}`).join('\n')}log(msg);await checkHashLibrary()}catch(e){log('同步图片Hash失败：'+e.message)}}
async function showLogs(){try{const res=await api('/api/logs');let html='<div class="table-scroll">最近日志 '+res.count+' 个：<table class="checktable"><tr><th>文件</th><th>时间</th><th>大小</th><th>路径</th></tr>';res.items.forEach(x=>{html+=`<tr><td>${x.name}</td><td>${x.modified}</td><td>${x.size_kb} KB</td><td><span class="hint">${x.path}</span></td></tr>`});html+='</table></div>';document.getElementById('pack_result').innerHTML=html}catch(e){document.getElementById('pack_result').innerHTML='<span class="bad">读取失败：'+e.message+'</span>'}}
async function syncForms(){try{const res=await api('/api/sync_forms',{page_id:v('page_id')});log(`Lead Forms 同步完成：Page ${v('page_id')}，总数 ${res.count}，识别 ${res.matched}\n缓存时间：${res.library.updated_at}`);await refreshOptions()}catch(e){log('同步失败：'+e.message)}}
async function preflight(){
  if(!QUEUE.length){alert('队列为空');return null}
  try{
    const res=await api('/api/preflight',{defaults:defaults(),items:QUEUE,country_mode:v('country_mode')});
    let ok=0,bad=0;
    let html='<div class="table-scroll"><table class="checktable"><tr><th>状态</th><th>产品/卖点/语种</th><th>国家</th><th>日预算</th><th>素材</th><th>表单</th><th>Campaign</th><th>落地页</th><th>问题</th></tr>';
    res.items.forEach(x=>{
      if(x.status==='OK')ok++;else bad++;
      const waitingNote=x.waiting_for_min_budget?`<div class="warn">低于 Meta 最低 $1.01，已暂缓投放<br>预计 ${escapeHtml(x.budget_eligible_date||'后续日期')} 重新检测</div>`:'';
      const blockedNote=(x.blocked_countries||[]).length?`<div class="bad">本季度禁投：${(x.blocked_countries||[]).map(escapeHtml).join('、')}</div>`:'';
      const budgetNote=x.budget_source==='quarter'?`<span class="budget-source">季度自动</span>${x.budget_info?`<div class="hint">剩余 $${Number(x.budget_info.remaining||0).toFixed(2)} / ${x.budget_info.remaining_days} 天</div>`:''}${waitingNote}`:x.budget_source==='manual'?'<span class="budget-source manual">手动</span>':'<span class="budget-source exhausted">未匹配</span>';
      html+=`<tr><td class="${x.status==='OK'?'ok':'warn'}">${x.status}</td><td>${escapeHtml(x.product)}<br>${escapeHtml(x.selling_point)}<br><span class="hint">${escapeHtml(x.asset_language||'DEFAULT')}</span></td><td>${(x.countries||[]).join(',')}${blockedNote}</td><td><b>${x.budget?'$'+escapeHtml(x.budget):'—'}</b><br>${budgetNote}</td><td><b>${escapeHtml(x.asset_language||'DEFAULT')}</b> · ${x.images} 张图 / ${x.copy?.primary_text?'有文案':'缺文案'}<div class="hint" title="${escapeHtml(x.asset_folder||'')}">严格匹配预算语种</div></td><td>${escapeHtml(x.form_name||'未匹配')}<br><span class="hint">${escapeHtml(x.form_id||'')}</span></td><td>${escapeHtml(x.campaign_id||'缺失')}</td><td><span class="hint">${escapeHtml(x.website_url||'')}</span></td><td>${(x.problems||[]).map(escapeHtml).join('<br>')||'-'}</td></tr>`;
    });
    html+='</table></div>';
    html='<p class="hint"><span class="ok">OK '+ok+'</span> / <span class="warn">需处理 '+bad+'</span>。检查项包含预算、素材、文案、表单、Campaign、落地页和本季度禁投地区。</p>'+html;
    document.getElementById('preflight_result').innerHTML=html;
    return res;
  }catch(e){document.getElementById('preflight_result').innerHTML='<span class="bad">检查失败：'+escapeHtml(e.message)+'</span>';return null}
}
function showRunErrorPopup(status){if(!status||status.running||!status.run_error)return;const error=status.run_error||{},key=String(status.run_id||status.finished_at||status.started_at||error.message||'run-error');if(LAST_RUN_ERROR_POPUP===key)return;LAST_RUN_ERROR_POPUP=key;const title=document.getElementById('run_error_title'),meta=document.getElementById('run_error_meta'),message=document.getElementById('run_error_message'),parts=[];if(error.excel_row)parts.push(`计划第 ${error.excel_row} 行`);if(Number(error.skipped_count||0)>0)parts.push(`已暂停后续 ${error.skipped_count} 项`);if(title)title.textContent=error.adset_name?`失败广告组：${error.adset_name}`:'投放运行错误';if(meta)meta.textContent=parts.join(' · ')||'后续计划已暂停';if(message)message.textContent=error.message||'投放脚本运行失败，请查看完整运行日志。';openDialog('dlg_run_error')}
function focusRunErrorLog(){closeDialog('dlg_run_error');switchPage('ads');const box=document.getElementById('status');if(box)box.scrollIntoView({behavior:'smooth',block:'center'})}
function updateRunStatusUi(status){const badge=document.getElementById('run_state_badge'),meta=document.getElementById('run_state_meta');if(!badge||!meta)return;const summary=status?.summary||{};if(status?.running){badge.textContent='正在投放';badge.className='run-state-badge running';meta.textContent=`开始于 ${status.started_at||'刚刚'} · 已完成 ${Number(summary.success||0)+Number(summary.dry_run||0)} / ${summary.planned||'—'}`}else if(status?.returncode===0&&status?.finished_at){badge.textContent='运行完成';badge.className='run-state-badge success';meta.textContent=`完成于 ${status.finished_at} · 成功 ${summary.success||0}${summary.dry_run?` · Dry Run ${summary.dry_run}`:''}`}else if(status?.returncode!==null&&status?.returncode!==undefined){badge.textContent='已暂停';badge.className='run-state-badge error';meta.textContent=`结束于 ${status.finished_at||'未知'} · 失败 ${summary.failed||1} · 跳过 ${summary.skipped||0}`}else{badge.textContent='等待运行';badge.className='run-state-badge idle';meta.textContent='尚未开始投放'}}
function showRunSummaryPopup(status){const summary=status?.summary||{},planned=Number(summary.planned||0);if(!status||status.running||Number(status.returncode)!==0||!planned)return;const key=String(status.run_id||status.finished_at||'run-summary');if(LAST_RUN_SUMMARY_POPUP===key)return;LAST_RUN_SUMMARY_POPUP=key;const values=[['计划',planned],['成功',Number(summary.success||0)],['Dry Run',Number(summary.dry_run||0)],['失败',Number(summary.failed||0)],['跳过',Number(summary.skipped||0)]];document.getElementById('run_summary_time').textContent=`${status.started_at||''} 至 ${status.finished_at||''}`;document.getElementById('run_summary_kpis').innerHTML=values.map(([label,value])=>`<div><span>${label}</span><b>${value}</b></div>`).join('');document.getElementById('run_summary_message').innerHTML=Number(summary.dry_run||0)?'Dry Run 检查已完成，没有创建真实广告。':'队列已执行完成。建议到“广告数据”确认花费、Leads 和 CPL，再根据自动化规则继续观察。';openDialog('dlg_run_summary')}
function focusRunSummaryLog(){closeDialog('dlg_run_summary');switchPage('ads');const box=document.getElementById('status');if(box)box.scrollIntoView({behavior:'smooth',block:'center'})}
async function generatePlan(run){if(!QUEUE.length){alert('队列为空');return}if(run){const mode=v('dry_run')==='yes'?'DRY RUN（不创建）':'LIVE 正式创建';const total=QUEUE.length;if(!confirm(`即将生成 plan.xlsx 并运行 run.py。\n队列：${total} 个广告组\n模式：${mode}\n默认广告组/广告状态：ACTIVE\n\n确认继续？`))return}try{const res=await api('/api/generate_plan',{defaults:defaults(),items:QUEUE,country_mode:v('country_mode')});log(`已生成 plan.xlsx\n行数：${res.rows}\n备份：${res.backup||'无'}`);if(run){const started=await api('/api/run',{});updateRunStatusUi(started.status||{});if(!started.started){showRunErrorPopup(started.status||{});throw new Error(started.status?.run_error?.message||'投放程序当前正在运行，未重复启动。')}log('已开始运行 run.py，日志会在这里实时刷新。');setTimeout(pollStatus,1000)}}catch(e){log('生成失败：'+e.message)}}async function checkStatus(){try{const s=await api('/api/status');log((s.output||'')+`\n\nreturncode: ${s.returncode}\nrunning: ${s.running}\nlog: ${s.log_path||''}`);updateRunStatusUi(s);showRunErrorPopup(s);showRunSummaryPopup(s);return s}catch(e){log('状态失败：'+e.message);return {running:false}}}async function pollStatus(){const s=await checkStatus();if(s.running){setTimeout(pollStatus,1500)}}document.addEventListener('change',e=>{if(['dry_run','language_mode'].includes(e.target.id))updateTopStatus()});refreshOptions();api('/api/status').then(updateRunStatusUi).catch(()=>{});
</script></body></html>'''


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): return
    def do_GET(self):
        path=urlparse(self.path).path; qs=parse_qs(urlparse(self.path).query)
        try:
            if path in {"/","/index.html"}: return tresp(self, HTML)
            if path=="/api/ping": return jresp(self,{"ok":True,"time":datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
            if path=="/api/options": return jresp(self,{"assets":scan_asset_packs(),"countries":COUNTRY_OPTIONS,"defaults":get_defaults(),"campaign_map":get_campaign_map(),"forms":load_forms_library(),"product_urls":load_product_urls(),"product_settings":load_product_settings(),"status":{"token":TOKEN_FILE.exists() or bool(os.getenv("META_ACCESS_TOKEN")),"asset_root":ASSET_ROOT.exists(),"forms_library":FORMS_LIBRARY_FILE.exists(),"default_ad_status":"ACTIVE"}})
            if path=="/api/asset_preview": return jresp(self, inspect_pack(unquote(qs.get("product",[""])[0]), unquote(qs.get("selling_point",[""])[0]), unquote(qs.get("asset_language",[""])[0])))
            if path=="/api/status": return jresp(self,last_run)
            if path=="/api/check_packs": return jresp(self, check_all_packs())
            if path=="/api/logs": return jresp(self, list_logs())
            if path=="/api/image_hash_status": return jresp(self, image_hash_status())
            if path=="/api/leads/options": return jresp(self,{"products":lead_products()})
            if path=="/api/leads/status": return jresp(self,last_lead_sync)
            if path=="/api/leads/files": return jresp(self,list_lead_files())
            if path=="/api/leads/download": return send_lead_file(self, qs.get("name",[""])[0])
            if path=="/api/insights": return jresp(self,fetch_adset_insights(qs.get("days",["7"])[0], qs.get("force",["0"])[0] in {"1","true","yes"}))
            if path=="/api/insights/export": return send_insights_excel(self, qs.get("days",["7"])[0], qs.get("campaign_id",[""])[0], qs.get("country",[""])[0])
            if path=="/api/report/export": return send_operating_report(self, qs.get("period",["week"])[0])
            if path=="/api/budget/status": return jresp(self,quarter_budget_status(qs.get("force",["0"])[0] in {"1","true","yes"}))
            if path=="/api/budget/template": return send_budget_template(self)
            if path=="/api/automation/settings": return jresp(self,automation_status_payload())
            if path=="/api/automation/status": return jresp(self,automation_status_payload())
            if path=="/api/automation/effectiveness": return jresp(self,automation_effectiveness())
            if path=="/asset":
                rel=unquote(qs.get("path",[""])[0]); p=(BASE_DIR/rel).resolve()
                if not str(p).startswith(str(BASE_DIR.resolve())) or not p.exists(): return jresp(self,{"error":"文件不存在 / Asset not found","detail":f"图片路径不存在或不在 ADS 文件夹内：{rel}","hint":"如果是图片预览失败，请检查素材包/产品/卖点/图片 里的图片是否还在，文件名不要包含特殊路径。"},404)
                data=p.read_bytes(); self.send_response(200); self.send_header("Content-Type","image/png"); self.send_header("Content-Length",str(len(data))); self.end_headers(); self.wfile.write(data); return
            return not_found_response(self, path)
        except Exception as e: return jresp(self,{"error":str(e),"traceback":traceback.format_exc(),"hint":"后端处理请求时报错。请把这段完整错误发给我，尤其是 traceback。"},400)
    def do_POST(self):
        path=urlparse(self.path).path
        try:
            payload=read_body(self)
            if path=="/api/sync_forms": return jresp(self, sync_forms(payload.get("page_id")))
            if path=="/api/preflight": return jresp(self,{"items":preflight(payload)})
            if path=="/api/generate_plan":
                rows=build_rows(payload)
                unresolved_languages=[row for row in rows if norm(row.get("asset_language")).upper() in {"", "DEFAULT", "AUTO"}]
                if unresolved_languages:
                    details=[f"{row.get('product')} / {','.join(parse_countries(row.get('countries')))}" for row in unresolved_languages]
                    raise RuntimeError("以下产品和地区没有从季度预算表匹配到素材语种，已停止生成。请先在预算表填写语种：\n"+"\n".join(details[:30]))
                invalid_assets=[]
                for row in rows:
                    pack=inspect_pack(row.get("product"),row.get("selling_point"),row.get("asset_language"))
                    if not pack.get("exists") or not pack.get("images") or not pack.get("copy_exists"):
                        invalid_assets.append(f"{row.get('product')} / {row.get('selling_point')} / {row.get('asset_language') or 'DEFAULT'}")
                if invalid_assets:
                    raise RuntimeError("以下项目没有严格匹配预算表语种的完整素材（图片+文案），已停止生成，且不会回退到 EN 或其他语种：\n"+"\n".join(invalid_assets[:30]))
                blocked=[row for row in rows if any(country in BLOCKED_COUNTRIES for country in parse_countries(row.get("countries")))]
                if blocked: raise RuntimeError("计划包含本季度禁投地区 TW/SG，已停止生成。请从队列移除台湾和新加坡；下季度再从预算表正式移除。")
                invalid=[row for row in rows if _number(row.get("daily_budget")) <= 0]
                if invalid: raise RuntimeError("有广告组没有可用预算。请导入当前季度预算表，或填写手动日预算，并先做投放前检查。")
                below_minimum=[row for row in rows if 0 < _number(row.get("daily_budget")) < META_MIN_DAILY_BUDGET]
                if below_minimum: raise RuntimeError(f"有广告组的日均预算低于 Meta 最低 ${META_MIN_DAILY_BUDGET:.2f}，已停止生成。请等待预算日均达到门槛后再投，或调整季度预算。")
                duplicate_map = find_duplicate_active_ads(rows)
                if duplicate_map:
                    details = []
                    for row_index, hits in duplicate_map.items():
                        row = rows[row_index]
                        for hit in hits:
                            details.append(f"{row.get('product')} / {row.get('selling_point')} / {','.join(hit.get('countries') or [])}：{hit.get('name') or hit.get('id')}")
                    raise RuntimeError("检测到相同产品地区和卖点的重复投放，已阻止生成：\n" + "\n".join(details[:20]))
                paths=write_plan(rows); return jresp(self,{"ok":True,"rows":len(rows),**paths})
            if path=="/api/create_pack": return jresp(self, create_asset_pack(payload.get("product"), payload.get("selling_point"), payload.get("asset_language") or "EN"))
            if path=="/api/sync_image_hashes": return jresp(self, sync_image_hashes(payload.get("ad_account_id") or get_defaults().get("ad_account_id")))
            if path=="/api/run": return jresp(self,{"ok":True,"started":run_ads_script(),"status":last_run})
            if path=="/api/leads/start":
                started=start_lead_sync(payload)
                return jresp(self,{"ok":started,"started":started,"status":last_lead_sync,"error":"客户同步正在运行，或启动条件未满足。" if not started else ""},200 if started else 409)
            if path=="/api/budget/import": return jresp(self,import_quarter_budget(payload))
            if path=="/api/automation/settings": return jresp(self,{"rules":save_automation_rules(payload)})
            if path=="/api/automation/preview": return jresp(self,evaluate_automation_rules(execute=False))
            if path=="/api/automation/run": return jresp(self,evaluate_automation_rules(execute=True))
            if path=="/api/automation/apply": return jresp(self,apply_automation_items(payload.get("keys") or []))
            if path=="/api/automation/rollback": return jresp(self,rollback_automation_action(payload.get("action_id")))
            return not_found_response(self, path)
        except Exception as e: return jresp(self,{"error":str(e),"traceback":traceback.format_exc(),"hint":"后端处理请求时报错。请把这段完整错误发给我，尤其是 traceback。"},400)

def main():
    os.chdir(BASE_DIR); CONFIG_DIR.mkdir(exist_ok=True); CACHE_DIR.mkdir(exist_ok=True); ASSET_ROOT.mkdir(exist_ok=True)
    threading.Thread(target=automation_scheduler, daemon=True).start()
    url=f"http://{HOST}:{PORT}"; print(f"ADS 前端已启动：{url}")
    if str(os.getenv("ADS_NO_BROWSER") or "").strip().lower() not in {"1", "true", "yes"}:
        threading.Timer(1.0,lambda:webbrowser.open(url)).start()
    ThreadingHTTPServer((HOST,PORT),Handler).serve_forever()
if __name__=="__main__": main()
