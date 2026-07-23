import copy
import hashlib
import json
import os
import random
import re
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook, Workbook

BASE_DIR = Path(__file__).resolve().parent
CONFIG_DIR = BASE_DIR / "config"
PLAN_FILE = BASE_DIR / "plan.xlsx"
API_VERSION = "v23.0"
DEFAULT_LOG_DIR = BASE_DIR / "日志"  # 日志默认写到当前 ADS 文件夹，不绑定某个电脑用户
CACHE_DIR = BASE_DIR / "_cache"
IMAGE_LIBRARY_FILE = "image_library.json"
CREATIVE_LIBRARY_FILE = "creative_library.json"
TARGETING_CACHE_FILE = "targeting_interest_cache.json"
SAVED_AUDIENCE_CACHE_FILE = "saved_audience_cache.json"
AUDIENCE_FILE = CONFIG_DIR / "audiences.json"
FORMS_LIBRARY_FILE = "forms_library.json"
PRODUCT_SETTINGS_FILE = CONFIG_DIR / "product_settings.json"
PRODUCT_URLS_FILE = CONFIG_DIR / "product_urls.json"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
DEFAULT_ASSET_ROOT = BASE_DIR / "素材包"
TEXT_FILE_NAMES = ["文案.txt", "copy.txt", "ad_copy.txt", "text.txt"]
IMAGE_FOLDER_NAMES = ["图片", "images", "image"]
BLOCKED_COUNTRIES = {"TW", "SG"}

COUNTRY_LOCAL_LANGUAGES = {
    "US": ["English"], "CA": ["English", "French"], "MX": ["Spanish"], "BR": ["Portuguese"], "CL": ["Spanish"], "CO": ["Spanish"], "PE": ["Spanish"], "AR": ["Spanish"],
    "GB": ["English"], "IE": ["English", "Irish"], "FR": ["French"], "DE": ["German"], "IT": ["Italian"], "ES": ["Spanish"], "PT": ["Portuguese"],
    "NL": ["Dutch"], "BE": ["French", "Dutch"], "LU": ["French", "German"], "CH": ["German", "French", "Italian"], "AT": ["German"],
    "PL": ["Polish"], "CZ": ["Czech"], "SK": ["Slovak"], "HU": ["Hungarian"], "RO": ["Romanian"], "BG": ["Bulgarian"],
    "HR": ["Croatian"], "SI": ["Slovenian"], "RS": ["Serbian"], "BA": ["Bosnian"], "ME": ["Serbian"], "MK": ["Macedonian"], "AL": ["Albanian"],
    "EE": ["Estonian"], "LV": ["Latvian"], "LT": ["Lithuanian"],
    "SE": ["Swedish"], "NO": ["Norwegian"], "FI": ["Finnish"], "DK": ["Danish"], "IS": ["Icelandic"],
    "GR": ["Greek"], "CY": ["Greek"], "MT": ["English"],
    "TR": ["Turkish"], "UA": ["Ukrainian"], "MD": ["Romanian"], "GE": ["Georgian"], "AM": ["Armenian"], "AZ": ["Azerbaijani"],
    "SA": ["Arabic"], "AE": ["Arabic", "English"], "QA": ["Arabic"], "KW": ["Arabic"], "OM": ["Arabic"], "BH": ["Arabic"], "JO": ["Arabic"], "IL": ["Hebrew"],
    "EG": ["Arabic"], "MA": ["Arabic", "French"], "DZ": ["Arabic", "French"], "TN": ["Arabic", "French"],
    "ZA": ["English"], "NG": ["English"], "KE": ["English"], "TZ": ["Swahili", "English"], "GH": ["English"], "ET": ["Amharic"],
    "IN": ["Hindi", "English"], "PK": ["Urdu", "English"], "BD": ["Bengali"], "LK": ["Sinhala", "Tamil"], "NP": ["Nepali"],
    "CN": ["Chinese"], "HK": ["Chinese", "English"], "TW": ["Chinese"], "JP": ["Japanese"], "KR": ["Korean"],
    "ID": ["Indonesian"], "MY": ["Malay", "English"], "SG": ["English", "Chinese"], "TH": ["Thai"], "VN": ["Vietnamese"], "PH": ["English"], "KH": ["Khmer"], "LA": ["Lao"], "MM": ["Burmese"],
    "AU": ["English"], "NZ": ["English"],
}

# 命名展示用，避免广告名里出现过长或不统一的语种文案。
LANGUAGE_DISPLAY_NAMES = {
    "English": "English", "Irish": "Irish", "French": "French", "German": "German", "Italian": "Italian", "Spanish": "Spanish", "Portuguese": "Portuguese", "Dutch": "Dutch",
    "Polish": "Polish", "Czech": "Czech", "Slovak": "Slovak", "Hungarian": "Hungarian", "Romanian": "Romanian", "Bulgarian": "Bulgarian", "Croatian": "Croatian", "Slovenian": "Slovenian", "Serbian": "Serbian", "Bosnian": "Bosnian", "Macedonian": "Macedonian", "Albanian": "Albanian",
    "Estonian": "Estonian", "Latvian": "Latvian", "Lithuanian": "Lithuanian", "Swedish": "Swedish", "Norwegian": "Norwegian", "Finnish": "Finnish", "Danish": "Danish", "Icelandic": "Icelandic", "Greek": "Greek",
    "Turkish": "Turkish", "Ukrainian": "Ukrainian", "Georgian": "Georgian", "Armenian": "Armenian", "Azerbaijani": "Azerbaijani", "Arabic": "Arabic", "Hebrew": "Hebrew",
    "Swahili": "Swahili", "Amharic": "Amharic", "Hindi": "Hindi", "Urdu": "Urdu", "Bengali": "Bengali", "Sinhala": "Sinhala", "Tamil": "Tamil", "Nepali": "Nepali",
    "Chinese": "Chinese", "Japanese": "Japanese", "Korean": "Korean", "Indonesian": "Indonesian", "Malay": "Malay", "Thai": "Thai", "Vietnamese": "Vietnamese", "Khmer": "Khmer", "Lao": "Lao", "Burmese": "Burmese",
}

# 自定义语言输入支持：可以填 English / French，也可以填 EN,FR,KR 这类代码。
LANGUAGE_CODE_ALIASES = {
    "EN": "English", "US": "English", "GB": "English",
    "FR": "French", "DE": "German", "IT": "Italian", "ES": "Spanish", "PT": "Portuguese",
    "KR": "Korean", "KO": "Korean", "JP": "Japanese", "JA": "Japanese",
    "ID": "Indonesian", "TH": "Thai", "VN": "Vietnamese", "TR": "Turkish", "PL": "Polish",
    "NL": "Dutch", "CZ": "Czech", "SK": "Slovak", "HR": "Croatian", "EE": "Estonian",
    "AR": "Arabic", "ZH": "Chinese", "CN": "Chinese", "TW": "Chinese", "HK": "Chinese",
    "MS": "Malay", "HI": "Hindi", "SV": "Swedish", "NO": "Norwegian", "FI": "Finnish", "DA": "Danish",
}

def parse_custom_language_names(value):
    """把用户手动输入的语言转成 Meta 可搜索的语言名称。

    支持：English,French / EN,FR / Korean+English / Spanish;Portuguese。
    """
    raw = str(value or "").strip()
    if not raw:
        return []
    parts = [p.strip() for p in re.split(r"[,+/;|\n]+", raw) if p.strip()]
    names = []
    for part in parts:
        key = re.sub(r"[^A-Za-z]+", "", part).upper()
        name = LANGUAGE_CODE_ALIASES.get(key) or part.strip()
        # 统一首字母大写，但保留类似 Portuguese 这类原有格式。
        if name and name not in names:
            names.append(name)
    return names

ERROR_HINTS = {
    "1885260": "创意内容被安全系统拦截：优先检查 Page ID、Lead Form ID、图片、文案、链接是否正确。",
    "1885998": "动态素材广告必须放在动态素材广告组下：确认 is_dynamic_creative=True。",
    "3390001": "缺少潜在客户表单：确认 Lead Form ID 正确，并且属于当前 Page。",
    "1885183": "Meta App 还在开发模式：需要切 Live 或使用已发布 App 的 Token。",
    "190": "Token 无效或过期：重新生成 Access Token。",
    "17": "广告账户 API 调用过多：新版会自动等待 300 秒后重试一次；建议减少重复同步和批量分批跑。",
    "1885949": "CTA/链接参数错误：不要传 deeplink_url/app_link；Lead Form 广告只需要 link + lead_gen_form_id。",
    "100": "参数错误：检查 Campaign ID、Page ID、Lead Form ID、图片 Hash、广告目标是否匹配。",
    "1870227": "Meta 要求明确设置 Advantage Audience：已在新版中自动传 targeting_automation.advantage_audience。",
}


def truthy(value):
    return str(value or "").strip().lower() in {"yes", "y", "true", "1", "是", "ready", "dry_run", "dryrun"}


def split_multi(value):
    if value is None:
        return []
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    parts = []
    for line in text.split("\n"):
        item = line.strip()
        if item:
            parts.append(item)
    return parts


def parse_country_codes(value):
    """支持 ID|NZ / KR,FR,PL / KR;FR;PL / 换行。返回去重后的大写国家代码列表。"""
    if value is None:
        return []
    text = str(value).strip().upper()
    if not text:
        return []
    parts = re.split(r"[|,;\s]+", text)
    countries = []
    for part in parts:
        code = part.strip().upper()
        if not code:
            continue
        if not re.fullmatch(r"[A-Z]{2}", code):
            raise RuntimeError(f"国家代码格式不正确：{code}。多个国家请写成 KR,FR,PL 或 ID|NZ")
        if code not in countries:
            countries.append(code)
    return countries


def country_label(countries):
    return "-".join(countries)


def clean_name_part(value, default=""):
    text = str(value or default).strip()
    # Meta 名称允许 &、空格、括号，这里只清掉换行和容易造成文件/日志混乱的分隔符。
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" _")


def official_language_names_for_countries(countries):
    """根据所选国家返回官方/本地语种名称；未知国家才回退 English。"""
    names = []
    for country_code in countries:
        langs = COUNTRY_LOCAL_LANGUAGES.get(country_code, [])
        for lang in langs:
            if lang not in names:
                names.append(lang)
    return names or ["English"]


def preferred_asset_language_names(row):
    """Return the budget/material language as the first delivery language."""
    value = str(row.get("asset_language") or row.get("copy_language") or "").strip()
    if not value or value.upper() in {"DEFAULT", "AUTO"}:
        return []
    return parse_custom_language_names(value)


def delivery_language_names(row, countries=None):
    """Budget language first, then local country languages, then English."""
    names = []
    for lang in preferred_asset_language_names(row):
        if lang not in names:
            names.append(lang)
    for lang in official_language_names_for_countries(countries or []):
        if lang not in names:
            names.append(lang)
    if "English" not in names:
        names.append("English")
    return names


def language_name_for_ad(row, countries=None):
    """广告命名里的语种字段。

    - language_mode=all 时显示 English (All)，表示不限制语种。
    - language_mode=auto 时按 countries 自动显示官方/本地语种。
    - language_mode=custom 时使用 custom_language，例如 EN,FR -> English+French。
    - 如果表里临时加 name_language，则允许手动覆盖命名显示。
    """
    manual = str(row.get("name_language") or "").strip()
    if manual:
        return clean_name_part(manual)
    language_mode = str(row.get("language_mode") or "auto").strip().lower()
    if language_mode == "all":
        return "English (All)"
    if language_mode == "custom":
        langs = parse_custom_language_names(row.get("custom_language"))
        if not langs:
            return "Custom Language"
    else:
        langs = delivery_language_names(row, countries or [])
    label_parts = [LANGUAGE_DISPLAY_NAMES.get(lang, lang) for lang in langs]
    return clean_name_part("+".join(label_parts), "English")



def load_product_settings():
    return load_json(Path(PRODUCT_SETTINGS_FILE), {"default_page_id": "", "product_lines": {}})


def product_line_for_product(product):
    product_key = normalize_product(product)
    settings = load_product_settings()
    for line_name, info in (settings.get("product_lines") or {}).items():
        products = [normalize_product(p) for p in info.get("products", [])]
        if product_key in products:
            return line_name
    # Fallback for older config files
    scanner_products = {"P1", "P2", "P2 Vision+", "S1", "S2", "V4e", "Model Web"}
    rtk_products = {"V10L", "V10a", "V10i", "V1t", "V1t 5W", "V4e Pro"}
    if product_key in scanner_products:
        return "Scanner"
    if product_key in rtk_products:
        return "RTK"
    return "MC"


def default_page_id_for_product(product):
    line = product_line_for_product(product)
    settings = load_product_settings()
    info = (settings.get("product_lines") or {}).get(line, {})
    return str(info.get("page_id") or settings.get("default_page_id") or "").strip()



def product_line_info_for_product(product):
    """Return the product line configuration for a product.
    Used for saved audience / Advantage Audience defaults.
    """
    line = product_line_for_product(product)
    settings = load_product_settings()
    return (settings.get("product_lines") or {}).get(line, {}) or {}

def audience_group_for_product(product):
    """命名字段里的产品线：MC / RTK / Scanner。"""
    return product_line_for_product(product)


def industry_name_for_ad(row):
    # 用户要求：原 Construction (industry) 字段改为 MC / RTK。
    # 如果表里临时加 name_industry，仍允许覆盖。
    return clean_name_part(row.get("name_industry"), audience_group_for_product(row.get("product")))


def audience_name_for_ad(row):
    return clean_name_part(row.get("name_audience"), "End User&Dealer")


def selling_point_name_for_ad(row):
    # 命名字段：原来的 End User&Dealer 位置改为卖点。
    # 只影响广告组/广告命名，不影响 Saved Audience 定向。
    value = row.get("selling_point") or row.get("asset_pack") or row.get("卖点") or row.get("素材包")
    return clean_name_part(value, "SellingPoint")


def material_name_for_ad(row):
    return clean_name_part(row.get("name_material"), "Picture")


def build_ad_name(row, countries):
    product = normalize_product(row.get("product")) or "PRODUCT"
    country = country_label(countries) if countries else "COUNTRY"
    date_part = datetime.now().strftime("%Y%m%d")
    return "_".join([
        "FJD",
        product,
        selling_point_name_for_ad(row),
        language_name_for_ad(row, countries),
        industry_name_for_ad(row),
        material_name_for_ad(row),
        country,
        date_part,
    ])


def now_stamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def run_suffix():
    return datetime.now().strftime("%Y%m%d") + f"{random.randint(0, 999):03d}"


def clean_id(value):
    if value is None:
        return ""
    text = str(value).strip()
    if ":" in text:
        return text.split(":")[-1]
    return text


def normalize_product(product):
    text = str(product or "").strip()
    compact = re.sub(r"[^A-Za-z0-9]+", "", text).upper()
    aliases = {
        "V4EPRO": "V4e Pro", "V4EMINI": "V4e Pro", "V10A": "V10a", "V10I": "V10i",
        "V10L": "V10L", "V1T": "V1t", "EG10": "EG10", "EC100": "EC100",
        "GC100": "GC100", "DC100": "DC100", "TERRAMIND": "TerraMind",
        "P1": "P1", "P2": "P2", "P2VISION": "P2 Vision+", "P2VISIONPLUS": "P2 Vision+",
        "P2VISION+": "P2 Vision+", "S1": "S1", "S2": "S2", "V4E": "V4e",
        "MODELWEB": "Model Web",
    }
    return aliases.get(compact, text)


def safe_save_workbook(workbook, target_path, fallback_dir=None, label="workbook"):
    target_path = Path(target_path)
    try:
        workbook.save(target_path)
        return str(target_path)
    except PermissionError:
        fallback_base = Path(fallback_dir or target_path.parent or ".")
        fallback_base.mkdir(parents=True, exist_ok=True)
        fallback_path = fallback_base / f"{target_path.stem}_autosave_{now_stamp()}{target_path.suffix}"
        workbook.save(fallback_path)
        print(f"WARNING: {label} 无法保存到 {target_path}，通常是因为文件正被 Excel 打开。")
        print(f"WARNING: 已改存到备份文件：{fallback_path}")
        return str(fallback_path)


def get_token():
    """优先读取环境变量；其次读取当前 ADS 文件夹里的 token.txt。
    前端/批量模式不能等待命令行输入，否则会表现为一直 running=true。
    """
    token = os.getenv("META_ACCESS_TOKEN")
    if token and token.strip():
        return token.strip()

    token_file = Path("token.txt")
    if token_file.exists():
        token = token_file.read_text(encoding="utf-8").strip()
        invalid_tokens = {"", "PASTE_TOKEN_HERE", "把你的Meta Token粘贴到这里"}
        if token and token not in invalid_tokens:
            return token

    raise RuntimeError(
        "缺少 Meta Access Token：请打开 ADS/token.txt，删除占位文字并粘贴有效 token 后保存。"
        "前端运行模式不会弹出输入框，否则程序会卡住。"
    )


def load_json(path, default):
    path = Path(path)
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def save_json(path, data):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")



def product_key_for_form(product):
    return re.sub(r"[^A-Za-z0-9]+", "", str(product or "")).upper()

FORM_LANG_FROM_NAME = {
    "english":"EN", "french":"FR", "spanish":"ES", "portuguese":"PT", "italian":"IT",
    "japanese":"JP", "korean":"KR", "russian":"RU", "turkish":"TR", "german":"DE",
    "dutch":"NL", "polish":"PL", "indonesian":"ID",
}

def preferred_form_lang_from_row(row, countries):
    if str(row.get("language_mode") or "").strip().lower() == "custom":
        names = parse_custom_language_names(row.get("custom_language"))
        if names:
            first = str(names[0]).strip()
            return FORM_LANG_FROM_NAME.get(first.lower(), re.sub(r"[^A-Za-z]+", "", first).upper()[:2]) or "EN"
    return choose_form_lang(countries)

def normalize_forms_library_for_lookup(lib):
    out = {"by_product": {}, "by_page": {}}
    def add(page_id, product, lang, item):
        if not isinstance(item, dict):
            return
        fid = str(item.get("id") or item.get("form_id") or "").strip()
        if not fid:
            return
        one = dict(item)
        one["id"] = fid
        one["form_id"] = fid
        if page_id:
            one["page_id"] = str(page_id)
        pkey = product_key_for_form(product)
        out["by_product"].setdefault(pkey, {}).setdefault(str(lang).upper(), []).append(one)
        if page_id:
            out["by_page"].setdefault(str(page_id), {"by_product": {}})["by_product"].setdefault(pkey, {}).setdefault(str(lang).upper(), []).append(one)
    if not isinstance(lib, dict):
        return out
    for product, langs in (lib.get("by_product") or {}).items():
        for lang, items in (langs or {}).items():
            if isinstance(items, dict):
                items = [items]
            for item in items or []:
                add(None, product, lang, item)
    for page_id, pdata in (lib.get("by_page") or {}).items():
        byp = (pdata or {}).get("by_product", {}) if isinstance(pdata, dict) else {}
        for product, langs in byp.items():
            for lang, items in (langs or {}).items():
                if isinstance(items, dict):
                    items = [items]
                for item in items or []:
                    add(page_id, product, lang, item)
    for page_id, pdata in lib.items():
        if str(page_id) in {"by_product", "by_page", "raw", "updated_at"}:
            continue
        if not isinstance(pdata, dict):
            continue
        for product, langs in pdata.items():
            if not isinstance(langs, dict):
                continue
            for lang, item in langs.items():
                if isinstance(item, list):
                    for one in item:
                        add(page_id, product, lang, one)
                else:
                    add(page_id, product, lang, item)
    return out

COUNTRY_TO_FORM_LANG = {
    "KR":"KR", "JP":"JP", "ES":"ES", "MX":"ES", "CO":"ES", "CL":"ES", "PE":"ES", "AR":"ES",
    "BR":"PT", "PT":"PT", "FR":"FR", "DE":"DE", "IT":"IT", "TR":"TR", "PL":"PL", "ID":"ID", "TH":"TH", "VN":"VN",
    "US":"EN", "GB":"EN", "IE":"EN", "CA":"EN", "AU":"EN", "NZ":"EN", "SG":"EN", "MY":"EN",
    "AE":"EN", "SA":"EN", "QA":"EN", "KW":"EN", "OM":"EN", "BH":"EN", "ZA":"EN", "NG":"EN", "KE":"EN", "PH":"EN", "IN":"EN",
}

def choose_form_lang(countries):
    langs = []
    for c in countries:
        lang = COUNTRY_TO_FORM_LANG.get(c, "EN")
        if lang not in langs:
            langs.append(lang)
    if len(langs) > 1:
        return "EN"
    return langs[0] if langs else "EN"

def latest_form(items):
    return sorted(items or [], key=lambda x: str(x.get("date") or x.get("created_time") or ""), reverse=True)[0] if items else None

def resolve_lead_form_id(row, countries):
    """lead_form_id 为空时，从 _cache/forms_library.json 里自动匹配。多语种合并默认用 EN 表单兜底。"""
    existing = clean_id(row.get("lead_form_id"))
    if existing:
        row["lead_form_id"] = existing
        return existing
    lib_raw = load_json(Path(CACHE_DIR) / FORMS_LIBRARY_FILE, {"by_product": {}})
    lib = normalize_forms_library_for_lookup(lib_raw)
    pkey = product_key_for_form(row.get("product"))
    lang = preferred_form_lang_from_row(row, countries)
    page_id = clean_id(row.get("page_id"))
    by_product = None
    if page_id:
        by_product = (lib.get("by_page", {}).get(page_id, {}) or {}).get("by_product")
    if not by_product:
        by_product = lib.get("by_product", {})
    product_forms = by_product.get(pkey, {})
    item = latest_form(product_forms.get(lang, [])) or latest_form(product_forms.get("EN", []))
    if not item:
        available = ",".join(sorted(product_forms.keys())) if product_forms else "无"
        raise RuntimeError(f"lead_form_id 为空，并且表单库没有匹配到 {row.get('product')} / {lang} 或 EN。当前产品可用表单语种：{available}。请检查 Page ID、产品名和 _cache/forms_library.json。")
    row["lead_form_id"] = str(item.get("id") or item.get("form_id"))
    print(f"Lead Form matched: {row.get('product')} / {lang} -> {item.get('name')} ({row['lead_form_id']})")
    return row["lead_form_id"]

def file_md5(path):
    h = hashlib.md5()
    with Path(path).open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def content_hash(data):
    raw = json.dumps(data, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def request_meta(method, endpoint, token, payload=None, params=None, retry_on_limit=True):
    """Meta API 请求。遇到广告账户调用限流 code=17 时，自动等待并重试一次。"""
    url = f"https://graph.facebook.com/{API_VERSION}/{endpoint}"
    def _send():
        if method == "GET":
            local_params = dict(params or {})
            local_params["access_token"] = token
            return requests.get(url, params=local_params, timeout=120)
        local_payload = dict(payload or {})
        local_payload["access_token"] = token
        return requests.post(
            url,
            data={k: json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v for k, v in local_payload.items()},
            timeout=120,
        )

    response = None
    for attempt in range(4):
        try:
            response = _send()
            break
        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError, requests.exceptions.Timeout) as exc:
            if attempt == 3:
                safe_error = re.sub(r"(?i)(access_token=)[^&\s]+", r"\1[REDACTED]", str(exc)).replace(token, "[REDACTED]")
                raise RuntimeError(f"连接 Meta 失败，已自动重试 4 次：{safe_error}") from exc
            wait_seconds = 2 ** (attempt + 1)
            print(f"连接 Meta 暂时中断，{wait_seconds} 秒后进行第 {attempt + 2}/4 次尝试...", flush=True)
            time.sleep(wait_seconds)
    assert response is not None
    try:
        data = response.json()
    except Exception:
        raise RuntimeError(f"Meta API returned non-JSON response: {response.text}")

    if "error" in data:
        err = data["error"]
        code = str(err.get("error_subcode") or err.get("code") or "")
        if retry_on_limit and str(err.get("code")) == "17":
            wait_seconds = 300
            print("Meta API 调用限流，自动等待 300 秒后重试一次。")
            time.sleep(wait_seconds)
            return request_meta(method, endpoint, token, payload=payload, params=params, retry_on_limit=False)
        hint = ERROR_HINTS.get(code, "")
        print("Meta API Error:")
        print(json.dumps(data, indent=2, ensure_ascii=False))
        user_title = str(err.get("error_user_title") or "").strip()
        user_message = str(err.get("error_user_msg") or "").strip()
        details = "：".join(part for part in (user_title, user_message) if part)
        if hint:
            details = " | ".join(part for part in (details, f"可能原因：{hint}") if part)
        message = str(err.get("message") or "Meta API error")
        raise RuntimeError(f"{message} | {details}" if details else message)
    return data


def upload_image(ad_account_id, token, image_path):
    url = f"https://graph.facebook.com/{API_VERSION}/{ad_account_id}/adimages"
    response = None
    for attempt in range(4):
        try:
            with Path(image_path).open("rb") as f:
                response = requests.post(url, data={"access_token": token}, files={"filename": f}, timeout=120)
            break
        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError, requests.exceptions.Timeout) as exc:
            if attempt == 3:
                safe_error = re.sub(r"(?i)(access_token=)[^&\s]+", r"\1[REDACTED]", str(exc)).replace(token, "[REDACTED]")
                raise RuntimeError(f"上传图片时连接 Meta 失败，已自动重试 4 次：{safe_error}") from exc
            wait_seconds = 2 ** (attempt + 1)
            print(f"上传连接暂时中断，{wait_seconds} 秒后进行第 {attempt + 2}/4 次尝试...", flush=True)
            time.sleep(wait_seconds)
    assert response is not None
    data = response.json()
    if "error" in data:
        err = data["error"]
        print("Meta API Error while uploading image:")
        print(json.dumps(data, indent=2, ensure_ascii=False))
        raise RuntimeError(err.get("message", "Unknown image upload error"))
    images = data.get("images", {})
    if not images:
        raise RuntimeError(f"No image hash returned after upload: {Path(image_path).name}")
    first_image = next(iter(images.values()))
    image_hash = first_image.get("hash")
    if not image_hash:
        raise RuntimeError(f"No hash found in upload response for: {Path(image_path).name}")
    return image_hash





def product_alias_key(value):
    return re.sub(r"[^A-Za-z0-9+]+", "", str(value or "")).upper()

def load_product_urls():
    return load_json(Path(PRODUCT_URLS_FILE), {"default_url":"https://www.fjdtrion.com/", "products":{}, "aliases":{}})

def resolve_product_url(product):
    data = load_product_urls()
    products = data.get("products", {})
    aliases = data.get("aliases", {})
    p = str(product or "").strip()
    if p in products:
        return products[p]
    compact = product_alias_key(p)
    for k, v in products.items():
        if product_alias_key(k) == compact:
            return v
    alias_target = aliases.get(compact) or aliases.get(p.upper())
    if alias_target and alias_target in products:
        return products[alias_target]
    return data.get("default_url") or "https://www.fjdtrion.com/"

def fill_product_url(row):
    manual = str(row.get("website_url") or "").strip()
    # 如果为空或只是官网首页，则自动替换成对应产品页。手动填了其他链接则保留。
    if (not manual) or manual.rstrip("/") == "https://www.fjdtrion.com":
        row["website_url"] = resolve_product_url(row.get("product"))
    return row

def normalize_header_name(name):
    text = str(name or "").strip()
    aliases = {
        "是否投放": "enabled", "启用": "enabled", "状态": "enabled",
        "产品": "product", "产品型号": "product",
        "卖点": "selling_point", "素材包": "selling_point", "素材包名": "selling_point",
        "素材语种": "asset_language", "文案语种": "asset_language", "asset language": "asset_language",
        "国家": "countries", "国家代码": "countries", "投放国家": "countries",
        "预算": "daily_budget", "日预算": "daily_budget",
        "广告账户": "ad_account_id", "广告账户id": "ad_account_id",
        "广告系列": "campaign_id", "广告系列id": "campaign_id", "campaign": "campaign_id",
        "主页": "page_id", "主页id": "page_id",
        "表单": "lead_form_id", "表单id": "lead_form_id",
        "链接": "website_url", "官网链接": "website_url", "落地页": "website_url",
        "cta": "cta_type", "CTA": "cta_type",
        "语种模式": "language_mode", "语言模式": "language_mode",
        "自定义语言": "custom_language", "指定语言": "custom_language", "custom language": "custom_language",
        "预检": "dry_run", "dryrun": "dry_run",
        "间隔": "delay_max_seconds", "随机间隔": "delay_max_seconds",
        "素材根目录": "asset_root", "素材包根目录": "asset_root",
    }
    return aliases.get(text, text)


def read_text_file(path):
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return Path(path).read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return Path(path).read_text(encoding="utf-8", errors="ignore")


def _copy_key_from_filename(path):
    stem = Path(path).stem.lower().replace("-", "_").replace(" ", "_")
    if any(word in stem for word in ("headline", "title", "标题")):
        return "headline"
    if any(word in stem for word in ("description", "desc", "描述")):
        return "description"
    return "primary_text"


def _plain_text_variants(text):
    normalized = str(text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not normalized:
        return []
    paragraphs = [re.sub(r"\s*\n\s*", " ", block).strip() for block in re.split(r"\n\s*\n+", normalized) if block.strip()]
    if len(paragraphs) > 1:
        return paragraphs
    return [line.strip() for line in normalized.split("\n") if line.strip()]


def _merge_copy_value(target, key, value):
    existing = [line.strip() for line in str(target.get(key, "")).split("\n") if line.strip()]
    incoming = [line.strip() for line in str(value or "").split("\n") if line.strip()]
    target[key] = "\n".join(existing + [line for line in incoming if line not in existing])


def parse_copy_text(text, default_key="primary_text"):
    """解析素材包里的文案文件。

    支持：
    正文1/正文2/正文3/正文4/正文5
    标题1/标题2/标题3/标题4/标题5
    描述1/描述2/描述3/描述4/描述5
    也支持 primary_text/headline/description。多条会用换行合并，run.py 会作为动态素材多条文案。
    """
    key_map = {
        "primary_text":"primary_text","body":"primary_text","text":"primary_text","正文":"primary_text","主文案":"primary_text",
        "headline":"headline","title":"headline","标题":"headline",
        "description":"description","desc":"description","描述":"description",
        "cta":"cta_type","cta_type":"cta_type","行动号召":"cta_type",
        "url_tags":"url_tags","utm":"url_tags","link":"website_url","链接":"website_url","website_url":"website_url"
    }
    res, cur = {}, None
    for raw in text.replace("\r\n","\n").replace("\r","\n").split("\n"):
        s = raw.strip()
        if not s:
            continue
        section = re.match(r"^(?:#{1,6}\s*)?[\[【]?\s*(正文|主文案|primary_text|body|text|标题|headline|title|描述|description|desc)\s*(?:\d{0,2})\s*[\]】]?$", s, re.I)
        if section:
            raw_label = section.group(1)
            label = raw_label.lower().replace(" ", "_")
            cur = key_map.get(label) or key_map.get(raw_label)
            if cur:
                res.setdefault(cur, "")
                continue
        m = re.match(r"^([A-Za-z_ ]+|[\u4e00-\u9fff]+)\s*(\d{0,2})\s*[:：]\s*(.*)$", s)
        if m:
            raw_label = m.group(1).strip()
            label = raw_label.lower().replace(" ", "_")
            label = re.sub(r"\d+$", "", label)
            key = key_map.get(label) or key_map.get(re.sub(r"\d+$", "", raw_label))
            if key:
                cur = key
                val = m.group(3).strip()
                if val:
                    res[key] = (res.get(key, "") + ("\n" if res.get(key) else "") + val).strip()
                else:
                    res.setdefault(key, "")
                continue
        if cur:
            res[cur] = (res.get(cur, "") + ("\n" if res.get(cur) else "") + s).strip()
    if not any(str(value or "").strip() for value in res.values()) and default_key:
        variants = _plain_text_variants(text)
        if variants:
            res = {default_key: "\n".join(variants)}
    return res

def find_asset_pack_folder(row):
    selling_point = str(row.get("selling_point") or row.get("asset_pack") or "").strip()
    if not selling_point:
        return None
    root = Path(str(row.get("asset_root") or DEFAULT_ASSET_ROOT).strip())
    if not root.is_absolute():
        root = BASE_DIR / root
    product = normalize_product(row.get("product"))
    asset_language = str(row.get("asset_language") or row.get("copy_language") or row.get("素材语种") or "").strip()
    base_candidates = [
        root / product / selling_point,
        root / str(row.get("product") or "").strip() / selling_point,
        root / selling_point,
        Path(selling_point),
    ]
    if asset_language and asset_language.upper() not in {"DEFAULT", "AUTO"}:
        for b in base_candidates:
            if not b.exists() or not b.is_dir():
                continue
            for child in b.iterdir():
                if child.is_dir() and child.name.casefold() == asset_language.casefold():
                    return child
        # 指定了预算表语种时必须严格命中对应目录，绝不回退到根目录或 EN。
        return root / product / selling_point / asset_language
    for c in base_candidates:
        if c.exists() and c.is_dir():
            return c
    return root / product / selling_point


def load_asset_pack_into_row(row):
    """如果表里填了 selling_point，则从 素材包/产品/卖点[/语种] 文件夹读取图片和文案。"""
    pack_folder = find_asset_pack_folder(row)
    if not pack_folder:
        return row
    if not pack_folder.exists():
        raise FileNotFoundError(f"找不到素材包文件夹：{pack_folder}")

    image_paths = []
    manual_image_file = str(row.get("image_file") or "").strip()
    manual_image_folder = str(row.get("image_folder") or "").strip()
    if not manual_image_file and not manual_image_folder:
        for folder_name in IMAGE_FOLDER_NAMES:
            img_dir = pack_folder / folder_name
            if img_dir.exists() and img_dir.is_dir():
                image_paths.extend(sorted([p for p in img_dir.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS]))
        if not image_paths:
            image_paths = sorted([p for p in pack_folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS])
        row["_asset_image_paths"] = image_paths
        row["image_folder"] = str(pack_folder)

    text_values = {}
    text_files = sorted([path for path in pack_folder.iterdir() if path.is_file() and path.suffix.lower() == ".txt"], key=lambda path: path.name.lower())
    priority_names = ["primary_text.txt", "body.txt", "正文.txt", "主文案.txt", "headline.txt", "title.txt", "标题.txt", "description.txt", "desc.txt", "描述.txt", *TEXT_FILE_NAMES]
    priority = {name.lower(): index for index, name in enumerate(priority_names)}
    text_files.sort(key=lambda path: (priority.get(path.name.lower(), 999), path.name.lower()))
    for path in text_files:
        parsed = parse_copy_text(read_text_file(path), default_key=_copy_key_from_filename(path))
        for key, value in parsed.items():
            if value:
                _merge_copy_value(text_values, key, value)
    for key, value in text_values.items():
        if value and not str(row.get(key) or "").strip():
            row[key] = value
    row["_asset_pack_folder"] = str(pack_folder)
    return row

def collect_image_paths(row):
    if row.get("_asset_image_paths"):
        return [Path(p) for p in row["_asset_image_paths"]]
    folder_text = str(row.get("image_folder") or "").strip()
    file_text = str(row.get("image_file") or "").strip()
    if not folder_text and not file_text:
        raise RuntimeError("没有找到素材包图片：请检查 素材包/产品/卖点/语种/图片/，不要再使用旧版 images 目录。")
    base = Path(folder_text.replace("\\", os.sep).replace("/", os.sep))

    paths = []
    if file_text:
        # 支持 image_file 一格里用换行/分号/逗号写多张图。
        for name in re.split(r"[,;\n]+", file_text):
            name = name.strip()
            if not name:
                continue
            p = Path(name.replace("\\", os.sep).replace("/", os.sep))
            if not p.is_absolute():
                p = base / p
            paths.append(p)
    else:
        if base.is_file():
            paths.append(base)
        elif base.is_dir():
            paths = sorted([p for p in base.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS])
        else:
            raise FileNotFoundError(f"找不到图片文件夹：{base}")

    if not paths:
        raise RuntimeError("没有找到图片：请检查 素材包/产品/卖点/语种/图片/ 是否有图片。")
    return paths


def get_or_upload_images(row, token, image_library, dry_run):
    ad_account_id = row["ad_account_id"]
    image_paths = collect_image_paths(row)
    hashes = []
    image_md5_list = []

    for image_path in image_paths:
        if not image_path.exists():
            raise FileNotFoundError(f"找不到图片：{image_path}")
        md5 = file_md5(image_path)
        image_md5_list.append(md5)
        file_name = image_path.name

        lib_item = image_library.get("by_md5", {}).get(md5)
        if lib_item and lib_item.get("hash"):
            print(f"Reuse image by MD5: {file_name} -> {lib_item['hash']}")
            hashes.append(lib_item["hash"])
            continue

        lib_item = image_library.get("by_filename", {}).get(file_name)
        if lib_item and lib_item.get("hash"):
            print(f"Reuse image by filename: {file_name} -> {lib_item['hash']}")
            hashes.append(lib_item["hash"])
            image_library.setdefault("by_md5", {})[md5] = lib_item
            continue

        if dry_run:
            print(f"DRY RUN: would upload image: {file_name}")
            hashes.append(f"DRYRUN_HASH_{len(hashes)+1}")
            continue

        print(f"Uploading image: {image_path}")
        image_hash = upload_image(ad_account_id, token, image_path)
        print(f"Uploaded: {file_name} -> {image_hash}")
        item = {
            "filename": file_name,
            "path": str(image_path),
            "md5": md5,
            "hash": image_hash,
            "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "ad_account_id": ad_account_id,
        }
        image_library.setdefault("by_md5", {})[md5] = item
        image_library.setdefault("by_filename", {})[file_name] = item
        hashes.append(image_hash)
    return hashes, image_md5_list


def search_locale_id(language_name, token):
    result = request_meta("GET", "search", token, params={"type": "adlocale", "q": language_name, "limit": 20})
    data = result.get("data", [])
    if not data:
        print(f"No locale found for {language_name}, skipped.")
        return None
    if language_name.lower() == "english":
        for item in data:
            if item.get("name", "").lower() == "english (uk)":
                return item["key"]
    for item in data:
        if item.get("name", "").lower() == language_name.lower():
            return item["key"]
    for item in data:
        if language_name.lower() in item.get("name", "").lower():
            return item["key"]
    return data[0]["key"]


def get_locale_ids_for_countries(countries, token, language_mode, custom_language="", preferred_language=""):
    mode = (language_mode or "auto").strip().lower()
    if mode == "all":
        return None
    if mode == "custom":
        language_names = parse_custom_language_names(custom_language)
        if not language_names:
            raise RuntimeError("language_mode=custom 但 custom_language 为空。请填写 English / French，或 EN,FR 这样的语言。")
    else:
        # auto 模式顺序固定为：预算表素材语种 -> 地区本地语种 -> English。
        language_names = []
        preferred = parse_custom_language_names(preferred_language)
        for lang in preferred + official_language_names_for_countries(countries) + ["English"]:
            if lang not in language_names:
                language_names.append(lang)
    locale_ids = []
    for lang in language_names:
        locale_id = search_locale_id(lang, token)
        if locale_id is not None:
            locale_ids.append(locale_id)
    locale_ids = list(dict.fromkeys(locale_ids))
    return locale_ids or None


def load_audience_config():
    return load_json(Path(AUDIENCE_FILE), {"products": {}})


def get_config_saved_audience_id(audience_config, product=None):
    """按产品读取 Saved Audience ID。

    - audiences.json 的 saved_audience_by_product 优先。
    - 如果产品在映射里明确配置为空字符串，表示不用 Saved Audience，改用产品关键词。
    - 其次读取 product_settings.json 里的产品线 saved_audience_id。
    - 最后才回退 default_saved_audience_id。
    """
    use_saved = audience_config.get("use_saved_audience", True)
    if str(use_saved).strip().lower() in {"0", "false", "no", "off", "不用"}:
        return ""

    product_key = normalize_product(product) if product else ""
    by_product = audience_config.get("saved_audience_by_product", {}) or {}
    if product_key in by_product:
        return str(by_product.get(product_key) or "").strip()

    settings = load_product_settings()
    line = product_line_for_product(product_key)
    line_info = (settings.get("product_lines") or {}).get(line, {})
    if "saved_audience_id" in line_info:
        return str(line_info.get("saved_audience_id") or "").strip()

    return str(
        audience_config.get("default_saved_audience_id")
        or audience_config.get("saved_audience_id")
        or ""
    ).strip()


def get_saved_audience_targeting(saved_audience_id, token, saved_audience_cache):
    """读取 Meta Saved Audience 的 targeting，并缓存到 _cache。
    这个 saved audience 可以是 ALL 地区；创建广告组时会用表里的 countries 覆盖地区。
    """
    saved_audience_id = str(saved_audience_id or "").strip()
    if not saved_audience_id:
        return None
    cached = saved_audience_cache.get(saved_audience_id)
    if cached and cached.get("targeting"):
        print(f"Audience: reuse saved audience cache {saved_audience_id} -> {cached.get('name', '')}")
        return copy.deepcopy(cached["targeting"])
    print(f"Audience: reading saved audience {saved_audience_id} from Meta...")
    result = request_meta("GET", saved_audience_id, token, params={"fields": "id,name,targeting"})
    targeting = result.get("targeting") or {}
    if not targeting:
        print(f"Audience: saved audience {saved_audience_id} 没有返回 targeting，将回退到产品关键词/基础受众。")
        return None
    saved_audience_cache[saved_audience_id] = {
        "id": result.get("id") or saved_audience_id,
        "name": result.get("name") or "",
        "targeting": targeting,
        "cached_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    print(f"Audience: loaded saved audience {saved_audience_id} -> {result.get('name', '')}")
    return copy.deepcopy(targeting)


def build_targeting(row, token, audience_config, targeting_cache, saved_audience_cache):
    countries = parse_country_codes(row["countries"])
    locale_ids = get_locale_ids_for_countries(
        countries,
        token,
        row.get("language_mode") or "auto",
        row.get("custom_language") or "",
        row.get("asset_language") or "",
    )

    saved_audience_id = get_config_saved_audience_id(audience_config, row.get("product"))
    saved_targeting = get_saved_audience_targeting(saved_audience_id, token, saved_audience_cache)
    if saved_targeting:
        targeting = saved_targeting
        print(f"Audience: {normalize_product(row.get('product'))} -> using Saved Audience ID {saved_audience_id} as base; countries will be overwritten by plan.xlsx.")
    else:
        targeting = {}

    # 从 Saved Audience 复制出来的 targeting 里有些展示字段不能直接复用。
    # 注意：Meta v23+ 要求 targeting_automation 放在 targeting 里面，不是 adset payload 顶层。
    targeting.pop("targeting_automation", None)
    targeting.pop("age_range", None)

    # 地区必须以表格为准；Saved Audience 里即使是 ALL，也会在这里覆盖掉。
    old_geo = targeting.get("geo_locations") or {}
    new_geo = {"countries": countries}
    if old_geo.get("location_types"):
        new_geo["location_types"] = old_geo.get("location_types")
    targeting["geo_locations"] = new_geo

    # 年龄：表格有就用表格；没有则保留 saved audience；还没有就默认 18-65。
    targeting["age_min"] = int(float(row.get("age_min") or targeting.get("age_min") or 18))
    targeting["age_max"] = int(float(row.get("age_max") or targeting.get("age_max") or 65))

    # 语言：auto 按所选国家官方/本地语种匹配；all 则不限制语言，并清掉 saved audience 旧语言。
    if locale_ids:
        targeting["locales"] = locale_ids
    else:
        targeting.pop("locales", None)

    # 默认使用 Saved Audience，不再强制叠加产品关键词，避免受众过窄。
    # 如果没有 Saved Audience，才使用 audiences.json 里的产品关键词。
    use_keywords_with_saved = str(audience_config.get("use_product_keywords_with_saved_audience", False)).strip().lower() in {"1", "true", "yes", "y", "on"}
    if (not saved_targeting) or use_keywords_with_saved:
        interests = get_interest_targets_for_product(row.get("product"), token, audience_config, targeting_cache)
        if interests:
            targeting["flexible_spec"] = [{"interests": interests}]

    return targeting


def get_product_keywords(product, audience_config):
    product_key = normalize_product(product)
    item = audience_config.get("products", {}).get(product_key)
    if not item:
        return product_key, []
    return product_key, [str(k).strip() for k in item.get("keywords", []) if str(k).strip()]


def search_interest(keyword, token, targeting_cache):
    cache_key = keyword.strip().lower()
    cached = targeting_cache.get(cache_key)
    if cached:
        return cached
    result = request_meta("GET", "search", token, params={"type": "adinterest", "q": keyword, "limit": 10})
    data = result.get("data", [])
    chosen = None
    for item in data:
        if item.get("id") and item.get("name", "").lower() == keyword.lower():
            chosen = {"id": item["id"], "name": item.get("name", keyword)}
            break
    if chosen is None:
        for item in data:
            if item.get("id"):
                chosen = {"id": item["id"], "name": item.get("name", keyword)}
                break
    targeting_cache[cache_key] = chosen or {}
    return chosen


def get_interest_targets_for_product(product, token, audience_config, targeting_cache):
    product_key, keywords = get_product_keywords(product, audience_config)
    if not keywords:
        print(f"Audience: {product_key} 没有配置关键词，使用国家+语言基础受众。")
        return []
    interests = []
    seen = set()
    print(f"Audience: {product_key} -> searching interests: {', '.join(keywords)}")
    for keyword in keywords:
        try:
            item = search_interest(keyword, token, targeting_cache)
            if not item:
                print(f"Audience keyword skipped: {keyword} not found")
                continue
            if item["id"] in seen:
                continue
            seen.add(item["id"])
            interests.append({"id": item["id"], "name": item["name"]})
            print(f"Audience interest matched: {keyword} -> {item['name']} ({item['id']})")
        except Exception as e:
            print(f"Audience keyword skipped: {keyword} -> {e}")
            continue
    if not interests:
        print(f"Audience: {product_key} 没有匹配到可用兴趣，使用国家+语言基础受众。")
    return interests[:10]


def adset_exists(campaign_id, token, adset_name):
    endpoint = f"{campaign_id}/adsets"
    params = {"fields": "id,name,status", "limit": 100}
    while True:
        result = request_meta("GET", endpoint, token, params=params)
        for item in result.get("data", []):
            if item.get("name") == adset_name:
                return item
        next_url = result.get("paging", {}).get("next")
        if not next_url:
            return None
        response = requests.get(next_url, timeout=120)
        data = response.json()
        if "error" in data:
            return None
        result = data
        params = {}


def create_dynamic_creative(row, token, image_hashes, image_md5_list, creative_library, dry_run):
    product = row.get("product") or "PRODUCT"
    page_id = row["page_id"]
    lead_form_id = row["lead_form_id"]
    link = row["website_url"]
    cta_type = str(row.get("cta_type") or "LEARN_MORE").strip().upper()
    bodies = split_multi(row.get("primary_text"))
    titles = split_multi(row.get("headline"))
    descriptions = split_multi(row.get("description"))
    if not bodies:
        raise RuntimeError("primary_text 为空。")
    if not titles:
        raise RuntimeError("headline 为空。")
    if not descriptions:
        descriptions = [" "]

    creative_key = content_hash({
        "cache_schema": "language-strict-v2",
        "product": product,
        "asset_language": str(row.get("asset_language") or "").strip().upper(),
        "asset_pack_folder": str(row.get("_asset_pack_folder") or ""),
        "ad_account_id": row["ad_account_id"],
        "page_id": page_id,
        "lead_form_id": lead_form_id,
        "link": link,
        "cta_type": cta_type,
        "url_tags": row.get("url_tags") or "",
        "image_md5": image_md5_list,
        "bodies": bodies,
        "titles": titles,
        "descriptions": descriptions,
    })
    existing = creative_library.get(creative_key)
    if existing and existing.get("creative_id"):
        print(f"Reuse Creative: {product} -> {existing['creative_id']}")
        return existing["creative_id"], creative_key
    if dry_run:
        print(f"DRY RUN: would create creative for: {product}")
        return f"DRYRUN_CREATIVE_{product}", creative_key

    asset_feed_spec = {
        "images": [{"hash": h} for h in image_hashes],
        "bodies": [{"text": t} for t in bodies],
        "titles": [{"text": t} for t in titles],
        "descriptions": [{"text": t} for t in descriptions],
        "link_urls": [{"website_url": link, "display_url": link.replace("https://", "").replace("http://", "").strip("/")}],
        "call_to_action_types": [cta_type],
        "call_to_actions": [{"type": cta_type, "value": {"link": link, "lead_gen_form_id": lead_form_id}}],
        "ad_formats": ["SINGLE_IMAGE"],
    }
    # 注意：不要把 website_url 写入 deeplink_url。
    # deeplink_url 会被 Meta 当作 app_link 处理，Lead Form 广告没有应用商店链接时会报错：
    # error_subcode 1885949 / 提供的应用链接没有应用商店网址。

    payload = {
        "name": f"{product}_creative_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
        "object_story_spec": {"page_id": page_id},
        "asset_feed_spec": asset_feed_spec,
        "url_tags": row.get("url_tags") or "",
    }
    result = request_meta("POST", f"{row['ad_account_id']}/adcreatives", token, payload=payload)
    creative_id = result["id"]
    creative_library[creative_key] = {
        "product": product,
        "asset_language": str(row.get("asset_language") or "").strip().upper(),
        "creative_id": creative_id,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    return creative_id, creative_key


def get_advantage_audience_flag(row):
    """Meta v23+ 要求创建广告组时必须明确启用或禁用 Advantage Audience。
    优先使用表格中的 advantage_audience；如果没填，则读取 product_settings.json 中产品线的默认值。
    Scanner 受众 Trion-scanner 默认启用 1；MC/RTK 默认 0。
    """
    value = row.get("advantage_audience")
    if value not in (None, ""):
        text = str(value).strip().lower()
        if text in {"1", "yes", "y", "true", "on", "enable", "enabled", "开", "开启"}:
            return 1
        return 0
    line_info = product_line_info_for_product(row.get("product"))
    line_value = line_info.get("advantage_audience")
    if line_value in (None, ""):
        return 0
    text = str(line_value).strip().lower()
    return 1 if text in {"1", "yes", "y", "true", "on", "enable", "enabled", "开", "开启"} else 0


def create_adset(row, token, adset_name, audience_config, targeting_cache, saved_audience_cache):
    budget = int(round(float(row["daily_budget"]) * 100))  # Meta API daily_budget uses cents; plan/frontend use USD dollars.
    targeting = build_targeting(row, token, audience_config, targeting_cache, saved_audience_cache)

    # Meta v23+ 要求这个字段明确放在 targeting 参数里：
    # targeting.targeting_automation.advantage_audience = 0 或 1
    targeting["targeting_automation"] = {"advantage_audience": get_advantage_audience_flag(row)}

    payload = {
        "name": adset_name,
        "campaign_id": row["campaign_id"],
        "daily_budget": budget,
        "billing_event": "IMPRESSIONS",
        "optimization_goal": "LEAD_GENERATION",
        "bid_strategy": "LOWEST_COST_WITHOUT_CAP",
        "destination_type": "ON_AD",
        "is_dynamic_creative": True,
        "status": str(row.get("adset_status") or "ACTIVE").strip().upper(),
        "targeting": targeting,
        "promoted_object": {"page_id": row["page_id"]},
    }
    result = request_meta("POST", f"{row['ad_account_id']}/adsets", token, payload=payload)
    return result["id"]


def create_ad(row, token, adset_id, creative_id, ad_name):
    payload = {
        "name": ad_name,
        "adset_id": adset_id,
        "creative": {"creative_id": creative_id},
        "status": str(row.get("ad_status") or "ACTIVE").strip().upper(),
    }
    result = request_meta("POST", f"{row['ad_account_id']}/ads", token, payload=payload)
    return result["id"]


def read_product_campaigns(wb):
    mapping = {}
    if "产品Campaign对照" not in wb.sheetnames:
        return mapping
    ws = wb["产品Campaign对照"]
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    if "product" not in headers or "campaign_id" not in headers:
        return mapping
    p_col = headers.index("product") + 1
    c_col = headers.index("campaign_id") + 1
    for r in range(2, ws.max_row + 1):
        product = normalize_product(ws.cell(r, p_col).value)
        campaign_id = clean_id(ws.cell(r, c_col).value)
        if product and campaign_id:
            mapping[product] = campaign_id
    return mapping


def build_rows(ws, campaign_map):
    headers = [normalize_header_name(cell.value) for cell in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        if all(ws.cell(r, c).value in (None, "") for c in range(1, len(headers) + 1)):
            continue
        item = {"_row": r}
        for c, h in enumerate(headers, start=1):
            if h:
                item[h] = ws.cell(r, c).value
        item["product"] = normalize_product(item.get("product"))
        item["campaign_id"] = clean_id(item.get("campaign_id")) or campaign_map.get(item["product"], "")
        for key in ["page_id", "lead_form_id"]:
            item[key] = clean_id(item.get(key))
        if not item.get("page_id"):
            item["page_id"] = default_page_id_for_product(item.get("product"))
        account = str(item.get("ad_account_id") or "").strip()
        if account and not account.startswith("act_"):
            account = "act_" + clean_id(account)
        item["ad_account_id"] = account
        fill_product_url(item)
        rows.append(item)
    return rows, headers


def write_cell(ws, row, headers, header_name, value):
    if header_name not in headers:
        return
    ws.cell(row, headers.index(header_name) + 1).value = value


def create_log_file():
    log_dir = Path(DEFAULT_LOG_DIR)
    log_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "日志"
    ws.append(["广告组名字", "广告状态", "广告预算", "广告建立状态"])
    path = log_dir / f"Meta广告建立日志_{now_stamp()}.xlsx"
    return wb, ws, path


def explain_status(status, error_text=""):
    if status in {"success", "dry_run"}:
        return status
    if error_text:
        text = str(error_text)
        for code, hint in ERROR_HINTS.items():
            if code in text:
                return f"failed - {hint}"
        return f"failed - {text[:120]}"
    return status


def maybe_delay(_max_seconds=None):
    # 投递节奏固定为 1–60 秒随机；即使旧计划表仍写着 300，也不会超过 60 秒。
    wait = random.randint(1, 60)
    print(f"Waiting {wait} seconds before next item...")
    time.sleep(wait)


def main():
    plan_path = Path(PLAN_FILE)
    if not plan_path.exists():
        raise FileNotFoundError(f"找不到 {PLAN_FILE}，请确认 run.py 和 plan.xlsx 在同一个文件夹。")
    print("正在读取 token.txt / 环境变量...", flush=True)
    token = get_token()
    print("Token 已读取，开始加载缓存和计划表...", flush=True)
    cache_dir = Path(CACHE_DIR)
    cache_dir.mkdir(exist_ok=True)
    image_library_path = cache_dir / IMAGE_LIBRARY_FILE
    creative_library_path = cache_dir / CREATIVE_LIBRARY_FILE
    targeting_cache_path = cache_dir / TARGETING_CACHE_FILE
    saved_audience_cache_path = cache_dir / SAVED_AUDIENCE_CACHE_FILE
    image_library = load_json(image_library_path, {"by_md5": {}, "by_filename": {}})
    creative_library = load_json(creative_library_path, {})
    targeting_cache = load_json(targeting_cache_path, {})
    saved_audience_cache = load_json(saved_audience_cache_path, {})
    audience_config = load_audience_config()

    wb = load_workbook(plan_path)
    sheet_name = "广告主表" if "广告主表" in wb.sheetnames else "投放计划"
    ws = wb[sheet_name]
    campaign_map = read_product_campaigns(wb)
    rows, headers = build_rows(ws, campaign_map)
    log_wb, log_ws, log_path = create_log_file()

    ready_rows = []
    for row in rows:
        flag = str(row.get("enabled") or "").strip().upper()
        if flag in {"Y", "YES", "READY", "DRY_RUN", "DRYRUN", "1", "TRUE"}:
            ready_rows.append(row)
    if not ready_rows:
        print("没有 enabled=Y / READY 的行需要处理。")
        return
    print("__ADS_RUN_START__" + json.dumps({"planned": len(ready_rows)}, ensure_ascii=False), flush=True)

    had_failures = False

    for idx, row in enumerate(ready_rows, start=1):
        stop_after_error = False
        excel_row = row["_row"]
        product = str(row.get("product") or "").strip()
        countries = parse_country_codes(row.get("countries"))
        blocked_countries = [country for country in countries if country in BLOCKED_COUNTRIES]
        adset_name = build_ad_name(row, countries)
        ad_name = adset_name
        budget_usd = float(row.get("daily_budget") or 0)
        budget = f"${budget_usd:,.2f}"
        dry_run = truthy(row.get("dry_run")) or str(row.get("enabled") or "").strip().upper() in {"DRY_RUN", "DRYRUN"}

        print(f"\n=== Processing row {excel_row}: {adset_name} ===")
        print(f"Run mode: {'DRY RUN / 只预检，不会创建广告' if dry_run else 'LIVE / 正式创建广告'}")
        try:
            if blocked_countries:
                raise RuntimeError(f"本季度禁投地区：{','.join(blocked_countries)}。请从计划中移除台湾和新加坡。")
            load_asset_pack_into_row(row)
            fill_product_url(row)
            resolve_lead_form_id(row, countries)
            required = [
                "product", "ad_account_id", "campaign_id", "daily_budget", "page_id",
                "lead_form_id", "countries", "image_folder", "primary_text", "headline", "website_url",
            ]
            for key in required:
                if row.get(key) in (None, ""):
                    raise RuntimeError(f"缺少必填字段：{key}")

            image_hashes, image_md5_list = get_or_upload_images(row, token, image_library, dry_run)
            creative_id, _ = create_dynamic_creative(row, token, image_hashes, image_md5_list, creative_library, dry_run)
            # plan.xlsx 不再自动写回，避免把 enabled 改成 DONE/FAILED。

            if dry_run:
                _ = build_targeting(row, token, audience_config, targeting_cache, saved_audience_cache)
            # plan.xlsx 不再自动写回，避免把 enabled 改成 DONE/FAILED。
                log_ws.append([adset_name, str(row.get("adset_status") or "ACTIVE").upper(), budget, "dry_run"])
                print(f"DRY RUN OK: {adset_name}")
                print("__ADS_RUN_ITEM__" + json.dumps({"status": "dry_run", "excel_row": excel_row, "adset_name": adset_name}, ensure_ascii=False), flush=True)
                continue

            if truthy(row.get("duplicate_check")):
                duplicate = adset_exists(row["campaign_id"], token, adset_name)
                if duplicate:
                    raise RuntimeError(f"发现同名广告组已存在：{adset_name} / {duplicate.get('id')}")

            adset_id = create_adset(row, token, adset_name, audience_config, targeting_cache, saved_audience_cache)
            ad_id = create_ad(row, token, adset_id, creative_id, ad_name)
            # plan.xlsx 不再自动写回，避免把 enabled 改成 DONE/FAILED。
            log_ws.append([adset_name, str(row.get("adset_status") or "ACTIVE").upper(), budget, "success"])
            print(f"SUCCESS: {adset_name} -> Ad Set {adset_id}, Ad {ad_id}")
            print("__ADS_RUN_ITEM__" + json.dumps({"status": "success", "excel_row": excel_row, "adset_name": adset_name}, ensure_ascii=False), flush=True)

        except Exception as e:
            had_failures = True
            stop_after_error = True
            err = str(e)
            skipped_count = max(0, len(ready_rows) - idx)
            # plan.xlsx 不再自动写回，避免把 enabled 改成 DONE/FAILED。
            log_ws.append([adset_name, str(row.get("adset_status") or "ACTIVE").upper(), budget, explain_status("failed", err)])
            print(f"FAILED: {adset_name}: {err}")
            print("__ADS_RUN_ITEM__" + json.dumps({"status": "failed", "excel_row": excel_row, "adset_name": adset_name}, ensure_ascii=False), flush=True)
            print("__ADS_RUN_ERROR__" + json.dumps({
                "excel_row": excel_row, "product": product, "adset_name": adset_name,
                "message": err, "skipped_count": skipped_count,
            }, ensure_ascii=False), flush=True)

        finally:
            save_json(image_library_path, image_library)
            save_json(creative_library_path, creative_library)
            save_json(targeting_cache_path, targeting_cache)
            save_json(saved_audience_cache_path, saved_audience_cache)
            # plan.xlsx 不再保存，保持原表不变。
            safe_save_workbook(log_wb, log_path, cache_dir, "日志")
            if idx < len(ready_rows) and not stop_after_error:
                maybe_delay(row.get("delay_max_seconds"))
        if stop_after_error:
            print(f"计划已暂停：第 {excel_row} 行失败，后续 {max(0, len(ready_rows) - idx)} 个广告组未执行。", flush=True)
            break

    # plan.xlsx 不再保存，保持原表不变。
    final_log_saved = safe_save_workbook(log_wb, log_path, cache_dir, "日志")
    print(f"Log saved: {final_log_saved}")
    if had_failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
